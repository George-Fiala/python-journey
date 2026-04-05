# =========================
# PO + GR One Flow (v1.9.7)
# - Create Purchase Order (PO) PDF  *** STYLED PDF — VARIANT D: CLEAN MINIMAL ***
# - Write to "Parts Ordered" monthly sheet
# - Add to "To Receive" (Goods Receipt backlog)
# - Mark as Received (updates both sheets)
# - Auto-open Outlook email with PDF attachment
# - Full Sync with main shared workbook
#
# Author: George Fiala
# GitHub: https://github.com/George-Fiala/python-journey
# License: MIT
#
# SETUP: Update the CONFIG section below with your own paths,
# company details, machine names, areas, and cost centre codes.
# =========================

import os
import sys
from pathlib import Path


def _reexec_into_local_venv() -> None:
    try:
        here = Path(__file__).resolve().parent
    except Exception:
        return
    venv_python = here / ".venv" / "Scripts" / "python.exe"
    if not venv_python.is_file():
        return
    try:
        current = Path(sys.executable).resolve()
        target = venv_python.resolve()
    except Exception:
        return
    if current != target:
        print(f"[boot] Re-launching with venv Python: {target}")
        os.execv(str(target), [str(target), str(__file__), *sys.argv[1:]])


if os.name == "nt":
    _reexec_into_local_venv()

from datetime import datetime, date
import json
import re
from difflib import get_close_matches

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPORTLAB_AVAILABLE = False
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.utils import ImageReader
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

OUTLOOK_AVAILABLE = False
try:
    import win32com.client
    OUTLOOK_AVAILABLE = True
except Exception:
    OUTLOOK_AVAILABLE = False


# ========= CONFIG — UPDATE THESE FOR YOUR SETUP =========
WORKBOOK_PATH       = r"C:\Users\YOUR_USERNAME\Desktop\Parts Ordered Local.xlsx"
SUPPLIERS_JSON_PATH = r"C:\Users\YOUR_USERNAME\Desktop\suppliers.json"
MAIN_WORKBOOK_PATH  = r"C:\path\to\shared\Parts Ordered.xlsx"
SYNC_FROM_YEAR      = 2026

INITIALS         = "XX"   # Your initials — used in PO number generation
SHEET_COST_CENTRE = "Cost centre"
TO_RECEIVE_SHEET  = "To Receive"
PO_PDF_OUTPUT_DIR = r"C:\Users\YOUR_USERNAME\Desktop\PO PDFs"
LOGO_PATH         = r"C:\Users\YOUR_USERNAME\Desktop\company_logo.png"

DEBUG = True

VAT_RATE_DEFAULT      = 0.20
PROMPT_VAT_RATE_EACH_PO = True

# ========= COMPANY DETAILS — UPDATE THESE =========
SHEARD_COMPANY_NAME = "Your Company Name"

SHEARD_BILL_TO = [
    "Your Company Ltd",
    "123 Example Street",
    "Your City",
    "AB1 2CD",
    "United Kingdom",
    "Company Reg: 00000000",
]

SHEARD_DELIVER_TO = [
    "Your Company Name",
    "123 Example Street",
    "Your City",
    "AB1 2CD",
    "United Kingdom",
]

CONTACT_NAME  = "Your Name"
CONTACT_EMAIL = "yourname@yourcompany.co.uk"
CONTACT_PHONE = "07700000000"

PDF_HEADER_BG = HexColor('#2C3E50') if REPORTLAB_AVAILABLE else None
PDF_ACCENT    = HexColor('#2980B9') if REPORTLAB_AVAILABLE else None
PDF_LIGHT_BG  = HexColor('#ECF0F1') if REPORTLAB_AVAILABLE else None
PDF_DARK_TEXT = HexColor('#2C3E50') if REPORTLAB_AVAILABLE else None

AUTO_OPEN_OUTLOOK = True

EMAIL_SUBJECT_TEMPLATE = "Purchase Order {po_number} — {company}"

EMAIL_BODY_TEMPLATE = """\
Dear {supplier_first_name},

Please find attached Purchase Order {po_number}.

Order Summary:
  Items: {items_summary}
  Net Total: \u00a3{net_total:.2f}
  Due Date: {due_date}

If you have any questions, please don't hesitate to contact me.

Best regards
"""


# ========= HELPERS =========
BACK = "__BACK__"


def is_back(s):
    return (s or "").strip().lower() == "back"


def input_text(prompt, allow_blank=False, allow_back=True):
    while True:
        raw = input(prompt).strip()
        if allow_back and is_back(raw):
            return BACK
        if raw == "" and allow_blank:
            return ""
        if raw == "" and not allow_blank:
            print("This field cannot be blank. Type a value or 'back'.")
            continue
        return raw


def input_positive_int(prompt, allow_blank=False, allow_back=True):
    while True:
        raw = input(prompt).strip()
        if allow_back and is_back(raw):
            return BACK
        if raw == "" and allow_blank:
            return ""
        if not raw.isdigit() or int(raw) <= 0:
            print("Enter a whole number (e.g., 1). Leave blank to finish.")
            continue
        return raw


def input_money(prompt, allow_blank=False, allow_back=True):
    while True:
        raw = input(prompt).strip()
        if allow_back and is_back(raw):
            return BACK
        if raw == "" and allow_blank:
            return ""
        if parse_currency_to_float(raw) is None:
            print("Enter a valid amount (e.g., 12.50).")
            continue
        return raw


def input_with_autocomplete(prompt, known_values, allow_blank=False, allow_back=True, min_chars=1):
    while True:
        raw = input(prompt).strip()
        if allow_back and is_back(raw):
            return BACK
        if raw == "" and allow_blank:
            return ""
        if raw == "" and not allow_blank:
            print("This field cannot be blank.")
            continue

        raw_lower = raw.lower()
        for val in known_values:
            if val.lower() == raw_lower:
                return val

        if len(raw) < min_chars:
            return raw

        prefix_matches = [v for v in known_values if v.lower().startswith(raw_lower)]
        substring_matches = [v for v in known_values if raw_lower in v.lower() and v not in prefix_matches]
        all_matches = prefix_matches + substring_matches

        if all_matches:
            print(f"\n  Matches for '{raw}':")
            for i, val in enumerate(all_matches[:9], 1):
                marker = "\u2192" if val in prefix_matches else "~"
                print(f"    {i}) {marker} {val}")
            print(f"    0) Use '{raw}' as typed")
            pick = input("  Pick 1-9 (or 0 to keep, Enter for 1): ").strip()
            if is_back(pick):
                return BACK
            if pick == "" or pick == "1":
                return all_matches[0]
            if pick == "0":
                return raw
            if pick.isdigit() and 1 <= int(pick) <= min(9, len(all_matches)):
                return all_matches[int(pick) - 1]
            return all_matches[0]

        fuzzy = get_close_matches(raw, known_values, n=5, cutoff=0.55)
        if fuzzy:
            print(f"\n  No exact match for '{raw}'. Did you mean:")
            for i, val in enumerate(fuzzy, 1):
                print(f"    {i}) {val}")
            print(f"    0) Use '{raw}' as typed")
            pick = input("  Pick 1-5 (or 0 to keep, Enter for 1): ").strip()
            if is_back(pick):
                return BACK
            if pick == "" or pick == "1":
                return fuzzy[0]
            if pick == "0":
                return raw
            if pick.isdigit() and 1 <= int(pick) <= min(5, len(fuzzy)):
                return fuzzy[int(pick) - 1]
            return fuzzy[0]

        return raw


def normalize_text(x):
    if x is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(x).strip().lower())


def parse_currency_to_float(text):
    t = (text or "").strip().replace("\u00a3", "").replace(",", "").strip()
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def parse_ddmmyyyy_to_date(text):
    t = (text or "").strip()
    if not t:
        return None
    try:
        return datetime.strptime(t, "%d/%m/%Y").date()
    except ValueError:
        return None


def ensure_dir(path):
    if path and not os.path.isdir(path):
        os.makedirs(path, exist_ok=True)


def load_or_create_workbook(path):
    while True:
        try:
            return load_workbook(path)
        except FileNotFoundError:
            return Workbook()
        except PermissionError:
            print(f"\nFile locked: {path}")
            input("Close it and press Enter to retry...")
        except Exception as e:
            print(f"\nError: {e}")
            input("Press Enter to retry...")


def safe_save_workbook(wb, path):
    while True:
        try:
            wb.save(path)
            return
        except PermissionError:
            print(f"\nFile locked: {path}")
            input("Close it and press Enter to retry...")
        except Exception as e:
            print(f"\nError saving: {e}")
            input("Press Enter to retry...")


def load_main_workbook_readonly(path):
    while True:
        try:
            return load_workbook(path, read_only=False, data_only=True)
        except PermissionError:
            print(f"\nMain workbook locked: {path}")
            input("Close it and press Enter to retry...")
        except Exception as e:
            print(f"\nError: {e}")
            input("Press Enter to retry...")


def get_month_sheet_name():
    return datetime.today().strftime("%B %Y")


def get_or_create_month_sheet(wb):
    name = get_month_sheet_name()
    if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet":
        ws = wb["Sheet"]
        ws.title = name
        return ws
    if name not in wb.sheetnames:
        return wb.create_sheet(name)
    return wb[name]


def to_float_cell(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip()
        if not s or s.startswith("="):
            return 0.0
        v = parse_currency_to_float(s)
        return float(v) if v is not None else 0.0
    return 0.0


def is_blank_cell(value):
    if value is None:
        return True
    if isinstance(value, str):
        return value.replace("\u00a0", " ").strip() == ""
    return False


def ensure_month_summary_cells(ws):
    for col, label in [("H", "Total spend \u00a3"), ("L", "Total Savings \u00a3")]:
        val_col = chr(ord(col) + 1)
        if ws[f"{col}1"].value is None:
            ws[f"{col}1"].value = label
            ws[f"{col}1"].font = Font(bold=True)
            ws[f"{col}1"].alignment = Alignment(horizontal="right")
        if ws[f"{val_col}1"].value is None:
            ws[f"{val_col}1"].value = 0.0
        ws[f"{val_col}1"].number_format = "\u00a3#,##0.00"
        ws[f"{val_col}1"].font = Font(bold=True)


def update_month_totals(ws_parts, header_row, colmap):
    ensure_month_summary_cells(ws_parts)
    data_start = header_row + 1
    total_cost = 0.0
    total_savings = 0.0
    cost_col = colmap.get("Cost")
    savings_col = colmap.get("Savings")
    po_col = colmap.get("PONumber")

    for r in range(data_start, ws_parts.max_row + 1):
        if po_col and is_blank_cell(ws_parts[f"{po_col}{r}"].value):
            continue
        if cost_col:
            total_cost += to_float_cell(ws_parts[f"{cost_col}{r}"].value)
        if savings_col:
            total_savings += to_float_cell(ws_parts[f"{savings_col}{r}"].value)

    ws_parts["I1"].value = float(total_cost)
    ws_parts["M1"].value = float(total_savings)


def auto_fit_columns(ws, min_width=8, max_width=70, scan_rows_limit=300):
    max_row = min(ws.max_row, scan_rows_limit)
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        best = max(
            len(str(ws.cell(row=r, column=c).value or ""))
            for r in range(1, max_row + 1)
        )
        ws.column_dimensions[letter].width = max(min_width, min(best + 2, max_width))


def set_sheet_view(ws, zoom=90):
    try:
        ws.sheet_view.zoomScale = int(zoom)
    except Exception:
        pass


# ========= HEADER DETECTION =========
HEADER_SYNONYMS = {
    "Machine":         {"machine"},
    "Area":            {"area", "areaslot", "slot"},
    "Reason":          {"reason"},
    "MechElecOther":   {"mechelecother", "mechelec", "mech", "elec", "other"},
    "Supplier":        {"supplier"},
    "PartsOrdered":    {"partsordered", "partsorder", "parts"},
    "JobNo":           {"jobno", "job", "jobnumber"},
    "PONumber":        {"ponumber", "purchaseorder", "po", "orderno"},
    "Cost":            {"cost", "costgbp"},
    "DateOrdered":     {"dateordered", "ordereddate", "ordered"},
    "DueDate":         {"duedate", "due"},
    "DeliveredDate":   {"delivereddate", "delivered"},
    "Savings":         {"costinitiatives", "initiatives", "savings"},
    "OriginalSupplier":{"originalsupplier", "origsupplier"},
    "CostCentre":      {"costcentre", "costcenter", "cc"},
}

REQUIRED_PARTS_FIELDS = [
    "Machine", "Area", "Reason", "MechElecOther", "Supplier",
    "PartsOrdered", "JobNo", "PONumber", "Cost", "DateOrdered",
    "DueDate", "Savings", "OriginalSupplier", "CostCentre", "DeliveredDate"
]


def detect_header_row_and_columns(ws, max_scan_rows=12):
    best_row, best_score, best_map = None, -1, {}
    max_cols = max(ws.max_column or 1, 30)

    for r in range(1, max_scan_rows + 1):
        row_values = [(c, normalize_text(ws.cell(row=r, column=c).value)) for c in range(1, max_cols + 1)]
        row_raw = [(c, str(ws.cell(row=r, column=c).value or "").strip()) for c in range(1, max_cols + 1)]
        field_to_col, score = {}, 0

        for field, synonyms in HEADER_SYNONYMS.items():
            for c, norm in row_values:
                if norm and norm in synonyms:
                    field_to_col[field] = get_column_letter(c)
                    score += 1
                    break

        if "Cost" not in field_to_col:
            for c, raw in row_raw:
                if raw in ("\u00a3", "\u00a3 "):
                    field_to_col["Cost"] = get_column_letter(c)
                    score += 1
                    break

        if score > best_score:
            best_score, best_row, best_map = score, r, field_to_col

    return (best_row, best_map) if best_score >= 6 else (None, {})


def init_parts_sheet_default(ws):
    headers = [
        ("A","Machine"),("B","Area"),("C","Reason"),("D","Mech/Elec/Other"),
        ("E","Supplier"),("F","Parts ordered"),("G","Job No"),("H","PO number"),
        ("I","Cost \u00a3"),("J","Date ordered"),("K","Due date"),("L","Delivered date"),
        ("M","Cost Initiatives"),("N","Original supplier"),("O","Cost Centre"),
    ]
    for col, label in headers:
        cell = ws[f"{col}2"]
        cell.value = label
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A3"


def ensure_parts_column_map(ws):
    header_row, colmap = detect_header_row_and_columns(ws)
    if not colmap:
        init_parts_sheet_default(ws)
        header_row = 2
        colmap = {
            "Machine":"A","Area":"B","Reason":"C","MechElecOther":"D",
            "Supplier":"E","PartsOrdered":"F","JobNo":"G","PONumber":"H",
            "Cost":"I","DateOrdered":"J","DueDate":"K","DeliveredDate":"L",
            "Savings":"M","OriginalSupplier":"N","CostCentre":"O",
        }
    return header_row or 2, colmap


def validate_required_parts_columns(colmap):
    missing = [k for k in REQUIRED_PARTS_FIELDS if k not in colmap]
    if missing:
        raise RuntimeError(f"\nERROR: Missing columns: {', '.join(missing)}")


def find_first_empty_row(ws, key_col_letter, start_row):
    r = start_row
    while not is_blank_cell(ws[f"{key_col_letter}{r}"].value):
        r += 1
    return r


# ========= MACHINE / AREA / COST CENTRE =========
# NOTE: Replace these example mappings with your own machine names,
# area names, and cost centre codes.

MACHINE_ALIAS_RAW = {
    "machine 1": "Machine 1",
    "m1":        "Machine 1",
    "machine 2": "Machine 2",
    "m2":        "Machine 2",
    "workshop":  "Workshop/Plant",
    "plant":     "Workshop/Plant",
    "h&s":       "Health & Safety",
    "health & safety": "Health & Safety",
}
MACHINE_ALIAS = {normalize_text(k): v for k, v in MACHINE_ALIAS_RAW.items()}

AREA_ALIAS_RAW = {
    "cons":        "Consumables",
    "consumables": "Consumables",
    "conveyor":    "Conveyors",
    "conveyors":   "Conveyors",
}
AREA_ALIAS = {normalize_text(k): v for k, v in AREA_ALIAS_RAW.items()}

KNOWN_MACHINES = [
    "Machine 1",
    "Machine 2",
    "Workshop/Plant",
    "Health & Safety",
    # Add your machines here...
]

KNOWN_AREAS = [
    "Consumables",
    "Conveyors",
    "Mech",
    "Elec",
    "Pneumatics",
    # Add your areas here...
]

KNOWN_MECH_ELEC = ["Mech", "Elec", "Pneumatics", "Other", "Hydraulics"]

KNOWN_REASONS = [
    "Breakdown", "Preventive Maintenance", "Planned Repair", "Stock Replenishment",
    "New Install", "Upgrade", "Health & Safety", "Consumables",
    "Project", "Modification", "Inspection",
]

# Replace with your own (Machine, Area) -> CostCentreCode mappings
CC_LOOKUP = {
    ("Machine 1", ""):            "100",
    ("Machine 1", "Consumables"): "100A",
    ("Machine 2", ""):            "200",
    ("Machine 2", "Consumables"): "200A",
    ("Workshop/Plant", ""):       "300",
    ("Health & Safety", ""):      "400",
}


def normalize_machine_name(raw):
    raw = (raw or "").strip()
    return MACHINE_ALIAS.get(normalize_text(raw), raw)


def normalize_area_name(raw, machine_canon=""):
    raw = (raw or "").strip()
    return AREA_ALIAS.get(normalize_text(raw), raw)


def _tolerant_match(a, b):
    if not a or not b:
        return False
    return a == b or a.startswith(b) or b.startswith(a) or a in b or b in a


def find_cost_centre(ws_costs, machine_input, area_input):
    m, a = normalize_text(machine_input), normalize_text(area_input)
    for row in ws_costs.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        mach, area, cc = row[0], row[1], row[2]
        if not mach or not cc:
            continue
        if _tolerant_match(normalize_text(mach), m) and _tolerant_match(normalize_text(area or ""), a):
            return str(cc)
    return None


def resolve_cost_centre(wb, machine, area, mech_elec):
    m, a = machine.strip(), area.strip()
    if (m, a) in CC_LOOKUP:
        return CC_LOOKUP[(m, a)]
    if (m, "") in CC_LOOKUP:
        return CC_LOOKUP[(m, "")]
    if SHEET_COST_CENTRE in wb.sheetnames:
        cc = find_cost_centre(wb[SHEET_COST_CENTRE], machine, area)
        if cc:
            return cc
    return "UNKNOWN"


def count_orders_today(ws, ordered_col, start_row):
    today = datetime.today().date()
    count = 0
    for r in range(start_row, ws.max_row + 1):
        val = ws[f"{ordered_col}{r}"].value
        if isinstance(val, datetime) and val.date() == today:
            count += 1
        elif isinstance(val, date) and val == today:
            count += 1
        elif isinstance(val, str):
            try:
                if datetime.strptime(val.strip(), "%d/%m/%Y").date() == today:
                    count += 1
            except ValueError:
                pass
    return count


def generate_po_number(cost_centre, initials, seq_base):
    now = datetime.now()
    return f"{cost_centre}-{initials}{now.day:02d}{now.month:02d}{seq_base + 1}{now.year % 100:02d}"


# ========= SUPPLIERS =========
def load_suppliers_db(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"suppliers": []}


def pick_supplier(path, query):
    db = load_suppliers_db(path)
    suppliers = db.get("suppliers", [])
    q_raw = (query or "").strip()

    if q_raw.lower() in ("manual", "add"):
        name = input_text("Supplier name (or 'back'): ", allow_blank=False, allow_back=True)
        return BACK if name == BACK else {"name": name}

    if not q_raw:
        return {"name": ""}

    q_norm = normalize_text(q_raw)
    found = [s for s in suppliers
             if q_norm in normalize_text(s.get("name") or "")
             or any(q_norm in normalize_text(a) for a in (s.get("aliases") or []))]

    if not found:
        names = [s.get("name", "") for s in suppliers if s.get("name")]
        suggestion = get_close_matches(q_raw, names, n=5, cutoff=0.60)
        if suggestion:
            print("No direct match. Did you mean:")
            for i, s in enumerate(suggestion, 1):
                print(f"  {i}) {s}")
            pick = input_text("Pick 1-5 or Enter for manual: ", allow_blank=True, allow_back=True)
            if pick == BACK:
                return BACK
            if pick.isdigit() and 1 <= int(pick) <= len(suggestion):
                for sup in suppliers:
                    if sup.get("name") == suggestion[int(pick) - 1]:
                        return sup
        manual = input_text("Supplier name (or 'back'): ", allow_blank=False, allow_back=True)
        return BACK if manual == BACK else {"name": manual}

    print("Matches:")
    for i, s in enumerate(found[:8], 1):
        print(f"  {i}) {s.get('name')} ({s.get('supplier_id','N/A')})")
    pick = input_text("Pick 1-8, Enter for first, or 'back': ", allow_blank=True, allow_back=True)
    if pick == BACK:
        return BACK
    if pick.isdigit() and 1 <= int(pick) <= min(8, len(found)):
        return found[int(pick) - 1]
    return found[0]


# ========= TO RECEIVE SHEET =========
TO_RECEIVE_HEADERS = [
    "PO Number", "Supplier", "Machine", "Area", "Reason", "Mech/Elec/Other",
    "Job No", "Items", "Total Cost", "Date Ordered", "Due Date", "Cost Centre",
    "Status", "Received Date", "Invoice Number", "Notes",
]


def ensure_to_receive_sheet(wb):
    if TO_RECEIVE_SHEET in wb.sheetnames:
        return wb[TO_RECEIVE_SHEET]
    ws = wb.create_sheet(TO_RECEIVE_SHEET)
    for i, h in enumerate(TO_RECEIVE_HEADERS, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    return ws


def to_receive_col_index(header_name):
    return TO_RECEIVE_HEADERS.index(header_name) + 1


def find_to_receive_row_by_po(ws, po_number):
    target = (po_number or "").strip()
    if not target:
        return None
    po_col = to_receive_col_index("PO Number")
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=po_col).value or "").strip() == target:
            return r
    return None


# ========= OUTLOOK EMAIL =========
def validate_supplier_email(supplier_obj):
    email = (supplier_obj.get("email") or "").strip()
    if not email:
        print(f"\n[Email] No email for supplier '{supplier_obj.get('name', '')}'.")
        return "" if input("Continue without email? (y/n): ").strip().lower() == "y" else None
    return email


def _extract_first_name(full_name):
    parts = (full_name or "").strip().split()
    return parts[0] if parts else "Sir/Madam"


def open_outlook_with_po(pdf_path, po_number, supplier_obj, order, totals):
    if not AUTO_OPEN_OUTLOOK or not OUTLOOK_AVAILABLE:
        return False
    if not os.path.isfile(pdf_path):
        return False

    supplier_email = validate_supplier_email(supplier_obj)
    if supplier_email is None:
        return False

    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        if supplier_email:
            mail.To = supplier_email
        mail.Subject = EMAIL_SUBJECT_TEMPLATE.format(po_number=po_number, company=SHEARD_COMPANY_NAME)

        contact = (supplier_obj.get("contact_name") or supplier_obj.get("name") or "").strip()
        body = EMAIL_BODY_TEMPLATE.format(
            supplier_first_name=_extract_first_name(contact),
            po_number=po_number,
            items_summary=order.get("items_summary", "See attached PDF"),
            net_total=totals.get("net_total", 0.0),
            due_date=order.get("due_raw", "TBC") or "TBC",
        )
        mail.Attachments.Add(os.path.abspath(pdf_path))
        mail.Display()
        body_html = body.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")
        mail.HTMLBody = f'<html><body style="font-family:Calibri,sans-serif;font-size:11pt;">{body_html}</body></html>'
        print(f"[Email] Outlook opened — PO {po_number}")
        return True
    except Exception as e:
        print(f"[Email] Failed: {e}")
        return False


# ========= PDF GENERATION =========
def try_register_dejavusans():
    for p in [r"C:\Windows\Fonts\DejaVuSans.ttf", r"C:\Windows\Fonts\Arial.ttf"]:
        if os.path.isfile(p):
            try:
                pdfmetrics.registerFont(TTFont("BodyFont", p))
                return "BodyFont"
            except Exception:
                pass
    return None


def require_reportlab_or_abort():
    if not REPORTLAB_AVAILABLE:
        print(f'\nERROR: Install ReportLab: "{sys.executable}" -m pip install reportlab')
        raise RuntimeError("ReportLab missing")


def generate_po_pdf(pdf_path, po_number, supplier_obj, order, items, totals):
    require_reportlab_or_abort()
    try:
        ensure_dir(os.path.dirname(pdf_path))
        header_bg = PDF_HEADER_BG or HexColor('#2C3E50')
        accent    = PDF_ACCENT    or HexColor('#2980B9')
        light_bg  = PDF_LIGHT_BG  or HexColor('#ECF0F1')
        dark_text = PDF_DARK_TEXT or HexColor('#2C3E50')

        font_normal = try_register_dejavusans() or "Helvetica"
        font_bold   = "Helvetica-Bold"

        c = canvas.Canvas(pdf_path, pagesize=A4)
        pw, ph = A4
        ml, mr = 30, 30
        cw = pw - ml - mr
        hb = ph - 80

        # Header
        if os.path.isfile(LOGO_PATH):
            try:
                img = ImageReader(LOGO_PATH)
                iw, ih = img.getSize()
                scale = min(150/iw, 55/ih)
                c.drawImage(LOGO_PATH, ml, ph - 20 - ih*scale, width=iw*scale, height=ih*scale, mask='auto')
            except Exception:
                c.setFillColor(dark_text); c.setFont(font_bold, 16)
                c.drawString(ml, ph - 40, SHEARD_COMPANY_NAME)
        else:
            c.setFillColor(dark_text); c.setFont(font_bold, 16)
            c.drawString(ml, ph - 40, SHEARD_COMPANY_NAME)

        c.setFillColor(dark_text)
        c.setFont(font_bold, 10); c.drawRightString(pw - mr, ph - 25, "PURCHASE ORDER")
        c.setFont(font_bold, 20); c.drawRightString(pw - mr, ph - 48, po_number)
        c.setFont(font_normal, 9); c.drawRightString(pw - mr, ph - 62, f"Date: {datetime.today().strftime('%d/%m/%Y')}")
        c.setStrokeColor(accent); c.setLineWidth(2)
        c.line(ml, hb, pw - mr, hb)

        y = hb - 25
        col_w = (cw - 20) / 3
        bh = 18
        lh = 13

        def banner(x, y, w, h, text):
            c.setFillColor(accent); c.rect(x, y - 3, w, h, fill=1, stroke=0)
            c.setFillColor(white); c.setFont(font_bold, 10)
            c.drawString(x + 5, y - 3 + (h - 10) / 2, text)

        banner(ml, y, col_w, bh, "Supplier:")
        c.setFillColor(dark_text); c.setFont(font_normal, 9)
        yy = y - 20
        c.drawString(ml + 5, yy, (supplier_obj.get("name") or order.get("supplier") or "").strip())
        for line in (supplier_obj.get("address_lines") or [])[:5]:
            yy -= lh; c.drawString(ml + 5, yy, str(line))
        for field in ("email", "phone"):
            val = supplier_obj.get(field, "")
            if val:
                yy -= lh; c.drawString(ml + 5, yy, val)

        x2 = ml + col_w + 10
        banner(x2, y, col_w, bh, "Bill To:")
        c.setFillColor(dark_text); c.setFont(font_normal, 9)
        yy = y - 20
        for line in SHEARD_BILL_TO:
            c.drawString(x2 + 5, yy, line); yy -= lh

        x3 = x2 + col_w + 10
        banner(x3, y, col_w, bh, "Deliver To:")
        c.setFillColor(dark_text); c.setFont(font_normal, 9)
        yy = y - 20
        for line in SHEARD_DELIVER_TO:
            c.drawString(x3 + 5, yy, line); yy -= lh

        yc = y - 100
        banner(ml, yc, cw, bh, "Contact:")
        c.setFillColor(dark_text); c.setFont(font_normal, 10)
        c.drawString(ml + 5, yc - 20, f"{CONTACT_NAME} | {CONTACT_PHONE} | {CONTACT_EMAIL}")

        yd = yc - 50
        c.setFillColor(light_bg); c.rect(ml, yd - 3, cw, 18, fill=1, stroke=0)
        c.setFillColor(dark_text)
        xp = ml + 5
        for label, key in [("Machine:", "machine"), ("Area:", "area"),
                            ("Reason:", "reason"), ("Mech/Elec/Other:", "mech_elec")]:
            c.setFont(font_bold, 9); c.drawString(xp, yd, label)
            lw = c.stringWidth(label, font_bold, 9)
            c.setFont(font_normal, 9); val = str(order.get(key, ""))
            c.drawString(xp + lw + 3, yd, val)
            xp += lw + c.stringWidth(val, font_normal, 9) + 18

        yd2 = yd - 22
        c.setFillColor(light_bg); c.rect(ml, yd2 - 3, cw, 18, fill=1, stroke=0)
        c.setFillColor(dark_text)
        xp = ml + 5
        for label, key in [("Job No:", "job_no"), ("Cost Centre:", "cost_centre"), ("Due Date:", "due_raw")]:
            c.setFont(font_bold, 9); c.drawString(xp, yd2, label)
            lw = c.stringWidth(label, font_bold, 9)
            c.setFont(font_normal, 9); val = str(order.get(key, ""))
            c.drawString(xp + lw + 3, yd2, val)
            xp += lw + c.stringWidth(val, font_normal, 9) + 18

        yi = yd2 - 35
        c.setFillColor(accent); c.setFont(font_bold, 12); c.drawString(ml, yi, "Items:")

        yth = yi - 25
        c.setFillColor(header_bg); c.rect(ml, yth - 5, cw, 20, fill=1, stroke=0)
        c.setFillColor(white); c.setFont(font_bold, 9)
        c.drawString(ml + 5, yth, "Qty")
        c.drawString(ml + 38, yth, "Part No")
        c.drawString(ml + 130, yth, "Description")
        c.drawRightString(pw - mr - 100, yth, "Unit \u00a3")
        c.drawRightString(pw - mr - 5, yth, "Line \u00a3")

        yr = yth - 22
        for idx, it in enumerate(items):
            if yr < 120:
                c.showPage(); yr = ph - 60
            if idx % 2 == 0:
                c.setFillColor(light_bg); c.rect(ml, yr - 7, cw, 20, fill=1, stroke=0)
            c.setFillColor(dark_text); c.setFont(font_normal, 9)
            c.drawString(ml + 5, yr, str(it.get("qty", "")))
            c.drawString(ml + 38, yr, (it.get("part_no", "") or "")[:18])
            c.drawString(ml + 130, yr, (it.get("desc", "") or "")[:45])
            c.drawRightString(pw - mr - 100, yr, f"{float(it.get('unit_price', 0.0)):.2f}")
            c.drawRightString(pw - mr - 5, yr, f"{float(it.get('line_total', 0.0)):.2f}")
            yr -= 20

        yt = yr - 20
        net   = float(totals.get("net_total", 0.0))
        vrate = float(totals.get("vat_rate", VAT_RATE_DEFAULT))
        vat   = float(totals.get("vat_amount", 0.0))
        gross = float(totals.get("gross_total", net + vat))

        c.setFillColor(dark_text)
        c.setFont(font_bold, 10); c.drawRightString(pw - mr - 100, yt, "Net Total \u00a3:")
        c.setFont(font_normal, 10); c.drawRightString(pw - mr - 5, yt, f"{net:.2f}")
        yt -= 18
        c.setFont(font_bold, 10); c.drawRightString(pw - mr - 100, yt, f"VAT ({vrate*100:.0f}%) \u00a3:")
        c.setFont(font_normal, 10); c.drawRightString(pw - mr - 5, yt, f"{vat:.2f}")
        yt -= 24
        gbw = 200
        c.setFillColor(header_bg); c.rect(pw - mr - gbw, yt - 7, gbw, 24, fill=1, stroke=0)
        c.setFillColor(white); c.setFont(font_bold, 11)
        c.drawRightString(pw - mr - 100, yt, "Gross Total \u00a3:")
        c.drawRightString(pw - mr - 5, yt, f"{gross:.2f}")

        c.save()
        return os.path.isfile(pdf_path) and os.path.getsize(pdf_path) > 0
    except Exception as e:
        print(f"PDF failed: {e}")
        import traceback; traceback.print_exc()
        return False


# ========= INPUT FLOW =========
def prompt_items():
    print("\nEnter line items. Blank Quantity to finish. 'back' to go back.")
    items = []

    while True:
        qty_raw = input_positive_int("Quantity (blank to finish): ", allow_blank=True, allow_back=True)
        if qty_raw == BACK:
            if items:
                removed = items.pop()
                print(f"Removed: x{removed['qty']} {removed['desc']}")
                continue
            return BACK
        if qty_raw == "":
            break

        qty = int(qty_raw)
        part_no = input_text("Part number (optional, Enter to skip): ", allow_blank=True, allow_back=True)
        if part_no == BACK:
            continue
        desc = input_text("Item description: ", allow_blank=False, allow_back=True)
        if desc == BACK:
            continue
        unit_raw = input_money("Unit price (\u00a3) (blank for line total): ", allow_blank=True, allow_back=True)
        if unit_raw == BACK:
            continue

        if unit_raw == "":
            line_raw = input_money("Line total (\u00a3): ", allow_blank=False, allow_back=True)
            if line_raw == BACK:
                continue
            line_total = float(parse_currency_to_float(line_raw) or 0.0)
            unit = (line_total / qty) if qty else 0.0
        else:
            unit = float(parse_currency_to_float(unit_raw) or 0.0)
            line_total = qty * unit

        items.append({"part_no": part_no, "desc": desc, "qty": qty,
                      "unit_price": float(unit), "line_total": float(line_total)})

    if not items:
        items.append({"part_no": "", "desc": "N/A", "qty": 1, "unit_price": 0.0, "line_total": 0.0})

    net_total = sum(i["line_total"] for i in items)

    while True:
        if PROMPT_VAT_RATE_EACH_PO:
            raw = input_text(f"VAT rate % (Enter for {int(VAT_RATE_DEFAULT*100)}): ",
                             allow_blank=True, allow_back=True)
            if raw == BACK:
                continue
            vat_rate = VAT_RATE_DEFAULT if raw == "" else float(raw.replace("%", "")) / 100.0
            if vat_rate < 0:
                print("VAT cannot be negative."); continue
            break
        else:
            vat_rate = VAT_RATE_DEFAULT; break

    vat_amount  = net_total * vat_rate
    gross_total = net_total + vat_amount
    return items, {"net_total": float(net_total), "vat_rate": float(vat_rate),
                   "vat_amount": float(vat_amount), "gross_total": float(gross_total)}


def collect_po_data():
    print("\n=== CREATE NEW PURCHASE ORDER ===")
    order, supplier_obj = {}, {"name": ""}

    steps = [
        ("machine", None),
        ("area", None),
        ("reason",       lambda: input_with_autocomplete("Reason: ",           KNOWN_REASONS,   allow_blank=False)),
        ("mech_elec",    lambda: input_with_autocomplete("Mech/Elec/Other: ",  KNOWN_MECH_ELEC, allow_blank=False)),
        ("supplier", None),
        ("items", None),
        ("job_no",       lambda: input_text("Job No (optional): ", allow_blank=True, allow_back=True)),
        ("due_raw",      lambda: input_text("Due date (dd/mm/yyyy) or blank: ", allow_blank=True, allow_back=True)),
        ("savings_raw",  lambda: input_text("Savings \u00a3 (optional): ", allow_blank=True, allow_back=True)),
        ("orig_supplier",lambda: input_text("Original supplier (optional): ", allow_blank=True, allow_back=True)),
    ]

    i, items, totals = 0, None, None

    while i < len(steps):
        key, fn = steps[i]

        if key == "machine":
            val = input_with_autocomplete("Machine: ", KNOWN_MACHINES, allow_blank=False)
            if val == BACK:
                i = max(0, i - 1); continue
            order["machine"] = val; i += 1; continue

        if key == "area":
            m = order.get("machine", "")
            machine_areas = [a for (mc, a) in CC_LOOKUP if mc == m and a != ""]
            if machine_areas:
                print(f"\n  Areas for '{m}':")
                for idx, a in enumerate(machine_areas, 1):
                    print(f"    {idx}) {a}  (CC: {CC_LOOKUP.get((m, a), '?')})")
                print(f"    0) Base machine only  (CC: {CC_LOOKUP.get((m, ''), '?')})")
                print(f"    M) Manual")
                pick = input("  Pick (Enter=0, or 'back'): ").strip()
                if is_back(pick):
                    i = max(0, i - 1); continue
                if pick == "" or pick == "0":
                    order["area"] = ""; i += 1; continue
                if pick.upper() == "M":
                    av = input_with_autocomplete("Area: ", KNOWN_AREAS, allow_blank=False)
                    if av == BACK: continue
                    order["area"] = av; i += 1; continue
                if pick.isdigit() and 1 <= int(pick) <= len(machine_areas):
                    order["area"] = machine_areas[int(pick) - 1]; i += 1; continue
                print("  Invalid."); continue
            else:
                av = input_with_autocomplete("Area: ", KNOWN_AREAS, allow_blank=True)
                if av == BACK:
                    i = max(0, i - 1); continue
                order["area"] = av; i += 1; continue

        if key == "supplier":
            q = input_text("Supplier search: ", allow_blank=False, allow_back=True)
            if q == BACK:
                i = max(0, i - 1); continue
            sup = pick_supplier(SUPPLIERS_JSON_PATH, q)
            if sup == BACK: continue
            supplier_obj = sup
            order["supplier"] = (supplier_obj.get("name", "") or "").strip()
            i += 1; continue

        if key == "items":
            res = prompt_items()
            if res == BACK:
                i = max(0, i - 1); continue
            items, totals = res
            order["items_summary"] = "; ".join(
                f"x{it['qty']} {('[' + it['part_no'] + '] ') if it.get('part_no') else ''}{it['desc']}"
                for it in items
            )[:250]
            i += 1; continue

        val = fn()
        if val == BACK:
            i = max(0, i - 1); continue
        if key == "due_raw" and val and parse_ddmmyyyy_to_date(val) is None:
            print("Invalid date. Use dd/mm/yyyy."); continue
        order[key] = val; i += 1

    return order, supplier_obj, items, totals


# ========= WRITE FUNCTIONS =========
def write_parts_order_row(ws_parts, header_row, colmap, order, po_number, cost_centre, net_total):
    data_start = header_row + 1
    row = find_first_empty_row(ws_parts, colmap["Machine"], data_start)

    def w(field, value):
        if field in colmap:
            ws_parts[f"{colmap[field]}{row}"] = value

    ws_parts[f"{colmap['DateOrdered']}{row}"] = datetime.today().date()
    ws_parts[f"{colmap['DateOrdered']}{row}"].number_format = "DD/MM/YYYY"
    w("Machine", order["machine"])
    w("Area", order["area"])
    w("Reason", order["reason"])
    w("MechElecOther", order["mech_elec"])
    w("Supplier", order["supplier"])
    w("PartsOrdered", order["items_summary"])
    if "PartsOrdered" in colmap:
        ws_parts[f"{colmap['PartsOrdered']}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    w("JobNo", order.get("job_no", ""))
    w("PONumber", po_number)
    w("CostCentre", cost_centre)

    if "Cost" in colmap:
        cost_cell = ws_parts[f"{colmap['Cost']}{row}"]
        cost_cell.value = float(net_total)
        cost_cell.number_format = "\u00a3#,##0.00"

    savings_val = parse_currency_to_float(order.get("savings_raw", ""))
    if savings_val is not None and "Savings" in colmap:
        sc = ws_parts[f"{colmap['Savings']}{row}"]
        sc.value = float(savings_val)
        sc.number_format = "\u00a3#,##0.00"

    if order.get("orig_supplier"):
        w("OriginalSupplier", order["orig_supplier"])

    due_dt = parse_ddmmyyyy_to_date(order.get("due_raw", ""))
    if due_dt and "DueDate" in colmap:
        ws_parts[f"{colmap['DueDate']}{row}"] = due_dt
        ws_parts[f"{colmap['DueDate']}{row}"].number_format = "DD/MM/YYYY"
    elif order.get("due_raw") and "DueDate" in colmap:
        w("DueDate", order["due_raw"])

    return row


def add_to_receive_row(ws_recv, po_number, order, cost_centre, net_total, totals):
    row = ws_recv.max_row + 1

    def set_cell(header, value):
        ws_recv.cell(row=row, column=to_receive_col_index(header), value=value)

    set_cell("PO Number", po_number)
    set_cell("Supplier", order["supplier"])
    set_cell("Machine", order["machine"])
    set_cell("Area", order["area"])
    set_cell("Reason", order["reason"])
    set_cell("Mech/Elec/Other", order["mech_elec"])
    set_cell("Job No", order.get("job_no", ""))
    set_cell("Items", order["items_summary"])

    cost_cell = ws_recv.cell(row=row, column=to_receive_col_index("Total Cost"), value=float(net_total))
    cost_cell.number_format = "\u00a3#,##0.00"

    date_cell = ws_recv.cell(row=row, column=to_receive_col_index("Date Ordered"), value=datetime.today().date())
    date_cell.number_format = "DD/MM/YYYY"

    due_val = parse_ddmmyyyy_to_date(order.get("due_raw", "")) or order.get("due_raw", "")
    due_cell = ws_recv.cell(row=row, column=to_receive_col_index("Due Date"), value=due_val)
    if isinstance(due_val, (date, datetime)):
        due_cell.number_format = "DD/MM/YYYY"

    set_cell("Cost Centre", cost_centre)
    set_cell("Status", "Ordered")
    set_cell("Notes", f"VAT {totals.get('vat_rate', VAT_RATE_DEFAULT)*100:.0f}% | Gross \u00a3{totals.get('gross_total', 0.0):.2f}")
    return row


def mark_received(wb, po_number, invoice_number=""):
    if TO_RECEIVE_SHEET not in wb.sheetnames:
        print("To Receive sheet not found.")
        return

    ws_recv = wb[TO_RECEIVE_SHEET]
    recv_row = find_to_receive_row_by_po(ws_recv, po_number)
    if not recv_row:
        print("PO number not found.")
        return

    ws_recv.cell(row=recv_row, column=to_receive_col_index("Status"), value="Received")
    rd_cell = ws_recv.cell(row=recv_row, column=to_receive_col_index("Received Date"),
                           value=datetime.today().date())
    rd_cell.number_format = "DD/MM/YYYY"
    if invoice_number:
        ws_recv.cell(row=recv_row, column=to_receive_col_index("Invoice Number"), value=invoice_number)

    for sheet_name in wb.sheetnames:
        if sheet_name in {TO_RECEIVE_SHEET, SHEET_COST_CENTRE}:
            continue
        ws = wb[sheet_name]
        header_row, colmap = detect_header_row_and_columns(ws)
        if not colmap or "PONumber" not in colmap or "DeliveredDate" not in colmap:
            continue
        start_row = (header_row or 2) + 1
        for r in range(start_row, ws.max_row + 1):
            if str(ws[f"{colmap['PONumber']}{r}"].value or "").strip() == po_number:
                dd_cell = ws[f"{colmap['DeliveredDate']}{r}"]
                dd_cell.value = datetime.today().date()
                dd_cell.number_format = "DD/MM/YYYY"
                print(f"Delivered date set in '{sheet_name}' row {r}.")
                return


# ========= SYNC =========
GREEN_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")


def _parse_date_flexible(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def _is_valid_month_sheet(sheet_name, min_year=SYNC_FROM_YEAR):
    parts = sheet_name.strip().split()
    if len(parts) >= 2:
        try:
            return int(parts[-1]) >= min_year
        except ValueError:
            pass
    return False


def _get_to_receive_po_numbers(wb):
    pos = set()
    if TO_RECEIVE_SHEET in wb.sheetnames:
        ws = wb[TO_RECEIVE_SHEET]
        po_col_idx = to_receive_col_index("PO Number")
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=po_col_idx).value
            if not is_blank_cell(val):
                pos.add(str(val).strip())
    return pos


def _get_monthly_sheet_po_index(wb):
    index = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in {TO_RECEIVE_SHEET, SHEET_COST_CENTRE}:
            continue
        ws = wb[sheet_name]
        header_row, colmap = detect_header_row_and_columns(ws)
        if not colmap or "PONumber" not in colmap:
            continue
        start_row = (header_row or 2) + 1
        pos = set()
        for r in range(start_row, ws.max_row + 1):
            val = ws[f"{colmap['PONumber']}{r}"].value
            if not is_blank_cell(val):
                pos.add(str(val).strip())
        index[sheet_name] = pos
    return index


def _write_order_to_monthly_sheet(ws_month, cmap, h_row, o, po_str):
    start_row = (h_row or 2) + 1
    data_row = find_first_empty_row(ws_month, cmap["Machine"], start_row)

    def w(field, value):
        if field in cmap:
            ws_month[f"{cmap[field]}{data_row}"] = value

    w("Machine", o["machine"]); w("Area", o["area"]); w("Reason", o["reason"])
    w("MechElecOther", o["mech_elec"]); w("Supplier", o["supplier"])
    w("PartsOrdered", o["parts"]); w("JobNo", o["job_no"]); w("PONumber", po_str)

    if "Cost" in cmap:
        cc = ws_month[f"{cmap['Cost']}{data_row}"]
        cc.value = float(o["cost"]); cc.number_format = "\u00a3#,##0.00"

    if o["date_ordered"] and "DateOrdered" in cmap:
        dc = ws_month[f"{cmap['DateOrdered']}{data_row}"]
        dc.value = o["date_ordered"]; dc.number_format = "DD/MM/YYYY"

    if o["delivered_date"] and "DeliveredDate" in cmap:
        ddc = ws_month[f"{cmap['DeliveredDate']}{data_row}"]
        ddc.value = o["delivered_date"]; ddc.number_format = "DD/MM/YYYY"; ddc.fill = GREEN_FILL

    if o["savings"] and "Savings" in cmap:
        sc = ws_month[f"{cmap['Savings']}{data_row}"]
        sc.value = float(o["savings"]); sc.number_format = "\u00a3#,##0.00"

    if "CostCentre" in cmap:
        mc = normalize_machine_name(o["machine"])
        ac = normalize_area_name(o["area"])
        cc_code = CC_LOOKUP.get((mc, ac)) or CC_LOOKUP.get((mc, ""), "")
        w("CostCentre", cc_code)

    return data_row


def sync_with_main_sheet():
    if not os.path.isfile(MAIN_WORKBOOK_PATH):
        print(f"\n[Sync] Main workbook not found:\n  {MAIN_WORKBOOK_PATH}")
        return

    print(f"\n{'='*60}\n  FULL SYNC WITH MAIN SHEET\n{'='*60}")

    try:
        main_wb = load_main_workbook_readonly(MAIN_WORKBOOK_PATH)
    except Exception as e:
        print(f"[Sync] Failed to open main workbook: {e}")
        return

    skip_sheets = {"cost centre", "sheet2", "to receive"}
    valid_sheets = [s for s in main_wb.sheetnames
                    if s.lower().strip() not in skip_sheets and _is_valid_month_sheet(s)]
    print(f"  Scanning {len(valid_sheets)} sheets...")

    main_orders, orders_by_sheet = {}, {}

    for sheet_name in valid_sheets:
        ws = main_wb[sheet_name]
        rows_data = list(ws.iter_rows(values_only=True))
        if len(rows_data) < 2:
            continue

        col_map, best_score, header_row_idx = {}, -1, None
        for r_idx in range(min(12, len(rows_data))):
            row = rows_data[r_idx]
            score, temp_map = 0, {}
            for c_idx, val in enumerate(row):
                norm = normalize_text(val)
                raw_val = str(val or "").strip()
                if "\u00a3" in raw_val and len(raw_val) <= 10 and "Cost" not in temp_map:
                    temp_map["Cost"] = c_idx; score += 1; continue
                for field, synonyms in HEADER_SYNONYMS.items():
                    if norm in synonyms and field not in temp_map:
                        temp_map[field] = c_idx; score += 1; break
            if score > best_score:
                col_map = temp_map; best_score = score; header_row_idx = r_idx

        if not col_map or "PONumber" not in col_map:
            continue

        count = 0
        for r_idx in range((header_row_idx or 0) + 1, len(rows_data)):
            row = rows_data[r_idx]
            if not row:
                continue

            def sg(field):
                if field in col_map and col_map[field] < len(row):
                    return row[col_map[field]]
                return None

            po_val = sg("PONumber")
            if is_blank_cell(po_val):
                continue
            po_str = str(po_val).strip()
            if not po_str:
                continue

            order = {
                "sheet": sheet_name,
                "machine": str(sg("Machine") or ""),
                "area": str(sg("Area") or ""),
                "reason": str(sg("Reason") or ""),
                "mech_elec": str(sg("MechElecOther") or ""),
                "supplier": str(sg("Supplier") or ""),
                "parts": str(sg("PartsOrdered") or ""),
                "job_no": str(sg("JobNo") or ""),
                "cost": to_float_cell(sg("Cost")),
                "date_ordered": _parse_date_flexible(sg("DateOrdered")),
                "due_date": _parse_date_flexible(sg("DueDate")),
                "delivered_date": _parse_date_flexible(sg("DeliveredDate")),
                "savings": to_float_cell(sg("Savings")),
                "orig_supplier": str(sg("OriginalSupplier") or ""),
            }
            main_orders[po_str] = order
            orders_by_sheet.setdefault(sheet_name, {})[po_str] = order
            count += 1
        print(f"  '{sheet_name}': {count} orders")

    try:
        main_wb.close()
    except Exception:
        pass

    if not main_orders:
        print("[Sync] No orders found on main sheet.")
        return

    print(f"[Sync] Total: {len(main_orders)} orders")

    wb = load_or_create_workbook(WORKBOOK_PATH)
    ws_recv = ensure_to_receive_sheet(wb)
    max_col = ws_recv.max_column or len(TO_RECEIVE_HEADERS)
    po_col_idx = to_receive_col_index("PO Number")
    status_col_idx = to_receive_col_index("Status")
    recv_date_col_idx = to_receive_col_index("Received Date")

    # STEP 1: Sync deliveries
    print(f"\n--- STEP 1: Sync Deliveries ---")
    delivered_main = {po: o for po, o in main_orders.items() if o["delivered_date"]}
    delivery_updated = 0

    for r in range(2, ws_recv.max_row + 1):
        po_val = ws_recv.cell(row=r, column=po_col_idx).value
        if is_blank_cell(po_val):
            continue
        po_str = str(po_val).strip()
        if po_str not in delivered_main:
            continue
        status_val = str(ws_recv.cell(row=r, column=status_col_idx).value or "").strip()
        if status_val.lower() == "received":
            for c in range(1, max_col + 1):
                ws_recv.cell(row=r, column=c).fill = GREEN_FILL
            continue
        ws_recv.cell(row=r, column=status_col_idx).value = "Received"
        recv_cell = ws_recv.cell(row=r, column=recv_date_col_idx)
        recv_cell.value = delivered_main[po_str]["delivered_date"]
        recv_cell.number_format = "DD/MM/YYYY"
        for c in range(1, max_col + 1):
            ws_recv.cell(row=r, column=c).fill = GREEN_FILL
        delivery_updated += 1
        print(f"  \u2713 {po_str} \u2192 Received")

    print(f"  Deliveries updated: {delivery_updated}")

    # STEP 2: Import missing
    print(f"\n--- STEP 2: Import Missing Orders ---")
    to_receive_pos = _get_to_receive_po_numbers(wb)
    missing = {po: o for po, o in main_orders.items() if po not in to_receive_pos}
    imported = 0

    if missing:
        print(f"  {len(missing)} orders not yet in To Receive.")
        choice = input("  Import all? (y/n): ").strip().lower()
        if choice in ("y", "yes"):
            for po_str, o in missing.items():
                row = ws_recv.max_row + 1
                def set_cell(header, value):
                    ws_recv.cell(row=row, column=to_receive_col_index(header), value=value)
                set_cell("PO Number", po_str)
                set_cell("Supplier", o["supplier"])
                set_cell("Machine", o["machine"])
                set_cell("Area", o["area"])
                set_cell("Items", o["parts"])
                cc_cell = ws_recv.cell(row=row, column=to_receive_col_index("Total Cost"),
                                       value=float(o["cost"]))
                cc_cell.number_format = "\u00a3#,##0.00"
                if o["date_ordered"]:
                    dc = ws_recv.cell(row=row, column=to_receive_col_index("Date Ordered"),
                                      value=o["date_ordered"])
                    dc.number_format = "DD/MM/YYYY"
                mc = normalize_machine_name(o["machine"])
                ac = normalize_area_name(o["area"])
                cc = CC_LOOKUP.get((mc, ac)) or CC_LOOKUP.get((mc, ""), "")
                set_cell("Cost Centre", cc)
                if o["delivered_date"]:
                    set_cell("Status", "Received")
                    rc = ws_recv.cell(row=row, column=to_receive_col_index("Received Date"),
                                      value=o["delivered_date"])
                    rc.number_format = "DD/MM/YYYY"
                    for c in range(1, max_col + 1):
                        ws_recv.cell(row=row, column=c).fill = GREEN_FILL
                else:
                    set_cell("Status", "Ordered")
                imported += 1
            print(f"  \u2713 Imported {imported}")
    else:
        print("  All orders already in To Receive.")

    # STEP 3: Sync monthly sheets
    print(f"\n--- STEP 3: Sync Monthly Sheets ---")
    monthly_po_index = _get_monthly_sheet_po_index(wb)
    monthly_written = 0

    for sheet_name, sheet_orders in orders_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue
        for po_str, o in sheet_orders.items():
            if sheet_name in monthly_po_index and po_str in monthly_po_index[sheet_name]:
                continue
            ws_month = wb[sheet_name]
            h_row, cmap = ensure_parts_column_map(ws_month)
            if "PONumber" not in cmap or "Machine" not in cmap:
                continue
            _write_order_to_monthly_sheet(ws_month, cmap, h_row, o, po_str)
            monthly_written += 1

    if monthly_written > 0:
        for sheet_name in orders_by_sheet:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                h_row, cmap = detect_header_row_and_columns(ws)
                if cmap:
                    update_month_totals(ws, h_row or 2, cmap)
                    auto_fit_columns(ws)

    print(f"  Monthly rows written: {monthly_written}")

    auto_fit_columns(ws_recv)
    safe_save_workbook(wb, WORKBOOK_PATH)
    print(f"\n{'='*60}\n  SYNC COMPLETE\n{'='*60}")
    print(f"  Deliveries:    {delivery_updated}")
    print(f"  Imported:      {imported}")
    print(f"  Monthly added: {monthly_written}")


# ========= MAIN MENU =========
def create_new_po():
    wb = load_or_create_workbook(WORKBOOK_PATH)
    ws_parts = get_or_create_month_sheet(wb)
    ws_recv = ensure_to_receive_sheet(wb)
    header_row, colmap = ensure_parts_column_map(ws_parts)
    try:
        validate_required_parts_columns(colmap)
    except RuntimeError as e:
        print(e); return

    order, supplier_obj, items, totals = collect_po_data()
    order["machine"] = normalize_machine_name(order["machine"])
    order["area"]    = normalize_area_name(order["area"], order["machine"])

    cost_centre = resolve_cost_centre(wb, order["machine"], order["area"], order["mech_elec"])
    order["cost_centre"] = cost_centre

    data_start = header_row + 1
    seq_base  = count_orders_today(ws_parts, colmap["DateOrdered"], data_start)
    po_number = generate_po_number(cost_centre, INITIALS, seq_base)
    net_total = totals["net_total"]

    parts_row = write_parts_order_row(ws_parts, header_row, colmap, order, po_number, cost_centre, net_total)
    recv_row  = add_to_receive_row(ws_recv, po_number, order, cost_centre, net_total, totals)

    update_month_totals(ws_parts, header_row, colmap)
    auto_fit_columns(ws_parts)
    auto_fit_columns(ws_recv)
    set_sheet_view(ws_parts, zoom=90)
    set_sheet_view(ws_recv, zoom=90)

    ensure_dir(PO_PDF_OUTPUT_DIR)
    pdf_path = os.path.join(PO_PDF_OUTPUT_DIR, f"{po_number}.pdf")
    pdf_ok = generate_po_pdf(pdf_path, po_number, supplier_obj, order, items, totals)

    safe_save_workbook(wb, WORKBOOK_PATH)

    print(f"\nSUCCESS")
    print(f"PO Number: {po_number}")
    print(f"Sheet:     '{ws_parts.title}' row {parts_row}")
    print(f"To Receive: row {recv_row}")

    if pdf_ok:
        print(f"PDF:       {pdf_path}")
        if AUTO_OPEN_OUTLOOK:
            open_outlook_with_po(pdf_path, po_number, supplier_obj, order, totals)
    else:
        print("PDF generation failed.")


def receive_po():
    wb = load_or_create_workbook(WORKBOOK_PATH)
    po_number = input_text("PO number to mark Received (or 'back'): ", allow_blank=False, allow_back=True)
    if po_number == BACK:
        return
    invoice_number = input_text("Invoice number (optional): ", allow_blank=True, allow_back=True)
    if invoice_number == BACK:
        invoice_number = ""
    mark_received(wb, po_number, invoice_number)
    safe_save_workbook(wb, WORKBOOK_PATH)
    print("Saved.")


def main():
    ensure_dir(PO_PDF_OUTPUT_DIR)
    while True:
        print("\n=== PO FLOW MENU ===")
        print("1) Create new Purchase Order (PO)")
        print("2) Mark PO as Received (Goods Receipt)")
        print("3) Full Sync with Main Sheet")
        print("4) Exit")
        choice = input("Choose 1-4: ").strip()

        if choice == "1":
            create_new_po()
        elif choice == "2":
            receive_po()
        elif choice == "3":
            sync_with_main_sheet()
        elif choice == "4":
            break
        else:
            print("Invalid choice.")


if __name__ == "__main__":
    main()
