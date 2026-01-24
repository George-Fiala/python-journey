from openpyxl import load_workbook, Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Alignment, Font
from datetime import datetime
from difflib import get_close_matches
import re
import time
import json

# ====== SETTINGS ======
FILE_PATH = r"C:\Users\george\OneDrive - S Sheard\Desktop\Parts ordered Home.xlsx"
SUPPLIERS_PATH = r"C:\Users\george\OneDrive - S Sheard\Desktop\Python project\suppliers.json"

SHEET_COSTS = "Cost centre"   # List s tabulkou Machine / Area / CostCentre
INITIALS = "GF"               # Tvé iniciály

# ====== COLUMNS CONFIGURATION ======
COL_MACHINE      = "A"
COL_AREA         = "B"
COL_REASON       = "C"
COL_MECH_ELEC    = "D"
COL_SUPPLIER     = "E"
COL_PARTS        = "F"
COL_JOB          = "G"
COL_PO           = "H"
COL_COST         = "I"   # Cost £
COL_ORDERED      = "J"   # Date ordered
COL_DUE          = "K"   # Due date
COL_DELIVERED    = "L"   # Delivered date
COL_SAVINGS      = "M"   # Cost Initiatives (Savings)
COL_ORIG_SUPPLIER= "N"   # Original Supplier
COL_COST_CENTER  = "O"   # Cost Centre

# --- ALIASY PRO STROJE ---
MACHINE_ALIAS = {
    "e1": "Emba 1", "e2": "Emba 2", "e3": "Emba 3", "e4": "Emba 4", "e5": "Emba 5", "e6": "Emba 6",
    "g1": "Gopfert 1", "g2": "Gopfert 2",
    "b1": "Bobst 1 Visionfold", "b2": "Bobst 2 ExpertFold",
    "a1": "Asahi 1", "a2": "Asahi 2",
    "plant": "Workshop/Plant", "workshop": "Workshop/Plant", "production": "Workshop/Plant"
}

AREA_ALIAS = {
    "cons": "consumables", "consum": "consumables", "consumable": "consumables"
}

# ====== POMOCNÉ FUNKCE (LOGIKA) ======

def normalize_machine_name(machine_raw):
    m = (machine_raw or "").strip().lower()
    return MACHINE_ALIAS.get(m, machine_raw)

def normalize_text(name):
    if name is None: return ""
    return re.sub(r"[^a-z0-9]+", "", str(name).strip().lower())

def is_workshop_plant_electrical(machine_any, mech_elec_raw):
    m = normalize_text(machine_any)
    me = normalize_text(mech_elec_raw)
    workshop_like = m in {"plant", "workshop", "workshopplant", "production", "prod"}
    electrical_like = "elect" in me or me in {"ele", "elec", "electrical"}
    return workshop_like and electrical_like

def normalize_area_name(area_raw, machine_canon=""):
    raw = (area_raw or "").strip()
    norm_area = normalize_text(raw)
    norm_machine = normalize_text(machine_canon)
    if norm_machine == normalize_text("Bobst 2 ExpertFold") and norm_area in {"ctech", "ctechrobopacker"}:
        return "C-Tech Robopacker"
    return AREA_ALIAS.get(norm_area, raw)

def load_or_create_workbook(file_path):
    while True:
        try:
            return load_workbook(file_path)
        except FileNotFoundError: return Workbook()
        except PermissionError:
            input(f"\n⚠️ Zavři soubor Excel na cestě: {file_path}\nPak zmáčkni Enter...")

def get_sheet_for_current_month(wb):
    full_name = datetime.today().strftime("%B %Y")
    if full_name not in wb.sheetnames:
        ws = wb.create_sheet(full_name) if wb.sheetnames[0] != "Sheet" else wb.active
        ws.title = full_name
        return ws
    return wb[full_name]

def init_parts_sheet(ws):
    """Inicializace hlaviček a vzorců pro součty."""
    # --- 1. SOUČTY (ŘÁDEK 1) ---
    if ws["H1"].value is None or ws["M1"].value is None:
        ws["H1"] = "Total spend £"
        ws["I1"] = f"=SUM({COL_COST}3:{COL_COST}1000)"
        ws["I1"].font = Font(bold=True)

        ws["L1"] = "Total Savings £"
        ws["M1"] = f"=SUM({COL_SAVINGS}3:{COL_SAVINGS}1000)"
        ws["L1"].alignment = Alignment(horizontal="right")
        ws["M1"].font = Font(bold=True, color="006100") # Zelená pro úspory!

    # --- 2. HLAVIČKY (ŘÁDEK 2) ---
    if ws[f"{COL_MACHINE}2"].value is None:
        headers = {
            COL_MACHINE: "Machine", COL_AREA: "Area", COL_REASON: "Reason",
            COL_MECH_ELEC: "Mech/Elec/Other", COL_SUPPLIER: "Supplier",
            COL_PARTS: "Parts ordered", COL_JOB: "Job No", COL_PO: "PO number",
            COL_COST: "Cost £", COL_ORDERED: "Date ordered", COL_DUE: "Due date",
            COL_DELIVERED: "Delivered date", COL_SAVINGS: "Cost Initiatives",
            COL_ORIG_SUPPLIER: "Original supplier", COL_COST_CENTER: "Cost Centre"
        }
        for col, title in headers.items():
            cell = ws[f"{col}2"]
            cell.value = title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            if col == COL_SAVINGS: cell.fill = PatternFill("solid", start_color="92D050")
            if col == COL_ORIG_SUPPLIER:
                cell.fill = PatternFill("solid", start_color="0070C0")
                cell.font = Font(bold=True, color="FFFFFF")

        widths = {"A": 16, "B": 18, "C": 18, "D": 18, "E": 24, "F": 40, "G": 14, "H": 20, "I": 12, "M": 16, "N": 20, "O": 16}
        for col, w in widths.items(): ws.column_dimensions[col].width = w

def find_cost_centre(ws_costs, machine_input, area_input):
    m_canon = normalize_machine_name(machine_input)
    a_norm = normalize_text(normalize_area_name(area_input, m_canon))
    m_norm = normalize_text(m_canon)

    for row in ws_costs.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[2]: continue
        if normalize_text(row[0]) == m_norm:
            if not row[1] or normalize_text(row[1]) == a_norm or normalize_text(row[1]) == "machine":
                return str(row[2]).strip()
    return "UNKNOWN"

def generate_po(cost_center, count_today):
    now = datetime.now()
    return f"{cost_center}-{INITIALS}{now.day:02d}{now.month:02d}{count_today + 1}{now.year % 100:02d}"

# =========================
# DATABÁZE DODAVATELŮ
# =========================

def load_suppliers_db(path):
    try:
        with open(path, "r", encoding="utf-8") as f: return json.load(f)
    except: return {"suppliers": []}

def suggest_suppliers(query, suppliers):
    q = query.lower().strip()
    matches = [s for s in suppliers if q in s['name'].lower() or any(q in a.lower() for a in s.get('aliases', []))]
    return matches[:8]

def choose_supplier_interactive(path):
    db = load_suppliers_db(path)
    while True:
        q = input("Supplier search (or 'manual'): ").strip()
        if q.lower() == 'manual': return {"name": input("Name: ")}
        results = suggest_suppliers(q, db['suppliers'])
        for i, s in enumerate(results, 1): print(f"  {i}) {s['name']}")
        choice = input("Pick number or Enter to search again: ")
        if choice.isdigit() and 1 <= int(choice) <= len(results): return results[int(choice)-1]

# =========================
# MAIN PROGRAM
# =========================

def parse_cost(val):
    try: return float(val.replace("£", "").replace(",", ""))
    except: return None

def main():
    print("=== New Order Entry ===")
    m_raw = input("Machine: ")
    a_raw = input("Area: ")
    reason = input("Reason: ")
    m_e = input("Mech/Elec/Other: ")
    sup_obj = choose_supplier_interactive(SUPPLIERS_PATH)
    parts = input("Parts ordered: ")
    job = input("Job No: ")
    cost = parse_cost(input("Cost £: "))
    savings = parse_cost(input("Savings £: "))
    orig_sup = input("Original Supplier: ")
    due = input("Due date (dd/mm/yyyy): ")

    wb = load_or_create_workbook(FILE_PATH)
    ws = get_sheet_for_current_month(wb)
    init_parts_sheet(ws)
    
    m_canon = normalize_machine_name(m_raw)
    a_canon = normalize_area_name(a_raw, m_canon)
    
    # Cost Centre logika
    if is_workshop_plant_electrical(m_canon, m_e): cc = "321C"
    else: cc = find_cost_centre(wb["Cost centre"], m_canon, a_canon)
    
    # Počet dnešních objednávek pro PO
    today_str = datetime.today().date()
    count_today = sum(1 for r in range(3, ws.max_row+1) if ws[f"J{r}"].value == today_str)
    po = generate_po(cc, count_today)

    row = 3
    while ws[f"A{row}"].value: row += 1

    # Zápis dat
    ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"] = m_canon, a_canon, reason
    ws[f"D{row}"], ws[f"E{row}"], ws[f"F{row}"] = m_e, sup_obj['name'], parts
    ws[f"G{row}"], ws[f"H{row}"], ws[f"O{row}"] = job, po, cc
    ws[f"J{row}"] = today_str
    ws[f"J{row}"].number_format = "DD/MM/YYYY"

    if cost:
        ws[f"I{row}"] = cost
        ws[f"I{row}"].number_format = '£#,##0.00'
    if savings:
        ws[f"M{row}"] = savings
        ws[f"M{row}"].number_format = '£#,##0.00'
    if orig_sup: ws[f"N{row}"] = orig_sup
    if due:
        try:
            ws[f"K{row}"] = datetime.strptime(due, "%d/%m/%Y").date()
            ws[f"K{row}"].number_format = "DD/MM/YYYY"
        except: pass

    # Podmíněné formátování (červená, když je po termínu)
    red_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
    rule = FormulaRule(formula=[f'AND(K{row}<TODAY(), K{row}<>"")'], fill=red_fill)
    ws.conditional_formatting.add(f"K{row}", rule)

    wb.save(FILE_PATH)
    print(f"\n✅ Hotovo! PO: {po}")

if __name__ == "__main__":
    main()