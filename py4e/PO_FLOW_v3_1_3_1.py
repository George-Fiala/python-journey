# =========================
# PO + GR One Flow (v3.1.3)
# - Create Purchase Order (PO) PDF  *** STYLED PDF — VARIANT D: CLEAN MINIMAL ***
# - Write to "Parts Ordered" monthly sheet
# - Add to "To Receive" (Goods Receipt (GR) backlog)
# - Mark as Received later (updates both sheets)
# - Auto-open Outlook email with PDF attachment
# - Full Sync with main shared workbook
# - AUTO PUSH to main SharePoint workbook via Microsoft Graph API (v3.0.0)

# v3.1.3 changes:
#   - FIX (main SharePoint, take 2): _graph_format_new_row now uses the
#     correct Graph API endpoints. Diagnostic testing on 22 May 2026
#     showed that PATCH /range with {"format": {"rowHeight": X}} nested
#     returns 400 InvalidArgument. The correct approach:
#       - PATCH /range(A:N)/format with {"wrapText": True}     for cells
#       - PATCH /range(N:N)/format with {"rowHeight": X}       for row
#       - POST  /range/format/autofitRows                      as fallback
#     All three return 200/204. Row now expands to fit wrapped text.
#
# v3.1.2 changes:
#   - FIX (local): Local workbook row height now auto-expands when parts
#     text is long. write_parts_order_row estimates visual lines from
#     text length + semicolon-separated items and sets row_dimensions.
# v3.1.1 changes:
# - Create Purchase Order (PO) PDF  *** STYLED PDF — VARIANT D: CLEAN MINIMAL ***
# - Write to "Parts Ordered" monthly sheet
# - Add to "To Receive" (Goods Receipt (GR) backlog)
# - Mark as Received later (updates both sheets)
# - Auto-open Outlook email with PDF attachment
# - Full Sync with main shared workbook
# - AUTO PUSH to main SharePoint workbook via Microsoft Graph API (v3.0.0)

# v3.1.1 changes:
#   - IMPROVE: New sheets created by _graph_ensure_main_sheets_through now
#     also get the two SUM formulas in row 1, matching all existing monthly
#     sheets. I1=SUM(I3:I5000) (total spend), M1=SUM(M3:M5000) (total
#     savings/cost initiatives). Both formatted as £#,##0.00. The formula
#     starts at 0 and live-updates as POs are added throughout the month.
#   - CHANGE: _graph_init_sheet_headers now does two PATCH calls:
#     one for row 2 headers, one for row 1 formulas + number format.
#
# v3.1.0 changes:
# - Create Purchase Order (PO) PDF  *** STYLED PDF — VARIANT D: CLEAN MINIMAL ***
# - Write to "Parts Ordered" monthly sheet
# - Add to "To Receive" (Goods Receipt (GR) backlog)
# - Mark as Received later (updates both sheets)
# - Auto-open Outlook email with PDF attachment
# - Full Sync with main shared workbook
# - AUTO PUSH to main SharePoint workbook via Microsoft Graph API (v3.0.0)

# v3.1.0 changes:
#   - NEW: If target monthly sheet does not exist in main workbook, the
#     push automatically creates ALL missing intermediate sheets in
#     strict chronological order (Jan → Feb → Mar → Apr) before writing
#     the PO. No months are skipped, maintaining correct tab bar order.
#     Each new sheet gets standard headers on row 2 matching existing
#     monthly sheets. Data starts at row 3 as per convention.
#   - Context: In practice, George creates next-year sheets manually at
#     year-end and books services from January. This is a safety net for
#     edge cases (e.g. service contract due April 2027 when main only
#     has through December 2026).
#   - NEW helpers: _graph_create_worksheet(), _graph_init_sheet_headers(),
#     _graph_ensure_main_sheets_through(), MAIN_SHEET_HEADERS constant.
#   - CHANGE: push_po_to_main no longer aborts when sheet is missing —
#     it creates the sheet(s) first, then writes.
#
# v3.0.4 changes:
# - Create Purchase Order (PO) PDF  *** STYLED PDF — VARIANT D: CLEAN MINIMAL ***
# - Write to "Parts Ordered" monthly sheet
# - Add to "To Receive" (Goods Receipt (GR) backlog)
# - Mark as Received later (updates both sheets)
# - Auto-open Outlook email with PDF attachment
# - Full Sync with main shared workbook
# - AUTO PUSH to main SharePoint workbook via Microsoft Graph API (v3.0.0)

# v3.0.4 changes:
#   - NEW: After Graph push, automatically applies Wrap Text + autofitRows
#     to the new row. Without this, long parts-ordered descriptions in
#     column F were truncated/invisible. With this, Excel expands the row
#     height to show the full description across multiple visual lines —
#     identical to rows entered manually in the main workbook.
#   - NEW helper: _graph_format_new_row() — PATCH wrapText + POST
#     autofitRows. Graceful: data is preserved even if formatting fails.
#   - Two additional Graph API calls per push (wrapText + autofitRows),
#     total ~7 calls per PO. Still typically completes in 2-4 seconds.
#
# v3.0.3 changes:
# - Create Purchase Order (PO) PDF  *** STYLED PDF — VARIANT D: CLEAN MINIMAL ***
# - Write to "Parts Ordered" monthly sheet
# - Add to "To Receive" (Goods Receipt (GR) backlog)
# - Mark as Received later (updates both sheets)
# - Auto-open Outlook email with PDF attachment
# - Full Sync with main shared workbook
# - AUTO PUSH to main SharePoint workbook via Microsoft Graph API (v3.0.0)

# v3.0.3 changes — race condition defense:
#   - NEW: _graph_verify_cell_empty() re-reads the target cell just before
#     writing to confirm it's still empty. Closes the ~1-3s race window
#     where another user could have started typing into the same row
#     between our scan and our PATCH.
#   - NEW: Retry loop in push_po_to_main — if a race is detected (cell now
#     has data), automatically find a new empty row and verify again. Up
#     to GRAPH_PUSH_MAX_RETRIES (3) attempts before giving up gracefully.
#   - Verification fails OPEN (treats as empty) on network errors to avoid
#     blocking legitimate writes if the Graph endpoint is flaky.
#   - Reduces race collision risk from ~0.5% per PO to ~0.05% (10x safer).
#
# v3.0.2 changes — BUG FIXES from first live test:
# - Create Purchase Order (PO) PDF  *** STYLED PDF — VARIANT D: CLEAN MINIMAL ***
# - Write to "Parts Ordered" monthly sheet
# - Add to "To Receive" (Goods Receipt (GR) backlog)
# - Mark as Received later (updates both sheets)
# - Auto-open Outlook email with PDF attachment
# - Full Sync with main shared workbook
# - AUTO PUSH to main SharePoint workbook via Microsoft Graph API (v3.0.0)

# v3.0.2 changes — BUG FIXES from first live test:
#   - FIX: Column F (Parts ordered) was empty in main. Used wrong order
#     dict key — "parts_ordered" instead of the actual "items_summary"
#     that collect_po_data() populates. Parts description now writes
#     correctly to main.
#   - FIX: Column M (Cost Initiatives) was always blank. Now reads
#     order["savings_raw"] and parses to float — matches local behaviour
#     where savings are stored when entered.
#   - FIX: Column N (Original supplier) was always blank. Now reads
#     order["orig_supplier"] — matches local behaviour where cross-
#     reference suppliers are stored.
#   - FIX: Dates rendered as Excel serials (e.g. 46162) instead of
#     'dd/mm/yyyy'. _graph_write_row now also sends numberFormat array
#     setting J/K/L to 'dd/mm/yyyy' and I/M to '£#,##0.00'.
#   - REMOVED: legacy "part_number" handling — already included in
#     items_summary by collect_po_data.
#
# v3.0.1 changes:
# - Create Purchase Order (PO) PDF  *** STYLED PDF — VARIANT D: CLEAN MINIMAL ***
# - Write to "Parts Ordered" monthly sheet
# - Add to "To Receive" (Goods Receipt (GR) backlog)
# - Mark as Received later (updates both sheets)
# - Auto-open Outlook email with PDF attachment
# - Full Sync with main shared workbook
# - AUTO PUSH to main SharePoint workbook via Microsoft Graph API (v3.0.0)

# v3.0.1 changes:
#   - FIX: Graph push now writes to the first EMPTY ROW (gap) in column A,
#     not blindly after the last used row. Critical fix — the main workbook
#     often has late entries from other users placed far below the
#     contiguous block of recent POs. Example: rows 101-114 filled, rows
#     115-124 empty, row 125 has a late entry from another user. v3.0.0
#     would have written to row 126 (wrong); v3.0.1 writes to row 115
#     (correct).
#   - Implementation: usedRange tells us how far to scan, then one batched
#     GET of column A finds the first gap. Falls back to appending after
#     last row if no gap exists.
#
# v3.0.0 changes — THE BIG ONE:
#   - NEW: Automatic push to main Parts Ordered.xlsx on SharePoint immediately
#     after every PO is created. Eliminates the manual end-of-day copy step
#     entirely. Uses Microsoft Graph API + MSAL OAuth (delegated auth as George).
#   - Azure AD app: "PO Flow - SharePoint Sync" registered 15 May 2026.
#     Admin consent granted by Sheard IT (Harry) 19 May 2026.
#   - Push is GRACEFUL: if it fails for any reason (no token, API error,
#     sheet missing, network issue), the local PO write always succeeds and
#     a clear warning is printed. Set GRAPH_PUSH_ENABLED=False to disable.
#   - Column mapping verified live against main workbook (19 May 2026):
#     A-G identical to local; H=Order No; I=Cost; J/K=dates as Excel serials;
#     L/M/N left blank on creation (received date, savings, orig supplier).
#   - NEW helpers: _date_to_excel_serial(), _graph_get_token_silent(),
#     _graph_create/close_session(), _graph_list_sheets(),
#     _graph_get_next_row(), _graph_write_row(), _graph_build_row(),
#     push_po_to_main().
#   - NEW optional imports: msal, requests (graceful if not installed).
#   - NEW constants: GRAPH_CLIENT_ID, GRAPH_TENANT_ID, GRAPH_DRIVE_ID,
#     GRAPH_FILE_ID, GRAPH_PUSH_ENABLED.
#
# v2.1.3 changes:
# - Create Purchase Order (PO) PDF  *** STYLED PDF — VARIANT D: CLEAN MINIMAL ***
# - Write to "Parts Ordered" monthly sheet
# - Add to "To Receive" (Goods Receipt (GR) backlog)
# - Mark as Received later (updates both sheets)
# - Auto-open Outlook email with PDF attachment
# - Full Sync with main shared workbook

# v2.1.3 changes:
#   - PERF: Main workbook PO scan is now ~5-10x faster via iter_rows(values_only=True)
#     instead of per-cell access. For Sheard's main workbook (~700+ POs across
#     years of monthly sheets), end-to-end load+scan should drop from
#     30-60s to a few seconds after the first SharePoint sync fetch.
#   - PERF: In-memory session cache for the main PO list (10 min TTL). Second
#     and subsequent calls to option 8 within a session reuse the cached set
#     instantly. User is prompted "Refresh main workbook cache? (y/N)" on
#     subsequent calls.
#   - UX: Progress output during main workbook scan — shows each sheet being
#     processed with running PO count, plus open/scan timings, so the user
#     can see the tool is alive during long operations.
#   - NEW import: column_index_from_string from openpyxl.utils (needed for
#     iter_rows column indexing).
#
# v2.1.2 changes:
#   - IMPROVE: Menu option 8 now filters out POs already in the main workbook
#     by default. Reads main read-only (non-blocking — does not hang if main
#     is locked) and skips local POs whose PO number is found in any main
#     monthly sheet. User can disable the filter with 'n' at the prompt for
#     full audit view.
#   - NEW helpers: _try_load_main_readonly() (non-blocking), 
#     _collect_main_po_numbers() (scans all valid month sheets in main).
#
# v2.1.1 changes:
#   - NEW: Menu option 8 "Show recent POs across all sheets". Lists every PO
#     ordered in the last N days (user-configurable, default 7) across ALL
#     monthly sheets — current month + any future months created by v2.1.0
#     future-routing. Grouped by sheet with cost subtotals and grand total.
#     Read-only — does not modify the workbook.
#   - Helper for manual end-of-day push to main workbook while Graph API
#     automation is pending admin consent in the Sheard tenant.
#
# v2.1.0 changes:
#   - NEW: Future-month routing. When Due Date is in a future month, the PO is
#     written to that future month's sheet instead of the current month's
#     sheet. Any missing intermediate monthly sheets between today and the
#     target month are auto-created with default Parts Ordered headers.
#     Date Ordered column always remains today's date (reflects reality).
#     Rationale: a service contract ordered today for July 2027 is a July 2027
#     cost — counting it in the current month would distort monthly totals.
#   - REFACTOR: create_new_po() now calls collect_po_data() BEFORE resolving
#     the target monthly sheet, so the due date is known when the sheet is
#     chosen. ensure_to_receive_sheet() still runs early (no due-date dep).
#   - NEW helpers: get_target_month_for_order(), ensure_month_sheets_through(),
#     get_or_create_sheet_for_date().
#
# v1.9.13 changes:
#   - FIX: PDF Part No overflow into Description column resolved with hybrid
#     shrink+wrap strategy. Long Part Numbers auto-shrink from 9pt → 8pt → 7pt
#     to fit; if still too wide, wrap onto multiple lines at 7pt. Row height
#     dynamically adjusts to the longer of Part No or Description lines.
#
# v1.9.9 changes:
#   - FIX: update_month_totals now counts rows with cost even if PO/date are
#     blank (fixes missing no-PO orders from totals — e.g. £3,455 discrepancy)
#   - FIX: to_float_cell now handles formula strings starting with "=" by
#     returning 0.0 with a debug warning instead of silently ignoring
#   - NEW: Menu option 7 — Recalculate All Totals. Walks every monthly sheet
#     and rewrites I1/M1 totals so they always match the actual data rows.
#
# v1.9.8 changes:
#   - FIX: PDF Part Number now displays FULL text (no truncation)
#   - FIX: PDF Description now wraps to multiple lines instead of truncating
#   - FIX: PDF item row height dynamically adjusts to content
#   - FIX: PDF alternating row backgrounds scale with multi-line content
#   - FIX: PDF page break logic accounts for variable row height
#   - FIX: PDF description no longer overflows into Unit £ column (60pt buffer)
#
# v1.9.7 changes:
#   - NEW: Part Number field added to item input (optional, Enter to skip)
#     Prompt order: Quantity → Part Number → Description → Price
#   - NEW: PDF table now shows Part No column between Qty and Description
#   - NEW: Part number included in items_summary when present (visible in Excel)
#
# v1.9.6 changes:
#   - FIX: Full Sync STEP 2D now detects orders moved between months on main
#     sheet and moves them on local sheets too (delete wrong-sheet copy).
#   - FIX: STEP 2E no-PO dedup now runs always (fixes old duplicate rows from
#     previous sync runs, e.g. 4x Asahi 1/2 problem).
#   - FIX: STEP 2F verification pass added — physically checks local sheets
#     after all writes and recovers any still-missing orders.
#
# v1.9.3 changes:
#   - FIX: Monthly sheet sync now runs INDEPENDENTLY of To Receive import.
#   - FIX: Improved "£" column detection in main workbook header scanning.
#   - FIX: Added per-sheet column mapping debug output during sync.
#   - FIX: Zero-cost sanity warning.
#   - FIX: Date sanity check.
#   - NEW: STEP 2C — Misrouted rows report.
#   - NEW: Monthly sheet totals are recalculated after sync writes.
#
# v1.9.2 changes:
#   - FIX: Monthly sheet import now routes by SOURCE SHEET from main workbook.
#   - NEW: Delivered Date cells get GREEN_FILL during sync.
#
# v1.9.1 changes:
#   - FIX: _read_all_orders_from_main now uses iter_rows()
#   - FIX: Added retry logic for locked main workbook
#   - FIX: Supplier email validation before opening Outlook
#   - FIX: Email body template now includes contact name/phone/email at end
#
# v1.8.0 changes:
#   - AUTOCOMPLETE for Machine, Area, Reason, Mech/Elec/Other fields
#
# v1.7.0 changes:
#   - Outlook integration
#
# v1.6.0 changes:
#   - Company logo added to PDF header
# =========================

# ========= BOOTSTRAP: auto-run inside local .venv (Windows) =========
import os
import sys
from pathlib import Path


def _reexec_into_local_venv() -> None:
    try:
        here = Path(__file__).resolve().parent
    except Exception:
        return

    venv_python = here / ".venv" / "Scripts" / "python.exe"
    if not venv_python.is_file():
        return

    try:
        current = Path(sys.executable).resolve()
        target = venv_python.resolve()
    except Exception:
        return

    if current != target:
        print(f"[boot] Re-launching with venv Python: {target}")
        os.execv(str(target), [str(target), str(__file__), *sys.argv[1:]])


if os.name == "nt":
    _reexec_into_local_venv()
# ============================================================

from datetime import datetime, date, timedelta
import json
import re
from difflib import get_close_matches

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string


# ========= OPTIONAL PDF (ReportLab) =========
REPORTLAB_AVAILABLE = False
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.utils import ImageReader
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

# ========= OPTIONAL GRAPH API (msal + requests) =========
# Required for automatic push to main SharePoint workbook (v3.0.0).
# Install: pip install msal requests
GRAPH_AVAILABLE = False
try:
    import msal
    import requests as _requests
    GRAPH_AVAILABLE = True
except ImportError:
    pass


# ========= OPTIONAL OUTLOOK (pywin32) =========
OUTLOOK_AVAILABLE = False
try:
    import win32com.client
    OUTLOOK_AVAILABLE = True
except Exception:
    OUTLOOK_AVAILABLE = False

# ========= OPTIONAL REORDER DETECTOR ========= NEW v2.0.0
DETECTOR_AVAILABLE = False
try:
    from data_parser import load_all_orders
    from detector import (
        calculate_order_patterns,
        get_reorder_alerts,
        get_monthly_forecast,
        print_alerts,
        print_forecast,
    )
    DETECTOR_AVAILABLE = True
except Exception as e:
    print(f"  [DEBUG] Detector import failed: {type(e).__name__}: {e}")
    DETECTOR_AVAILABLE = False


# ========= CONFIG =========
WORKBOOK_PATH = r"C:\Users\george\OneDrive - S Sheard\Desktop\Parts ordered Home.xlsx"
SUPPLIERS_JSON_PATH = r"C:\Users\george\OneDrive - S Sheard\Desktop\Python project\suppliers.json"

MAIN_WORKBOOK_PATH = r"C:\Users\george\S Sheard\Engineers and Maintenance - Documents\Parts order files\Parts Ordered.xlsx"
SYNC_FROM_YEAR = 2026

INITIALS = "GF"
SHEET_COST_CENTRE = "Cost centre"
TO_RECEIVE_SHEET = "To Receive"
PO_PDF_OUTPUT_DIR = r"C:\Users\george\OneDrive - S Sheard\Desktop\PO PDFs"

LOGO_PATH = r"C:\Users\george\OneDrive - S Sheard\Desktop\Python project\sheard_logo.png"

DEBUG = True

# ========= MICROSOFT GRAPH API (v3.0.0) =========
# Azure AD app: "PO Flow - SharePoint Sync"
# Registered: 15 May 2026 — S Sheard tenant
# Permission: Files.ReadWrite.All (Delegated) — admin consent granted by Harry (IT)
GRAPH_CLIENT_ID  = "d60d0d19-cd0b-48e2-875e-3ae208f04af8"
GRAPH_TENANT_ID  = "ffcd164c-1ae6-4fe7-8d5a-8fdf2adf22d0"
GRAPH_AUTHORITY  = f"https://login.microsoftonline.com/{GRAPH_TENANT_ID}"
GRAPH_SCOPES     = ["Files.ReadWrite.All"]
GRAPH_TOKEN_CACHE_PATH = Path(os.path.expanduser("~")) / ".po_flow_token_cache.json"

# SharePoint file location (discovered Phase 2 — 19 May 2026)
GRAPH_DRIVE_ID   = "b!qExCVLQ55E6ijUtRuCtzuEFkecDGXEdAom0xQ2iTxoRQQ1SkVR5KS51oJwgG3Bx5"
GRAPH_FILE_ID    = "015VXJEIMCE5QI2MPQRVFY24OVML7SRS5S"
GRAPH_BASE       = "https://graph.microsoft.com/v1.0"
GRAPH_WB_BASE    = f"{GRAPH_BASE}/drives/{GRAPH_DRIVE_ID}/items/{GRAPH_FILE_ID}/workbook"

# Set to False to disable Graph push without removing the code
GRAPH_PUSH_ENABLED = True

# ========= VAT SETTINGS =========
VAT_RATE_DEFAULT = 0.20
PROMPT_VAT_RATE_EACH_PO = True

# ========= COMPANY DETAILS =========
SHEARD_COMPANY_NAME = "Sheard Packaging"

SHEARD_BILL_TO = [
    "S Sheard & Son Ltd",
    "Solar Works, Church Street",
    "Greetland, Halifax",
    "HX4 8EG",
    "United Kingdom",
    "Company Reg: 01101020",
]

SHEARD_DELIVER_TO = [
    "Sheard Packaging",
    "Solar Works, Church Street",
    "Greetland, Halifax",
    "HX4 8EG",
    "United Kingdom",
]

CONTACT_NAME = "George Fiala"
CONTACT_EMAIL = "george@sheard.co.uk"
CONTACT_PHONE = "07768818702"

# ========= PDF STYLE SETTINGS =========
PDF_HEADER_BG    = HexColor('#2C3E50') if REPORTLAB_AVAILABLE else None
PDF_ACCENT       = HexColor('#2980B9') if REPORTLAB_AVAILABLE else None
PDF_LIGHT_BG     = HexColor('#ECF0F1') if REPORTLAB_AVAILABLE else None
PDF_DARK_TEXT     = HexColor('#2C3E50') if REPORTLAB_AVAILABLE else None


# ========= OUTLOOK EMAIL SETTINGS =========
AUTO_OPEN_OUTLOOK = True

EMAIL_SUBJECT_TEMPLATE = "Purchase Order {po_number} — Sheard Packaging"

EMAIL_BODY_TEMPLATE = """\
Dear {supplier_first_name},

Please find attached Purchase Order {po_number}.

Order Summary:
  Supplier: {supplier_name}
  Items: {items_summary}
  Net Total: £{net_total:.2f}
  Due Date: {due_date}

If you have any questions, please don't hesitate to contact me.

Best regards
"""


# ========= INPUT HELPERS =========
BACK = "__BACK__"


def is_back(s: str) -> bool:
    return (s or "").strip().lower() == "back"


def input_text(prompt: str, allow_blank: bool = False, allow_back: bool = True) -> str:
    while True:
        raw = input(prompt).strip()
        if allow_back and is_back(raw):
            return BACK
        if raw == "" and allow_blank:
            return ""
        if raw == "" and not allow_blank:
            print("This field cannot be blank. Type a value or 'back'.")
            continue
        return raw


def input_positive_int(prompt: str, allow_blank: bool = False, allow_back: bool = True) -> str:
    while True:
        raw = input(prompt).strip()
        if allow_back and is_back(raw):
            return BACK
        if raw == "" and allow_blank:
            return ""
        if not raw.isdigit() or int(raw) <= 0:
            print("Enter a whole number (e.g., 1). Leave blank to finish. Type 'back' to go back.")
            continue
        return raw


def input_money(prompt: str, allow_blank: bool = False, allow_back: bool = True) -> str:
    while True:
        raw = input(prompt).strip()
        if allow_back and is_back(raw):
            return BACK
        if raw == "" and allow_blank:
            return ""
        val = parse_currency_to_float(raw)
        if val is None:
            print("Enter a valid amount (e.g., 12.50). Type 'back' to go back.")
            continue
        return raw


# ====================================================================
# AUTOCOMPLETE INPUT — v1.8.0
# ====================================================================
def input_with_autocomplete(prompt: str, known_values: list,
                            allow_blank: bool = False,
                            allow_back: bool = True,
                            min_chars: int = 1) -> str:
    while True:
        raw = input(prompt).strip()

        if allow_back and is_back(raw):
            return BACK

        if raw == "" and allow_blank:
            return ""
        if raw == "" and not allow_blank:
            print("This field cannot be blank. Type a value or 'back'.")
            continue

        raw_lower = raw.lower()
        for val in known_values:
            if val.lower() == raw_lower:
                return val

        if len(raw) < min_chars:
            return raw

        prefix_matches = [v for v in known_values if v.lower().startswith(raw_lower)]
        substring_matches = [v for v in known_values
                             if raw_lower in v.lower() and v not in prefix_matches]
        all_matches = prefix_matches + substring_matches

        if all_matches:
            print(f"\n  Matches for '{raw}':")
            for i, val in enumerate(all_matches[:9], 1):
                marker = "→" if val in prefix_matches else "~"
                print(f"    {i}) {marker} {val}")
            print(f"    0) Use '{raw}' as typed")

            pick = input("  Pick 1-9 (or 0 to keep, Enter for 1): ").strip()

            if is_back(pick):
                return BACK
            if pick == "" or pick == "1":
                return all_matches[0]
            if pick == "0":
                return raw
            if pick.isdigit() and 1 <= int(pick) <= min(9, len(all_matches)):
                return all_matches[int(pick) - 1]
            return all_matches[0]

        fuzzy = get_close_matches(raw, known_values, n=5, cutoff=0.55)

        if fuzzy:
            print(f"\n  No exact match for '{raw}'. Did you mean:")
            for i, val in enumerate(fuzzy, 1):
                print(f"    {i}) {val}")
            print(f"    0) Use '{raw}' as typed")

            pick = input("  Pick 1-5 (or 0 to keep, Enter for 1): ").strip()

            if is_back(pick):
                return BACK
            if pick == "" or pick == "1":
                return fuzzy[0]
            if pick == "0":
                return raw
            if pick.isdigit() and 1 <= int(pick) <= min(5, len(fuzzy)):
                return fuzzy[int(pick) - 1]
            return fuzzy[0]

        return raw


# ========= GENERAL HELPERS =========
def normalize_text(x) -> str:
    if x is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(x).strip().lower())


def parse_currency_to_float(text: str):
    t = (text or "").strip()
    if not t:
        return None
    t = t.replace("£", "").replace(",", "").strip()
    try:
        return float(t)
    except ValueError:
        return None


def parse_ddmmyyyy_to_date(text: str):
    t = (text or "").strip()
    if not t:
        return None
    try:
        return datetime.strptime(t, "%d/%m/%Y").date()
    except ValueError:
        return None


def ensure_dir(path: str):
    if path and not os.path.isdir(path):
        os.makedirs(path, exist_ok=True)


def load_or_create_workbook(path: str):
    while True:
        try:
            return load_workbook(path)
        except FileNotFoundError:
            print(f"Workbook not found. Creating new workbook at: {path}")
            return Workbook()
        except PermissionError:
            print(f"\nFile is open and cannot be read: {path}")
            input("Close it and press Enter to retry...")
        except Exception as e:
            print(f"\nError while loading workbook: {e}")
            input("Press Enter to retry...")


def safe_save_workbook(wb, path: str):
    while True:
        try:
            wb.save(path)
            return
        except PermissionError:
            print(f"\nFile is open and cannot be saved: {path}")
            input("Close it and press Enter to retry...")
        except Exception as e:
            print(f"\nError while saving workbook: {e}")
            input("Press Enter to retry...")


def load_main_workbook_readonly(path: str):
    while True:
        try:
            print("  (Loading with read_only=False for reliable formula values...)")
            return load_workbook(path, read_only=False, data_only=True)
        except PermissionError:
            print(f"\nMain workbook is locked: {path}")
            input("Close it in Excel and press Enter to retry...")
        except Exception as e:
            print(f"\nError opening main workbook: {e}")
            input("Press Enter to retry or Ctrl+C to cancel...")


def get_month_sheet_name() -> str:
    return datetime.today().strftime("%B %Y")


def get_or_create_month_sheet(wb):
    name = get_month_sheet_name()

    if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet":
        ws = wb["Sheet"]
        ws.title = name
        return ws

    if name not in wb.sheetnames:
        return wb.create_sheet(name)
    return wb[name]


# v2.1.0: Future-month routing helpers
def _month_sheet_name_for(d) -> str:
    """Build the 'Month YYYY' sheet name for a given date/datetime."""
    return d.strftime("%B %Y")


def _add_one_month(d: date) -> date:
    """Return the first day of the month after d."""
    if d.month == 12:
        return date(d.year + 1, 1, 1)
    return date(d.year, d.month + 1, 1)


def get_target_month_for_order(order: dict) -> date:
    """
    Decide which month's sheet should receive this PO.

    Rule: if order has a valid due_raw (dd/mm/yyyy) AND that month is strictly
    AFTER the current month, return the 1st of the due-date month. Otherwise
    return the 1st of the current month.
    """
    today = datetime.today().date()
    current_first = date(today.year, today.month, 1)

    due_raw = order.get("due_raw", "") or ""
    if not due_raw:
        return current_first

    due_dt = parse_ddmmyyyy_to_date(due_raw)
    if not due_dt:
        return current_first

    due_first = date(due_dt.year, due_dt.month, 1)
    if due_first > current_first:
        return due_first
    return current_first


def ensure_month_sheets_through(wb, target_first: date) -> list:
    """
    Make sure every monthly sheet from the current month up to and including
    target_first exists. Newly created sheets get the default Parts Ordered
    header layout via init_parts_sheet_default(). Returns the list of sheet
    names that were created in this call (empty if nothing needed creating).
    """
    today = datetime.today().date()
    cur = date(today.year, today.month, 1)
    created = []

    while cur <= target_first:
        name = _month_sheet_name_for(cur)

        if name not in wb.sheetnames:
            # Reuse the default "Sheet" if this is a fresh workbook
            if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet":
                ws = wb["Sheet"]
                ws.title = name
            else:
                ws = wb.create_sheet(name)
            init_parts_sheet_default(ws)
            created.append(name)

        cur = _add_one_month(cur)

    return created


def get_or_create_sheet_for_date(wb, target_first: date):
    """
    Return the worksheet for target_first's month, creating it (and any
    intermediate months from today up to target_first) if needed.
    """
    ensure_month_sheets_through(wb, target_first)
    name = _month_sheet_name_for(target_first)
    return wb[name]


def _month_sheet_sort_key(sheet_name: str) -> tuple:
    """Return (year, month) tuple for chronological sorting of 'Month YYYY' names."""
    try:
        d = datetime.strptime(sheet_name.strip(), "%B %Y")
        return (d.year, d.month)
    except ValueError:
        return (9999, 12)  # unparseable -> sort to end


def _is_future_sheet(sheet_name: str) -> bool:
    """True if the sheet name refers to a month strictly after the current one."""
    try:
        sheet_dt = datetime.strptime(sheet_name.strip(), "%B %Y").date()
        today = datetime.today().date()
        today_first = date(today.year, today.month, 1)
        return sheet_dt > today_first
    except ValueError:
        return False


def _try_load_main_readonly():
    """
    v2.1.2: Non-blocking variant of load_main_workbook_readonly.
    Returns (workbook, error_msg). On success: (wb, None). On failure: (None, msg).
    Used by helpers that want to read main opportunistically without blocking
    the user when the file is locked.
    """
    if not os.path.exists(MAIN_WORKBOOK_PATH):
        return None, f"Main workbook not found at {MAIN_WORKBOOK_PATH}"
    try:
        wb = load_workbook(MAIN_WORKBOOK_PATH, read_only=True, data_only=True)
        return wb, None
    except PermissionError:
        return None, "Main workbook is locked (open in Excel by someone)"
    except Exception as e:
        return None, f"Could not open main workbook: {e}"


def _collect_main_po_numbers(main_wb, verbose: bool = True) -> set:
    """
    v2.1.3: Scan all monthly sheets in the main workbook and return the set
    of PO numbers found. Uses iter_rows(values_only=True) for ~5-10x speedup
    over per-cell access in openpyxl read_only mode.
    """
    po_set = set()
    valid_sheets = [s for s in main_wb.sheetnames if _is_valid_month_sheet(s)]

    if verbose:
        print(f"[main] Scanning {len(valid_sheets)} monthly sheet(s)...")

    scan_start = datetime.now()

    for i, sheet_name in enumerate(valid_sheets, 1):
        ws = main_wb[sheet_name]
        header_row, colmap = detect_header_row_and_columns(ws)
        if not colmap or not header_row:
            if verbose:
                print(f"[main]   [{i}/{len(valid_sheets)}] {sheet_name}: skipped (no headers)")
            continue

        po_col_letter = colmap.get("PONumber")
        if not po_col_letter:
            if verbose:
                print(f"[main]   [{i}/{len(valid_sheets)}] {sheet_name}: skipped (no PO column)")
            continue

        po_col_idx = column_index_from_string(po_col_letter)
        data_start = header_row + 1

        # iter_rows with values_only is dramatically faster than per-cell access
        # in read_only mode — streams cell values without creating Cell objects.
        sheet_pos_before = len(po_set)
        for row_tuple in ws.iter_rows(
            min_row=data_start,
            min_col=po_col_idx,
            max_col=po_col_idx,
            values_only=True,
        ):
            val = row_tuple[0]
            if not is_blank_cell(val):
                po_set.add(str(val).strip())
        sheet_count = len(po_set) - sheet_pos_before

        if verbose:
            print(f"[main]   [{i}/{len(valid_sheets)}] {sheet_name}: +{sheet_count} PO(s)")

    elapsed = (datetime.now() - scan_start).total_seconds()
    if verbose:
        print(f"[main] Done — {len(po_set)} unique PO(s) found in {elapsed:.1f}s")
    return po_set


# v2.1.3: Session-level cache for main workbook PO list
_MAIN_PO_CACHE = None  # (po_set, timestamp) or None
_MAIN_PO_CACHE_TTL_SECONDS = 600  # 10 minutes


def _get_main_po_numbers_cached(force_refresh: bool = False):
    """
    Return (po_set, error_msg). Uses an in-memory cache to avoid re-scanning
    the (potentially large) main workbook on every call within a session.
    Cache TTL is 10 minutes. Pass force_refresh=True to bypass.
    """
    global _MAIN_PO_CACHE

    if not force_refresh and _MAIN_PO_CACHE is not None:
        po_set, cached_at = _MAIN_PO_CACHE
        age = (datetime.now() - cached_at).total_seconds()
        if age < _MAIN_PO_CACHE_TTL_SECONDS:
            print(f"[main] Using cached PO list ({len(po_set)} PO(s), cached {int(age)}s ago)")
            return po_set, None
        else:
            print(f"[main] Cache expired ({int(age)}s old, TTL {_MAIN_PO_CACHE_TTL_SECONDS}s)")

    print("[main] Opening main workbook (this can take up to a minute for large files)...")
    open_start = datetime.now()
    main_wb, err = _try_load_main_readonly()
    if main_wb is None:
        return None, err
    print(f"[main] Workbook opened in {(datetime.now() - open_start).total_seconds():.1f}s")

    try:
        po_set = _collect_main_po_numbers(main_wb, verbose=True)
    finally:
        try:
            main_wb.close()
        except Exception:
            pass

    _MAIN_PO_CACHE = (po_set, datetime.now())
    return po_set, None


# ========= MONTH SUMMARY + AUTOFIT =========
def to_float_cell(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return 0.0
        # v1.9.9: warn on formula cells instead of silently returning 0
        if s.startswith("="):
            if DEBUG:
                print(f"      [to_float_cell] WARNING: formula cell '{s[:40]}' → 0.0")
            return 0.0
        v = parse_currency_to_float(s)
        return float(v) if v is not None else 0.0
    return 0.0


def is_blank_cell(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        s = value.replace("\u00A0", " ").strip()
        return s == ""
    return False


def ensure_month_summary_cells(ws):
    if ws["H1"].value is None:
        ws["H1"].value = "Total spend £"
        ws["H1"].font = Font(bold=True)
        ws["H1"].alignment = Alignment(horizontal="right")

    if ws["I1"].value is None:
        ws["I1"].value = 0.0
    ws["I1"].number_format = "£#,##0.00"
    ws["I1"].font = Font(bold=True)

    if ws["L1"].value is None:
        ws["L1"].value = "Total Savings £"
        ws["L1"].font = Font(bold=True)
        ws["L1"].alignment = Alignment(horizontal="right")

    if ws["M1"].value is None:
        ws["M1"].value = 0.0
    ws["M1"].number_format = "£#,##0.00"
    ws["M1"].font = Font(bold=True)


def update_month_totals(ws_parts, header_row, colmap):
    """
    v1.9.9 FIX: Now counts ALL rows that contain data, not just rows with
    a PO number or date.  A row is 'real' if ANY of these is true:
      - PO number is not blank
      - Date ordered is not blank
      - Cost value is non-zero
      - Machine is not blank (catches no-PO manual entries)
    This ensures orders without a PO number are included in the total.
    """
    ensure_month_summary_cells(ws_parts)

    data_start = header_row + 1
    total_cost = 0.0
    total_savings = 0.0

    cost_col = colmap.get("Cost")
    savings_col = colmap.get("Savings")
    po_col = colmap.get("PONumber")
    date_col = colmap.get("DateOrdered")
    machine_col = colmap.get("Machine")

    for r in range(data_start, ws_parts.max_row + 1):
        real_row = False

        # Check PO number
        if po_col:
            po_val = ws_parts[f"{po_col}{r}"].value
            if not is_blank_cell(po_val):
                real_row = True

        # Check date ordered
        if not real_row and date_col:
            dt_val = ws_parts[f"{date_col}{r}"].value
            if not is_blank_cell(dt_val):
                real_row = True

        # v1.9.9: Check if row has a non-zero cost (catches no-PO orders)
        if not real_row and cost_col:
            cost_val = to_float_cell(ws_parts[f"{cost_col}{r}"].value)
            if cost_val != 0.0:
                real_row = True

        # v1.9.9: Check if row has a machine name (catches manual entries)
        if not real_row and machine_col:
            mach_val = ws_parts[f"{machine_col}{r}"].value
            if not is_blank_cell(mach_val):
                real_row = True

        if not real_row:
            continue

        if cost_col:
            total_cost += to_float_cell(ws_parts[f"{cost_col}{r}"].value)
        if savings_col:
            total_savings += to_float_cell(ws_parts[f"{savings_col}{r}"].value)

    ws_parts["I1"].value = float(total_cost)
    ws_parts["M1"].value = float(total_savings)


def auto_fit_columns(ws, min_width=8, max_width=70, scan_rows_limit=300):
    max_row = min(ws.max_row, scan_rows_limit)
    max_col = ws.max_column

    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        best = 0
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            if len(s) > best:
                best = len(s)
        width = max(min_width, min(best + 2, max_width))
        ws.column_dimensions[letter].width = width


def set_sheet_view(ws, zoom=90):
    try:
        ws.sheet_view.zoomScale = int(zoom)
    except Exception:
        pass


# ========= HEADER DETECTION =========
HEADER_SYNONYMS = {
    "Machine": {"machine"},
    "Area": {"area", "areaslot", "slot", "areaslotlocation"},
    "Reason": {"reason"},
    "MechElecOther": {"mechelecother", "mechelec", "mech", "elec", "other", "mechanicalelectrical"},
    "Supplier": {"supplier"},
    "PartsOrdered": {"partsordered", "partsorder", "parts"},
    "JobNo": {"jobno", "job", "jobnumber"},
    "PONumber": {"ponumber", "purchaseorder", "po", "orderno", "ordernumber"},
    "Cost": {"cost", "costgbp", "costpound"},
    "DateOrdered": {"dateordered", "ordereddate", "dateorder", "ordered"},
    "DueDate": {"duedate", "due"},
    "DeliveredDate": {"delivereddate", "delivered", "fullydelivered"},
    "Savings": {"costinitiatives", "initiatives", "savings"},
    "OriginalSupplier": {"originalsupplier", "origsupplier"},
    "CostCentre": {"costcentre", "costcenter", "cc"},
}

REQUIRED_PARTS_FIELDS = [
    "Machine", "Area", "Reason", "MechElecOther", "Supplier",
    "PartsOrdered", "JobNo", "PONumber", "Cost", "DateOrdered",
    "DueDate", "Savings", "OriginalSupplier", "CostCentre", "DeliveredDate"
]


def detect_header_row_and_columns(ws, max_scan_rows: int = 12):
    best_row = None
    best_score = -1
    best_map = {}
    max_cols = max(ws.max_column or 1, 30)

    for r in range(1, max_scan_rows + 1):
        row_values = [(c, normalize_text(ws.cell(row=r, column=c).value)) for c in range(1, max_cols + 1)]
        row_raw = [(c, str(ws.cell(row=r, column=c).value or "").strip()) for c in range(1, max_cols + 1)]

        field_to_col = {}
        score = 0

        for field, synonyms in HEADER_SYNONYMS.items():
            for c, norm in row_values:
                if norm and norm in synonyms:
                    field_to_col[field] = get_column_letter(c)
                    score += 1
                    break

        if "Cost" not in field_to_col:
            for c, raw in row_raw:
                if raw == "£" or raw == "£ ":
                    field_to_col["Cost"] = get_column_letter(c)
                    score += 1
                    break

        if score > best_score:
            best_score = score
            best_row = r
            best_map = field_to_col

    if best_score >= 6:
        return best_row, best_map
    return None, {}


def init_parts_sheet_default(ws):
    headers = [
        ("A", "Machine"),
        ("B", "Area"),
        ("C", "Reason"),
        ("D", "Mech/Elec/Other"),
        ("E", "Supplier"),
        ("F", "Parts ordered"),
        ("G", "Job No"),
        ("H", "PO number"),
        ("I", "Cost £"),
        ("J", "Date ordered"),
        ("K", "Due date"),
        ("L", "Delivered date"),
        ("M", "Cost Initiatives"),
        ("N", "Original supplier"),
        ("O", "Cost Centre"),
    ]
    header_row = 2
    for col, label in headers:
        cell = ws[f"{col}{header_row}"]
        cell.value = label
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A3"


def ensure_parts_column_map(ws):
    header_row, colmap = detect_header_row_and_columns(ws)
    if not colmap:
        init_parts_sheet_default(ws)
        header_row = 2
        colmap = {
            "Machine": "A",
            "Area": "B",
            "Reason": "C",
            "MechElecOther": "D",
            "Supplier": "E",
            "PartsOrdered": "F",
            "JobNo": "G",
            "PONumber": "H",
            "Cost": "I",
            "DateOrdered": "J",
            "DueDate": "K",
            "DeliveredDate": "L",
            "Savings": "M",
            "OriginalSupplier": "N",
            "CostCentre": "O",
        }
    return header_row or 2, colmap


def validate_required_parts_columns(colmap: dict):
    missing = [k for k in REQUIRED_PARTS_FIELDS if k not in colmap or not colmap.get(k)]
    if missing:
        raise RuntimeError(
            "\nERROR: Parts Ordered sheet is missing required columns.\n"
            f"Missing: {', '.join(missing)}\n"
            "Fix your header names or update HEADER_SYNONYMS. Aborting to avoid wrong-column writes."
        )


def find_first_empty_row(ws, key_col_letter: str, start_row: int) -> int:
    r = start_row
    while True:
        if is_blank_cell(ws[f"{key_col_letter}{r}"].value):
            return r
        r += 1


# ========= COST CENTRE LOGIC =========
MACHINE_ALIAS_RAW = {
    "plant": "Workshop/Plant",
    "workshop": "Workshop/Plant",
    "prod": "Workshop/Plant",
    "production": "Workshop/Plant",
    "site": "Workshop/Plant",
    "emba 1": "Emba 1",
    "emba1": "Emba 1",
    "e1": "Emba 1",
    "emba 170": "Emba 1",
    "emba170": "Emba 1",
    "1st 170": "Emba 1",
    "emba no 1": "Emba 1",
    "emba no.1": "Emba 1",
    "emba 2": "Emba 2",
    "emba2": "Emba 2",
    "e2": "Emba 2",
    "2nd 170": "Emba 2",
    "emba no 2": "Emba 2",
    "emba no.2": "Emba 2",
    "emba 3": "Emba 3",
    "emba3": "Emba 3",
    "e3": "Emba 3",
    "3rd 170": "Emba 3",
    "emba no 3": "Emba 3",
    "emba no.3": "Emba 3",
    "emba 4": "Emba 4",
    "emba4": "Emba 4",
    "e4": "Emba 4",
    "emba 245": "Emba 4",
    "emba245": "Emba 4",
    "1st 245": "Emba 4",
    "emba no 4": "Emba 4",
    "emba no.4": "Emba 4",
    "emba 5": "Emba 5",
    "emba5": "Emba 5",
    "e5": "Emba 5",
    "2nd 245": "Emba 5",
    "emba no 5": "Emba 5",
    "emba no.5": "Emba 5",
    "emba 6": "Emba 6",
    "emba6": "Emba 6",
    "e6": "Emba 6",
    "3rd 245": "Emba 6",
    "emba no 6": "Emba 6",
    "emba no.6": "Emba 6",
    "twin out": "Emba 6",
    "bobst 1": "Bobst 1 Visionfold",
    "bobst1": "Bobst 1 Visionfold",
    "b1": "Bobst 1 Visionfold",
    "visionfold": "Bobst 1 Visionfold",
    "bost 1": "Bobst 1 Visionfold",
    "bost1": "Bobst 1 Visionfold",
    "bosbt 1": "Bobst 1 Visionfold",
    "bosbt1": "Bobst 1 Visionfold",
    "bobts 1": "Bobst 1 Visionfold",
    "bobts1": "Bobst 1 Visionfold",
    "bobst 2": "Bobst 2 ExpertFold",
    "bobst2": "Bobst 2 ExpertFold",
    "b2": "Bobst 2 ExpertFold",
    "expertfold": "Bobst 2 ExpertFold",
    "bost 2": "Bobst 2 ExpertFold",
    "bost2": "Bobst 2 ExpertFold",
    "bosbt 2": "Bobst 2 ExpertFold",
    "bosbt2": "Bobst 2 ExpertFold",
    "bobts 2": "Bobst 2 ExpertFold",
    "bobts2": "Bobst 2 ExpertFold",
    "vega": "Vega Gluer",
    "vega gluer": "Vega Gluer",
    "gopfert": "Göpfert",
    "göpfert": "Göpfert",
    "gopfert 1": "Göpfert",
    "göpfert 1": "Göpfert",
    "gop": "Göpfert",
    "gop 1": "Göpfert",
    "gopfert 2": "Göpfert 2",
    "göpfert 2": "Göpfert 2",
    "gopfert2": "Göpfert 2",
    "göpfert2": "Göpfert 2",
    "gop 2": "Göpfert 2",
    "gop2": "Göpfert 2",
    "asahi 1": "Asahi 1",
    "asahi1": "Asahi 1",
    "asahi no 1": "Asahi 1",
    "asahi no.1": "Asahi 1",
    "asahi 2": "Asahi 2",
    "asahi2": "Asahi 2",
    "asahi no 2": "Asahi 2",
    "asahi no.2": "Asahi 2",
    "mosca pp1": "Mosca PalletPress 1",
    "mosca palletpress 1": "Mosca PalletPress 1",
    "mosca 1": "Mosca PalletPress 1",
    "mosca1": "Mosca PalletPress 1",
    "palletpress 1": "Mosca PalletPress 1",
    "pallet press 1": "Mosca PalletPress 1",
    "mosca pp2": "Mosca PalletPress 2",
    "mosca palletpress 2": "Mosca PalletPress 2",
    "mosca 2": "Mosca PalletPress 2",
    "mosca2": "Mosca PalletPress 2",
    "palletpress 2": "Mosca PalletPress 2",
    "pallet press 2": "Mosca PalletPress 2",
    "palletpress": "Mosca PalletPress 2",
    "pallet press": "Mosca PalletPress 2",
    "mosca pp3": "Mosca PalletPress 3",
    "mosca palletpress 3": "Mosca PalletPress 3",
    "mosca 3": "Mosca PalletPress 3",
    "mosca3": "Mosca PalletPress 3",
    "palletpress 3": "Mosca PalletPress 3",
    "pallet press 3": "Mosca PalletPress 3",
    "mosca pp4": "Mosca PalletPress 4",
    "mosca palletpress 4": "Mosca PalletPress 4",
    "mosca 4": "Mosca PalletPress 4",
    "mosca4": "Mosca PalletPress 4",
    "palletpress 4": "Mosca PalletPress 4",
    "pallet press 4": "Mosca PalletPress 4",
    "robopac 1": "Robopac Wrapper 1 (Bay 6)",
    "robopac1": "Robopac Wrapper 1 (Bay 6)",
    "robopac bay 6": "Robopac Wrapper 1 (Bay 6)",
    "robo 1": "Robopac Wrapper 1 (Bay 6)",
    "robopac 2": "Robopac Wrapper 2 (Bay 6)",
    "robopac2": "Robopac Wrapper 2 (Bay 6)",
    "robo 2": "Robopac Wrapper 2 (Bay 6)",
    "robopac 3": "Robopac Wrapper 3 (Bay 6)",
    "robopac3": "Robopac Wrapper 3 (Bay 6)",
    "robo 3": "Robopac Wrapper 3 (Bay 6)",
    "robopac bay 7": "Robopac Wrapper (Bay 7)",
    "robopac 4": "Robopac Wrapper (Bay 7)",
    "robopac4": "Robopac Wrapper (Bay 7)",
    "transfer car bay 7": "Transfer Car (Bay 7)",
    "transfer car 7": "Transfer Car (Bay 7)",
    "transfer car bay 9": "Transfer Car (Bay 9)",
    "transfer car 9": "Transfer Car (Bay 9)",
    "transfer car": "Transfer Car (Bay 7)",
    "mill": "Mill",
    "mill lift": "Mill Lift",
    "avanti": "Avanti Conveyor Lines",
    "avanti conveyor": "Avanti Conveyor Lines",
    "avanti conveyors": "Avanti Conveyor Lines",
    "avanti conveyours": "Avanti Conveyor Lines",
    "avanti conveyor lines": "Avanti Conveyor Lines",
    "machine-avanti": "Avanti Conveyor Lines",
    "machine avanti": "Avanti Conveyor Lines",
    "avanti bidi": "Avanti Bidi",
    "avantibidi": "Avanti Bidi",
    "avanti-bidi": "Avanti Bidi",
    "bidi": "Avanti Bidi",
    "conveyor": "Conveyors",
    "conveyors": "Conveyors",
    "conveyour": "Conveyors",
    "conveyours": "Conveyors",
    "health & safety": "Health & Safety",
    "health and safety": "Health & Safety",
    "health safety": "Health & Safety",
    "health&safety": "Health & Safety",
    "healthandsafety": "Health & Safety",
    "healthsafety": "Health & Safety",
    "h&s": "Health & Safety",
    "h & s": "Health & Safety",
    "h and s": "Health & Safety",
    "hs": "Health & Safety",
    "healt & safety": "Health & Safety",
    "healt and safety": "Health & Safety",
    "healt safety": "Health & Safety",
    "healt&safety": "Health & Safety",
    "healtsafety": "Health & Safety",
    "health safet": "Health & Safety",
    "healt safet": "Health & Safety",
    "health nad safety": "Health & Safety",
    "healt nad safety": "Health & Safety",
    "health and safet": "Health & Safety",
    "healt and safet": "Health & Safety",
    "h&s site": "Health & Safety",
    "plan repairs carton": "Plan Repairs Carton",
    "plan repairs": "Plan Repairs Carton",
    "planned repairs carton": "Plan Repairs Carton",
    "plan repair carton": "Plan Repairs Carton",
    "plan repair": "Plan Repairs Carton",
    "prc": "Plan Repairs Carton",
    "waste": "Waste Repairs",
    "waste repairs": "Waste Repairs",
    "waste bailer": "Waste Bailer",
    "bailer": "Waste Bailer",
    "waste extraction": "Waste Extraction",
    "property repair": "Property Repair",
    "property": "Property Repair",
    "prop repair": "Property Repair",
    "propertyrepair": "Property Repair",
    "prop": "Property Repair",
    "training": "Training",
    "service contracts": "Service Contracts",
    "service contract": "Service Contracts",
}
MACHINE_ALIAS = {normalize_text(k): v for k, v in MACHINE_ALIAS_RAW.items()}

AREA_ALIAS_RAW = {
    "cons": "Consumables",
    "consum": "Consumables",
    "consumable": "Consumables",
    "consumables": "Consumables",
    "strapper": "Strapper",
    "isb strapper": "Strapper",
    "isb": "Strapper",
    "mosca strapper": "Strapper",
    "signode strapper": "Strapper",
    "signode isb strapper": "Strapper",
    "signode isb": "Strapper",
    "palletiser": "Palletiser",
    "palletizer": "Palletiser",
    "alliance palletiser": "Palletiser",
    "alliance pall": "Palletiser",
    "ducker palletiser": "Palletiser",
    "pre feeder": "Pre Feeder",
    "prefeeder": "Pre Feeder",
    "pre-feeder": "Pre Feeder",
    "alliance prefeeder": "Pre Feeder",
    "alliance pre feeder": "Pre Feeder",
    "wooson prefeeder": "Pre Feeder",
    "wooson": "Pre Feeder",
    "ducker pre-feeder": "Pre Feeder",
    "ducker prefeeder": "Pre Feeder",
    "ducker pre feeder": "Pre Feeder",
    "stacker": "Stacker",
    "breakers": "Breakers",
    "ducker breakers": "Breakers",
    "c-tech": "C-Tech Robopacker",
    "ctech": "C-Tech Robopacker",
    "c tech": "C-Tech Robopacker",
    "c-tech robopacker": "C-Tech Robopacker",
    "robopacker": "C-Tech Robopacker",
    "die cutter": "Die Cutter",
    "die cut": "Die Cutter",
    "diecutter": "Die Cutter",
    "load former": "Load Former",
    "loadformer": "Load Former",
    "scissor lifts": "Scissor Lifts",
    "scissor lift": "Scissor Lifts",
    "sig sorter": "Sig Sorter",
    "sorter": "Sig Sorter",
    "feeder conveyor": "Feeder Conveyor",
    "gluer": "Gluer",
    "folder gluer": "Folder Gluer",
    "conveyor": "Conveyors",
    "conveyors": "Conveyors",
    "conveyour": "Conveyors",
    "conveyours": "Conveyors",
    "avanti": "Avanti Conveyors",
    "avanti conveyor": "Avanti Conveyors",
    "avanti conveyors": "Avanti Conveyors",
    "avanti conveyours": "Avanti Conveyors",
    "avanti bidi": "Avanti Bidi",
    "avantibidi": "Avanti Bidi",
    "avanti-bidi": "Avanti Bidi",
    "bidi": "Avanti Bidi",
    "multiple misclics": "Multiple Misc Lics",
    "multiple misc lics": "Multiple Misc Lics",
    "misc lics": "Multiple Misc Lics",
    "line 1": "Line 1",
    "line1": "Line 1",
    "l1": "Line 1",
    "line01": "Line 1",
    "line 2": "Line 2",
    "line2": "Line 2",
    "l2": "Line 2",
    "line02": "Line 2",
    "mech": "Mech",
    "elec": "Elec",
    "pneumatics": "Pneumatics",
}
AREA_ALIAS = {normalize_text(k): v for k, v in AREA_ALIAS_RAW.items()}


# ====================================================================
# KNOWN VALUES FOR AUTOCOMPLETE — v1.8.0
# ====================================================================
KNOWN_MACHINES = [
    "Emba 1", "Emba 2", "Emba 3", "Emba 4", "Emba 5", "Emba 6",
    "Bobst 1 Visionfold", "Bobst 2 ExpertFold",
    "Göpfert", "Göpfert 2",
    "Vega Gluer",
    "Asahi 1", "Asahi 2",
    "Mosca PalletPress 1", "Mosca PalletPress 2", "Mosca PalletPress 3", "Mosca PalletPress 4",
    "Robopac Wrapper 1 (Bay 6)", "Robopac Wrapper 2 (Bay 6)", "Robopac Wrapper 3 (Bay 6)", "Robopac Wrapper (Bay 7)",
    "Transfer Car (Bay 7)", "Transfer Car (Bay 9)",
    "Avanti Conveyor Lines", "Avanti Bidi", "Conveyors",
    "Mill", "Mill Lift",
    "Workshop/Plant", "Health & Safety", "Plan Repairs Carton",
    "Waste Repairs", "Waste Bailer", "Waste Extraction",
    "Property Repair", "Training", "Service Contracts",
]

KNOWN_AREAS = [
    "Strapper", "Palletiser", "Pre Feeder", "Stacker", "Breakers",
    "Die Cutter", "Load Former", "Scissor Lifts", "Sig Sorter",
    "Feeder Conveyor", "Gluer", "Folder Gluer", "C-Tech Robopacker",
    "Consumables", "Conveyors", "Avanti Conveyors", "Avanti Bidi",
    "Multiple Misc Lics", "Line 1", "Line 2", "Mech", "Elec", "Pneumatics",
]

KNOWN_MECH_ELEC = ["Mech", "Elec", "Pneumatics", "Other", "Hydraulics"]

KNOWN_REASONS = [
    "Breakdown", "Preventive Maintenance", "Planned Repair", "Stock Replenishment",
    "New Install", "Upgrade", "Health & Safety", "Consumables",
    "Project", "Modification", "Inspection",
]


def normalize_machine_name(machine_raw: str) -> str:
    raw = (machine_raw or "").strip()
    return MACHINE_ALIAS.get(normalize_text(raw), raw)


def normalize_area_name(area_raw: str, machine_canon: str = "") -> str:
    raw = (area_raw or "").strip()
    return AREA_ALIAS.get(normalize_text(raw), raw)


def _tolerant_match(a: str, b: str) -> bool:
    if not a or not b:
        return False
    return (a == b) or a.startswith(b) or b.startswith(a) or (a in b) or (b in a)


def find_cost_centre(ws_costs, machine_input: str, area_input: str):
    m = normalize_text(machine_input)
    a = normalize_text(area_input)
    for row in ws_costs.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        mach, area, cc = row[0], row[1], row[2]
        if not mach or not cc:
            continue
        mach_norm = normalize_text(mach)
        area_norm = normalize_text(area or "")
        if _tolerant_match(mach_norm, m) and _tolerant_match(area_norm, a):
            return str(cc)
    return None


# ====================================================================
# CC_LOOKUP — Complete Cost Centre table
# ====================================================================
CC_LOOKUP = {
    ("Emba 1", ""):             "335",
    ("Emba 1", "Strapper"):     "335A",
    ("Emba 1", "Pre Feeder"):   "335B",
    ("Emba 1", "Palletiser"):   "335C",
    ("Emba 2", ""):             "349",
    ("Emba 2", "Strapper"):     "349A",
    ("Emba 2", "Palletiser"):   "349B",
    ("Emba 3", ""):             "354",
    ("Emba 3", "Pre Feeder"):   "354A",
    ("Emba 3", "Strapper"):     "354B",
    ("Emba 3", "Palletiser"):   "354C",
    ("Emba 4", ""):             "350",
    ("Emba 4", "Pre Feeder"):   "350A",
    ("Emba 4", "Strapper"):     "350B",
    ("Emba 4", "Load Former"):  "350C",
    ("Emba 4", "Palletiser"):   "350D",
    ("Emba 5", ""):             "353",
    ("Emba 5", "Pre Feeder"):   "353A",
    ("Emba 5", "Strapper"):     "353B",
    ("Emba 5", "Palletiser"):   "353C",
    ("Emba 6", ""):             "366",
    ("Emba 6", "Pre Feeder"):   "366A",
    ("Emba 6", "Strapper"):     "366B",
    ("Emba 6", "Sig Sorter"):   "366C",
    ("Emba 6", "Palletiser"):   "366D",
    ("Bobst 1 Visionfold", ""):             "327",
    ("Bobst 1 Visionfold", "Strapper"):     "327A",
    ("Bobst 1 Visionfold", "Palletiser"):   "327B",
    ("Bobst 2 ExpertFold", ""):             "358",
    ("Bobst 2 ExpertFold", "Strapper"):     "358A",
    ("Bobst 2 ExpertFold", "Palletiser"):   "358B",
    ("Bobst 2 ExpertFold", "C-Tech Robopacker"): "358C",
    ("Vega Gluer", ""):             "346",
    ("Vega Gluer", "Strapper"):     "346A",
    ("Vega Gluer", "Palletiser"):   "346B",
    ("Göpfert", ""):            "362",
    ("Göpfert", "Pre Feeder"):  "362A",
    ("Göpfert", "Stacker"):     "362B",
    ("Göpfert", "Breakers"):    "362C",
    ("Göpfert", "Palletiser"):  "362D",
    ("Göpfert 2", ""):            "371",
    ("Göpfert 2", "Pre Feeder"):  "371A",
    ("Göpfert 2", "Stacker"):     "371B",
    ("Göpfert 2", "Breakers"):    "371C",
    ("Göpfert 2", "Palletiser"):  "371D",
    ("Asahi 1", ""):                "372",
    ("Asahi 1", "Die Cutter"):      "372",
    ("Asahi 1", "Breakers"):        "372A",
    ("Asahi 1", "Pre Feeder"):      "372B",
    ("Asahi 1", "Palletiser"):      "372C",
    ("Asahi 1", "Feeder Conveyor"): "372D",
    ("Asahi 2", ""):                "351",
    ("Asahi 2", "Die Cutter"):      "351",
    ("Asahi 2", "Breakers"):        "351A",
    ("Asahi 2", "Scissor Lifts"):   "351B",
    ("Mosca PalletPress 1", ""):    "363",
    ("Mosca PalletPress 2", ""):    "369",
    ("Mosca PalletPress 3", ""):    "364",
    ("Mosca PalletPress 4", ""):    "356",
    ("Robopac Wrapper 1 (Bay 6)", ""):  "363A",
    ("Robopac Wrapper 2 (Bay 6)", ""):  "369A",
    ("Robopac Wrapper 3 (Bay 6)", ""):  "364A",
    ("Robopac Wrapper (Bay 7)", ""):    "356A",
    ("Transfer Car (Bay 7)", ""):   "356B",
    ("Transfer Car (Bay 9)", ""):   "367",
    ("Avanti Conveyor Lines", ""):  "355",
    ("Avanti Bidi", ""):            "355A",
    ("Conveyors", ""):              "355",
    ("Mill", ""):       "373",
    ("Mill Lift", ""):  "374",
    ("Workshop/Plant", "Consumables"):      "334",
    ("Workshop/Plant", "Conveyors"):        "355",
    ("Workshop/Plant", "Avanti Conveyors"): "355",
    ("Workshop/Plant", "Multiple Misc Lics"): "355",
    ("Health & Safety", ""):    "333",
    ("Training", ""):           "322",
    ("Service Contracts", ""):  "323",
    ("Waste Repairs", ""):      "320",
    ("Waste Bailer", ""):       "299",
    ("Waste Extraction", ""):   "370",
    ("Property Repair", ""):    "523",
}


def resolve_cost_centre(wb, machine: str, area: str, mech_elec: str) -> str:
    m = machine.strip()
    a = area.strip()
    me = mech_elec.strip()
    me_norm = normalize_text(me)

    if (m, a) in CC_LOOKUP:
        return CC_LOOKUP[(m, a)]
    if (m, "") in CC_LOOKUP:
        return CC_LOOKUP[(m, "")]

    if normalize_text(m) in {"workshopplant", "plant", "workshop", "site", "production", "prod"}:
        if "elec" in me_norm or "ele" in me_norm:
            return "321B"
        if "pneu" in me_norm:
            return "321C"
        if "mech" in me_norm:
            return "321C"
        return "321C"

    if normalize_text(m) in {"planrepairscarton", "planrepairs", "plannedrepairscarton",
                              "planrepaircarton", "planrepair", "prc"}:
        if "elec" in me_norm or "ele" in me_norm:
            return "321B"
        if "pneu" in me_norm:
            return "321C"
        return "321A"

    hs_norm = normalize_text(m)
    if ("health" in hs_norm and "safe" in hs_norm) or hs_norm in {"hs", "healthsafety"}:
        return "333"

    if SHEET_COST_CENTRE in wb.sheetnames:
        cc = find_cost_centre(wb[SHEET_COST_CENTRE], machine, area)
        if cc:
            return cc

    return "UNKNOWN"


# v2.0.0: PO sequence now counts ALL POs of the day across ALL cost centres.
# First PO of the day = 1, second = 2, third = 3, regardless of cost centre.
# Duplicates are impossible by design — each new PO always gets max+1.
def find_next_po_seq(cost_centre: str, initials: str,
                     day: int, month: int, year_2digit: int) -> int:
    """
    Scan local + main workbooks for PO numbers matching pattern:
      {ANY_cost_centre}-{initials}{DD}{MM}{seq}{YY}
    Counts sequence across ALL cost centres for the day.
    Returns next free seq (max found + 1, or 1 if none found).

    Note: cost_centre parameter is kept for signature compatibility but
    is NOT used in the pattern — the sequence is global for the day.
    """
    # Pattern matches ANY cost centre (digits) — sequence is global for the day
    pattern = re.compile(
        rf"^\d+-{re.escape(initials)}"
        rf"{day:02d}{month:02d}(\d+){year_2digit:02d}$"
    )

    used_seqs = set()
    workbook_paths = [WORKBOOK_PATH, MAIN_WORKBOOK_PATH]

    for wb_path in workbook_paths:
        if not os.path.isfile(wb_path):
            continue
        try:
            wb = load_workbook(wb_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                if sheet_name.lower().strip() in {"cost centre", "to receive", "sheet2"}:
                    continue
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell is None:
                            continue
                        s = str(cell).strip()
                        m = pattern.match(s)
                        if m:
                            used_seqs.add(int(m.group(1)))
            try:
                wb.close()
            except Exception:
                pass
        except PermissionError:
            print(f"  [PO seq] Workbook locked, skipping: {os.path.basename(wb_path)}")
        except Exception as e:
            print(f"  [PO seq] Could not scan {os.path.basename(wb_path)}: {e}")

    return max(used_seqs) + 1 if used_seqs else 1


def generate_po_number(cost_centre: str, initials: str, seq_base: int = None) -> str:
    """
    v2.0.0: Sequence is now global per day across all cost centres.
    seq_base parameter is IGNORED (kept for backward compatibility).
    """
    now = datetime.now()
    seq = find_next_po_seq(
        cost_centre=cost_centre,
        initials=initials,
        day=now.day,
        month=now.month,
        year_2digit=now.year % 100,
    )
    return f"{cost_centre}-{initials}{now.day:02d}{now.month:02d}{seq}{now.year % 100:02d}"

# ========= SUPPLIERS =========
def load_suppliers_db(path: str):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"suppliers": []}


def pick_supplier(path: str, query: str):
    db = load_suppliers_db(path)
    suppliers = db.get("suppliers", [])

    q_raw = (query or "").strip()

    if q_raw.lower() == "manual":
        name = input_text("Manual supplier name (or 'back'): ", allow_blank=False, allow_back=True)
        if name == BACK:
            return BACK
        return {"name": name}

    if q_raw.lower() == "add":
        name = input_text("New supplier name (or 'back'): ", allow_blank=False, allow_back=True)
        if name == BACK:
            return BACK
        return {"name": name}

    if not q_raw:
        return {"name": ""}

    q_norm = normalize_text(q_raw)

    def matches(s):
        name_norm = normalize_text(s.get("name") or "")
        alias_norms = [normalize_text(a) for a in (s.get("aliases") or [])]
        supplier_id_norm = normalize_text(s.get("supplier_id") or "")
        return (
            (q_norm in name_norm)
            or any(q_norm in a for a in alias_norms)
            or (q_norm in supplier_id_norm)
        )

    found = [s for s in suppliers if matches(s)]

    if not found:
        names = [s.get("name", "") for s in suppliers if s.get("name")]
        suggestion = get_close_matches(q_raw, names, n=5, cutoff=0.60)

        if suggestion:
            print("No direct match. Did you mean:")
            for i, s in enumerate(suggestion, 1):
                print(f"  {i}) {s}")
            pick = input_text("Pick 1-5, Enter for manual, or type 'back': ", allow_blank=True, allow_back=True)
            if pick == BACK:
                return BACK
            if pick.isdigit() and 1 <= int(pick) <= len(suggestion):
                chosen = suggestion[int(pick) - 1]
                for sup in suppliers:
                    if sup.get("name") == chosen:
                        return sup

        manual = input_text("No matches. Manual supplier name (or 'back'): ", allow_blank=False, allow_back=True)
        if manual == BACK:
            return BACK
        return {"name": manual}

    print("Matches:")
    for i, s in enumerate(found[:8], 1):
        print(f"  {i}) {s.get('name')} ({s.get('supplier_id','N/A')})")

    pick = input_text("Pick 1-8, Enter for first, or type 'back': ", allow_blank=True, allow_back=True)
    if pick == BACK:
        return BACK
    if pick.isdigit() and 1 <= int(pick) <= min(8, len(found)):
        return found[int(pick) - 1]
    return found[0]


# ========= TO RECEIVE (GR BACKLOG) =========
TO_RECEIVE_HEADERS = [
    "PO Number", "Supplier", "Machine", "Area", "Reason", "Mech/Elec/Other",
    "Job No", "Items", "Total Cost", "Date Ordered", "Due Date", "Cost Centre",
    "Status", "Received Date", "Invoice Number", "Notes",
]


def ensure_to_receive_sheet(wb):
    if TO_RECEIVE_SHEET in wb.sheetnames:
        ws = wb[TO_RECEIVE_SHEET]
    else:
        ws = wb.create_sheet(TO_RECEIVE_SHEET)

    if ws["A1"].value is None:
        for i, h in enumerate(TO_RECEIVE_HEADERS, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
    return ws


def to_receive_col_index(header_name: str) -> int:
    return TO_RECEIVE_HEADERS.index(header_name) + 1


def find_to_receive_row_by_po(ws, po_number: str):
    target = (po_number or "").strip()
    if not target:
        return None
    po_col = to_receive_col_index("PO Number")
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=po_col).value or "").strip() == target:
            return r
    return None


# ====================================================================
# OUTLOOK EMAIL
# ====================================================================
def validate_supplier_email(supplier_obj):
    email = (supplier_obj.get("email") or "").strip()
    if not email:
        print(f"\n[Email] Supplier '{supplier_obj.get('name', 'Unknown')}' has no email address.")
        choice = input("Continue without email? (y/n): ").strip().lower()
        if choice != 'y':
            return None
        return ""
    if '@' not in email or '.' not in email.split('@')[-1]:
        print(f"[Email] Warning: '{email}' doesn't look like a valid email.")
        choice = input("Continue with this email? (y/n): ").strip().lower()
        if choice != 'y':
            return None
    return email


def _extract_first_name(full_name: str) -> str:
    name = (full_name or "").strip()
    if not name:
        return "Sir/Madam"
    parts = name.split()
    return parts[0] if parts else "Sir/Madam"


def open_outlook_with_po(pdf_path: str, po_number: str, supplier_obj: dict,
                         order: dict, totals: dict) -> bool:
    if not AUTO_OPEN_OUTLOOK:
        return False
    if not OUTLOOK_AVAILABLE:
        print("\n[Email] pywin32 not available. Install with:")
        print(f'  "{sys.executable}" -m pip install pywin32')
        print("Skipping Outlook email.")
        return False
    if not os.path.isfile(pdf_path):
        print(f"[Email] PDF not found at {pdf_path}. Skipping email.")
        return False

    supplier_email = validate_supplier_email(supplier_obj)
    if supplier_email is None:
        print("You can manually attach the PDF and email the supplier.")
        return False

    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)

        if supplier_email:
            mail.To = supplier_email
        mail.Subject = EMAIL_SUBJECT_TEMPLATE.format(po_number=po_number)

        supplier_contact = (supplier_obj.get("contact_name") or "").strip()
        if not supplier_contact:
            supplier_contact = (supplier_obj.get("name") or "Sir/Madam").strip()
        supplier_first_name = _extract_first_name(supplier_contact)

        due_date = order.get("due_raw", "TBC") or "TBC"
        items_summary = order.get("items_summary", "See attached PDF")
        net_total = totals.get("net_total", 0.0)

        body_text = EMAIL_BODY_TEMPLATE.format(
            supplier_first_name=supplier_first_name,
            po_number=po_number,
            supplier_name=supplier_obj.get("name", "").strip() or "N/A",
            items_summary=items_summary,
            net_total=net_total,
            due_date=due_date,
        )

        mail.Attachments.Add(os.path.abspath(pdf_path))
        mail.Display()

        body_html_lines = body_text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        body_html_lines = body_html_lines.replace("\n", "<br>")
        mail.HTMLBody = (
            '<html><body style="font-family: Calibri, Arial, sans-serif; font-size: 11pt;">'
            f'{body_html_lines}'
            '</body></html>'
        )

        print(f"[Email] Outlook email opened with PO {po_number} attached.")
        if supplier_email:
            print(f"[Email] To: {supplier_email}")
        else:
            print("[Email] No supplier email found — To field is empty. Fill it manually.")
        return True

    except Exception as e:
        print(f"[Email] Failed to open Outlook: {e}")
        print("Make sure Outlook is running. You can attach the PDF manually.")
        return False


# ====================================================================
# PDF GENERATION — VARIANT D: CLEAN MINIMAL (v1.9.8)
# ====================================================================
def try_register_dejavusans():
    font_paths = [
        r"C:\Windows\Fonts\DejaVuSans.ttf",
        r"C:\Windows\Fonts\dejavusans.ttf",
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\Arial.ttf",
    ]
    for p in font_paths:
        if os.path.isfile(p):
            try:
                pdfmetrics.registerFont(TTFont("BodyFont", p))
                return "BodyFont"
            except Exception:
                pass
    return None


def require_reportlab_or_abort():
    if REPORTLAB_AVAILABLE:
        return
    print("\nERROR: ReportLab is not available in this Python environment, so PDF cannot be created.")
    print("Install it into THIS Python (venv) like this:")
    print(f'  "{sys.executable}" -m pip install reportlab')
    print("Then run the script again.")
    raise RuntimeError("ReportLab missing")


def _pdf_draw_section_banner(c, x, y, w, h, text, bg_color, text_color, font_bold, font_size):
    c.setFillColor(bg_color)
    c.rect(x, y, w, h, fill=1, stroke=0)
    c.setFillColor(text_color)
    c.setFont(font_bold, font_size)
    c.drawString(x + 5, y + (h - font_size) / 2, text)


def _pdf_draw_detail_row(c, x, y, w, h, bg_color):
    c.setFillColor(bg_color)
    c.rect(x, y, w, h, fill=1, stroke=0)


def _pdf_wrap_text(c_obj, text, font_name, font_size, max_width):
    """Split text into lines that fit within max_width pixels."""
    if not text:
        return [""]
    words = text.split()
    lines = []
    current_line = ""
    for word in words:
        test = f"{current_line} {word}".strip()
        if c_obj.stringWidth(test, font_name, font_size) <= max_width:
            current_line = test
        else:
            if current_line:
                lines.append(current_line)
            # If a single word is wider than max_width, put it on its own line
            current_line = word
    if current_line:
        lines.append(current_line)
    return lines if lines else [""]


def generate_po_pdf(pdf_path: str, po_number: str, supplier_obj: dict,
                    order: dict, items: list, totals: dict) -> bool:
    require_reportlab_or_abort()

    try:
        ensure_dir(os.path.dirname(pdf_path))

        header_bg  = PDF_HEADER_BG  or HexColor('#2C3E50')
        accent     = PDF_ACCENT     or HexColor('#2980B9')
        light_bg   = PDF_LIGHT_BG   or HexColor('#ECF0F1')
        dark_text  = PDF_DARK_TEXT   or HexColor('#2C3E50')

        custom_font = try_register_dejavusans()
        font_normal = custom_font or "Helvetica"
        font_bold   = "Helvetica-Bold"

        c = canvas.Canvas(pdf_path, pagesize=A4)
        w, h = A4
        margin_left   = 30
        margin_right  = 30
        content_width = w - margin_left - margin_right

        header_bottom = h - 80

        if os.path.isfile(LOGO_PATH):
            try:
                logo_img = ImageReader(LOGO_PATH)
                iw, ih = logo_img.getSize()
                max_logo_w, max_logo_h = 150, 55
                scale = min(max_logo_w / iw, max_logo_h / ih)
                draw_w = iw * scale
                draw_h = ih * scale
                logo_x = margin_left
                logo_y = h - 20 - draw_h
                c.drawImage(LOGO_PATH, logo_x, logo_y, width=draw_w, height=draw_h, mask='auto')
            except Exception as e:
                print(f"[PDF] Logo could not be loaded: {e}")
                c.setFillColor(dark_text)
                c.setFont(font_bold, 16)
                c.drawString(margin_left, h - 40, SHEARD_COMPANY_NAME)
        else:
            c.setFillColor(dark_text)
            c.setFont(font_bold, 16)
            c.drawString(margin_left, h - 40, SHEARD_COMPANY_NAME)

        c.setFillColor(dark_text)
        c.setFont(font_bold, 10)
        c.drawRightString(w - margin_right, h - 25, "PURCHASE ORDER")
        c.setFont(font_bold, 20)
        c.drawRightString(w - margin_right, h - 48, po_number)
        c.setFont(font_normal, 9)
        c.drawRightString(w - margin_right, h - 62, f"Date: {datetime.today().strftime('%d/%m/%Y')}")

        c.setStrokeColor(accent)
        c.setLineWidth(2)
        c.line(margin_left, header_bottom, w - margin_right, header_bottom)
        c.setStrokeColor(header_bg)
        c.setLineWidth(0.5)
        c.line(margin_left, header_bottom - 3, w - margin_right, header_bottom - 3)

        y_section = header_bottom - 25
        col_w = (content_width - 20) / 3
        banner_h_small = 18
        line_h = 13

        x_sup = margin_left
        _pdf_draw_section_banner(c, x_sup, y_section - 3, col_w, banner_h_small, "Supplier:", accent, white, font_bold, 10)

        supplier_name = (supplier_obj.get("name") or order.get("supplier") or "").strip()
        supplier_addr = supplier_obj.get("address_lines") or []
        supplier_email = supplier_obj.get("email", "")
        supplier_phone = supplier_obj.get("phone", "")

        c.setFillColor(dark_text)
        c.setFont(font_normal, 9)
        yy = y_section - 20
        c.drawString(x_sup + 5, yy, supplier_name)
        yy -= line_h
        for line in supplier_addr[:5]:
            c.drawString(x_sup + 5, yy, str(line))
            yy -= line_h
        if supplier_email:
            c.drawString(x_sup + 5, yy, supplier_email)
            yy -= line_h
        if supplier_phone:
            c.drawString(x_sup + 5, yy, supplier_phone)

        x_bill = x_sup + col_w + 10
        _pdf_draw_section_banner(c, x_bill, y_section - 3, col_w, banner_h_small, "Bill To:", accent, white, font_bold, 10)
        c.setFillColor(dark_text)
        c.setFont(font_normal, 9)
        yy = y_section - 20
        for line in SHEARD_BILL_TO:
            c.drawString(x_bill + 5, yy, line)
            yy -= line_h

        x_del = x_bill + col_w + 10
        _pdf_draw_section_banner(c, x_del, y_section - 3, col_w, banner_h_small, "Deliver To:", accent, white, font_bold, 10)
        c.setFillColor(dark_text)
        c.setFont(font_normal, 9)
        yy = y_section - 20
        for line in SHEARD_DELIVER_TO:
            c.drawString(x_del + 5, yy, line)
            yy -= line_h

        y_contact = y_section - 100
        _pdf_draw_section_banner(c, margin_left, y_contact - 3, content_width, banner_h_small, "Contact:", accent, white, font_bold, 10)
        c.setFillColor(dark_text)
        c.setFont(font_normal, 10)
        c.drawString(margin_left + 5, y_contact - 20, f"{CONTACT_NAME} | {CONTACT_PHONE} | {CONTACT_EMAIL}")

        y_det = y_contact - 50
        row_h = 18

        _pdf_draw_detail_row(c, margin_left, y_det - 3, content_width, row_h, light_bg)
        c.setFillColor(dark_text)
        x_pos = margin_left + 5
        for label, key in [("Machine:", "machine"), ("Area:", "area"), ("Reason:", "reason"), ("Mech/Elec/Other:", "mech_elec")]:
            c.setFont(font_bold, 9)
            c.drawString(x_pos, y_det, label)
            lw = c.stringWidth(label, font_bold, 9)
            c.setFont(font_normal, 9)
            val = str(order.get(key, ""))
            c.drawString(x_pos + lw + 3, y_det, val)
            x_pos += lw + c.stringWidth(val, font_normal, 9) + 18

        y_det2 = y_det - 22
        _pdf_draw_detail_row(c, margin_left, y_det2 - 3, content_width, row_h, light_bg)
        c.setFillColor(dark_text)
        x_pos = margin_left + 5
        for label, key in [("Job No:", "job_no"), ("Cost Centre:", "cost_centre"), ("Due Date:", "due_raw")]:
            c.setFont(font_bold, 9)
            c.drawString(x_pos, y_det2, label)
            lw = c.stringWidth(label, font_bold, 9)
            c.setFont(font_normal, 9)
            val = str(order.get(key, ""))
            c.drawString(x_pos + lw + 3, y_det2, val)
            x_pos += lw + c.stringWidth(val, font_normal, 9) + 18

        y_items = y_det2 - 35
        c.setFillColor(accent)
        c.setFont(font_bold, 12)
        c.drawString(margin_left, y_items, "Items:")

        article_number = supplier_obj.get("article_number", "")
        if article_number:
            y_article = y_items - 18
            c.setFillColor(dark_text)
            c.setFont(font_bold, 10)
            c.drawString(margin_left, y_article, "Article Number:")
            c.setFont(font_normal, 10)
            c.drawString(margin_left + 100, y_article, str(article_number))
            y_items = y_article

        y_table_header = y_items - 25
        c.setFillColor(header_bg)
        c.rect(margin_left, y_table_header - 5, content_width, 20, fill=1, stroke=0)
        c.setFillColor(white)
        c.setFont(font_bold, 9)
        # v1.9.8: Column positions — Part No gets more space
        col_qty_x = margin_left + 5
        col_partno_x = margin_left + 38
        col_desc_x = margin_left + 130
        col_unit_x = w - margin_right - 100
        col_line_x = w - margin_right - 5

        c.drawString(col_qty_x, y_table_header, "Qty")
        c.drawString(col_partno_x, y_table_header, "Part No")
        c.drawString(col_desc_x, y_table_header, "Description")
        c.drawRightString(col_unit_x, y_table_header, "Unit £")
        c.drawRightString(col_line_x, y_table_header, "Line £")

        # v1.9.8: Item rows with full Part No and wrapped Description
        y_row = y_table_header - 22
        line_height_pdf = 12
        # Available width for description text
        # col_unit_x is where the RIGHT edge of unit price text sits,
        # so we need ~60pt extra gap for the price digits themselves
        desc_max_width = col_unit_x - col_desc_x - 60

        # v1.9.13: Available width for Part No (column between partno_x and desc_x)
        partno_max_width = col_desc_x - col_partno_x - 5  # 5pt safety gap

        for idx, it in enumerate(items):
            part_no_full = (it.get("part_no", "") or "")
            desc_full = (it.get("desc", "") or "")

            # v1.9.13: Hybrid Part No sizing — try shrink first, then wrap
            #   1. Try font 9 → if fits, use it
            #   2. Try font 8 → if fits, use it
            #   3. Try font 7 → if fits, use it (minimum readable)
            #   4. Otherwise wrap at font 7 onto multiple lines
            partno_font_size = 9
            partno_lines = [part_no_full]

            if part_no_full:
                for try_size in [9, 8, 7]:
                    if c.stringWidth(part_no_full, font_normal, try_size) <= partno_max_width:
                        partno_font_size = try_size
                        partno_lines = [part_no_full]
                        break
                else:
                    # Doesn't fit even at size 7 — wrap at size 7
                    partno_font_size = 7
                    partno_lines = _pdf_wrap_text(c, part_no_full, font_normal,
                                                  partno_font_size, partno_max_width)

            # Wrap description into multiple lines if needed
            desc_lines = _pdf_wrap_text(c, desc_full, font_normal, 9, desc_max_width)

            # Row height must accommodate the LONGER of Part No lines or Description lines
            row_lines = max(1, len(desc_lines), len(partno_lines))
            row_height = row_lines * line_height_pdf + 8

            # Page break check — use actual row height
            if y_row - row_height < 80:
                c.showPage()
                y_row = h - 60

            # Alternating row background — scaled to actual height
            if idx % 2 == 0:
                _pdf_draw_detail_row(c, margin_left,
                                     y_row - row_height + line_height_pdf - 3,
                                     content_width, row_height, light_bg)

            c.setFillColor(dark_text)
            c.setFont(font_normal, 9)

            # Qty — first line
            c.drawString(col_qty_x, y_row, str(it.get("qty", "")))

            # v1.9.13: Part No — possibly shrunk and/or multi-line
            c.setFont(font_normal, partno_font_size)
            for li, pn_line in enumerate(partno_lines):
                c.drawString(col_partno_x, y_row - (li * line_height_pdf), pn_line)
            c.setFont(font_normal, 9)  # Reset to default for other cells

            # Description — multi-line wrap
            for li, desc_line in enumerate(desc_lines):
                c.drawString(col_desc_x, y_row - (li * line_height_pdf), desc_line)

            # Prices — first line
            c.drawRightString(col_unit_x, y_row,
                              f"{float(it.get('unit_price', 0.0)):.2f}")
            c.drawRightString(col_line_x, y_row,
                              f"{float(it.get('line_total', 0.0)):.2f}")

            y_row -= row_height

        y_totals = y_row - 20
        net_total   = float(totals.get("net_total", 0.0))
        vat_rate    = float(totals.get("vat_rate", VAT_RATE_DEFAULT))
        vat_amount  = float(totals.get("vat_amount", 0.0))
        gross_total = float(totals.get("gross_total", net_total + vat_amount))

        c.setFillColor(dark_text)
        c.setFont(font_bold, 10)
        c.drawRightString(w - margin_right - 100, y_totals, "Net Total £:")
        c.setFont(font_normal, 10)
        c.drawRightString(w - margin_right - 5, y_totals, f"{net_total:.2f}")

        y_totals -= 18
        c.setFont(font_bold, 10)
        c.drawRightString(w - margin_right - 100, y_totals, f"VAT ({vat_rate*100:.0f}%) £:")
        c.setFont(font_normal, 10)
        c.drawRightString(w - margin_right - 5, y_totals, f"{vat_amount:.2f}")

        y_totals -= 24
        gross_banner_w = 200
        gross_banner_x = w - margin_right - gross_banner_w
        c.setFillColor(header_bg)
        c.rect(gross_banner_x, y_totals - 7, gross_banner_w, 24, fill=1, stroke=0)
        c.setFillColor(white)
        c.setFont(font_bold, 11)
        c.drawRightString(w - margin_right - 100, y_totals, "Gross Total £:")
        c.drawRightString(w - margin_right - 5, y_totals, f"{gross_total:.2f}")

        c.save()
        return os.path.isfile(pdf_path) and os.path.getsize(pdf_path) > 0

    except Exception as e:
        print(f"PDF generation failed: {e}")
        import traceback
        traceback.print_exc()
        return False


# ========= INPUT FLOW =========
def prompt_items():
    print("\nEnter line items. Leave Quantity blank to finish. Type 'back' to go back.")
    items = []

    while True:
        qty_raw = input_positive_int("Quantity (blank to finish): ", allow_blank=True, allow_back=True)
        if qty_raw == BACK:
            if items:
                removed = items.pop()
                print(f"Removed last item: x{removed['qty']} {removed['desc']}")
                continue
            return BACK
        if qty_raw == "":
            break

        qty = int(qty_raw)

        # v1.9.7: Part number — optional, Enter to skip
        part_no = input_text("Part number (optional, Enter to skip): ", allow_blank=True, allow_back=True)
        if part_no == BACK:
            continue

        desc = input_text("Item description: ", allow_blank=False, allow_back=True)
        if desc == BACK:
            continue

        unit_raw = input_money("Unit price (£) (blank if you only know line total): ", allow_blank=True, allow_back=True)
        if unit_raw == BACK:
            continue

        if unit_raw == "":
            line_raw = input_money("Line total (£): ", allow_blank=False, allow_back=True)
            if line_raw == BACK:
                continue
            line_total = float(parse_currency_to_float(line_raw) or 0.0)
            unit = (line_total / qty) if qty else 0.0
        else:
            unit = float(parse_currency_to_float(unit_raw) or 0.0)
            line_total = qty * unit

        # v1.9.7: part_no stored in item dict
        items.append({
            "part_no": part_no,
            "desc": desc,
            "qty": qty,
            "unit_price": float(unit),
            "line_total": float(line_total),
        })

    if not items:
        items.append({"part_no": "", "desc": "N/A", "qty": 1, "unit_price": 0.0, "line_total": 0.0})

    net_total = sum(i["line_total"] for i in items)

    while True:
        if PROMPT_VAT_RATE_EACH_PO:
            raw = input_text(f"VAT rate % (Enter for {int(VAT_RATE_DEFAULT*100)}): ", allow_blank=True, allow_back=True)
            if raw == BACK:
                print("Type VAT rate or press Enter for default.")
                continue
            if raw == "":
                vat_rate = VAT_RATE_DEFAULT
                break
            raw = raw.replace("%", "").strip()
            try:
                vat_rate = float(raw) / 100.0
                if vat_rate < 0:
                    print("VAT cannot be negative.")
                    continue
                break
            except ValueError:
                print("Invalid VAT. Try again.")
        else:
            vat_rate = VAT_RATE_DEFAULT
            break

    vat_amount = net_total * vat_rate
    gross_total = net_total + vat_amount

    totals = {
        "net_total": float(net_total),
        "vat_rate": float(vat_rate),
        "vat_amount": float(vat_amount),
        "gross_total": float(gross_total),
    }
    return items, totals


def collect_po_data():
    print("\n=== CREATE NEW PURCHASE ORDER (PO) ===")
    order = {}
    supplier_obj = {"name": ""}

    steps = [
        ("machine", None),
        ("area", None),
        ("reason",       lambda: input_with_autocomplete("Reason: ",           KNOWN_REASONS,   allow_blank=False)),
        ("mech_elec",    lambda: input_with_autocomplete("Mech/Elec/Other: ",  KNOWN_MECH_ELEC, allow_blank=False)),
        ("supplier", None),
        ("items", None),
        ("job_no", lambda: input_text("Job No (optional): ", allow_blank=True, allow_back=True)),
        ("due_raw", lambda: input_text("Due date (dd/mm/yyyy) or blank: ", allow_blank=True, allow_back=True)),
        ("savings_raw", lambda: input_text("Cost initiatives / Savings £ (optional): ", allow_blank=True, allow_back=True)),
        ("orig_supplier", lambda: input_text("Original supplier (optional): ", allow_blank=True, allow_back=True)),
    ]

    i = 0
    items = None
    totals = None

    while i < len(steps):
        key, fn = steps[i]

        if key == "machine":
            val = input_with_autocomplete("Machine: ", KNOWN_MACHINES, allow_blank=False)
            if val == BACK:
                i = max(0, i - 1)
                continue
            order["machine"] = val
            i += 1
            continue

        if key == "area":
            machine_chosen = order.get("machine", "")
            machine_areas = [area for (m, area) in CC_LOOKUP if m == machine_chosen and area != ""]

            if machine_areas:
                print(f"\n  Areas for '{machine_chosen}':")
                for idx, area_name in enumerate(machine_areas, 1):
                    cc_code = CC_LOOKUP.get((machine_chosen, area_name), "?")
                    print(f"    {idx}) {area_name}  (CC: {cc_code})")
                base_cc = CC_LOOKUP.get((machine_chosen, ""), "?")
                print(f"    0) — Base machine only  (CC: {base_cc})")
                print(f"    M) Manual / other area")

                pick = input("  Pick number (Enter for 0, or 'back'): ").strip()

                if is_back(pick):
                    i = max(0, i - 1)
                    continue
                if pick == "" or pick == "0":
                    order["area"] = ""
                    i += 1
                    continue
                if pick.upper() == "M":
                    area_val = input_with_autocomplete("Area (manual): ", KNOWN_AREAS, allow_blank=False)
                    if area_val == BACK:
                        continue
                    order["area"] = area_val
                    i += 1
                    continue
                if pick.isdigit() and 1 <= int(pick) <= len(machine_areas):
                    order["area"] = machine_areas[int(pick) - 1]
                    i += 1
                    continue
                print("  Invalid choice. Try again.")
                continue
            else:
                print(f"\n  No predefined areas for '{machine_chosen}'. Enter area manually:")
                area_val = input_with_autocomplete("Area: ", KNOWN_AREAS, allow_blank=True)
                if area_val == BACK:
                    i = max(0, i - 1)
                    continue
                order["area"] = area_val
                i += 1
                continue

        if key == "supplier":
            q = input_text("Supplier search (type 'manual' or 'add'): ", allow_blank=False, allow_back=True)
            if q == BACK:
                i = max(0, i - 1)
                continue
            sup = pick_supplier(SUPPLIERS_JSON_PATH, q)
            if sup == BACK:
                continue
            supplier_obj = sup
            order["supplier"] = (supplier_obj.get("name", "") or "").strip()
            i += 1
            continue

        if key == "items":
            res = prompt_items()
            if res == BACK:
                i = max(0, i - 1)
                continue
            items, totals = res
            # v1.9.7: include part_no in summary when present
            summary_parts = []
            for it in items:
                pn = it.get("part_no", "")
                prefix = f"[{pn}] " if pn else ""
                summary_parts.append(f"x{it['qty']} {prefix}{it['desc']}")
            order["items_summary"] = "; ".join(summary_parts)[:250]
            i += 1
            continue

        val = fn()
        if val == BACK:
            i = max(0, i - 1)
            continue

        if key == "due_raw" and val:
            if parse_ddmmyyyy_to_date(val) is None:
                print("Invalid date. Use dd/mm/yyyy or leave blank. Type 'back' to return.")
                continue

        order[key] = val
        i += 1

    return order, supplier_obj, items, totals


# ========= WRITE FUNCTIONS =========
def write_parts_order_row(ws_parts, header_row, colmap, order: dict, po_number: str, cost_centre: str, net_total: float):
    data_start = header_row + 1
    row = find_first_empty_row(ws_parts, colmap["Machine"], data_start)

    ws_parts[f"{colmap['DateOrdered']}{row}"] = datetime.today().date()
    ws_parts[f"{colmap['DateOrdered']}{row}"].number_format = "DD/MM/YYYY"
    ws_parts[f"{colmap['Machine']}{row}"] = order["machine"]
    ws_parts[f"{colmap['Area']}{row}"] = order["area"]
    ws_parts[f"{colmap['Reason']}{row}"] = order["reason"]
    ws_parts[f"{colmap['MechElecOther']}{row}"] = order["mech_elec"]
    ws_parts[f"{colmap['Supplier']}{row}"] = order["supplier"]
    ws_parts[f"{colmap['PartsOrdered']}{row}"] = order["items_summary"]

    parts_cell = ws_parts[f"{colmap['PartsOrdered']}{row}"]
    parts_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # v3.1.2: auto-fit row height for wrapped parts text.
    # openpyxl sets wrap_text but doesn't expand row height — text wraps
    # inside the cell but only the first line is visible at default 15pt.
    # We estimate visual lines from char count + semicolon separators
    # and set row height accordingly.
    items_text = order.get("items_summary", "") or ""
    if items_text:
        chars_per_line = 70  # approximate column F width in characters
        items = items_text.split("; ")
        total_lines = 0
        for item in items:
            n_chars = len(item)
            total_lines += max(1, (n_chars + chars_per_line - 1) // chars_per_line)
        # 15pt per line is Excel's default row height
        ws_parts.row_dimensions[row].height = max(15, total_lines * 15)

    ws_parts[f"{colmap['JobNo']}{row}"] = order.get("job_no", "")
    ws_parts[f"{colmap['PONumber']}{row}"] = po_number
    ws_parts[f"{colmap['CostCentre']}{row}"] = cost_centre

    ws_parts[f"{colmap['Cost']}{row}"] = float(net_total)
    ws_parts[f"{colmap['Cost']}{row}"].number_format = "£#,##0.00"

    savings_val = parse_currency_to_float(order.get("savings_raw", ""))
    if savings_val is not None:
        ws_parts[f"{colmap['Savings']}{row}"] = float(savings_val)
        ws_parts[f"{colmap['Savings']}{row}"].number_format = "£#,##0.00"

    if order.get("orig_supplier"):
        ws_parts[f"{colmap['OriginalSupplier']}{row}"] = order["orig_supplier"]

    due_dt = parse_ddmmyyyy_to_date(order.get("due_raw", ""))
    if due_dt:
        ws_parts[f"{colmap['DueDate']}{row}"] = due_dt
        ws_parts[f"{colmap['DueDate']}{row}"].number_format = "DD/MM/YYYY"
    elif order.get("due_raw"):
        ws_parts[f"{colmap['DueDate']}{row}"] = order["due_raw"]

    return row


def add_to_receive_row(ws_recv, po_number: str, order: dict, cost_centre: str, net_total: float, totals: dict):
    row = ws_recv.max_row + 1

    def set_cell(header, value):
        ws_recv.cell(row=row, column=to_receive_col_index(header), value=value)

    set_cell("PO Number", po_number)
    set_cell("Supplier", order["supplier"])
    set_cell("Machine", order["machine"])
    set_cell("Area", order["area"])
    set_cell("Reason", order["reason"])
    set_cell("Mech/Elec/Other", order["mech_elec"])
    set_cell("Job No", order.get("job_no", ""))
    set_cell("Items", order["items_summary"])

    cost_cell = ws_recv.cell(row=row, column=to_receive_col_index("Total Cost"), value=float(net_total))
    cost_cell.number_format = "£#,##0.00"

    date_cell = ws_recv.cell(row=row, column=to_receive_col_index("Date Ordered"), value=datetime.today().date())
    date_cell.number_format = "DD/MM/YYYY"

    due_val = parse_ddmmyyyy_to_date(order.get("due_raw", "")) or order.get("due_raw", "")
    due_cell = ws_recv.cell(row=row, column=to_receive_col_index("Due Date"), value=due_val)
    if isinstance(due_val, (date, datetime)):
        due_cell.number_format = "DD/MM/YYYY"

    set_cell("Cost Centre", cost_centre)
    set_cell("Status", "Ordered")
    set_cell("Received Date", "")
    set_cell("Invoice Number", "")
    set_cell("Notes", f"VAT {totals.get('vat_rate', VAT_RATE_DEFAULT)*100:.0f}% | Gross £{totals.get('gross_total', 0.0):.2f}")

    return row


def mark_received(wb, po_number: str, invoice_number: str = ""):
    if TO_RECEIVE_SHEET not in wb.sheetnames:
        print("To Receive sheet not found.")
        return

    ws_recv = wb[TO_RECEIVE_SHEET]
    recv_row = find_to_receive_row_by_po(ws_recv, po_number)
    if not recv_row:
        print("PO number not found in To Receive.")
        return

    ws_recv.cell(row=recv_row, column=to_receive_col_index("Status"), value="Received")
    rd_cell = ws_recv.cell(row=recv_row, column=to_receive_col_index("Received Date"), value=datetime.today().date())
    rd_cell.number_format = "DD/MM/YYYY"

    if invoice_number:
        ws_recv.cell(row=recv_row, column=to_receive_col_index("Invoice Number"), value=invoice_number)

    for sheet_name in wb.sheetnames:
        if sheet_name in {TO_RECEIVE_SHEET, SHEET_COST_CENTRE}:
            continue
        ws = wb[sheet_name]
        header_row, colmap = detect_header_row_and_columns(ws)
        if not colmap or "PONumber" not in colmap or "DeliveredDate" not in colmap:
            continue

        start_row = (header_row or 2) + 1
        for r in range(start_row, ws.max_row + 1):
            val = str(ws[f"{colmap['PONumber']}{r}"].value or "").strip()
            if val == po_number:
                dd_cell = ws[f"{colmap['DeliveredDate']}{r}"]
                dd_cell.value = datetime.today().date()
                dd_cell.number_format = "DD/MM/YYYY"
                print(f"Delivered date updated in sheet '{sheet_name}' row {r}.")
                return

    print("Received status updated. Delivered date not updated (PO not found in monthly sheets).")


# ====================================================================
# FULL SYNC WITH MAIN SHEET — v1.9.4 (major refactor)
# ====================================================================
GREEN_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")


def _parse_date_flexible(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def _is_valid_month_sheet(sheet_name, min_year=SYNC_FROM_YEAR):
    s = sheet_name.strip()
    parts = s.split()
    if len(parts) >= 2:
        try:
            year = int(parts[-1])
            return year >= min_year
        except ValueError:
            pass
    return False


def _read_all_orders_from_main(main_wb):
    all_orders = {}
    no_po_orders = {}
    orders_by_sheet = {}
    nopo_by_sheet = {}
    duplicate_warnings = []   # v1.9.10 — true errors (same PO, same supplier)
    split_warnings = []        # v1.9.12 — split orders (same PO, diff supplier)
    skip_sheets = {"cost centre", "sheet2", "to receive"}

    valid_sheets = []
    skipped_sheets = []
    for sheet_name in main_wb.sheetnames:
        if sheet_name.lower().strip() in skip_sheets:
            continue
        if _is_valid_month_sheet(sheet_name):
            valid_sheets.append(sheet_name)
        else:
            skipped_sheets.append(sheet_name)

    if skipped_sheets:
        print(f"  Skipping {len(skipped_sheets)} older/non-month sheets (before {SYNC_FROM_YEAR})")
    print(f"  Scanning {len(valid_sheets)} sheets: {', '.join(valid_sheets)}")

    for sheet_name in valid_sheets:
        print(f"    Reading '{sheet_name}'...", end=" ", flush=True)

        try:
            ws = main_wb[sheet_name]

            rows_data = []
            for row in ws.iter_rows(values_only=True):
                rows_data.append(row)

            if len(rows_data) < 2:
                print("empty sheet, skipped.")
                continue

            header_row_idx = None
            col_map = {}

            for r_idx in range(min(12, len(rows_data))):
                row = rows_data[r_idx]
                score = 0
                temp_map = {}

                for c_idx, cell_val in enumerate(row):
                    norm = normalize_text(cell_val)
                    raw_val = str(cell_val or "").strip()
                    if not norm and not raw_val:
                        continue

                    if ("£" in raw_val and len(raw_val) <= 10
                            and "Cost" not in temp_map):
                        temp_map["Cost"] = c_idx
                        score += 1
                        continue

                    for field, synonyms in HEADER_SYNONYMS.items():
                        if norm in synonyms and field not in temp_map:
                            temp_map[field] = c_idx
                            score += 1
                            break

                if score > len(col_map):
                    col_map = temp_map
                    header_row_idx = r_idx

            if not col_map or "PONumber" not in col_map:
                print("no PO column found, skipped.")
                continue

            if "Cost" not in col_map:
                mapped_cols = set(col_map.values())
                max_c = max(len(r) for r in rows_data if r) if rows_data else 0

                best_cost_col = None
                best_numeric_count = 0

                for c_idx in range(max_c):
                    if c_idx in mapped_cols:
                        continue

                    numeric_count = 0
                    for d_idx in range(header_row_idx + 1, min(header_row_idx + 20, len(rows_data))):
                        d_row = rows_data[d_idx]
                        if not d_row or c_idx >= len(d_row):
                            continue
                        val = d_row[c_idx]
                        if val is None:
                            continue
                        if isinstance(val, (int, float)):
                            numeric_count += 1
                        elif isinstance(val, str):
                            parsed = parse_currency_to_float(val)
                            if parsed is not None:
                                numeric_count += 1

                    if numeric_count >= 3 and numeric_count > best_numeric_count:
                        best_numeric_count = numeric_count
                        best_cost_col = c_idx

                if best_cost_col is not None:
                    col_map["Cost"] = best_cost_col
                    print(f"(auto-detected Cost at col index {best_cost_col}) ", end="", flush=True)

            if DEBUG:
                mapped_fields = {f: c_idx for f, c_idx in col_map.items()}
                print(f"\n      Column map: {mapped_fields}")

                print(f"      --- RAW CELL DUMP (first 3 data rows) ---")
                header_row = rows_data[header_row_idx] if header_row_idx is not None else ()
                print(f"      HEADER ROW (idx {header_row_idx}):")
                for field, c_idx in sorted(col_map.items(), key=lambda x: x[1]):
                    raw_h = header_row[c_idx] if c_idx < len(header_row) else "OUT_OF_RANGE"
                    print(f"        col[{c_idx}] = {repr(raw_h):30} → mapped as '{field}'")

                dump_count = 0
                for d_idx in range(header_row_idx + 1, len(rows_data)):
                    d_row = rows_data[d_idx]
                    if not d_row:
                        continue
                    po_c = col_map.get("PONumber")
                    if po_c is not None and po_c < len(d_row) and not is_blank_cell(d_row[po_c]):
                        dump_count += 1
                        po_raw = d_row[po_c]
                        cost_c = col_map.get("Cost")
                        cost_raw = d_row[cost_c] if (cost_c is not None and cost_c < len(d_row)) else "NO_COL"
                        savings_c = col_map.get("Savings")
                        savings_raw = d_row[savings_c] if (savings_c is not None and savings_c < len(d_row)) else "NO_COL"
                        print(f"      ROW {d_idx}: PO={repr(po_raw):25} "
                              f"Cost[{cost_c}]={repr(cost_raw):20} "
                              f"type={type(cost_raw).__name__:8} "
                              f"Savings[{savings_c}]={repr(savings_raw)}")
                        if dump_count >= 3:
                            break
                print(f"      --- END RAW DUMP ---")
                print(f"      ", end="")

            count = 0
            zero_cost_count = 0
            date_anomaly_count = 0
            nopo_count = 0

            for r_idx in range(header_row_idx + 1, len(rows_data)):
                row = rows_data[r_idx]
                if not row or len(row) <= col_map["PONumber"]:
                    continue

                po_val = row[col_map["PONumber"]]
                if is_blank_cell(po_val):
                    def safe_get_nopo(field):
                        if field in col_map and col_map[field] < len(row):
                            return row[col_map[field]]
                        return None

                    nopo_supplier = str(safe_get_nopo("Supplier") or "").strip()
                    nopo_parts = str(safe_get_nopo("PartsOrdered") or "").strip()
                    nopo_cost = to_float_cell(safe_get_nopo("Cost"))

                    if nopo_supplier or nopo_parts:
                        nopo_order = {
                            "sheet": sheet_name,
                            "machine": str(safe_get_nopo("Machine") or ""),
                            "area": str(safe_get_nopo("Area") or ""),
                            "reason": str(safe_get_nopo("Reason") or ""),
                            "mech_elec": str(safe_get_nopo("MechElecOther") or ""),
                            "supplier": nopo_supplier,
                            "parts": nopo_parts,
                            "job_no": str(safe_get_nopo("JobNo") or ""),
                            "cost": nopo_cost,
                            "date_ordered": _parse_date_flexible(safe_get_nopo("DateOrdered")),
                            "due_date": _parse_date_flexible(safe_get_nopo("DueDate")),
                            "delivered_date": _parse_date_flexible(safe_get_nopo("DeliveredDate")),
                            "savings": to_float_cell(safe_get_nopo("Savings")),
                            "orig_supplier": str(safe_get_nopo("OriginalSupplier") or ""),
                        }
                        nopo_machine_norm = normalize_text(nopo_order["machine"])[:15]
                        date_str = ""
                        if nopo_order["date_ordered"]:
                            date_str = nopo_order["date_ordered"].strftime("%d%m%Y")
                        fallback_key = f"NOPO-{nopo_machine_norm}-{normalize_text(nopo_supplier)[:15]}-{date_str}"
                        no_po_orders[fallback_key] = nopo_order
                        if sheet_name not in nopo_by_sheet:
                            nopo_by_sheet[sheet_name] = {}
                        nopo_by_sheet[sheet_name][fallback_key] = nopo_order
                        nopo_count += 1
                    continue

                po_str = str(po_val).strip()
                if not po_str:
                    continue

                def safe_get(field):
                    if field in col_map and col_map[field] < len(row):
                        return row[col_map[field]]
                    return None

                cost_val = to_float_cell(safe_get("Cost"))
                date_ordered_val = _parse_date_flexible(safe_get("DateOrdered"))
                delivered_date_val = _parse_date_flexible(safe_get("DeliveredDate"))

                if cost_val == 0.0:
                    zero_cost_count += 1

                if (date_ordered_val and delivered_date_val
                        and delivered_date_val < date_ordered_val):
                    date_anomaly_count += 1

                order = {
                    "sheet": sheet_name,
                    "machine": str(safe_get("Machine") or ""),
                    "area": str(safe_get("Area") or ""),
                    "reason": str(safe_get("Reason") or ""),
                    "mech_elec": str(safe_get("MechElecOther") or ""),
                    "supplier": str(safe_get("Supplier") or ""),
                    "parts": str(safe_get("PartsOrdered") or ""),
                    "job_no": str(safe_get("JobNo") or ""),
                    "cost": cost_val,
                    "date_ordered": date_ordered_val,
                    "due_date": _parse_date_flexible(safe_get("DueDate")),
                    "delivered_date": delivered_date_val,
                    "savings": to_float_cell(safe_get("Savings")),
                    "orig_supplier": str(safe_get("OriginalSupplier") or ""),
                }

                

            
                # v1.9.12: Three categories of duplicates
                #   1. Same PO + same sheet + same supplier   = TRUE ERROR (ignore second)
                #   2. Same PO + same sheet + diff supplier   = SPLIT ORDER (sync both via composite key)
                #   3. Same PO + diff sheets                  = RECURRING SERVICE (silent, last wins)

                supplier_norm = normalize_text(order["supplier"])

                if sheet_name in orders_by_sheet and po_str in orders_by_sheet[sheet_name]:
                    existing = orders_by_sheet[sheet_name][po_str]
                    existing_supplier_norm = normalize_text(existing["supplier"])

                    if supplier_norm == existing_supplier_norm:
                        # Category 1: TRUE ERROR — same PO, same supplier, same sheet
                        duplicate_warnings.append({
                            "type": "error",
                            "po": po_str,
                            "sheet": sheet_name,
                            "first": {
                                "supplier": existing["supplier"],
                                "cost": existing["cost"],
                                "parts": existing["parts"][:50],
                            },
                            "duplicate": {
                                "supplier": order["supplier"],
                                "cost": order["cost"],
                                "parts": order["parts"][:50],
                            },
                        })
                        # Skip — keep first version
                        continue
                    else:
                        # Category 2: SPLIT ORDER — same PO, different suppliers
                        # Sync both rows; warn only as FYI
                        split_warnings.append({
                            "po": po_str,
                            "sheet": sheet_name,
                            "first": {
                                "supplier": existing["supplier"],
                                "cost": existing["cost"],
                                "parts": existing["parts"][:50],
                            },
                            "second": {
                                "supplier": order["supplier"],
                                "cost": order["cost"],
                                "parts": order["parts"][:50],
                            },
                        })
                        # Don't skip — store both under composite keys
                        composite_key = f"{po_str}|{supplier_norm}"
                        all_orders[composite_key] = order
                        orders_by_sheet[sheet_name][composite_key] = order
                        # Also reindex the first one under composite key for consistency
                        first_composite = f"{po_str}|{existing_supplier_norm}"
                        if first_composite != po_str:
                            all_orders[first_composite] = existing
                            orders_by_sheet[sheet_name][first_composite] = existing
                        count += 1
                        continue

                # Normal case — first time we see this PO on this sheet
                # (Category 3 cross-sheet recurring handled implicitly: last sheet wins in all_orders)
                all_orders[po_str] = order

                if sheet_name not in orders_by_sheet:
                    orders_by_sheet[sheet_name] = {}
                orders_by_sheet[sheet_name][po_str] = order

                count += 1

            msg = f"{count} orders found."
            if nopo_count > 0:
                msg += f" ({nopo_count} without PO number)"
            print(msg)

            if count > 0 and zero_cost_count > count * 0.3:
                print(f"      ⚠ WARNING: {zero_cost_count}/{count} orders have £0.00 cost "
                      f"— possible formula cells or wrong column mapping!")
            if date_anomaly_count > 0:
                print(f"      ⚠ WARNING: {date_anomaly_count} orders have Delivered Date "
                      f"BEFORE Date Ordered — possible column swap!")

        except Exception as e:
            print(f"ERROR reading sheet: {e}")
            continue

    # v1.9.12: POSTPROCESS — Extract split orders into separate structure.
    # Composite keys (po|supplier) must NOT leak into the rest of the sync flow.
    # all_orders gets cleaned to use only real PO numbers (with "last wins" semantics).
    # split_orders is a separate list of (sheet, real_po, supplier_norm, order_dict)
    # for special handling in sync_with_main_sheet.

    split_orders = []  # List of (sheet, real_po, supplier_norm, order_dict)
    cleaned_all_orders = {}
    cleaned_orders_by_sheet = {}

    for key, order in all_orders.items():
        if "|" in key:
            real_po, supplier_norm = key.split("|", 1)
            split_orders.append((order["sheet"], real_po, supplier_norm, order))
            # Also store under real PO (last wins — sync_with_main_sheet treats
            # split orders separately, so this is fine for non-split logic)
            cleaned_all_orders[real_po] = order
        else:
            cleaned_all_orders[key] = order

    for sheet_name, sheet_dict in orders_by_sheet.items():
        cleaned_orders_by_sheet[sheet_name] = {}
        for key, order in sheet_dict.items():
            if "|" in key:
                real_po, _ = key.split("|", 1)
                cleaned_orders_by_sheet[sheet_name][real_po] = order
            else:
                cleaned_orders_by_sheet[sheet_name][key] = order

    return (cleaned_all_orders, no_po_orders, cleaned_orders_by_sheet,
            nopo_by_sheet, duplicate_warnings, split_warnings, split_orders)


def _get_local_po_numbers(wb):
    local_pos = set()

    if TO_RECEIVE_SHEET in wb.sheetnames:
        ws = wb[TO_RECEIVE_SHEET]
        po_col_idx = to_receive_col_index("PO Number")
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=po_col_idx).value
            if not is_blank_cell(val):
                local_pos.add(str(val).strip())

    for sheet_name in wb.sheetnames:
        if sheet_name in {TO_RECEIVE_SHEET, SHEET_COST_CENTRE}:
            continue
        ws = wb[sheet_name]
        header_row, colmap = detect_header_row_and_columns(ws)
        if not colmap or "PONumber" not in colmap:
            continue
        start_row = (header_row or 2) + 1
        for r in range(start_row, ws.max_row + 1):
            val = ws[f"{colmap['PONumber']}{r}"].value
            if not is_blank_cell(val):
                local_pos.add(str(val).strip())

    return local_pos


def _get_to_receive_po_numbers(wb):
    pos = set()
    if TO_RECEIVE_SHEET in wb.sheetnames:
        ws = wb[TO_RECEIVE_SHEET]
        po_col_idx = to_receive_col_index("PO Number")
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=po_col_idx).value
            if not is_blank_cell(val):
                pos.add(str(val).strip())
    return pos


def _get_monthly_sheet_po_index(wb):
    index = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in {TO_RECEIVE_SHEET, SHEET_COST_CENTRE}:
            continue
        ws = wb[sheet_name]
        header_row, colmap = detect_header_row_and_columns(ws)
        if not colmap or "PONumber" not in colmap:
            continue
        start_row = (header_row or 2) + 1
        pos = set()
        for r in range(start_row, ws.max_row + 1):
            val = ws[f"{colmap['PONumber']}{r}"].value
            if not is_blank_cell(val):
                pos.add(str(val).strip())
        index[sheet_name] = pos
    return index


def _find_local_duplicates(wb):
    if TO_RECEIVE_SHEET not in wb.sheetnames:
        return []

    ws = wb[TO_RECEIVE_SHEET]
    po_col_idx = to_receive_col_index("PO Number")
    po_counts = {}

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=po_col_idx).value
        if not is_blank_cell(val):
            po_str = str(val).strip()
            po_counts[po_str] = po_counts.get(po_str, 0) + 1

    return [(po, cnt) for po, cnt in po_counts.items() if cnt > 1]



def _write_order_to_monthly_sheet(ws_month, cmap, h_row, o, po_str):
    start_row = (h_row or 2) + 1
    data_row = find_first_empty_row(ws_month, cmap["Machine"], start_row)

    ws_month[f"{cmap['Machine']}{data_row}"] = o["machine"]
    if "Area" in cmap:
        ws_month[f"{cmap['Area']}{data_row}"] = o["area"]
    if "Reason" in cmap:
        ws_month[f"{cmap['Reason']}{data_row}"] = o["reason"]
    if "MechElecOther" in cmap:
        ws_month[f"{cmap['MechElecOther']}{data_row}"] = o["mech_elec"]
    if "Supplier" in cmap:
        ws_month[f"{cmap['Supplier']}{data_row}"] = o["supplier"]
    if "PartsOrdered" in cmap:
        ws_month[f"{cmap['PartsOrdered']}{data_row}"] = o["parts"]
    if "JobNo" in cmap:
        ws_month[f"{cmap['JobNo']}{data_row}"] = o["job_no"]

    # v1.9.12: po_str is always a real PO here (composite keys cleaned upstream)
    ws_month[f"{cmap['PONumber']}{data_row}"] = po_str

    if "Cost" in cmap:
        cost_cell = ws_month[f"{cmap['Cost']}{data_row}"]
        cost_cell.value = float(o["cost"])
        cost_cell.number_format = "£#,##0.00"

    if o["date_ordered"] and "DateOrdered" in cmap:
        d_cell = ws_month[f"{cmap['DateOrdered']}{data_row}"]
        d_cell.value = o["date_ordered"]
        d_cell.number_format = "DD/MM/YYYY"

    if o["due_date"] and "DueDate" in cmap:
        ws_month[f"{cmap['DueDate']}{data_row}"] = o["due_date"]
        ws_month[f"{cmap['DueDate']}{data_row}"].number_format = "DD/MM/YYYY"

    if o["delivered_date"] and "DeliveredDate" in cmap:
        dd_cell = ws_month[f"{cmap['DeliveredDate']}{data_row}"]
        dd_cell.value = o["delivered_date"]
        dd_cell.number_format = "DD/MM/YYYY"
        dd_cell.fill = GREEN_FILL

    if o["savings"] and "Savings" in cmap:
        ws_month[f"{cmap['Savings']}{data_row}"] = float(o["savings"])
        ws_month[f"{cmap['Savings']}{data_row}"].number_format = "£#,##0.00"

    if o["orig_supplier"] and "OriginalSupplier" in cmap:
        ws_month[f"{cmap['OriginalSupplier']}{data_row}"] = o["orig_supplier"]

    if "CostCentre" in cmap:
        machine_canon = normalize_machine_name(o["machine"])
        area_canon = normalize_area_name(o["area"])
        cc = ""
        if (machine_canon, area_canon) in CC_LOOKUP:
            cc = CC_LOOKUP[(machine_canon, area_canon)]
        elif (machine_canon, "") in CC_LOOKUP:
            cc = CC_LOOKUP[(machine_canon, "")]
        ws_month[f"{cmap['CostCentre']}{data_row}"] = cc

    return data_row


def _delete_row_from_monthly_sheet(wb, sheet_name, po_str):
    if sheet_name not in wb.sheetnames:
        return False

    ws = wb[sheet_name]
    header_row, colmap = detect_header_row_and_columns(ws)
    if not colmap or "PONumber" not in colmap:
        return False

    start_row = (header_row or 2) + 1
    for r in range(start_row, ws.max_row + 1):
        val = str(ws[f"{colmap['PONumber']}{r}"].value or "").strip()
        if val == po_str:
            ws.delete_rows(r)
            return True
    return False


def sync_with_main_sheet():
    if not os.path.isfile(MAIN_WORKBOOK_PATH):
        print(f"\n[Sync] Main workbook not found at:\n  {MAIN_WORKBOOK_PATH}")
        print("Check the path or make sure OneDrive is synced.")
        return

    print(f"\n{'='*60}")
    print(f"  FULL SYNC WITH MAIN SHEET")
    print(f"{'='*60}")
    print(f"\n[Sync] Opening main workbook (this may take a moment)...")
    print(f"  {MAIN_WORKBOOK_PATH}", flush=True)

    try:
        main_wb = load_main_workbook_readonly(MAIN_WORKBOOK_PATH)
        print(f"  Opened OK. Sheets found: {len(main_wb.sheetnames)}")
    except (KeyboardInterrupt, Exception) as e:
        print(f"[Sync] Failed to open main workbook: {e}")
        return

    main_orders, no_po_orders, orders_by_sheet, nopo_by_sheet, duplicate_warnings, split_warnings, split_orders = _read_all_orders_from_main(main_wb)
    try:
        main_wb.close()
    except Exception:
        pass
    print(f"[Sync] Found {len(main_orders)} orders with PO + {len(no_po_orders)} without PO in main workbook.")

    # v1.9.12: FYI report for SPLIT ORDERS (same PO, different suppliers, same sheet)
    # These are synced normally via composite key — both rows preserved.
    if split_warnings:
        print(f"\n{'~'*60}")
        print(f"  ℹ  SPLIT ORDERS DETECTED (same PO, different suppliers)  ℹ")
        print(f"{'~'*60}")
        print(f"  Found {len(split_warnings)} PO number(s) shared by multiple suppliers.")
        print(f"  These ARE being synced — both rows preserved with their own costs.")
        print(f"  FYI only: each PO number should ideally be unique per supplier.\n")

        for i, sp in enumerate(split_warnings, 1):
            print(f"  {i}) PO: {sp['po']}   [Sheet: {sp['sheet']}]")
            print(f"     ✓ {sp['first']['supplier']:25} £{sp['first']['cost']:>9.2f}")
            print(f"        Parts: {sp['first']['parts']}")
            print(f"     ✓ {sp['second']['supplier']:25} £{sp['second']['cost']:>9.2f}")
            print(f"        Parts: {sp['second']['parts']}")
            print()
        print(f"{'~'*60}\n")

    # v1.9.11/v1.9.12: STOP warning for TRUE duplicates (same PO + same supplier)
    if duplicate_warnings:
        print(f"\n{'!'*60}")
        print(f"  ⚠  TRUE DUPLICATE PO NUMBERS DETECTED  ⚠")
        print(f"{'!'*60}")
        print(f"  Found {len(duplicate_warnings)} PO number(s) duplicated with the SAME supplier.")
        print(f"  These are DATA ENTRY ERRORS — same order entered twice.")
        print(f"  Sync will use the FIRST occurrence only — second is IGNORED.\n")

        for i, dup in enumerate(duplicate_warnings, 1):
            print(f"  {i}) PO: {dup['po']}   [Sheet: {dup['sheet']}]")
            print(f"     ✓ KEEPING:  {dup['first']['supplier']:25} £{dup['first']['cost']:>9.2f}")
            print(f"        Parts: {dup['first']['parts']}")
            print(f"     ✗ IGNORING: {dup['duplicate']['supplier']:25} £{dup['duplicate']['cost']:>9.2f}")
            print(f"        Parts: {dup['duplicate']['parts']}")
            print()

        print(f"  RECOMMENDATION: Delete the duplicate row on the main sheet.")
        print(f"{'!'*60}\n")

        choice = input("  Continue sync anyway? (y/n): ").strip().lower()
        if choice not in ("y", "yes"):
            print("  Sync cancelled. Fix duplicates and try again.")
            return
        print("  Continuing with first-occurrence-only data...\n")

    if main_orders:
        zero_count = sum(1 for o in main_orders.values() if o["cost"] == 0.0)
        total_count = len(main_orders)
        if total_count > 0 and zero_count > total_count * 0.5:
            print(f"\n  ⚠ {zero_count}/{total_count} orders have £0.00 cost.")
            print(f"  Checking if Cost column contains FORMULAS (data_only=False)...")
            try:
                main_wb_formulas = load_workbook(MAIN_WORKBOOK_PATH, read_only=False, data_only=False)
                for sn in main_wb_formulas.sheetnames:
                    if not _is_valid_month_sheet(sn):
                        continue
                    ws_check = main_wb_formulas[sn]
                    formula_cells = 0
                    value_cells = 0
                    none_cells = 0
                    sample_formulas = []
                    for row in ws_check.iter_rows(min_row=2, max_row=min(20, ws_check.max_row), values_only=False):
                        for cell in row:
                            raw = cell.value
                            if raw is not None and isinstance(raw, str) and raw.startswith("="):
                                formula_cells += 1
                                if len(sample_formulas) < 3:
                                    sample_formulas.append(f"    {cell.coordinate}: {raw}")
                            elif raw is not None:
                                value_cells += 1
                            else:
                                none_cells += 1

                    if formula_cells > 0:
                        print(f"\n  ✘ FOUND FORMULAS on sheet '{sn}'!")
                        print(f"    Formula cells: {formula_cells}, Value cells: {value_cells}, Empty: {none_cells}")
                        for sf in sample_formulas:
                            print(sf)
                        print(f"\n  FIX: Open the main workbook in Excel, press Ctrl+S, then run sync again.")
                    else:
                        print(f"\n  ℹ No formulas found on '{sn}'. Check the column mapping in the RAW DUMP above.")
                    break

                try:
                    main_wb_formulas.close()
                except Exception:
                    pass
            except Exception as e:
                print(f"  Could not check for formulas: {e}")

    if not main_orders:
        print("[Sync] Main workbook appears empty. Nothing to sync.")
        return

    wb = load_or_create_workbook(WORKBOOK_PATH)
    ws_recv = ensure_to_receive_sheet(wb)
    max_col = ws_recv.max_column or len(TO_RECEIVE_HEADERS)

    po_col_idx = to_receive_col_index("PO Number")
    status_col_idx = to_receive_col_index("Status")
    recv_date_col_idx = to_receive_col_index("Received Date")

    # STEP 1: DELIVERY SYNC
    print(f"\n--- STEP 1: Sync Deliveries ---")

    delivered_main = {po: o for po, o in main_orders.items() if o["delivered_date"]}
    delivery_updated = 0
    delivery_already = 0

    for r in range(2, ws_recv.max_row + 1):
        po_val = ws_recv.cell(row=r, column=po_col_idx).value
        if is_blank_cell(po_val):
            continue

        po_str = str(po_val).strip()
        status_val = str(ws_recv.cell(row=r, column=status_col_idx).value or "").strip()

        if po_str in delivered_main:
            del_date = delivered_main[po_str]["delivered_date"]

            if status_val.lower() == "received":
                delivery_already += 1
                for c in range(1, max_col + 1):
                    ws_recv.cell(row=r, column=c).fill = GREEN_FILL
                continue

            ws_recv.cell(row=r, column=status_col_idx).value = "Received"
            recv_cell = ws_recv.cell(row=r, column=recv_date_col_idx)
            recv_cell.value = del_date
            recv_cell.number_format = "DD/MM/YYYY"

            for c in range(1, max_col + 1):
                ws_recv.cell(row=r, column=c).fill = GREEN_FILL

            delivery_updated += 1
            print(f"  ✓ {po_str} → Received {del_date.strftime('%d/%m/%Y')}")

    if delivery_updated > 0 or delivery_already > 0:
        for sheet_name in wb.sheetnames:
            if sheet_name in {TO_RECEIVE_SHEET, SHEET_COST_CENTRE}:
                continue
            ws = wb[sheet_name]
            header_row, colmap = detect_header_row_and_columns(ws)
            if not colmap or "PONumber" not in colmap or "DeliveredDate" not in colmap:
                continue
            start_row = (header_row or 2) + 1
            for r in range(start_row, ws.max_row + 1):
                po_val = str(ws[f"{colmap['PONumber']}{r}"].value or "").strip()
                if po_val in delivered_main:
                    dd_cell = ws[f"{colmap['DeliveredDate']}{r}"]
                    if is_blank_cell(dd_cell.value):
                        dd_cell.value = delivered_main[po_val]["delivered_date"]
                        dd_cell.number_format = "DD/MM/YYYY"
                    dd_cell.fill = GREEN_FILL

    print(f"  Deliveries updated: {delivery_updated}")
    print(f"  Already received: {delivery_already}")

    # STEP 2A: IMPORT MISSING ORDERS INTO TO RECEIVE
    print(f"\n--- STEP 2A: Import Missing Orders to To Receive ---")

    to_receive_pos = _get_to_receive_po_numbers(wb)
    # v1.9.12: For split orders (composite keys), check real PO number against
    # local To Receive, but preserve composite key in the dict for later use
    missing_from_to_receive = {}
    for po, o in main_orders.items():
        real_po = po.split("|")[0] if "|" in po else po
        if real_po not in to_receive_pos:
            missing_from_to_receive[po] = o
    imported_to_receive = 0

    if not missing_from_to_receive:
        print("  All main sheet orders already exist in To Receive. Nothing to import.")
    else:
        print(f"  Found {len(missing_from_to_receive)} orders on main sheet not in To Receive:")
        for idx, (po, o) in enumerate(sorted(missing_from_to_receive.items()), 1):
            delivered_mark = " ✓DEL" if o["delivered_date"] else ""
            # v1.9.12: Show real PO (not composite key) in listing
            display_po = po.split("|")[0] if "|" in po else po
            print(f"    {idx:3}) {display_po:25} | {o['supplier']:20} | £{o['cost']:>8.2f} | {o['machine']}{delivered_mark}")
            if idx >= 15 and len(missing_from_to_receive) > 15:
                print(f"    ... and {len(missing_from_to_receive) - 15} more")
                break

        choice = input(f"\n  Import all {len(missing_from_to_receive)} into To Receive? (y/n): ").strip().lower()

        if choice in ("y", "yes"):
            for po_str, o in missing_from_to_receive.items():
                row = ws_recv.max_row + 1

                def set_cell(header, value):
                    ws_recv.cell(row=row, column=to_receive_col_index(header), value=value)

                # v1.9.12: Extract real PO from composite key for split orders
                real_po = po_str.split("|")[0] if "|" in po_str else po_str
                set_cell("PO Number", real_po)
                set_cell("Supplier", o["supplier"])
                set_cell("Machine", o["machine"])
                set_cell("Area", o["area"])
                set_cell("Reason", o["reason"])
                set_cell("Mech/Elec/Other", o["mech_elec"])
                set_cell("Job No", o["job_no"])
                set_cell("Items", o["parts"])

                cost_cell = ws_recv.cell(row=row, column=to_receive_col_index("Total Cost"),
                                         value=float(o["cost"]))
                cost_cell.number_format = "£#,##0.00"

                if o["date_ordered"]:
                    date_cell = ws_recv.cell(row=row, column=to_receive_col_index("Date Ordered"),
                                              value=o["date_ordered"])
                    date_cell.number_format = "DD/MM/YYYY"

                if o["due_date"]:
                    due_cell = ws_recv.cell(row=row, column=to_receive_col_index("Due Date"),
                                             value=o["due_date"])
                    due_cell.number_format = "DD/MM/YYYY"

                machine_canon = normalize_machine_name(o["machine"])
                area_canon = normalize_area_name(o["area"])
                cc = ""
                if (machine_canon, area_canon) in CC_LOOKUP:
                    cc = CC_LOOKUP[(machine_canon, area_canon)]
                elif (machine_canon, "") in CC_LOOKUP:
                    cc = CC_LOOKUP[(machine_canon, "")]
                set_cell("Cost Centre", cc)

                if o["delivered_date"]:
                    set_cell("Status", "Received")
                    recv_cell = ws_recv.cell(row=row, column=to_receive_col_index("Received Date"),
                                              value=o["delivered_date"])
                    recv_cell.number_format = "DD/MM/YYYY"
                    set_cell("Notes", "Imported from main sheet (delivered)")
                    for c in range(1, max_col + 1):
                        ws_recv.cell(row=row, column=c).fill = GREEN_FILL
                else:
                    set_cell("Status", "Ordered")
                    set_cell("Received Date", "")
                    set_cell("Notes", "Imported from main sheet")

                set_cell("Invoice Number", "")
                imported_to_receive += 1

            print(f"  ✓ Imported {imported_to_receive} orders into To Receive")
        else:
            print("  Skipped To Receive import.")

    to_receive_cost_updated = 0
    cost_col_idx = to_receive_col_index("Total Cost")
    for r in range(2, ws_recv.max_row + 1):
        po_val = ws_recv.cell(row=r, column=po_col_idx).value
        if is_blank_cell(po_val):
            continue
        po_str = str(po_val).strip()
        if po_str not in main_orders:
            continue
        local_cost = to_float_cell(ws_recv.cell(row=r, column=cost_col_idx).value)
        main_cost = main_orders[po_str]["cost"]
        if main_cost != 0.0 and abs(local_cost - main_cost) > 0.005:
            cost_cell = ws_recv.cell(row=r, column=cost_col_idx, value=float(main_cost))
            cost_cell.number_format = "£#,##0.00"
            to_receive_cost_updated += 1
    if to_receive_cost_updated > 0:
        print(f"  ✓ Updated cost on {to_receive_cost_updated} existing To Receive rows.")

    # STEP 2B: SYNC MONTHLY SHEETS
    print(f"\n--- STEP 2B: Sync Monthly Sheets ---")

    monthly_po_index = _get_monthly_sheet_po_index(wb)

    monthly_written = 0
    monthly_existed = 0
    monthly_no_sheet = 0
    monthly_cost_updated = 0

    for sheet_name, sheet_orders in orders_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            monthly_no_sheet += len(sheet_orders)
            continue

        for po_str, o in sheet_orders.items():
            # v1.9.12: Composite keys are cleaned in _read_all_orders_from_main —
            # po_str here is always a real PO number. Split orders are handled
            # separately below (see STEP 2B-SPLIT).
            if sheet_name in monthly_po_index and po_str in monthly_po_index[sheet_name]:
                monthly_existed += 1

                ws_month = wb[sheet_name]
                h_row, cmap = detect_header_row_and_columns(ws_month)
                if cmap and "PONumber" in cmap:
                    start_row = (h_row or 2) + 1
                    for r in range(start_row, ws_month.max_row + 1):
                        if str(ws_month[f"{cmap['PONumber']}{r}"].value or "").strip() == po_str:
                            updated_fields = []

                            if "Cost" in cmap and o["cost"] != 0.0:
                                cost_cell = ws_month[f"{cmap['Cost']}{r}"]
                                local_cost = to_float_cell(cost_cell.value)
                                if abs(local_cost - o["cost"]) > 0.005:
                                    cost_cell.value = float(o["cost"])
                                    cost_cell.number_format = "£#,##0.00"
                                    monthly_cost_updated += 1
                                    updated_fields.append(f"Cost £{local_cost:.2f}→£{o['cost']:.2f}")

                            if "Savings" in cmap and o["savings"] and o["savings"] != 0.0:
                                sav_cell = ws_month[f"{cmap['Savings']}{r}"]
                                local_sav = to_float_cell(sav_cell.value)
                                if abs(local_sav - o["savings"]) > 0.005:
                                    sav_cell.value = float(o["savings"])
                                    sav_cell.number_format = "£#,##0.00"
                                    updated_fields.append(f"Savings→£{o['savings']:.2f}")

                            if "DeliveredDate" in cmap and o["delivered_date"]:
                                dd_cell = ws_month[f"{cmap['DeliveredDate']}{r}"]
                                if is_blank_cell(dd_cell.value):
                                    dd_cell.value = o["delivered_date"]
                                    dd_cell.number_format = "DD/MM/YYYY"
                                    dd_cell.fill = GREEN_FILL
                                    updated_fields.append("Delivered")

                            if updated_fields and DEBUG:
                                print(f"    ✎ {po_str} [{sheet_name}]: {', '.join(updated_fields)}")

                            break
                continue

            ws_month = wb[sheet_name]
            h_row, cmap = ensure_parts_column_map(ws_month)
            if "PONumber" not in cmap or "Machine" not in cmap:
                continue

            _write_order_to_monthly_sheet(ws_month, cmap, h_row, o, po_str)
            monthly_written += 1
            if DEBUG:
                print(f"    + {po_str} → {sheet_name} (£{o['cost']:.2f})")

    if monthly_written > 0 or monthly_cost_updated > 0:
        sheets_updated = set(sn for sn in orders_by_sheet.keys() if sn in wb.sheetnames)
        for sheet_name in sheets_updated:
            ws_month = wb[sheet_name]
            h_row, cmap = detect_header_row_and_columns(ws_month)
            if cmap:
                update_month_totals(ws_month, h_row or 2, cmap)
                auto_fit_columns(ws_month)

    # STEP 2B-SPLIT: SYNC SPLIT ORDERS (same PO, different suppliers)
    # v1.9.12: Each (real_po, supplier) combination is matched against local
    # rows on the correct sheet using BOTH po + supplier as the key.
    if split_orders:
        print(f"\n--- STEP 2B-SPLIT: Sync Split Orders ({len(split_orders)} pairs) ---")

        split_written = 0
        split_updated = 0

        for sheet_name, real_po, supplier_norm, o in split_orders:
            if sheet_name not in wb.sheetnames:
                continue

            ws_month = wb[sheet_name]
            h_row, cmap = ensure_parts_column_map(ws_month)
            if "PONumber" not in cmap or "Machine" not in cmap:
                continue

            start_row = (h_row or 2) + 1
            sup_col = cmap.get("Supplier")

            # Look for an EXISTING local row with this PO + this supplier
            found_row = None
            for r in range(start_row, ws_month.max_row + 1):
                local_po = str(ws_month[f"{cmap['PONumber']}{r}"].value or "").strip()
                if local_po != real_po:
                    continue
                local_sup = normalize_text(ws_month[f"{sup_col}{r}"].value) if sup_col else ""
                if local_sup == supplier_norm:
                    found_row = r
                    break

            if found_row is None:
                # No matching row — add new
                _write_order_to_monthly_sheet(ws_month, cmap, h_row, o, real_po)
                split_written += 1
                if DEBUG:
                    print(f"    + SPLIT {real_po} ({o['supplier']}) → {sheet_name} (£{o['cost']:.2f})")
            else:
                # Update existing
                updated_fields = []
                if "Cost" in cmap and o["cost"] != 0.0:
                    cost_cell = ws_month[f"{cmap['Cost']}{found_row}"]
                    local_cost = to_float_cell(cost_cell.value)
                    if abs(local_cost - o["cost"]) > 0.005:
                        cost_cell.value = float(o["cost"])
                        cost_cell.number_format = "£#,##0.00"
                        split_updated += 1
                        updated_fields.append(f"Cost £{local_cost:.2f}→£{o['cost']:.2f}")
                if "DeliveredDate" in cmap and o["delivered_date"]:
                    dd_cell = ws_month[f"{cmap['DeliveredDate']}{found_row}"]
                    if is_blank_cell(dd_cell.value):
                        dd_cell.value = o["delivered_date"]
                        dd_cell.number_format = "DD/MM/YYYY"
                        dd_cell.fill = GREEN_FILL
                        updated_fields.append("Delivered")
                if updated_fields and DEBUG:
                    print(f"    ✎ SPLIT {real_po} ({o['supplier']}) [{sheet_name}]: {', '.join(updated_fields)}")

        # Recalc totals for sheets affected by split orders
        if split_written > 0 or split_updated > 0:
            affected_sheets = set(s for s, _, _, _ in split_orders if s in wb.sheetnames)
            for sn in affected_sheets:
                ws = wb[sn]
                h_row, cmap = detect_header_row_and_columns(ws)
                if cmap:
                    update_month_totals(ws, h_row or 2, cmap)
                    auto_fit_columns(ws)

        print(f"  Split orders — written: {split_written}, updated: {split_updated}")

    # STEP 2C: MISROUTED ROWS REPORT
    print(f"\n--- STEP 2C: Misrouted Rows Check ---")

    main_sheet_pos = {}
    for sheet_name, sheet_orders in orders_by_sheet.items():
        main_sheet_pos[sheet_name] = set(sheet_orders.keys())

    misrouted_rows = []

    for sheet_name in wb.sheetnames:
        if sheet_name in {TO_RECEIVE_SHEET, SHEET_COST_CENTRE}:
            continue
        if not _is_valid_month_sheet(sheet_name):
            continue

        ws = wb[sheet_name]
        header_row, colmap = detect_header_row_and_columns(ws)
        if not colmap or "PONumber" not in colmap:
            continue

        expected_pos = main_sheet_pos.get(sheet_name, set())
        start_row = (header_row or 2) + 1

        for r in range(start_row, ws.max_row + 1):
            po_val = ws[f"{colmap['PONumber']}{r}"].value
            if is_blank_cell(po_val):
                continue
            po_str = str(po_val).strip()

            if po_str not in expected_pos:
                cost_val = 0.0
                if "Cost" in colmap:
                    cost_val = to_float_cell(ws[f"{colmap['Cost']}{r}"].value)
                misrouted_rows.append((sheet_name, r, po_str, cost_val))

    if misrouted_rows:
        suspicious = [(s, r, po, c) for s, r, po, c in misrouted_rows if c == 0.0]
        other = [(s, r, po, c) for s, r, po, c in misrouted_rows if c != 0.0]

        if suspicious:
            print(f"\n  ⚠ {len(suspicious)} rows with £0.00 cost on WRONG monthly sheet:")
            print(f"    Delete them manually in Excel if needed:")
            for s, r, po, c in suspicious[:20]:
                print(f"      {s:20} row {r:4} → {po}")
            if len(suspicious) > 20:
                print(f"      ... and {len(suspicious) - 20} more")

        if other:
            print(f"\n  ℹ {len(other)} rows on local sheets but not on corresponding main sheet month")
            for s, r, po, c in other[:10]:
                print(f"      {s:20} row {r:4} → {po:25} £{c:.2f}")
            if len(other) > 10:
                print(f"      ... and {len(other) - 10} more")
    else:
        print("  ✓ No misrouted rows found.")

    # STEP 2D: MOVE ORDERS BETWEEN MONTHS
    print(f"\n--- STEP 2D: Detect & Fix Moved Orders ---")

    monthly_po_index = _get_monthly_sheet_po_index(wb)

    local_po_all_sheets = {}
    for sheet_name, pos in monthly_po_index.items():
        if not _is_valid_month_sheet(sheet_name):
            continue
        for po in pos:
            if po not in local_po_all_sheets:
                local_po_all_sheets[po] = []
            local_po_all_sheets[po].append(sheet_name)

    po_correct_sheets = {}
    for sheet_name, sheet_orders in orders_by_sheet.items():
        for po_str in sheet_orders:
            if po_str not in po_correct_sheets:
                po_correct_sheets[po_str] = set()
            po_correct_sheets[po_str].add(sheet_name)

    wrong_sheet_copies = []
    for po_str, correct_sheets in po_correct_sheets.items():
        local_sheets = local_po_all_sheets.get(po_str, [])
        for ls in local_sheets:
            if ls not in correct_sheets:
                wrong_sheet_copies.append((po_str, ls, ", ".join(sorted(correct_sheets))))

    removed_count = 0
    cleared_count = 0

    if wrong_sheet_copies:
        print(f"\n  Found {len(wrong_sheet_copies)} orders on WRONG local sheet:")
        for po_str, wrong_sh, correct_sh in wrong_sheet_copies[:15]:
            print(f"    {po_str:25} | ON: {wrong_sh:15} → SHOULD BE: {correct_sh}")
        if len(wrong_sheet_copies) > 15:
            print(f"    ... and {len(wrong_sheet_copies) - 15} more")

        choice = input(f"\n  Remove these {len(wrong_sheet_copies)} wrong-sheet copies? (y/n): ").strip().lower()
        if choice in ("y", "yes"):
            rows_by_sheet = {}
            for po_str, wrong_sh, _ in wrong_sheet_copies:
                if wrong_sh not in wb.sheetnames:
                    continue
                ws = wb[wrong_sh]
                h_row, cmap = detect_header_row_and_columns(ws)
                if not cmap or "PONumber" not in cmap:
                    continue
                start_row = (h_row or 2) + 1
                for r in range(start_row, ws.max_row + 1):
                    val = str(ws[f"{cmap['PONumber']}{r}"].value or "").strip()
                    if val == po_str:
                        if wrong_sh not in rows_by_sheet:
                            rows_by_sheet[wrong_sh] = []
                        rows_by_sheet[wrong_sh].append((r, po_str))
                        break

            for sheet_name, row_list in rows_by_sheet.items():
                ws = wb[sheet_name]
                for r, po_str in sorted(row_list, reverse=True):
                    try:
                        ws.delete_rows(r)
                        removed_count += 1
                        if DEBUG:
                            print(f"    ✗ Deleted {po_str} from '{sheet_name}' row {r}")
                    except Exception as e:
                        if DEBUG:
                            print(f"    ⚠ delete_rows failed for {po_str} row {r}: {e}")
                        for c in range(1, (ws.max_column or 15) + 1):
                            ws.cell(row=r, column=c).value = None
                        cleared_count += 1

                h_row, cmap = detect_header_row_and_columns(ws)
                if cmap:
                    update_month_totals(ws, h_row or 2, cmap)
                    auto_fit_columns(ws)

            print(f"  ✓ Removed {removed_count} rows, cleared {cleared_count} rows.")
        else:
            print("  Skipped.")
    else:
        print("  ✓ All orders are on the correct monthly sheets.")

    total_actions = removed_count + cleared_count

    # STEP 2E: ORDERS WITHOUT PO NUMBER
    print(f"\n--- STEP 2E: Orders Without PO Number ---")

    nopo_written = 0
    nopo_skipped = 0
    nopo_deduped = 0

    if nopo_by_sheet:
        total_nopo = sum(len(v) for v in nopo_by_sheet.values())
        print(f"  Found {total_nopo} no-PO orders across {len(nopo_by_sheet)} sheets.")

        for target_sheet, sheet_nopo in nopo_by_sheet.items():
            if target_sheet not in wb.sheetnames:
                nopo_skipped += len(sheet_nopo)
                continue

            ws_month = wb[target_sheet]
            h_row, cmap = ensure_parts_column_map(ws_month)
            if "Machine" not in cmap:
                nopo_skipped += len(sheet_nopo)
                continue

            for fkey, o in sheet_nopo.items():
                start_row = (h_row or 2) + 1
                o_supplier_norm = normalize_text(o["supplier"])
                o_parts_norm = normalize_text(o["parts"])[:40]
                o_machine_norm = normalize_text(o["machine"])
                o_cost = o["cost"]
                already_exists = False

                sup_col = cmap.get("Supplier")
                parts_col = cmap.get("PartsOrdered")
                cost_col = cmap.get("Cost")
                mach_col = cmap.get("Machine")

                for r in range(start_row, ws_month.max_row + 1):
                    local_mach = normalize_text(ws_month[f"{mach_col}{r}"].value) if mach_col else ""
                    if not local_mach:
                        continue

                    local_sup = normalize_text(ws_month[f"{sup_col}{r}"].value) if sup_col else ""
                    local_parts = normalize_text(ws_month[f"{parts_col}{r}"].value)[:40] if parts_col else ""
                    local_cost = to_float_cell(ws_month[f"{cost_col}{r}"].value) if cost_col else 0.0

                    if o_supplier_norm and o_parts_norm:
                        if o_supplier_norm == local_sup and o_parts_norm == local_parts:
                            already_exists = True
                            break

                    if not o_supplier_norm and o_machine_norm and o_parts_norm:
                        if o_machine_norm == local_mach and o_parts_norm == local_parts:
                            already_exists = True
                            break

                    if o_machine_norm and o_cost > 0 and not o_parts_norm:
                        if o_machine_norm == local_mach and abs(local_cost - o_cost) < 1.0:
                            already_exists = True
                            break

                if already_exists:
                    nopo_skipped += 1
                    continue

                data_row = find_first_empty_row(ws_month, cmap["Machine"], start_row)

                ws_month[f"{cmap['Machine']}{data_row}"] = o["machine"]
                if "Area" in cmap:
                    ws_month[f"{cmap['Area']}{data_row}"] = o["area"]
                if "Reason" in cmap:
                    ws_month[f"{cmap['Reason']}{data_row}"] = o["reason"]
                if "MechElecOther" in cmap:
                    ws_month[f"{cmap['MechElecOther']}{data_row}"] = o["mech_elec"]
                if "Supplier" in cmap:
                    ws_month[f"{cmap['Supplier']}{data_row}"] = o["supplier"]
                if "PartsOrdered" in cmap:
                    ws_month[f"{cmap['PartsOrdered']}{data_row}"] = o["parts"]
                if "JobNo" in cmap:
                    ws_month[f"{cmap['JobNo']}{data_row}"] = o["job_no"]

                if "Cost" in cmap and o["cost"]:
                    cost_cell = ws_month[f"{cmap['Cost']}{data_row}"]
                    cost_cell.value = float(o["cost"])
                    cost_cell.number_format = "£#,##0.00"

                if o["date_ordered"] and "DateOrdered" in cmap:
                    d_cell = ws_month[f"{cmap['DateOrdered']}{data_row}"]
                    d_cell.value = o["date_ordered"]
                    d_cell.number_format = "DD/MM/YYYY"

                if o["due_date"] and "DueDate" in cmap:
                    ws_month[f"{cmap['DueDate']}{data_row}"] = o["due_date"]
                    ws_month[f"{cmap['DueDate']}{data_row}"].number_format = "DD/MM/YYYY"

                if o["delivered_date"] and "DeliveredDate" in cmap:
                    dd_cell = ws_month[f"{cmap['DeliveredDate']}{data_row}"]
                    dd_cell.value = o["delivered_date"]
                    dd_cell.number_format = "DD/MM/YYYY"
                    dd_cell.fill = GREEN_FILL

                if o["savings"] and "Savings" in cmap:
                    ws_month[f"{cmap['Savings']}{data_row}"] = float(o["savings"])
                    ws_month[f"{cmap['Savings']}{data_row}"].number_format = "£#,##0.00"

                if o["orig_supplier"] and "OriginalSupplier" in cmap:
                    ws_month[f"{cmap['OriginalSupplier']}{data_row}"] = o["orig_supplier"]

                nopo_written += 1
                if DEBUG:
                    print(f"    ✓ {o['machine'][:15]:15} | {o['supplier'][:20]:20} | {o['parts'][:35]:35} | £{o['cost']:.2f} → {target_sheet}")

        if nopo_written > 0:
            for sn in nopo_by_sheet.keys():
                if sn in wb.sheetnames:
                    ws = wb[sn]
                    h_row, cmap = detect_header_row_and_columns(ws)
                    if cmap:
                        update_month_totals(ws, h_row or 2, cmap)

        print(f"  Written: {nopo_written}, Already existed: {nopo_skipped}")

    else:
        print("  No orders without PO found on main sheet.")

    # Dedup no-PO rows (runs always)
    for sheet_name in wb.sheetnames:
        if sheet_name in {TO_RECEIVE_SHEET, SHEET_COST_CENTRE}:
            continue
        if not _is_valid_month_sheet(sheet_name):
            continue

        ws = wb[sheet_name]
        h_row, cmap = detect_header_row_and_columns(ws)
        if not cmap or "Machine" not in cmap:
            continue

        po_col = cmap.get("PONumber")
        mach_col = cmap.get("Machine")
        parts_col = cmap.get("PartsOrdered")
        cost_col = cmap.get("Cost")
        start_row = (h_row or 2) + 1

        seen_signatures = set()
        rows_to_remove = []

        for r in range(start_row, ws.max_row + 1):
            po_val = ws[f"{po_col}{r}"].value if po_col else None
            if not is_blank_cell(po_val):
                continue
            machine = normalize_text(ws[f"{mach_col}{r}"].value) if mach_col else ""
            parts = normalize_text(ws[f"{parts_col}{r}"].value)[:40] if parts_col else ""
            cost = to_float_cell(ws[f"{cost_col}{r}"].value) if cost_col else 0.0
            if not machine:
                continue
            sig = f"{machine}|{parts}|{cost:.2f}"
            if sig in seen_signatures:
                rows_to_remove.append(r)
            else:
                seen_signatures.add(sig)

        for r in sorted(rows_to_remove, reverse=True):
            try:
                ws.delete_rows(r)
                nopo_deduped += 1
            except Exception:
                for c in range(1, (ws.max_column or 15) + 1):
                    ws.cell(row=r, column=c).value = None
                nopo_deduped += 1

        if rows_to_remove:
            h_row2, cmap2 = detect_header_row_and_columns(ws)
            if cmap2:
                update_month_totals(ws, h_row2 or 2, cmap2)

    if nopo_deduped > 0:
        print(f"  ✓ Removed {nopo_deduped} duplicate no-PO rows.")

    # STEP 2F: VERIFICATION PASS
    print(f"\n--- STEP 2F: Verification Pass ---")

    verify_missing = 0
    verify_written = 0

    for sheet_name, sheet_orders in orders_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws_month = wb[sheet_name]
        h_row, cmap = detect_header_row_and_columns(ws_month)
        if not cmap or "PONumber" not in cmap:
            continue

        start_row = (h_row or 2) + 1

        local_pos_on_sheet = set()
        for r in range(start_row, ws_month.max_row + 1):
            val = str(ws_month[f"{cmap['PONumber']}{r}"].value or "").strip()
            if val:
                local_pos_on_sheet.add(val)

        for po_str, o in sheet_orders.items():
            if po_str not in local_pos_on_sheet:
                verify_missing += 1
                if "Machine" in cmap:
                    _write_order_to_monthly_sheet(ws_month, cmap, h_row, o, po_str)
                    verify_written += 1
                    if DEBUG and verify_written <= 20:
                        print(f"    ✓ RECOVERED {po_str} → {sheet_name} (£{o['cost']:.2f})")

    if verify_missing > 0:
        affected = set(orders_by_sheet.keys()) & set(wb.sheetnames)
        for sn in affected:
            ws = wb[sn]
            h_row, cmap = detect_header_row_and_columns(ws)
            if cmap:
                update_month_totals(ws, h_row or 2, cmap)
                auto_fit_columns(ws)

        print(f"  ⚠ Found {verify_missing} POs missing from correct local sheet!")
        print(f"  ✓ Recovered {verify_written} orders.")
    else:
        print("  ✓ All main POs verified on correct local sheets.")

    # STEP 3: HEALTH CHECK
    print(f"\n--- STEP 3: Health Check ---")

    local_pos_refreshed = _get_local_po_numbers(wb)

    dupes = _find_local_duplicates(wb)
    if dupes:
        print(f"\n  ⚠ DUPLICATES in your To Receive sheet:")
        for po, cnt in dupes:
            print(f"    {po} appears {cnt} times")
    else:
        print(f"  ✓ No duplicates found in To Receive.")

    main_po_set = set(main_orders.keys())
    only_local = local_pos_refreshed - main_po_set

    if only_local:
        print(f"\n  ⚠ {len(only_local)} orders on YOUR sheet but NOT on main sheet:")
        for po in sorted(only_local):
            print(f"    → {po}")
        print("  (These might need to be added to the main sheet manually)")
    else:
        print(f"  ✓ All your local orders exist on the main sheet.")

    auto_fit_columns(ws_recv)
    safe_save_workbook(wb, WORKBOOK_PATH)

    print(f"\n{'='*60}")
    print(f"  SYNC COMPLETE")
    print(f"{'='*60}")
    print(f"  Main sheet orders:    {len(main_orders)} (+{len(no_po_orders)} no-PO)")
    print(f"  Deliveries synced:    {delivery_updated}")
    print(f"  To Receive imported:  {imported_to_receive}")
    print(f"  To Receive costs upd: {to_receive_cost_updated}")
    print(f"  Monthly rows written: {monthly_written}")
    print(f"  Monthly costs updated:{monthly_cost_updated}")
    print(f"  Orders fixed (2D):    {total_actions}")
    print(f"  No-PO orders written: {nopo_written}")
    if nopo_deduped > 0:
        print(f"  No-PO dupes removed:  {nopo_deduped}")
    print(f"  Verified/recovered:   {verify_written}")
    print(f"  Misrouted rows:       {len(misrouted_rows)}")
    print(f"  Duplicates found:     {len(dupes)}")
    print(f"  Missing from main:    {len(only_local)}")


# ====================================================================
# RECALCULATE ALL TOTALS — v1.9.9
# ====================================================================
def recalculate_all_totals():
    """
    Walk every monthly sheet in the local workbook, recalculate the
    Total Spend and Total Savings cells (I1 / M1), and save.
    Useful as a standalone sanity check — no sync with main required.
    """
    print(f"\n{'='*60}")
    print(f"  RECALCULATE ALL MONTHLY SHEET TOTALS")
    print(f"{'='*60}")

    wb = load_or_create_workbook(WORKBOOK_PATH)
    sheets_updated = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in {TO_RECEIVE_SHEET, SHEET_COST_CENTRE}:
            continue

        ws = wb[sheet_name]
        header_row, colmap = detect_header_row_and_columns(ws)

        if not colmap or "Cost" not in colmap:
            continue

        # Read old values before recalc
        old_cost = to_float_cell(ws["I1"].value) if ws["I1"].value is not None else None
        old_savings = to_float_cell(ws["M1"].value) if ws["M1"].value is not None else None

        update_month_totals(ws, header_row or 2, colmap)
        auto_fit_columns(ws)

        new_cost = to_float_cell(ws["I1"].value)
        new_savings = to_float_cell(ws["M1"].value)

        cost_changed = old_cost is None or abs((old_cost or 0) - new_cost) > 0.005
        savings_changed = old_savings is None or abs((old_savings or 0) - new_savings) > 0.005

        if cost_changed or savings_changed:
            old_c = f"£{old_cost:,.2f}" if old_cost is not None else "N/A"
            old_s = f"£{old_savings:,.2f}" if old_savings is not None else "N/A"
            print(f"  ✎ {sheet_name:20}"
                  f"  Cost: {old_c} → £{new_cost:,.2f}"
                  f"  Savings: {old_s} → £{new_savings:,.2f}")
            sheets_updated += 1
        else:
            print(f"  ✓ {sheet_name:20}  Cost: £{new_cost:,.2f}  Savings: £{new_savings:,.2f}  (no change)")

    safe_save_workbook(wb, WORKBOOK_PATH)

    print(f"\n  Done. {sheets_updated} sheet(s) had totals updated.")
    print(f"  Workbook saved: {WORKBOOK_PATH}")


# ========= DIAGNOSTIC TOOLS =========
def diagnose_main_workbook():
    if not os.path.isfile(MAIN_WORKBOOK_PATH):
        print(f"Main workbook not found at: {MAIN_WORKBOOK_PATH}")
        return

    print(f"\n{'='*60}")
    print(f"  DIAGNOSTIC: RAW CELL DUMP OF MAIN WORKBOOK")
    print(f"{'='*60}")

    print(f"\n--- PASS 1: data_only=True (cached values) ---")
    try:
        wb1 = load_workbook(MAIN_WORKBOOK_PATH, read_only=False, data_only=True)
    except Exception as e:
        print(f"Failed to open: {e}")
        return

    for sn in wb1.sheetnames:
        if not _is_valid_month_sheet(sn):
            continue

        print(f"\n  Sheet: '{sn}'")
        ws = wb1[sn]

        row_num = 0
        data_rows_shown = 0
        for row in ws.iter_rows(max_row=20, values_only=False):
            row_num += 1
            cells_text = []
            for cell in row:
                coord = cell.coordinate
                val = cell.value
                cells_text.append(f"{coord}={repr(val)}")

            is_header = row_num <= 3
            has_data = any(cell.value is not None for cell in row)

            if is_header or (has_data and data_rows_shown < 5):
                if not is_header:
                    data_rows_shown += 1
                tag = "HDR" if is_header else "DAT"
                print(f"    [{tag} row {row_num}]")
                for ct in cells_text:
                    if "=None" not in ct:
                        print(f"      {ct}")

        break

    try:
        wb1.close()
    except Exception:
        pass

    print(f"\n--- PASS 2: data_only=False (raw formulas) ---")
    try:
        wb2 = load_workbook(MAIN_WORKBOOK_PATH, read_only=False, data_only=False)
    except Exception as e:
        print(f"Failed to open: {e}")
        return

    for sn in wb2.sheetnames:
        if not _is_valid_month_sheet(sn):
            continue

        print(f"\n  Sheet: '{sn}'")
        ws = wb2[sn]

        row_num = 0
        data_rows_shown = 0
        for row in ws.iter_rows(max_row=20, values_only=False):
            row_num += 1
            cells_text = []
            for cell in row:
                coord = cell.coordinate
                val = cell.value
                if val is not None and isinstance(val, str) and val.startswith("="):
                    cells_text.append(f"  ⚡ {coord} = FORMULA: {val}")
                elif val is not None:
                    cells_text.append(f"     {coord} = {repr(val)}")

            is_header = row_num <= 3
            has_formulas = any("FORMULA" in ct for ct in cells_text)

            if is_header or (has_formulas and data_rows_shown < 5):
                if not is_header:
                    data_rows_shown += 1
                print(f"    [row {row_num}]")
                for ct in cells_text:
                    if "FORMULA" in ct:
                        print(f"    {ct}")

        break

    try:
        wb2.close()
    except Exception:
        pass

    print(f"\n{'='*60}")
    print(f"  Copy the output above and send it to Claude.")
    print(f"{'='*60}")


def compare_month():
    if not os.path.isfile(MAIN_WORKBOOK_PATH):
        print(f"Main workbook not found at: {MAIN_WORKBOOK_PATH}")
        return

    wb_local = load_or_create_workbook(WORKBOOK_PATH)
    months = [s for s in wb_local.sheetnames
              if s not in {TO_RECEIVE_SHEET, SHEET_COST_CENTRE} and _is_valid_month_sheet(s)]

    if not months:
        print("No monthly sheets found.")
        return

    print("\nAvailable months:")
    for i, m in enumerate(months, 1):
        print(f"  {i}) {m}")

    pick = input(f"Pick 1-{len(months)}: ").strip()
    if not pick.isdigit() or int(pick) < 1 or int(pick) > len(months):
        print("Invalid choice.")
        return

    target_month = months[int(pick) - 1]
    print(f"\n{'='*70}")
    print(f"  COMPARING: {target_month}")
    print(f"{'='*70}")

    print(f"\nOpening main workbook...")
    main_wb = load_main_workbook_readonly(MAIN_WORKBOOK_PATH)

    if target_month not in main_wb.sheetnames:
        print(f"Sheet '{target_month}' not found in main workbook.")
        try:
            main_wb.close()
        except Exception:
            pass
        return

    ws_main = main_wb[target_month]
    main_rows = []
    for row in ws_main.iter_rows(values_only=True):
        main_rows.append(row)

    header_row_idx = None
    col_map = {}
    for r_idx in range(min(12, len(main_rows))):
        row = main_rows[r_idx]
        score = 0
        temp_map = {}
        for c_idx, cell_val in enumerate(row):
            norm = normalize_text(cell_val)
            raw_val = str(cell_val or "").strip()
            if not norm and not raw_val:
                continue
            if "£" in raw_val and len(raw_val) <= 10 and "Cost" not in temp_map:
                temp_map["Cost"] = c_idx
                score += 1
                continue
            for field, synonyms in HEADER_SYNONYMS.items():
                if norm in synonyms and field not in temp_map:
                    temp_map[field] = c_idx
                    score += 1
                    break
        if score > len(col_map):
            col_map = temp_map
            header_row_idx = r_idx

    if "Cost" not in col_map and header_row_idx is not None:
        mapped_cols = set(col_map.values())
        max_c = max(len(r) for r in main_rows if r) if main_rows else 0
        best_cost_col = None
        best_nc = 0
        for c_idx in range(max_c):
            if c_idx in mapped_cols:
                continue
            nc = 0
            for d_idx in range(header_row_idx + 1, min(header_row_idx + 20, len(main_rows))):
                d_row = main_rows[d_idx]
                if d_row and c_idx < len(d_row):
                    val = d_row[c_idx]
                    if isinstance(val, (int, float)):
                        nc += 1
            if nc >= 3 and nc > best_nc:
                best_nc = nc
                best_cost_col = c_idx
        if best_cost_col is not None:
            col_map["Cost"] = best_cost_col

    main_orders_month = {}
    main_no_po = []
    main_total_cost = 0.0

    if header_row_idx is not None:
        def sg(field, row):
            if field in col_map and col_map[field] < len(row):
                return row[col_map[field]]
            return None

        for r_idx in range(header_row_idx + 1, len(main_rows)):
            row = main_rows[r_idx]
            if not row:
                continue

            supplier = str(sg("Supplier", row) or "").strip()
            parts = str(sg("PartsOrdered", row) or "").strip()
            cost = to_float_cell(sg("Cost", row))
            machine = str(sg("Machine", row) or "").strip()

            if not supplier and not parts and not machine:
                continue

            main_total_cost += cost

            po_val = sg("PONumber", row) if "PONumber" in col_map else None
            po_str = str(po_val).strip() if not is_blank_cell(po_val) else ""

            entry = {
                "row": r_idx + 1,
                "po": po_str,
                "machine": machine,
                "supplier": supplier,
                "parts": parts[:60],
                "cost": cost,
            }

            if po_str:
                main_orders_month[po_str] = entry
            else:
                main_no_po.append(entry)

    try:
        main_wb.close()
    except Exception:
        pass

    ws_local = wb_local[target_month]
    h_row, lmap = detect_header_row_and_columns(ws_local)
    if not lmap:
        print("Cannot detect columns on local sheet.")
        return

    local_orders = {}
    local_no_po = []
    local_total_cost = 0.0
    start_row = (h_row or 2) + 1

    for r in range(start_row, ws_local.max_row + 1):
        po_val = ws_local[f"{lmap.get('PONumber', 'H')}{r}"].value if "PONumber" in lmap else None
        po_str = str(po_val).strip() if not is_blank_cell(po_val) else ""

        machine = str(ws_local[f"{lmap.get('Machine', 'A')}{r}"].value or "").strip() if "Machine" in lmap else ""
        supplier = str(ws_local[f"{lmap.get('Supplier', 'E')}{r}"].value or "").strip() if "Supplier" in lmap else ""
        parts = str(ws_local[f"{lmap.get('PartsOrdered', 'F')}{r}"].value or "").strip()[:60] if "PartsOrdered" in lmap else ""
        cost = 0.0
        if "Cost" in lmap:
            cost = to_float_cell(ws_local[f"{lmap['Cost']}{r}"].value)

        if not machine and not supplier and not parts:
            continue

        local_total_cost += cost

        entry = {"row": r, "po": po_str, "machine": machine, "supplier": supplier, "parts": parts, "cost": cost}

        if po_str:
            local_orders[po_str] = entry
        else:
            local_no_po.append(entry)

    print(f"\n  MAIN:  {len(main_orders_month)} orders with PO + {len(main_no_po)} without PO = {len(main_orders_month) + len(main_no_po)} total")
    print(f"  LOCAL: {len(local_orders)} orders with PO + {len(local_no_po)} without PO = {len(local_orders) + len(local_no_po)} total")
    print(f"\n  MAIN total cost:  £{main_total_cost:,.2f}")
    print(f"  LOCAL total cost: £{local_total_cost:,.2f}")
    print(f"  DIFFERENCE:       £{main_total_cost - local_total_cost:,.2f}")

    main_only = set(main_orders_month.keys()) - set(local_orders.keys())
    if main_only:
        main_only_cost = sum(main_orders_month[po]["cost"] for po in main_only)
        print(f"\n  --- ON MAIN BUT MISSING FROM LOCAL ({len(main_only)} orders, £{main_only_cost:,.2f}) ---")
        for po in sorted(main_only):
            e = main_orders_month[po]
            print(f"    {po:25} | {e['supplier']:20} | £{e['cost']:>9.2f} | {e['parts'][:40]}")

    local_only = set(local_orders.keys()) - set(main_orders_month.keys())
    if local_only:
        local_only_cost = sum(local_orders[po]["cost"] for po in local_only)
        print(f"\n  --- ON LOCAL BUT NOT ON MAIN ({len(local_only)} orders, £{local_only_cost:,.2f}) ---")
        for po in sorted(local_only):
            e = local_orders[po]
            print(f"    {po:25} | {e['supplier']:20} | £{e['cost']:>9.2f} | {e['parts'][:40]}")

    both = set(main_orders_month.keys()) & set(local_orders.keys())
    cost_diffs = []
    for po in both:
        mc = main_orders_month[po]["cost"]
        lc = local_orders[po]["cost"]
        if abs(mc - lc) > 0.01:
            cost_diffs.append((po, mc, lc))

    if cost_diffs:
        total_diff = sum(mc - lc for _, mc, lc in cost_diffs)
        print(f"\n  --- COST DIFFERENCES ({len(cost_diffs)} orders, net £{total_diff:,.2f}) ---")
        for po, mc, lc in sorted(cost_diffs, key=lambda x: abs(x[1]-x[2]), reverse=True):
            print(f"    {po:25} | MAIN: £{mc:>9.2f} | LOCAL: £{lc:>9.2f} | diff: £{mc-lc:>+9.2f}")

    if main_no_po:
        nopo_cost = sum(e["cost"] for e in main_no_po)
        print(f"\n  --- MAIN NO-PO ORDERS ({len(main_no_po)} orders, £{nopo_cost:,.2f}) ---")
        for e in main_no_po:
            print(f"    row {e['row']:3} | {e['supplier']:20} | £{e['cost']:>9.2f} | {e['parts'][:40]}")

    print(f"\n  {'='*60}")
    print(f"  COST BREAKDOWN OF DIFFERENCE (£{main_total_cost - local_total_cost:,.2f}):")
    main_only_c = sum(main_orders_month[po]["cost"] for po in main_only) if main_only else 0
    local_only_c = sum(local_orders[po]["cost"] for po in local_only) if local_only else 0
    cost_diff_c = sum(mc - lc for _, mc, lc in cost_diffs) if cost_diffs else 0
    nopo_c = sum(e["cost"] for e in main_no_po) if main_no_po else 0
    local_nopo_c = sum(e["cost"] for e in local_no_po) if local_no_po else 0
    print(f"    Missing from local (with PO):  +£{main_only_c:,.2f}")
    print(f"    Extra on local (not on main):  -£{local_only_c:,.2f}")
    print(f"    Cost differences:              {'+' if cost_diff_c >= 0 else ''}£{cost_diff_c:,.2f}")
    print(f"    Main no-PO orders:             +£{nopo_c:,.2f}")
    print(f"    Local no-PO orders:            -£{local_nopo_c:,.2f}")
    accounted = main_only_c - local_only_c + cost_diff_c + nopo_c - local_nopo_c
    print(f"    ─────────────────────────────")
    print(f"    Accounted:                     £{accounted:,.2f}")
    print(f"    Actual difference:             £{main_total_cost - local_total_cost:,.2f}")
    unaccounted = (main_total_cost - local_total_cost) - accounted
    if abs(unaccounted) > 0.01:
        print(f"    Unaccounted:                   £{unaccounted:,.2f}")


def reorder_check():
    """Run the Recurring Order Detector and show alerts + forecast."""
    if not DETECTOR_AVAILABLE:
        print("\n  [Reorder] Detector modules not found.")
        print("  Make sure data_parser.py and detector.py are in the same folder as this script.")
        return

    source = MAIN_WORKBOOK_PATH
    if not os.path.exists(source):
        print(f"\n  [Reorder] Main Parts Ordered file not found:")
        print(f"  {source}")
        return

    print("\n╔══════════════════════════════════════════════════════╗")
    print("║     RECURRING ORDER DETECTOR v0.1.0                 ║")
    print("║     Sheard Packaging — Engineering Stores            ║")
    print(f"║     {datetime.now().strftime('%A %d %B %Y, %H:%M'):<44}║")
    print("╚══════════════════════════════════════════════════════╝")

    print("\n  Loading order data (this may take a moment)...")
    df = load_all_orders(filepath=source)

    if df.empty:
        print("  No data loaded!")
        return

    print("\n  Analysing order patterns...")
    patterns = calculate_order_patterns(df)

    services = patterns[patterns['order_type'] == 'SERVICE']
    parts = patterns[patterns['order_type'] == 'PARTS']

    print(f"\n  Found {len(patterns)} recurring patterns "
          f"({len(services)} services, {len(parts)} parts)")

    for label, subset in [('PARTS', parts), ('SERVICES', services)]:
        print(f"\n  {label}:")
        print(f"    OVERDUE:     {(subset['status'] == 'OVERDUE').sum()}")
        print(f"    DUE SOON:    {(subset['status'] == 'DUE SOON').sum()}")
        print(f"    OK:          {(subset['status'] == 'OK').sum()}")
        print(f"    INACTIVE:    {(subset['status'] == 'INACTIVE').sum()}")

    print("\n")
    parts_alerts = get_reorder_alerts(parts, ['OVERDUE', 'DUE SOON'])
    print("  ┌─────────────────────────────────────┐")
    print("  │   PARTS — REORDER ALERTS            │")
    print("  └─────────────────────────────────────┘")
    print_alerts(parts_alerts)

    print()
    service_alerts = get_reorder_alerts(services, ['OVERDUE', 'DUE SOON'])
    print("  ┌─────────────────────────────────────┐")
    print("  │   SERVICES — OVERDUE/DUE SOON       │")
    print("  └─────────────────────────────────────┘")
    print_alerts(service_alerts, max_rows=20)

    print()
    forecast = get_monthly_forecast(parts)
    print_forecast(forecast)

    # Optional Excel export
    export_choice = input("\n  Export full report to Excel? (y/n): ").strip().lower()
    if export_choice in ('y', 'yes'):
        try:
            import pandas as pd

            report_dir = PO_PDF_OUTPUT_DIR  # same folder as PO PDFs
            report_path = os.path.join(report_dir, 'reorder_report.xlsx')
            ensure_dir(report_dir)

            export_cols = [
                'status', 'order_type', 'machine', 'supplier', 'total_orders',
                'avg_interval_days', 'last_order_date', 'days_since_last',
                'reorder_score', 'est_next_order', 'avg_price', 'last_price',
                'total_spend', 'projected_monthly', 'last_description',
                'common_description', 'last_order_no',
            ]

            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                pa = parts[parts['status'].isin(['OVERDUE', 'DUE SOON'])].sort_values(
                    'reorder_score', ascending=False)
                if not pa.empty:
                    pa[export_cols].to_excel(writer, sheet_name='Parts Alerts', index=False)

                sa = services[services['status'] != 'INACTIVE'].sort_values(
                    'reorder_score', ascending=False)
                if not sa.empty:
                    sa[export_cols].to_excel(writer, sheet_name='Service Alerts', index=False)

                active = patterns[patterns['status'] != 'INACTIVE']
                active[export_cols].to_excel(writer, sheet_name='All Active', index=False)

                mf = forecast['by_machine'].reset_index()
                mf.columns = ['Machine', 'Projected Monthly £']
                mf.to_excel(writer, sheet_name='Forecast Machine', index=False)

                sf = forecast['by_supplier'].reset_index()
                sf.columns = ['Supplier', 'Projected Monthly £']
                sf.to_excel(writer, sheet_name='Forecast Supplier', index=False)

            print(f"\n  Report saved: {report_path}")

        except Exception as e:
            print(f"\n  Export failed: {e}")

    print("\n  Done.")


# ========= MAIN MENU =========
def create_new_po():
    wb = load_or_create_workbook(WORKBOOK_PATH)
    ws_recv = ensure_to_receive_sheet(wb)

    # v2.1.0: collect order data FIRST so we know the due date before choosing
    # which monthly sheet to write to.
    order, supplier_obj, items, totals = collect_po_data()

    order["machine"] = normalize_machine_name(order["machine"])
    order["area"] = normalize_area_name(order["area"], order["machine"])

    cost_centre = resolve_cost_centre(wb, order["machine"], order["area"], order["mech_elec"])
    order["cost_centre"] = cost_centre

    if DEBUG:
        print(f"\nMachine (canonical): {order['machine']}")
        print(f"Area (canonical): {order['area']}")
        print(f"Cost Centre result: {cost_centre}")

    # v2.1.0: Determine target month from due date. If due date is in a future
    # month, write the PO to that future sheet (creating intermediate sheets
    # as needed). Otherwise stick with the current month.
    target_first = get_target_month_for_order(order)
    today_first = date(datetime.today().year, datetime.today().month, 1)

    created_sheets = ensure_month_sheets_through(wb, target_first)
    if created_sheets:
        print(f"\nCreated {len(created_sheets)} new monthly sheet(s): "
              f"{', '.join(created_sheets)}")

    ws_parts = get_or_create_sheet_for_date(wb, target_first)

    if target_first != today_first:
        print(f"\nFuture-month routing: PO will be written to sheet "
              f"'{ws_parts.title}' (Date Ordered = today).")

    header_row, colmap = ensure_parts_column_map(ws_parts)
    try:
        validate_required_parts_columns(colmap)
    except RuntimeError as e:
        print(e)
        return

    data_start = header_row + 1

    if DEBUG:
        print("\nDetected Parts Ordered columns:")
        for k in sorted(colmap.keys()):
            print(f"  {k:14} -> {colmap[k]}")
        print(f"Header row: {header_row}, data starts: {data_start}")

    # v1.9.12: PO sequence is now computed inside generate_po_number
    # by scanning local + main workbooks. seq_base no longer needed.
    po_number = generate_po_number(cost_centre, INITIALS)

    net_total = totals["net_total"]

    parts_row = write_parts_order_row(ws_parts, header_row, colmap, order, po_number, cost_centre, net_total)
    recv_row = add_to_receive_row(ws_recv, po_number, order, cost_centre, net_total, totals)

    update_month_totals(ws_parts, header_row, colmap)
    auto_fit_columns(ws_parts)
    auto_fit_columns(ws_recv)
    set_sheet_view(ws_parts, zoom=90)
    set_sheet_view(ws_recv, zoom=90)

    ensure_dir(PO_PDF_OUTPUT_DIR)
    pdf_path = os.path.join(PO_PDF_OUTPUT_DIR, f"{po_number}.pdf")
    pdf_ok = generate_po_pdf(pdf_path, po_number, supplier_obj, order, items, totals)

    safe_save_workbook(wb, WORKBOOK_PATH)

    # v3.0.0: Push to main SharePoint workbook automatically via Graph API
    push_po_to_main(order, po_number, net_total, ws_parts.title)

    print("\nSUCCESS")
    print(f"Purchase Order (PO): {po_number}")
    print(f"Written to sheet '{ws_parts.title}' row {parts_row}")
    print(f"Added to '{TO_RECEIVE_SHEET}' row {recv_row}")

    if pdf_ok:
        print(f"PDF saved: {pdf_path}")
        if AUTO_OPEN_OUTLOOK:
            print("\nOpening Outlook email...")
            email_ok = open_outlook_with_po(pdf_path, po_number, supplier_obj, order, totals)
            if not email_ok:
                print("You can attach the PDF manually and email the supplier.")
    else:
        print("PDF failed. Check the error above (or install ReportLab).")


def receive_po():
    wb = load_or_create_workbook(WORKBOOK_PATH)
    po_number = input_text("Enter PO number to mark as Received (or 'back'): ", allow_blank=False, allow_back=True)
    if po_number == BACK:
        return
    invoice_number = input_text("Invoice number (optional, or 'back'): ", allow_blank=True, allow_back=True)
    if invoice_number == BACK:
        invoice_number = ""
    mark_received(wb, po_number, invoice_number)
    safe_save_workbook(wb, WORKBOOK_PATH)
    print("Saved.")


# ===========================================================================
# v3.0.0: Microsoft Graph API — automatic push to main SharePoint workbook
# ===========================================================================

def _date_to_excel_serial(d) -> int:
    """Convert Python date/datetime to Excel serial. Excel epoch = 30 Dec 1899."""
    if d is None:
        return None
    if isinstance(d, datetime):
        d = d.date()
    excel_epoch = date(1899, 12, 30)
    return (d - excel_epoch).days


def _graph_col_letter(n: int) -> str:
    """1-based column index to Excel letter. 1=A, 26=Z, 27=AA."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def _graph_get_token_silent():
    """Acquire Graph API token silently from MSAL cache. Returns token or None."""
    if not GRAPH_AVAILABLE:
        return None
    try:
        cache = msal.SerializableTokenCache()
        if not GRAPH_TOKEN_CACHE_PATH.exists():
            return None
        cache.deserialize(GRAPH_TOKEN_CACHE_PATH.read_text(encoding="utf-8"))
        app = msal.PublicClientApplication(
            GRAPH_CLIENT_ID, authority=GRAPH_AUTHORITY, token_cache=cache,
        )
        accounts = app.get_accounts()
        if not accounts:
            return None
        result = app.acquire_token_silent(GRAPH_SCOPES, account=accounts[0])
        if cache.has_state_changed:
            GRAPH_TOKEN_CACHE_PATH.write_text(cache.serialize(), encoding="utf-8")
        return result.get("access_token") if result else None
    except Exception:
        return None


def _graph_session_headers(token: str, session_id: str) -> dict:
    return {"Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "workbook-session-id": session_id}


def _graph_create_session(token: str, persist: bool = True):
    """Open Excel workbook session. Returns session_id or None."""
    try:
        r = _requests.post(
            f"{GRAPH_WB_BASE}/createSession",
            headers={"Authorization": f"Bearer {token}",
                     "Content-Type": "application/json"},
            json={"persistChanges": persist}, timeout=30,
        )
        if r.status_code in (200, 201):
            return r.json().get("id")
    except Exception:
        pass
    return None


def _graph_close_session(token: str, session_id: str) -> None:
    """Close Excel session cleanly."""
    try:
        _requests.post(f"{GRAPH_WB_BASE}/closeSession",
                       headers=_graph_session_headers(token, session_id),
                       json={}, timeout=30)
    except Exception:
        pass


def _graph_list_sheets(token: str, session_id: str) -> list:
    """Return list of worksheet names in main workbook."""
    try:
        r = _requests.get(f"{GRAPH_WB_BASE}/worksheets",
                          headers=_graph_session_headers(token, session_id),
                          timeout=30)
        if r.status_code == 200:
            return [ws["name"] for ws in r.json().get("value", [])]
    except Exception:
        pass
    return []


# v3.1.0: Sheet creation helpers for main workbook
# Main workbook column headers (row 2), verified 19 May 2026:
MAIN_SHEET_HEADERS = [
    "Machine", "area", "Reason", "Mechanical / Electrical",
    "Supplier", "Parts ordered", "Job No", "Order No",
    "", "Ordered", "Due", "Fully Delivered",
    "Cost Initiatives", "Original supplier",
]


def _graph_create_worksheet(token: str, session_id: str,
                             sheet_name: str) -> bool:
    """
    v3.1.0: Create a new blank worksheet in the main workbook.
    Sheets are always added at the end — call in chronological order
    to maintain correct tab ordering (Jan → Feb → Mar → ...).
    """
    try:
        r = _requests.post(
            f"{GRAPH_WB_BASE}/worksheets/add",
            headers=_graph_session_headers(token, session_id),
            json={"name": sheet_name},
            timeout=30,
        )
        return r.status_code in (200, 201)
    except Exception:
        return False


def _graph_init_sheet_headers(token: str, session_id: str,
                               sheet_name: str) -> bool:
    """
    v3.1.1: Write standard headers to row 2 AND SUM formulas to row 1
    of a new monthly sheet, matching the format of all existing sheets.

    Row 1: I1=total spend formula, M1=total savings formula
    Row 2: column headers (Machine, area, Reason, ...)
    Row 3+: data (PO rows)
    """
    sheet_q = _requests.utils.quote(sheet_name)

    # Step 1: Write column headers to row 2
    try:
        r1 = _requests.patch(
            f"{GRAPH_WB_BASE}/worksheets/{sheet_q}"
            f"/range(address='A2:N2')",
            headers=_graph_session_headers(token, session_id),
            json={"values": [MAIN_SHEET_HEADERS]},
            timeout=30,
        )
        if r1.status_code != 200:
            return False
    except Exception:
        return False

    # Step 2: Write SUM formulas to row 1 (I1=total spend, M1=total savings)
    # 14 columns (A:N). Formula cells: I(index 8) and M(index 12).
    row1_formulas = [""] * 14
    row1_formulas[8]  = "=SUM(I3:I5000)"    # I1: total monthly spend
    row1_formulas[12] = "=SUM(M3:M5000)"    # M1: total cost initiatives/savings

    row1_formats = [""] * 14
    row1_formats[8]  = "\u00a3#,##0.00"    # £ format for I1
    row1_formats[12] = "\u00a3#,##0.00"    # £ format for M1

    try:
        r2 = _requests.patch(
            f"{GRAPH_WB_BASE}/worksheets/{sheet_q}"
            f"/range(address='A1:N1')",
            headers=_graph_session_headers(token, session_id),
            json={
                "formulas":     [row1_formulas],
                "numberFormat": [row1_formats],
            },
            timeout=30,
        )
        return r2.status_code == 200
    except Exception:
        return False


def _graph_ensure_main_sheets_through(token: str, session_id: str,
                                       existing_sheets: list,
                                       target_sheet_name: str) -> list:
    """
    v3.1.0: Create all missing monthly sheets from the month after the
    last existing monthly sheet up to and including target_sheet_name.

    Key behaviour (matches local ensure_month_sheets_through):
      - Creates sheets in strict chronological order (Jan → Feb → Mar → ...)
      - Never skips a month — no gaps in the tab bar
      - Each new sheet gets headers on row 2 matching the existing format
      - Returns list of sheet names that were created

    Example: main has sheets through December 2026, target = April 2027.
    Creates: January 2027, February 2027, March 2027, April 2027 — in that
    order — so the tab bar stays chronological.

    George's note: in practice, next-year sheets are created manually at
    year-end, so this rarely fires. It's a safety net for edge cases.
    """
    try:
        target_dt = datetime.strptime(target_sheet_name.strip(),
                                      "%B %Y").date()
    except ValueError:
        print(f"[push] Cannot parse target sheet name '{target_sheet_name}'")
        return []

    # Find the latest monthly sheet that already exists in main
    latest_existing = date(datetime.today().year,
                           datetime.today().month, 1)  # fallback
    for s in existing_sheets:
        try:
            d = datetime.strptime(s.strip(), "%B %Y").date()
            if d > latest_existing:
                latest_existing = d
        except ValueError:
            continue

    # Iterate month by month from latest+1 to target (inclusive)
    cur = _add_one_month(latest_existing)
    target_first = date(target_dt.year, target_dt.month, 1)
    created = []

    while cur <= target_first:
        sheet_name = cur.strftime("%B %Y")
        if sheet_name not in existing_sheets:
            print(f"[push] Creating missing sheet '{sheet_name}' in main...")
            ok = _graph_create_worksheet(token, session_id, sheet_name)
            if ok:
                _graph_init_sheet_headers(token, session_id, sheet_name)
                existing_sheets.append(sheet_name)
                created.append(sheet_name)
                print(f"[push] '{sheet_name}' created ✓")
            else:
                print(f"[push] WARNING: Could not create '{sheet_name}' "
                      f"— aborting sheet creation.")
                break
        cur = _add_one_month(cur)

    return created


def _graph_get_next_row(token: str, session_id: str, sheet_name: str) -> int:
    """
    v3.0.1: Return the first row where column A (Machine) is empty, starting
    from row 3 (after header in row 2).

    Why not just usedRange.rowCount + 1? Because the main workbook often has
    GAPS in the data — late entries get placed far below the contiguous block
    of recent POs. Example seen 19 May 2026: rows 101-114 filled with current
    POs, rows 115-124 empty, row 125 has a late entry from another user.
    The correct target is row 115 (first gap), not row 126.

    Algorithm:
      1. GET usedRange → know how far to scan
      2. GET column A from row 3 to last_row (one batched call)
      3. Return the first row where column A is empty/whitespace
      4. If no gap found within usedRange → append after (rowCount + 1)
    """
    try:
        # Step 1: Get used range to bound the scan
        sheet_q = _requests.utils.quote(sheet_name)
        used_r = _requests.get(
            f"{GRAPH_WB_BASE}/worksheets/{sheet_q}/usedRange",
            headers=_graph_session_headers(token, session_id),
            timeout=30,
        )
        if used_r.status_code != 200:
            return 3
        last_row = used_r.json().get("rowCount", 2)

        # Empty sheet → write right after header
        if last_row < 3:
            return 3

        # Step 2: Read column A values from row 3 to last_row in one call
        a_r = _requests.get(
            f"{GRAPH_WB_BASE}/worksheets/{sheet_q}"
            f"/range(address='A3:A{last_row}')",
            headers=_graph_session_headers(token, session_id),
            timeout=30,
        )
        if a_r.status_code != 200:
            return last_row + 1  # safe fallback

        values = a_r.json().get("values", [])
        for i, row_arr in enumerate(values):
            cell = row_arr[0] if row_arr else None
            # Empty if None, empty string, or whitespace-only string
            if cell is None or cell == "":
                return 3 + i
            if isinstance(cell, str) and not cell.strip():
                return 3 + i

        # Step 3: No gap found → append after the last used row
        return last_row + 1

    except Exception:
        return 3


def _graph_verify_cell_empty(token: str, session_id: str,
                              sheet_name: str, row_index: int) -> bool:
    """
    v3.0.3: Race condition defense. Just before writing to A{row_index},
    re-read that single cell to confirm it's still empty. Closes the small
    window (~1-3 seconds) where another user could have started typing into
    the same row between our usedRange scan and our PATCH.

    Returns True if empty (safe to write), False if data is now there.
    Fails OPEN (returns True) on network errors — we don't want a flaky
    Graph endpoint to block legitimate writes.
    """
    try:
        sheet_q = _requests.utils.quote(sheet_name)
        r = _requests.get(
            f"{GRAPH_WB_BASE}/worksheets/{sheet_q}"
            f"/range(address='A{row_index}')",
            headers=_graph_session_headers(token, session_id),
            timeout=15,
        )
        if r.status_code != 200:
            return True  # fail open
        vals = r.json().get("values", [[None]])
        cell = vals[0][0] if vals and vals[0] else None
        if cell is None or cell == "":
            return True
        if isinstance(cell, str) and not cell.strip():
            return True
        return False  # cell now has data — race detected
    except Exception:
        return True  # fail open


# v3.0.3: how many times to re-find an empty row if a race is detected
GRAPH_PUSH_MAX_RETRIES = 3


def _graph_write_row(token: str, session_id: str, sheet_name: str,
                     row_index: int, values: list,
                     number_formats: list = None) -> bool:
    """
    Write a single row to the main workbook at row_index (1-based).
    v3.0.2: also applies per-cell number formats so dates render as
    'dd/mm/yyyy' (not raw serials like 46162) and cost as '£#,##0.00'.
    """
    last_col = _graph_col_letter(len(values))
    address = f"A{row_index}:{last_col}{row_index}"
    body = {"values": [values]}
    if number_formats:
        body["numberFormat"] = [number_formats]
    try:
        r = _requests.patch(
            f"{GRAPH_WB_BASE}/worksheets/"
            f"{_requests.utils.quote(sheet_name)}"
            f"/range(address='{address}')",
            headers=_graph_session_headers(token, session_id),
            json=body, timeout=30,
        )
        return r.status_code == 200
    except Exception:
        return False


def _graph_build_row(order: dict, po_number: str, net_total: float) -> list:
    """
    Build the 14-column row (A:N) for the main Parts Ordered workbook.

    v3.0.2 FIX: corrected order dict key names to match what collect_po_data()
    actually populates:
      - "items_summary" (not "parts_ordered") for column F
      - "savings_raw" → parsed float for column M (Cost Initiatives)
      - "orig_supplier" for column N (Original supplier)

    Column mapping (verified 19 May 2026 against live main workbook):
      A Machine  B area  C Reason  D Mech/Elec  E Supplier  F Parts ordered
      G Job No   H Order No  I Cost(£)  J Ordered(serial)  K Due(serial)
      L Fully Delivered(empty)  M Cost Initiatives  N Original supplier
    """
    today = datetime.today().date()

    # F: Parts description (already includes part numbers, built by collect_po_data)
    parts_desc = order.get("items_summary", "") or ""

    # K: Due date as Excel serial
    due_dt = parse_ddmmyyyy_to_date(order.get("due_raw", "") or "")

    # M: Savings (parse from raw string — empty if not set)
    savings_val = parse_currency_to_float(order.get("savings_raw", "") or "")
    savings_cell = float(savings_val) if savings_val is not None else ""

    # N: Original supplier (only set when user provides cross-reference)
    orig_supplier = order.get("orig_supplier", "") or ""

    return [
        order.get("machine", ""),                            # A Machine
        order.get("area", ""),                               # B area
        order.get("reason", ""),                             # C Reason
        order.get("mech_elec", ""),                          # D Mech/Elec
        order.get("supplier", ""),                           # E Supplier
        parts_desc,                                          # F Parts ordered
        order.get("job_no", ""),                             # G Job No
        po_number,                                           # H Order No
        net_total,                                           # I Cost
        _date_to_excel_serial(today),                        # J Ordered
        _date_to_excel_serial(due_dt) if due_dt else "",     # K Due
        "",                                                  # L Fully Delivered
        savings_cell,                                        # M Cost Initiatives
        orig_supplier,                                       # N Original supplier
    ]


# v3.0.2: Per-column number formats for main workbook writes.
# None = leave whatever the cell currently has (Graph treats null as no-change).
GRAPH_ROW_NUMBER_FORMATS = [
    None,           # A Machine
    None,           # B area
    None,           # C Reason
    None,           # D Mech/Elec
    None,           # E Supplier
    None,           # F Parts ordered
    None,           # G Job No
    None,           # H Order No
    "£#,##0.00",    # I Cost — currency with two decimals
    "dd/mm/yyyy",   # J Ordered — date
    "dd/mm/yyyy",   # K Due — date
    "dd/mm/yyyy",   # L Fully Delivered — pre-format for when George fills it later
    "£#,##0.00",    # M Cost Initiatives — currency
    None,           # N Original supplier
]


def _graph_format_new_row(token: str, session_id: str,
                          sheet_name: str, row_index: int,
                          items_text: str = "") -> bool:
    """
    v3.1.3 final — apply wrap text + explicit row height after row write.

    Diagnostic learning (22 May 2026): Excel Online's Graph API only
    accepts format updates via dedicated /format endpoints. Sending
    {"format": {"rowHeight": X}} as part of a range PATCH returns
    400 InvalidArgument. The /range/format endpoint and full-row
    /range(N:N)/format work correctly.

    Three calls:
      1. PATCH /range(A:N)/format    → wrapText: true on the cells
      2. PATCH /range(N:N)/format    → rowHeight: calculated on the row
      3. POST  /range/format/autofitRows → belt-and-braces (only grows)

    Row height calculation matches the local openpyxl fix: ~70 chars
    per visual line in column F, 15pt per line (Excel default).
    """
    sheet_q = _requests.utils.quote(sheet_name)
    cells_address = f"A{row_index}:N{row_index}"
    row_address = f"{row_index}:{row_index}"

    # Calculate target row height from items text length
    chars_per_line = 70
    if items_text:
        items = items_text.split("; ")
        total_lines = sum(
            max(1, (len(it) + chars_per_line - 1) // chars_per_line)
            for it in items
        )
    else:
        total_lines = 1
    row_height = max(15, total_lines * 15)

    overall_ok = True

    # Step 1: wrapText on cells (via /format endpoint)
    try:
        r1 = _requests.patch(
            f"{GRAPH_WB_BASE}/worksheets/{sheet_q}"
            f"/range(address='{cells_address}')/format",
            headers=_graph_session_headers(token, session_id),
            json={"wrapText": True},
            timeout=20,
        )
        if r1.status_code != 200:
            print(f"[push] wrapText PATCH returned {r1.status_code}")
            overall_ok = False
    except Exception as e:
        print(f"[push] wrapText PATCH exception: {e}")
        overall_ok = False

    # Step 2: rowHeight on the row (via full-row /format endpoint)
    try:
        r2 = _requests.patch(
            f"{GRAPH_WB_BASE}/worksheets/{sheet_q}"
            f"/range(address='{row_address}')/format",
            headers=_graph_session_headers(token, session_id),
            json={"rowHeight": row_height},
            timeout=20,
        )
        if r2.status_code != 200:
            print(f"[push] rowHeight PATCH returned {r2.status_code}")
            overall_ok = False
    except Exception as e:
        print(f"[push] rowHeight PATCH exception: {e}")
        overall_ok = False

    # Step 3: autofitRows fallback — only grows, never shrinks, so safe
    try:
        _requests.post(
            f"{GRAPH_WB_BASE}/worksheets/{sheet_q}"
            f"/range(address='{cells_address}')/format/autofitRows",
            headers=_graph_session_headers(token, session_id),
            json={}, timeout=20,
        )
    except Exception:
        pass  # non-critical — height already explicitly set

    return overall_ok

    return True


def push_po_to_main(order: dict, po_number: str, net_total: float,
                    target_sheet: str) -> bool:
    """
    v3.0.0: Push a newly created PO row to the main Parts Ordered workbook
    on SharePoint via Microsoft Graph API.

    Graceful by design — never blocks or aborts PO creation:
      - GRAPH_PUSH_ENABLED = False  → silent skip
      - msal/requests missing       → warning, manual copy reminder
      - Token unavailable           → warning, manual copy reminder
      - Sheet missing in main       → warning, manual copy reminder
      - Any API error               → warning, manual copy reminder
    """
    if not GRAPH_PUSH_ENABLED:
        return False

    if not GRAPH_AVAILABLE:
        print("\n[push] SKIP: pip install msal requests first.")
        return False

    print(f"\n[push] Pushing {po_number} to main workbook...")

    token = _graph_get_token_silent()
    if not token:
        print("[push] WARNING: No auth token. Run oauth_test.py once to authenticate.")
        print("[push] Manual copy to main workbook required.")
        return False

    session_id = _graph_create_session(token, persist=True)
    if not session_id:
        print("[push] WARNING: Could not open Excel session on SharePoint.")
        print("[push] Manual copy to main workbook required.")
        return False

    try:
        sheets = _graph_list_sheets(token, session_id)
        if target_sheet not in sheets:
            # v3.1.0: Sheet missing → create all intermediate + target in order
            print(f"[push] Sheet '{target_sheet}' not in main workbook.")
            created = _graph_ensure_main_sheets_through(
                token, session_id, sheets, target_sheet)
            if target_sheet not in sheets:
                print(f"[push] WARNING: Could not create '{target_sheet}'. "
                      f"Manual copy required.")
                return False
            if created:
                print(f"[push] Created {len(created)} new sheet(s): "
                      f"{', '.join(created)}")

        next_row = _graph_get_next_row(token, session_id, target_sheet)

        # v3.0.3: race condition defense — verify the target row is still
        # empty just before writing. Closes the small race window where
        # another user could have started typing into the same row between
        # our usedRange scan and our PATCH.
        for attempt in range(GRAPH_PUSH_MAX_RETRIES):
            if _graph_verify_cell_empty(token, session_id,
                                         target_sheet, next_row):
                break  # Safe to write
            print(f"[push] Row {next_row} no longer empty (race detected, "
                  f"attempt {attempt + 1}/{GRAPH_PUSH_MAX_RETRIES}). "
                  f"Finding new row...")
            next_row = _graph_get_next_row(token, session_id, target_sheet)
        else:
            # All retries exhausted — give up
            print(f"[push] WARNING: Could not find a stable empty row after "
                  f"{GRAPH_PUSH_MAX_RETRIES} attempts. Other users may be "
                  f"actively editing.")
            print(f"[push] Manual copy to main workbook required.")
            return False

        row_values = _graph_build_row(order, po_number, net_total)
        success = _graph_write_row(token, session_id, target_sheet,
                                   next_row, row_values,
                                   number_formats=GRAPH_ROW_NUMBER_FORMATS)
        if success:
            # v3.1.2: pass items_summary so format_new_row can calculate
            # explicit row height rather than relying on unreliable autofitRows.
            items_text = order.get("items_summary", "") or ""
            fmt_ok = _graph_format_new_row(token, session_id,
                                           target_sheet, next_row,
                                           items_text=items_text)
            if fmt_ok:
                print(f"[push] ✓ Pushed to main '{target_sheet}' row {next_row} "
                      f"— no manual copy needed.")
            else:
                print(f"[push] ✓ Pushed to main '{target_sheet}' row {next_row} "
                      f"(data written, wrap text formatting skipped).")
            return True
        else:
            print("[push] WARNING: Write failed. Manual copy required.")
            return False

    except Exception as e:
        print(f"[push] WARNING: {e}. Manual copy required.")
        return False
    finally:
        _graph_close_session(token, session_id)


# v2.1.1: Show recent POs across all monthly sheets
def show_recent_pos_across_sheets():
    """
    Walk every monthly sheet in the local workbook, find POs whose Date Ordered
    falls in the last N days, and print them grouped by sheet.

    Designed as a manual end-of-day helper: with v2.1.0 future-month routing,
    POs may land on multiple sheets (current + future months). This makes it
    easy to see at a glance what needs to be copied to the main workbook
    without hunting through individual sheets.

    Read-only: does not modify the workbook.
    """
    # 1) Ask for the time window
    print("\n--- Recent POs across all sheets ---")
    raw = input("How many days back? (default 7, press Enter to accept): ").strip()
    if not raw:
        days_back = 7
    else:
        try:
            days_back = int(raw)
            if days_back < 1 or days_back > 365:
                print(f"Out of range (1-365); using 7.")
                days_back = 7
        except ValueError:
            print(f"Invalid input '{raw}'; using 7.")
            days_back = 7

    # v2.1.2: Ask whether to filter out POs already in the main workbook
    filter_raw = input("Filter out POs already in main workbook? (Y/n, default Y): ").strip().lower()
    filter_pushed = filter_raw not in ("n", "no")

    today = datetime.today().date()
    cutoff = today - timedelta(days=days_back - 1)  # inclusive: today minus (N-1)

    print(f"\nSearching POs ordered between {cutoff.strftime('%d/%m/%Y')} "
          f"and {today.strftime('%d/%m/%Y')} (inclusive)...")

    # v2.1.2/3: Build set of PO numbers already in main (for filtering)
    # v2.1.3: Uses session cache + optimized iter_rows scan
    main_po_set = set()
    main_status = ""
    if filter_pushed:
        # Ask whether to force refresh the cache (default: use cache if valid)
        if _MAIN_PO_CACHE is not None:
            refresh_raw = input("Refresh main workbook cache? (y/N, default N): ").strip().lower()
            force_refresh = refresh_raw in ("y", "yes")
        else:
            force_refresh = False

        po_set_or_none, err = _get_main_po_numbers_cached(force_refresh=force_refresh)
        if po_set_or_none is not None:
            main_po_set = po_set_or_none
            main_status = f"Found {len(main_po_set)} PO(s) in main workbook (will be filtered out)."
        else:
            main_status = f"WARNING: {err}. Showing all POs (filter disabled)."
            filter_pushed = False
        print(main_status)

    # 2) Load the workbook read-only
    if not os.path.exists(WORKBOOK_PATH):
        print(f"\nERROR: Workbook not found at {WORKBOOK_PATH}")
        return

    try:
        wb = load_workbook(WORKBOOK_PATH, data_only=True)
    except Exception as e:
        print(f"\nERROR loading workbook: {e}")
        return

    # 3) Identify monthly sheets
    month_sheets = [s for s in wb.sheetnames if _is_valid_month_sheet(s)]
    if not month_sheets:
        print("\nNo monthly sheets found in the workbook.")
        return

    # 4) Scan each sheet for in-window POs
    results_by_sheet = {}   # sheet_name -> list of dicts
    total_found = 0
    total_skipped_pushed = 0  # v2.1.2: counter for filtered POs
    skipped_sheets = []

    for sheet_name in month_sheets:
        ws = wb[sheet_name]
        # Use detect_ directly (not ensure_) to keep this READ-ONLY
        header_row, colmap = detect_header_row_and_columns(ws)
        if not colmap or not header_row:
            skipped_sheets.append(sheet_name)
            continue

        po_col = colmap.get("PONumber")
        date_col = colmap.get("DateOrdered")
        if not po_col or not date_col:
            skipped_sheets.append(sheet_name)
            continue

        supp_col = colmap.get("Supplier")
        mach_col = colmap.get("Machine")
        cost_col = colmap.get("Cost")
        due_col = colmap.get("DueDate")

        rows_in_window = []
        data_start = header_row + 1
        max_row = ws.max_row or data_start

        for r in range(data_start, max_row + 1):
            po_val = ws[f"{po_col}{r}"].value
            if is_blank_cell(po_val):
                continue

            date_val = ws[f"{date_col}{r}"].value
            ordered_date = None
            if isinstance(date_val, datetime):
                ordered_date = date_val.date()
            elif isinstance(date_val, date):
                ordered_date = date_val
            elif isinstance(date_val, str):
                ordered_date = parse_ddmmyyyy_to_date(date_val)

            if not ordered_date:
                continue

            if cutoff <= ordered_date <= today:
                # v2.1.2: skip if already pushed to main
                po_str = str(po_val).strip()
                if filter_pushed and po_str in main_po_set:
                    total_skipped_pushed += 1
                    continue

                supp = ws[f"{supp_col}{r}"].value if supp_col else ""
                mach = ws[f"{mach_col}{r}"].value if mach_col else ""
                cost = ws[f"{cost_col}{r}"].value if cost_col else 0
                due_val = ws[f"{due_col}{r}"].value if due_col else None

                # Normalise due date for display
                due_display = ""
                if isinstance(due_val, (datetime, date)):
                    due_display = due_val.strftime("%d/%m/%Y") if isinstance(due_val, datetime) else due_val.strftime("%d/%m/%Y")
                elif isinstance(due_val, str) and due_val.strip():
                    due_display = due_val.strip()

                try:
                    cost_f = float(cost) if cost is not None else 0.0
                except (TypeError, ValueError):
                    cost_f = 0.0

                rows_in_window.append({
                    "row": r,
                    "po": str(po_val).strip(),
                    "date_ordered": ordered_date,
                    "due_display": due_display,
                    "supplier": str(supp or "").strip(),
                    "machine": str(mach or "").strip(),
                    "cost": cost_f,
                })

        if rows_in_window:
            # Sort within sheet by date asc, then by PO
            rows_in_window.sort(key=lambda r: (r["date_ordered"], r["po"]))
            results_by_sheet[sheet_name] = rows_in_window
            total_found += len(rows_in_window)

    # 5) Print results
    if not results_by_sheet:
        if filter_pushed and total_skipped_pushed > 0:
            print(f"\nNo POs to copy — all {total_skipped_pushed} in window are already in main workbook.")
        else:
            print(f"\nNo POs found ordered in the last {days_back} day(s).")
        if skipped_sheets:
            print(f"(Skipped {len(skipped_sheets)} sheet(s) with no readable headers.)")
        return

    bar = "=" * 95
    print(f"\n{bar}")
    title = f"POs ordered in last {days_back} day(s) — {total_found} to copy across {len(results_by_sheet)} sheet(s)"
    if filter_pushed and total_skipped_pushed > 0:
        title += f"  (+ {total_skipped_pushed} already in main, skipped)"
    print(title)
    print(bar)

    grand_total = 0.0
    current_month_name = datetime.today().strftime("%B %Y")

    sorted_sheet_names = sorted(results_by_sheet.keys(), key=_month_sheet_sort_key)

    for sheet_name in sorted_sheet_names:
        rows = results_by_sheet[sheet_name]
        sheet_total = sum(r["cost"] for r in rows)
        grand_total += sheet_total

        if sheet_name == current_month_name:
            marker = " (CURRENT MONTH)"
        elif _is_future_sheet(sheet_name):
            marker = " (FUTURE — service contracts)"
        else:
            marker = " (past month)"

        print(f"\n--- {sheet_name}{marker}  |  {len(rows)} PO(s)  |  £{sheet_total:,.2f} ---")
        # Header
        print(f"  {'PO Number':<18} {'Ordered':<11} {'Age':<10} "
              f"{'Due':<11} {'Supplier':<25} {'Machine':<18} {'Cost':>11}")
        print(f"  {'-'*18} {'-'*11} {'-'*10} {'-'*11} {'-'*25} {'-'*18} {'-'*11}")

        for r in rows:
            age = (today - r["date_ordered"]).days
            if age == 0:
                age_lbl = "TODAY"
            elif age == 1:
                age_lbl = "yesterday"
            else:
                age_lbl = f"{age}d ago"

            print(f"  {r['po'][:18]:<18} "
                  f"{r['date_ordered'].strftime('%d/%m/%Y'):<11} "
                  f"{age_lbl:<10} "
                  f"{r['due_display'][:11]:<11} "
                  f"{r['supplier'][:25]:<25} "
                  f"{r['machine'][:18]:<18} "
                  f"£{r['cost']:>10,.2f}")

    print(f"\n{bar}")
    if filter_pushed and total_skipped_pushed > 0:
        print(f"GRAND TOTAL TO COPY: {total_found} PO(s)  |  £{grand_total:,.2f}")
        print(f"(Already in main: {total_skipped_pushed} PO(s) — not shown)")
    else:
        print(f"GRAND TOTAL: {total_found} PO(s)  |  £{grand_total:,.2f}")
    print(bar)

    if skipped_sheets:
        print(f"\n(Skipped {len(skipped_sheets)} sheet(s) with no readable headers: "
              f"{', '.join(skipped_sheets[:5])}"
              f"{'...' if len(skipped_sheets) > 5 else ''})")

    print("\nTip: copy these entries to the main Parts Ordered workbook to keep")
    print("month-end totals in sync. Once Graph API admin consent is granted,")
    print("this push will happen automatically.")


def main():
    ensure_dir(PO_PDF_OUTPUT_DIR)

    while True:
        
        print("\n=== PO FLOW v3.1.3 ===")
        print("1) Create new Purchase Order (PO)")
        print("2) Mark Purchase Order (PO) as Received (Goods Receipt (GR))")
        print("3) Full Sync with Main Sheet (deliveries + import + health check)")
        print("4) Diagnose Main Workbook (raw cell dump)")
        print("5) Compare Month (main vs local)")
        print("6) Recalculate All Totals (fix monthly sheet I1/M1 cells)")
        print("7) Reorder Check (analyse recurring orders)")
        print("8) Show recent POs across all sheets (manual push-to-main helper)")
        print("9) Exit")
        choice = input("Choose 1-9: ").strip()

        if choice == "1":
            create_new_po()
        elif choice == "2":
            receive_po()
        elif choice == "3":
            sync_with_main_sheet()
        elif choice == "4":
            diagnose_main_workbook()
        elif choice == "5":
            compare_month()
        elif choice == "6":
            recalculate_all_totals()
        elif choice == "7":
            reorder_check()
        elif choice == "8":
            show_recent_pos_across_sheets()
        elif choice == "9":
            break
        else:
            print("Invalid choice.")

if __name__ == "__main__":
    main()