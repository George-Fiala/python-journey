"""
PO Flow GUI — Alpha v0.1
Desktop application for Purchase Order workflow automation.

Requirements:
  - PO_FLOW_v3_1_1.py in the same directory (or update PO_FLOW_FILE below)
  - pip install msal requests openpyxl reportlab

To run WITHOUT a black terminal window appearing:
  pythonw.exe po_flow_gui.py   (use pythonw, not python)

Or double-click po_flow_gui.pyw after renaming.
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import threading
import os
import sys
import json
from pathlib import Path
from datetime import datetime, date

# ──────────────────────────────────────────────
# IMPORT FROM MAIN PO FLOW FILE
# ──────────────────────────────────────────────
PO_FLOW_FILE = "PO_FLOW_v3_1_3.py"

PO_FLOW_LOADED = False
_pf = None

try:
    import importlib.util
    _dir = os.path.dirname(os.path.abspath(__file__))
    _path = os.path.join(_dir, PO_FLOW_FILE)
    if not os.path.exists(_path):
        raise FileNotFoundError(f"Cannot find {PO_FLOW_FILE} in {_dir}")
    spec = importlib.util.spec_from_file_location("po_flow_core", _path)
    _pf = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(_pf)
    PO_FLOW_LOADED = True
except Exception as _e:
    PO_FLOW_LOADED = False
    _LOAD_ERROR = str(_e)

# ──────────────────────────────────────────────
# DESIGN TOKENS
# ──────────────────────────────────────────────
C = {
    "bg_dark":   "#1C2B3A",   # Header / dark sections
    "bg_mid":    "#243447",   # Sidebar / secondary
    "bg_light":  "#F0F3F7",   # Main content background
    "bg_card":   "#FFFFFF",   # Card / form background
    "primary":   "#2471A3",   # Primary blue
    "primary_h": "#1A5276",   # Primary hover
    "success":   "#1E8449",   # Green
    "warning":   "#D68910",   # Amber
    "error":     "#C0392B",   # Red
    "accent":    "#CA6F1E",   # Orange accent (industrial)
    "txt_dark":  "#1C2B3A",   # Main text
    "txt_mid":   "#5D6D7E",   # Secondary text
    "txt_light": "#FDFEFE",   # Light text on dark bg
    "border":    "#D5DBDB",   # Input borders
    "header_h":  "60",        # Header height px (as string for grid)
}

FONT = {
    "title":   ("Segoe UI", 16, "bold"),
    "heading": ("Segoe UI", 12, "bold"),
    "body":    ("Segoe UI", 10),
    "small":   ("Segoe UI", 9),
    "mono":    ("Consolas",  10),
    "btn":     ("Segoe UI", 10, "bold"),
    "status":  ("Segoe UI", 9),
}

REASONS = [
    "Stock Replenishment", "Repair", "Breakdown", "Service",
    "Project", "job/stock", "Consumables", "Other"
]
MECH_ELEC = ["Mech", "Elec", "Mech/ Elec", "Pneumatic", "Other"]


# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────
def get_suppliers_list() -> list:
    """Load supplier names from the JSON file used by PO Flow.
    Handles structure: {"suppliers": [{"name": "...", ...}, ...]}
    """
    if not PO_FLOW_LOADED:
        return []
    try:
        sup_path = getattr(_pf, "SUPPLIERS_JSON_PATH", None)
        if sup_path and os.path.exists(sup_path):
            with open(sup_path, encoding="utf-8") as f:
                data = json.load(f)
            # Most common structure: {"suppliers": [{"name": ...}, ...]}
            if isinstance(data, dict) and isinstance(data.get("suppliers"), list):
                return sorted(s.get("name", "") for s in data["suppliers"]
                              if s.get("name"))
            # Alt structure: plain list of supplier dicts
            if isinstance(data, list):
                return sorted(s.get("name", "") for s in data if s.get("name"))
            # Fallback: dict of name → details
            if isinstance(data, dict):
                return sorted(data.keys())
    except Exception as e:
        print(f"[gui] Could not load suppliers: {e}")
    return []


def get_pending_pos() -> list:
    """
    Read the To Receive sheet and return list of pending POs (Status != Received).
    Each dict has: po_number, supplier, machine, area, date_ordered, cost, sheet_hint
    """
    if not PO_FLOW_LOADED:
        return []
    try:
        from openpyxl import load_workbook
        pf = _pf
        wb_path = getattr(pf, "WORKBOOK_PATH", None)
        if not wb_path or not os.path.exists(wb_path):
            return []
        wb = load_workbook(wb_path, data_only=True)
        if pf.TO_RECEIVE_SHEET not in wb.sheetnames:
            return []
        ws = wb[pf.TO_RECEIVE_SHEET]

        # Use TO_RECEIVE_HEADERS to find column indices
        headers = getattr(pf, "TO_RECEIVE_HEADERS", [])
        def col(name):
            return headers.index(name) + 1 if name in headers else None

        po_col = col("PO Number")
        sup_col = col("Supplier")
        mach_col = col("Machine")
        area_col = col("Area")
        date_col = col("Date Ordered")
        cost_col = col("Total Cost")
        status_col = col("Status")
        if not (po_col and status_col):
            return []

        result = []
        for r in range(2, (ws.max_row or 2) + 1):
            status = (ws.cell(row=r, column=status_col).value or "").strip().lower()
            if status == "received":
                continue
            po = ws.cell(row=r, column=po_col).value
            if not po:
                continue
            date_v = ws.cell(row=r, column=date_col).value if date_col else None
            date_str = ""
            if isinstance(date_v, (datetime, date)):
                d = date_v.date() if isinstance(date_v, datetime) else date_v
                date_str = d.strftime("%d/%m/%Y")
            elif isinstance(date_v, str):
                date_str = date_v[:10]
            cost_v = ws.cell(row=r, column=cost_col).value if cost_col else 0
            try:
                cost_f = float(cost_v) if cost_v else 0.0
            except (TypeError, ValueError):
                cost_f = 0.0
            result.append({
                "po_number":    str(po).strip(),
                "supplier":     str(ws.cell(row=r, column=sup_col).value or "")
                                if sup_col else "",
                "machine":      str(ws.cell(row=r, column=mach_col).value or "")
                                if mach_col else "",
                "area":         str(ws.cell(row=r, column=area_col).value or "")
                                if area_col else "",
                "date_ordered": date_str,
                "cost":         cost_f,
            })
        # Sort newest first
        result.sort(key=lambda x: x["date_ordered"], reverse=True)
        return result
    except Exception as e:
        print(f"[gui] get_pending_pos error: {e}")
        return []


def mark_received_workflow(po_number: str, invoice_number: str = "") -> dict:
    """
    Mark PO as received in local workbook + push to main SharePoint.
    Returns dict with results.
    """
    pf = _pf
    today = date.today()
    out = {
        "success":     False,
        "po_number":   po_number,
        "sheet":       None,
        "row":         None,
        "to_recv_ok":  False,
        "push_ok":     False,
        "error":       None,
    }

    try:
        wb = pf.load_or_create_workbook(pf.WORKBOOK_PATH)

        # Step 1: Find PO in any monthly sheet, update L cell + green fill
        from openpyxl.styles import PatternFill
        green = getattr(pf, "GREEN_FILL", None)
        if green is None:
            green = PatternFill(start_color="92D050",
                                end_color="92D050",
                                fill_type="solid")

        found = False
        for sheet_name in wb.sheetnames:
            if sheet_name in {pf.TO_RECEIVE_SHEET, "Cost centre", "Sheet2"}:
                continue
            ws = wb[sheet_name]
            header_row, colmap = pf.detect_header_row_and_columns(ws)
            if not colmap or "PONumber" not in colmap \
               or "DeliveredDate" not in colmap:
                continue
            start_row = (header_row or 2) + 1
            for r in range(start_row, (ws.max_row or start_row) + 1):
                val = str(ws[f"{colmap['PONumber']}{r}"].value or "").strip()
                if val == po_number:
                    dd_cell = ws[f"{colmap['DeliveredDate']}{r}"]
                    dd_cell.value = today
                    dd_cell.number_format = "DD/MM/YYYY"
                    dd_cell.fill = green
                    out["sheet"] = sheet_name
                    out["row"] = r
                    found = True
                    break
            if found:
                break

        if not found:
            out["error"] = f"PO '{po_number}' not found in any monthly sheet."
            return out

        # Step 2: Update To Receive sheet (status + received date + invoice)
        if pf.TO_RECEIVE_SHEET in wb.sheetnames:
            ws_recv = wb[pf.TO_RECEIVE_SHEET]
            recv_row = pf.find_to_receive_row_by_po(ws_recv, po_number)
            if recv_row:
                ws_recv.cell(row=recv_row,
                             column=pf.to_receive_col_index("Status"),
                             value="Received")
                rd = ws_recv.cell(row=recv_row,
                                  column=pf.to_receive_col_index("Received Date"),
                                  value=today)
                rd.number_format = "DD/MM/YYYY"
                if invoice_number:
                    ws_recv.cell(row=recv_row,
                                 column=pf.to_receive_col_index("Invoice Number"),
                                 value=invoice_number)
                out["to_recv_ok"] = True

        # Step 3: Save local workbook
        pf.safe_save_workbook(wb, pf.WORKBOOK_PATH)

        # Step 4: Push to main SharePoint (best-effort, graceful)
        try:
            out["push_ok"] = push_received_to_main(
                po_number, out["sheet"], today)
        except Exception as e:
            print(f"[gui] push_received_to_main error: {e}")

        out["success"] = True
        return out

    except Exception as e:
        out["error"] = str(e)
        return out


def push_received_to_main(po_number: str, sheet_name: str,
                          delivered_date) -> bool:
    """
    Update the Fully Delivered (L) cell in the main SharePoint workbook
    for the given PO. Sets today's date + green fill.

    v0.2 — extensive logging to diagnose failures. Run with python.exe
    (not pythonw.exe) to see the diagnostic output in the terminal.
    """
    pf = _pf
    if not getattr(pf, "GRAPH_AVAILABLE", False):
        print("[push-recv] GRAPH_AVAILABLE is False — skipping")
        return False
    try:
        import requests as _rq

        print(f"[push-recv] Pushing received marker for {po_number} "
              f"to '{sheet_name}'...")

        token = pf._graph_get_token_silent()
        if not token:
            print("[push-recv] FAIL: could not get token")
            return False
        session_id = pf._graph_create_session(token)
        if not session_id:
            print("[push-recv] FAIL: could not create session")
            return False

        try:
            sheet_q = _rq.utils.quote(sheet_name)

            # Step 1: get used range to know how far to scan
            used = _rq.get(
                f"{pf.GRAPH_WB_BASE}/worksheets/{sheet_q}/usedRange",
                headers=pf._graph_session_headers(token, session_id),
                timeout=30,
            )
            if used.status_code != 200:
                print(f"[push-recv] FAIL: usedRange returned "
                      f"{used.status_code} — sheet '{sheet_name}' may not "
                      f"exist in main workbook")
                return False
            last_row = used.json().get("rowCount", 2)
            if last_row < 3:
                print(f"[push-recv] FAIL: sheet has no data "
                      f"(last_row={last_row})")
                return False

            # Step 2: read column H to find the PO
            h_range = _rq.get(
                f"{pf.GRAPH_WB_BASE}/worksheets/{sheet_q}"
                f"/range(address='H3:H{last_row}')",
                headers=pf._graph_session_headers(token, session_id),
                timeout=30,
            )
            if h_range.status_code != 200:
                print(f"[push-recv] FAIL: column H read returned "
                      f"{h_range.status_code}")
                return False

            values = h_range.json().get("values", [])
            target_row = None
            po_norm = po_number.strip()
            scanned = 0
            for i, row in enumerate(values):
                cell = row[0] if row else None
                if cell is None or cell == "":
                    continue
                scanned += 1
                if str(cell).strip() == po_norm:
                    target_row = i + 3
                    break

            if not target_row:
                # Helpful diagnostic: show first few POs we did find
                sample = [str(r[0]).strip() for r in values[:10]
                          if r and r[0]]
                print(f"[push-recv] FAIL: PO '{po_norm}' not found in "
                      f"'{sheet_name}' (scanned {scanned} non-empty rows). "
                      f"Sample of what's there: {sample}")
                return False

            print(f"[push-recv] Found PO at row {target_row}")

            # Step 3: single PATCH with value + numberFormat + format.fill
            # Combining everything into one PATCH is more reliable than
            # two separate calls (format.fill PATCH was sometimes failing).
            serial = pf._date_to_excel_serial(delivered_date)
            patch_body = {
                "values":       [[serial]],
                "numberFormat": [["dd/mm/yyyy"]],
                "format": {
                    "fill": {"color": "#92D050"}
                }
            }
            r1 = _rq.patch(
                f"{pf.GRAPH_WB_BASE}/worksheets/{sheet_q}"
                f"/range(address='L{target_row}')",
                headers=pf._graph_session_headers(token, session_id),
                json=patch_body,
                timeout=30,
            )

            if r1.status_code == 200:
                print(f"[push-recv] ✓ Updated L{target_row} with date "
                      f"+ green fill (single PATCH)")
                return True

            # Fallback: combined PATCH didn't work — split into two calls
            print(f"[push-recv] Combined PATCH returned {r1.status_code}, "
                  f"trying split approach...")

            # First: just value + numberFormat
            r1a = _rq.patch(
                f"{pf.GRAPH_WB_BASE}/worksheets/{sheet_q}"
                f"/range(address='L{target_row}')",
                headers=pf._graph_session_headers(token, session_id),
                json={"values": [[serial]],
                      "numberFormat": [["dd/mm/yyyy"]]},
                timeout=30,
            )
            if r1a.status_code != 200:
                print(f"[push-recv] FAIL: value PATCH returned "
                      f"{r1a.status_code}, body={r1a.text[:300]}")
                return False

            # Then: fill via dedicated /format/fill endpoint
            r1b = _rq.patch(
                f"{pf.GRAPH_WB_BASE}/worksheets/{sheet_q}"
                f"/range(address='L{target_row}')/format/fill",
                headers=pf._graph_session_headers(token, session_id),
                json={"color": "#92D050"},
                timeout=30,
            )
            if r1b.status_code != 200:
                print(f"[push-recv] Value written but fill PATCH returned "
                      f"{r1b.status_code}, body={r1b.text[:300]}")
                # Value was set, just no green fill — partial success
                return True  # treat as success — date is what matters most

            print(f"[push-recv] ✓ Updated L{target_row} with date "
                  f"+ green fill (split PATCH)")
            return True

        finally:
            try:
                pf._graph_close_session(token, session_id)
            except Exception:
                pass

    except Exception as e:
        print(f"[push-recv] EXCEPTION: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return False


def save_suppliers_db(suppliers_list: list) -> tuple:
    """
    Atomically save the suppliers list to the JSON file used by PO Flow.
    Creates a .bak backup of the previous version before overwriting.
    Returns (success: bool, error_message: str).
    """
    if not PO_FLOW_LOADED:
        return False, "PO Flow core not loaded"
    try:
        import shutil
        sup_path = getattr(_pf, "SUPPLIERS_JSON_PATH", None)
        if not sup_path:
            return False, "SUPPLIERS_JSON_PATH not configured"

        db = {"suppliers": suppliers_list}
        tmp_path = sup_path + ".tmp"
        bak_path = sup_path + ".bak"

        # Backup current file first (best-effort)
        if os.path.exists(sup_path):
            try:
                shutil.copy2(sup_path, bak_path)
            except Exception:
                pass

        # Write to .tmp then atomic rename
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(db, f, indent=2, ensure_ascii=False)
        os.replace(tmp_path, sup_path)
        return True, ""
    except Exception as e:
        return False, str(e)


def load_full_suppliers_db() -> list:
    """Return the full list of supplier dicts (all fields, not just names)."""
    if not PO_FLOW_LOADED:
        return []
    try:
        sup_path = getattr(_pf, "SUPPLIERS_JSON_PATH", None)
        if sup_path and os.path.exists(sup_path):
            with open(sup_path, encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict) and isinstance(data.get("suppliers"), list):
                return data["suppliers"]
            if isinstance(data, list):
                return data
    except Exception as e:
        print(f"[gui] Could not load suppliers DB: {e}")
    return []


def get_machines_list() -> list:
    """Return machine names. Derived from CC_LOOKUP keys so it stays in
    sync with whatever load_extended_cc_from_workbook adds at startup."""
    if PO_FLOW_LOADED:
        try:
            cc = getattr(_pf, "CC_LOOKUP", {})
            if cc:
                return sorted({m for (m, _) in cc.keys() if m})
        except Exception:
            pass
    return [
        "Workshop/Plant", "Service Contracts", "Emba 1", "Emba 2",
        "Emba 3", "Emba 4", "Emba 5", "Bobst 2 ExpertFold",
        "Asahi 1", "Asahi 2", "Vega Gluer", "Göpfert",
        "Robopac 3", "Mosca 3", "Compressor", "Mill",
    ]


def load_extended_cc_from_workbook() -> int:
    """
    Parse the 'Cost centre' sheet in the local workbook and add any
    missing entries to _pf.CC_LOOKUP at runtime. This way the GUI always
    picks up new machines added to the Cost centre sheet without needing
    to modify the main Python file.

    The sheet has a 2D grid of (Code, Name) column pairs grouped under
    "Bay" headers. We scan every cell pair: if left looks like a cost
    centre code (e.g. '350' or '350A'), the right cell is the name.

    Returns the number of new entries added.
    """
    if not PO_FLOW_LOADED:
        return 0
    try:
        import re
        from openpyxl import load_workbook
        wb_path = getattr(_pf, "WORKBOOK_PATH", None)
        if not wb_path or not os.path.exists(wb_path):
            return 0
        wb = load_workbook(wb_path, data_only=True)
        if "Cost centre" not in wb.sheetnames:
            return 0
        ws = wb["Cost centre"]

        code_re = re.compile(r"^\d{3}[A-Z]?$")
        cc_lookup = getattr(_pf, "CC_LOOKUP", None)
        if cc_lookup is None:
            return 0

        # First pass: collect all (code, name) pairs
        pairs = []  # list of (code_str, name_str)
        for row in ws.iter_rows(values_only=True):
            row_list = list(row)
            for i in range(len(row_list) - 1):
                code = row_list[i]
                name = row_list[i + 1]
                if code is None or name is None:
                    continue
                code_str = (str(int(code)) if isinstance(code, (int, float))
                            else str(code).strip())
                if not code_re.match(code_str):
                    continue
                name_str = str(name).strip()
                if name_str and not name_str.lower() in ("name", "code"):
                    pairs.append((code_str, name_str))

        # Second pass: build parent (3-digit code) → machine name map
        parent_map = {}
        for code, name in pairs:
            if len(code) == 3:
                parent_map[code] = name

        # Third pass: add to CC_LOOKUP
        added = 0
        for code, name in pairs:
            if len(code) == 3:
                # Machine root (general / no sub-area)
                key = (name, "")
                if key not in cc_lookup:
                    cc_lookup[key] = code
                    added += 1
            else:
                # Sub-area code: 3 digits + 1 letter
                parent_code = code[:3]
                parent_name = parent_map.get(parent_code)
                if not parent_name:
                    # Orphan code — register as standalone
                    key = (name, "")
                    if key not in cc_lookup:
                        cc_lookup[key] = code
                        added += 1
                    continue
                # Derive area = name minus parent prefix
                area = name
                if name.lower().startswith(parent_name.lower()):
                    area = name[len(parent_name):].strip(" -:()")
                if not area:
                    area = name
                key = (parent_name, area)
                if key not in cc_lookup:
                    cc_lookup[key] = code
                    added += 1

        return added
    except Exception as e:
        print(f"[gui] Could not load cost centre sheet: {e}")
        return 0


def get_areas_for_machine(machine: str) -> list:
    """
    Return the list of valid areas for a given machine, derived from
    CC_LOOKUP. E.g. 'Emba 1' returns ['(machine itself)', 'Pre Feeder',
    'Strapper', 'Palletiser']. The '(machine itself)' option maps to
    empty string area (the general cost centre for the machine).
    Falls back to KNOWN_AREAS if machine not found.
    """
    if not PO_FLOW_LOADED:
        return []
    try:
        cc_lookup = getattr(_pf, "CC_LOOKUP", {})
        known_areas = getattr(_pf, "KNOWN_AREAS", [])
        if machine:
            all_areas = {a for (m, a) in cc_lookup if m == machine}
            non_empty = sorted(a for a in all_areas if a)
            if "" in all_areas:
                return [GENERAL_AREA_LABEL] + non_empty
            elif non_empty:
                return non_empty
        return sorted(set(known_areas))
    except Exception:
        return []


# Label used in the area dropdown for the empty/general option.
# Stored as this string in the combobox, but converted to "" before passing
# to resolve_cost_centre (which expects empty string for general area).
GENERAL_AREA_LABEL = "(machine itself)"


def load_user_config() -> dict:
    cfg_path = Path.home() / ".po_flow_gui_config.json"
    defaults = {
        "name": getattr(_pf, "USER_NAME", "Unknown User") if PO_FLOW_LOADED else "User",
        "initials": getattr(_pf, "INITIALS", "XX") if PO_FLOW_LOADED else "XX",
    }
    if cfg_path.exists():
        try:
            with open(cfg_path, encoding="utf-8") as f:
                saved = json.load(f)
            defaults.update(saved)
        except Exception:
            pass
    return defaults


def save_user_config(cfg: dict):
    cfg_path = Path.home() / ".po_flow_gui_config.json"
    with open(cfg_path, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2)


# ──────────────────────────────────────────────
# STYLED BUTTON
# ──────────────────────────────────────────────
class ActionButton(tk.Frame):
    """A big colourful action button with icon text and subtitle."""

    def __init__(self, parent, icon, title, subtitle, color,
                 command=None, **kw):
        super().__init__(parent, bg=C["bg_card"], **kw)
        self.command = command
        self.color = color
        self.config(cursor="hand2")

        # Color strip on left
        strip = tk.Frame(self, bg=color, width=5)
        strip.pack(side="left", fill="y")

        # Content
        content = tk.Frame(self, bg=C["bg_card"], padx=14, pady=12)
        content.pack(side="left", fill="both", expand=True)

        icon_lbl = tk.Label(content, text=icon, font=("Segoe UI Emoji", 20),
                            bg=C["bg_card"], fg=color)
        icon_lbl.pack(side="left", padx=(0, 12))

        text_frame = tk.Frame(content, bg=C["bg_card"])
        text_frame.pack(side="left", fill="x", expand=True)

        tk.Label(text_frame, text=title, font=FONT["heading"],
                 bg=C["bg_card"], fg=C["txt_dark"],
                 anchor="w").pack(fill="x")
        tk.Label(text_frame, text=subtitle, font=FONT["small"],
                 bg=C["bg_card"], fg=C["txt_mid"],
                 anchor="w").pack(fill="x")

        arrow = tk.Label(content, text="›", font=("Segoe UI", 20),
                         bg=C["bg_card"], fg=C["border"])
        arrow.pack(side="right")

        # Hover effects
        for w in (self, content, icon_lbl, text_frame, arrow,
                  strip):
            w.bind("<Enter>", self._on_enter)
            w.bind("<Leave>", self._on_leave)
            w.bind("<Button-1>", self._on_click)
        for child in text_frame.winfo_children():
            child.bind("<Enter>", self._on_enter)
            child.bind("<Leave>", self._on_leave)
            child.bind("<Button-1>", self._on_click)

    def _on_enter(self, e=None):
        self.config(bg=C["bg_light"])
        for w in self.winfo_children():
            try:
                w.config(bg=C["bg_light"])
            except Exception:
                pass

    def _on_leave(self, e=None):
        self.config(bg=C["bg_card"])
        for w in self.winfo_children():
            try:
                w.config(bg=C["bg_card"])
            except Exception:
                pass

    def _on_click(self, e=None):
        if self.command:
            self.command()


# ──────────────────────────────────────────────
# DASHBOARD METRICS (live from main SharePoint)
# ──────────────────────────────────────────────
# Sheard's financial year runs 1 May → 30 April.
# We read column I (Cost) and column M (Cost Initiatives) from each
# monthly sheet in the current FY and aggregate them.

def _get_fy_start_year(today=None) -> int:
    """Return the year in which the current Sheard FY started.
    e.g. on 22 May 2026 → 2026 (FY26/27 started 1 May 2026).
         on 15 Mar 2026 → 2025 (FY25/26 started 1 May 2025).
    """
    today = today or date.today()
    return today.year if today.month >= 5 else today.year - 1


def get_fy_label(today=None) -> str:
    """e.g. 'FY26/27' for the current Sheard financial year."""
    start = _get_fy_start_year(today)
    return f"FY{start % 100:02d}/{(start + 1) % 100:02d}"


def get_current_fy_sheets(today=None) -> list:
    """List of monthly sheet names for the current FY, in chronological
    order, only including months that have already started (no future).
    e.g. on 22 May 2026 → ['May 2026']
         on 15 Aug 2026 → ['May 2026', 'June 2026', 'July 2026',
                            'August 2026']
    """
    today = today or date.today()
    start_year = _get_fy_start_year(today)
    sheets = []
    for offset in range(12):
        month_num = ((5 - 1 + offset) % 12) + 1  # 5..12, 1..4
        year = start_year if month_num >= 5 else start_year + 1
        if (year, month_num) > (today.year, today.month):
            break  # don't include future months
        month_name = date(year, month_num, 1).strftime("%B %Y")
        sheets.append(month_name)
    return sheets


def get_current_month_label(today=None) -> str:
    """e.g. 'May 2026'."""
    return (today or date.today()).strftime("%B %Y")


def _days_in_month(d) -> int:
    """How many days in the calendar month containing date d?"""
    import calendar
    return calendar.monthrange(d.year, d.month)[1]


def _read_sheet_totals(token, session_id, sheet_name) -> dict | None:
    """Read sum of column I (Cost) and column M (Cost Initiatives) from
    a sheet in the main workbook. Returns {'cost_sum': X, 'savings_sum':
    Y} or None if sheet doesn't exist / read fails."""
    import requests as _rq
    pf = _pf
    sheet_q = _rq.utils.quote(sheet_name)

    # Step 1: get used range bounds (just rowCount to keep response small)
    try:
        r_used = _rq.get(
            f"{pf.GRAPH_WB_BASE}/worksheets/{sheet_q}"
            f"/usedRange(valuesOnly=true)?$select=rowCount",
            headers=pf._graph_session_headers(token, session_id),
            timeout=20,
        )
        if r_used.status_code != 200:
            # Sheet doesn't exist or access denied → treat as no data
            return None
        last_row = r_used.json().get("rowCount", 0)
        if last_row < 3:
            return {"cost_sum": 0.0, "savings_sum": 0.0}
    except Exception as e:
        print(f"[dash] usedRange failed for '{sheet_name}': {e}")
        return None

    # Step 2: read I3:M{last_row} — 5 columns
    try:
        r = _rq.get(
            f"{pf.GRAPH_WB_BASE}/worksheets/{sheet_q}"
            f"/range(address='I3:M{last_row}')?$select=values",
            headers=pf._graph_session_headers(token, session_id),
            timeout=30,
        )
        if r.status_code != 200:
            return None
        values = r.json().get("values", [])

        cost_sum = 0.0
        savings_sum = 0.0
        for row in values:
            # row layout: [I, J, K, L, M]
            if row and len(row) >= 1:
                try:
                    v = row[0]
                    if v not in (None, ""):
                        cost_sum += float(v)
                except (ValueError, TypeError):
                    pass
            if row and len(row) >= 5:
                try:
                    v = row[4]
                    if v not in (None, ""):
                        savings_sum += float(v)
                except (ValueError, TypeError):
                    pass

        return {"cost_sum": cost_sum, "savings_sum": savings_sum}
    except Exception as e:
        print(f"[dash] range read failed for '{sheet_name}': {e}")
        return None


def fetch_dashboard_metrics() -> dict:
    """
    Fetch live financial metrics from the main SharePoint workbook.
    Aggregates spend across all FY sheets that exist (chronological from
    1 May to current month).

    Returns:
      {
        "ok":                   bool,
        "current_month_label":  "May 2026",
        "current_month_spend":  14267.50,
        "current_month_savings": 820.00,
        "fy_label":             "FY26/27",
        "fy_total_spend":       14267.50,
        "fetched_at":           datetime,
        "sheets_read":          ["May 2026"],
        "error":                None or str,
      }
    """
    today = date.today()
    result = {
        "ok": False,
        "current_month_label":  get_current_month_label(today),
        "current_month_spend":  0.0,
        "current_month_savings": 0.0,
        "fy_label":             get_fy_label(today),
        "fy_total_spend":       0.0,
        "fetched_at":           datetime.now(),
        "sheets_read":          [],
        "error":                None,
    }

    if not PO_FLOW_LOADED:
        result["error"] = "PO Flow core not loaded"
        return result

    pf = _pf
    if not getattr(pf, "GRAPH_AVAILABLE", False):
        result["error"] = "Graph API not available"
        return result

    try:
        token = pf._graph_get_token_silent()
        if not token:
            result["error"] = "Could not authenticate to Graph"
            return result
        session_id = pf._graph_create_session(token)
        if not session_id:
            result["error"] = "Could not create Graph session"
            return result

        try:
            current_month_label = result["current_month_label"]
            fy_sheets = get_current_fy_sheets(today)
            for sheet_name in fy_sheets:
                sheet_data = _read_sheet_totals(token, session_id,
                                                sheet_name)
                if sheet_data is None:
                    continue  # sheet missing — skip silently
                result["sheets_read"].append(sheet_name)
                result["fy_total_spend"] += sheet_data["cost_sum"]
                if sheet_name == current_month_label:
                    result["current_month_spend"] = sheet_data["cost_sum"]
                    result["current_month_savings"] = sheet_data["savings_sum"]
            result["ok"] = True
        finally:
            try:
                pf._graph_close_session(token, session_id)
            except Exception:
                pass

    except Exception as e:
        result["error"] = f"{type(e).__name__}: {e}"
        import traceback
        traceback.print_exc()

    return result


# ──────────────────────────────────────────────
# MAIN APPLICATION WINDOW
# ──────────────────────────────────────────────
class POFlowApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.user_cfg = load_user_config()
        self._status_job = None
        self._refresh_job = None        # tkinter after() handle for auto-refresh
        self.metric_widgets = []        # populated by _build_dashboard
        self.last_metrics = None
        self._setup_window()
        self._build_ui()
        if not PO_FLOW_LOADED:
            self.root.after(200, self._warn_no_core)
        else:
            # Extend CC_LOOKUP with any machines from the Cost centre sheet
            # that aren't in the main file's hardcoded dictionary.
            try:
                added = load_extended_cc_from_workbook()
                if added:
                    print(f"[gui] Loaded {added} extra cost centre entries "
                          f"from Cost centre sheet")
            except Exception as e:
                print(f"[gui] CC extension failed: {e}")
            # Kick off the dashboard auto-refresh cycle.
            self.root.after(500, self._schedule_auto_refresh)

    def _setup_window(self):
        self.root.title("PO Flow")
        self.root.geometry("900x740")    # taller to accommodate dashboard
        self.root.minsize(800, 620)
        self.root.configure(bg=C["bg_light"])
        try:
            self.root.iconbitmap(default="")
        except Exception:
            pass
        self.root.option_add("*Font", FONT["body"])

    # ── UI STRUCTURE ──────────────────────────
    def _build_ui(self):
        # ─ Header ─────────────────────────────
        hdr = tk.Frame(self.root, bg=C["bg_dark"], height=70)
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)

        tk.Label(hdr, text="PO Flow",
                 font=("Segoe UI", 22, "bold"),
                 bg=C["bg_dark"], fg=C["txt_light"]
                 ).place(x=20, rely=0.5, anchor="w")

        tk.Label(hdr, text="Purchase Order Automation",
                 font=FONT["small"], bg=C["bg_dark"],
                 fg="#7F8C8D").place(x=130, rely=0.65, anchor="w")

        # User pill (top-right)
        user_frame = tk.Frame(hdr, bg=C["bg_mid"],
                              padx=12, pady=6)
        user_frame.place(relx=1.0, rely=0.5, anchor="e", x=-20)

        initials = self.user_cfg.get("initials", "XX")
        name = self.user_cfg.get("name", "User")
        tk.Label(user_frame,
                 text=f"  {initials}  ",
                 font=("Segoe UI", 10, "bold"),
                 bg=C["primary"], fg="white",
                 padx=4).pack(side="left")
        tk.Label(user_frame, text=f"  {name}",
                 font=FONT["small"],
                 bg=C["bg_mid"], fg=C["txt_light"]).pack(side="left")

        # Alpha badge
        tk.Label(hdr, text=" ALPHA ",
                 font=("Segoe UI", 8, "bold"),
                 bg=C["accent"], fg="white",
                 padx=4, pady=2).place(relx=1.0, rely=0.18,
                                       anchor="e", x=-20)

        # ─ Main content area ──────────────────
        main = tk.Frame(self.root, bg=C["bg_light"])
        main.pack(fill="both", expand=True, padx=20, pady=16)

        # ─ Dashboard (3 financial metric cards) ──
        self._build_dashboard(main)

        # Title
        tk.Label(main, text="What would you like to do?",
                 font=FONT["heading"], bg=C["bg_light"],
                 fg=C["txt_mid"]).pack(anchor="w", pady=(8, 10))

        # ─ Action buttons grid ────────────────
        grid = tk.Frame(main, bg=C["bg_light"])
        grid.pack(fill="both", expand=True)
        grid.columnconfigure(0, weight=1)
        grid.columnconfigure(1, weight=1)

        buttons_data = [
            ("📋", "Create New PO",
             "Generate PDF, write to Excel, push to SharePoint",
             C["primary"], self._create_po, 0, 0),
            ("✅", "Mark as Received",
             "Record goods receipt for a pending PO",
             C["success"], self._mark_received, 0, 1),
            ("📊", "Recent POs",
             "View POs created in the last 7 days across all sheets",
             C["warning"], self._show_recent, 1, 0),
            ("🏢", "Manage Suppliers",
             "Add, edit or search suppliers",
             C["accent"], self._manage_suppliers, 1, 1),
            ("🔄", "Full Sync",
             "Sync with main SharePoint workbook",
             C["txt_mid"], self._coming_soon, 2, 0),
            ("⚙️", "Settings",
             "Configure your name, initials and preferences",
             "#8E44AD", self._settings, 2, 1),
        ]

        for icon, title, sub, color, cmd, row, col in buttons_data:
            btn = ActionButton(grid, icon, title, sub, color,
                               command=cmd,
                               relief="flat", bd=0)
            btn.grid(row=row, column=col, padx=6, pady=6,
                     sticky="nsew")
            grid.rowconfigure(row, weight=1)

        # ─ Status bar ─────────────────────────
        status_bar = tk.Frame(self.root, bg=C["bg_dark"], height=30)
        status_bar.pack(fill="x", side="bottom")
        status_bar.pack_propagate(False)

        self._status_var = tk.StringVar(
            value="Ready  •  PO Flow Alpha v0.1")
        tk.Label(status_bar, textvariable=self._status_var,
                 font=FONT["status"], bg=C["bg_dark"],
                 fg="#7F8C8D").place(x=12, rely=0.5, anchor="w")

        ver = f"Core: {PO_FLOW_FILE}" if PO_FLOW_LOADED else "⚠ Core not loaded"
        tk.Label(status_bar, text=ver,
                 font=FONT["status"], bg=C["bg_dark"],
                 fg="#7F8C8D").place(relx=1.0, rely=0.5,
                                     anchor="e", x=-12)

    # ── DASHBOARD (live metrics from main SharePoint) ──
    def _build_dashboard(self, parent):
        """Build the 3 financial metric cards + refresh control row."""
        section = tk.Frame(parent, bg=C["bg_light"])
        section.pack(fill="x", pady=(0, 4))

        # Three cards side by side
        cards_row = tk.Frame(section, bg=C["bg_light"])
        cards_row.pack(fill="x")

        card_configs = [
            ("month",    "💷", C["primary"],
             get_current_month_label() + " Spend"),
            ("savings",  "💡", C["success"], "Cost Initiatives"),
            ("fy_total", "📊", C["accent"],
             get_fy_label() + " Total"),
        ]

        for key, icon, color, default_label in card_configs:
            card = tk.Frame(cards_row, bg=C["bg_card"],
                             highlightthickness=1,
                             highlightbackground=C["border"])
            card.pack(side="left", fill="both", expand=True, padx=3)

            top = tk.Frame(card, bg=C["bg_card"])
            top.pack(fill="x", padx=12, pady=(10, 0))
            tk.Label(top, text=icon,
                     font=("Segoe UI Emoji", 14),
                     bg=C["bg_card"]).pack(side="left")
            label_w = tk.Label(top, text=default_label,
                               font=FONT["small"],
                               bg=C["bg_card"], fg=C["txt_mid"])
            label_w.pack(side="left", padx=8)

            value_w = tk.Label(card, text="—",
                               font=("Segoe UI", 17, "bold"),
                               bg=C["bg_card"], fg=color,
                               anchor="w")
            value_w.pack(fill="x", padx=12, pady=(2, 0))

            sub_w = tk.Label(card, text="loading...",
                             font=FONT["small"],
                             bg=C["bg_card"], fg=C["txt_mid"],
                             anchor="w")
            sub_w.pack(fill="x", padx=12, pady=(0, 10))

            self.metric_widgets.append({
                "key":      key,
                "label":    label_w,
                "value":    value_w,
                "subtitle": sub_w,
            })

        # Refresh control row below cards
        ctrl = tk.Frame(section, bg=C["bg_light"])
        ctrl.pack(fill="x", pady=(4, 0))

        self.last_refresh_label = tk.Label(ctrl, text="",
                                            font=FONT["small"],
                                            bg=C["bg_light"],
                                            fg=C["txt_mid"])
        self.last_refresh_label.pack(side="right", padx=8)

        self.refresh_btn = tk.Button(ctrl, text="🔄 Refresh",
                                      font=FONT["small"],
                                      bg=C["bg_light"], fg=C["primary"],
                                      activebackground=C["bg_light"],
                                      relief="flat", cursor="hand2",
                                      command=self.refresh_metrics)
        self.refresh_btn.pack(side="right")

    def refresh_metrics(self):
        """Trigger a background fetch of dashboard metrics.
        Safe to call from any thread / event handler. UI updates happen
        in the main thread via root.after()."""
        if not PO_FLOW_LOADED:
            return
        # Visual feedback: disable button, show "loading..." in subtitles
        try:
            self.refresh_btn.config(state="disabled",
                                     text="🔄 Loading...")
            for w in self.metric_widgets:
                w["subtitle"].config(text="loading...",
                                      fg=C["txt_mid"])
            self.last_refresh_label.config(text="")
        except Exception:
            pass  # widgets may not exist yet
        threading.Thread(target=self._do_fetch_metrics,
                         daemon=True).start()

    def _do_fetch_metrics(self):
        """Background: actually fetch from Graph API."""
        result = fetch_dashboard_metrics()
        try:
            self.root.after(0,
                            lambda r=result: self._on_metrics_loaded(r))
        except Exception:
            pass  # window might be closing

    def _on_metrics_loaded(self, result):
        """Main thread: update card UI with fetched values."""
        self.last_metrics = result
        try:
            self.refresh_btn.config(state="normal", text="🔄 Refresh")
        except Exception:
            return

        if not result["ok"]:
            for w in self.metric_widgets:
                w["value"].config(text="—")
                w["subtitle"].config(text="(unavailable)", fg=C["error"])
            self.last_refresh_label.config(
                text=f"⚠ {result.get('error', 'failed')[:40]}",
                fg=C["error"])
            return

        today = date.today()
        for w in self.metric_widgets:
            key = w["key"]
            if key == "month":
                w["label"].config(
                    text=f"{result['current_month_label']} Spend")
                w["value"].config(
                    text=f"£{result['current_month_spend']:,.2f}")
                w["subtitle"].config(
                    text=f"so far, day {today.day} of "
                         f"{_days_in_month(today)}",
                    fg=C["txt_mid"])
            elif key == "savings":
                w["label"].config(text="Cost Initiatives")
                w["value"].config(
                    text=f"£{result['current_month_savings']:,.2f}")
                w["subtitle"].config(text="this month", fg=C["txt_mid"])
            elif key == "fy_total":
                n_sheets = len(result["sheets_read"])
                w["label"].config(text=f"{result['fy_label']} Total")
                w["value"].config(
                    text=f"£{result['fy_total_spend']:,.2f}")
                w["subtitle"].config(
                    text=f"across {n_sheets} month(s)",
                    fg=C["txt_mid"])

        fetched = result["fetched_at"].strftime("%H:%M")
        self.last_refresh_label.config(
            text=f"Updated: {fetched}", fg=C["txt_mid"])

    def _schedule_auto_refresh(self):
        """Refresh now and every 10 minutes thereafter while app runs."""
        self.refresh_metrics()
        # 10 min = 600,000 ms
        self._refresh_job = self.root.after(10 * 60 * 1000,
                                              self._schedule_auto_refresh)

    def set_status(self, msg: str, color: str = None, clear_after: int = 0):
        self._status_var.set(msg)
        if clear_after:
            if self._status_job:
                self.root.after_cancel(self._status_job)
            self._status_job = self.root.after(
                clear_after * 1000,
                lambda: self._status_var.set("Ready  •  PO Flow Alpha v0.1"))

    # ── ACTIONS ───────────────────────────────
    def _warn_no_core(self):
        messagebox.showwarning(
            "Core not loaded",
            f"Could not load {PO_FLOW_FILE}.\n\n"
            f"Error: {_LOAD_ERROR}\n\n"
            "Make sure both files are in the same folder.",
            parent=self.root
        )

    def _create_po(self):
        if not PO_FLOW_LOADED:
            messagebox.showerror("Error",
                                 "PO Flow core not loaded.",
                                 parent=self.root)
            return
        CreatePODialog(self.root, self)

    def _mark_received(self):
        if not PO_FLOW_LOADED:
            messagebox.showerror("Error",
                                 "PO Flow core not loaded.",
                                 parent=self.root)
            return
        MarkAsReceivedDialog(self.root, self)

    def _manage_suppliers(self):
        if not PO_FLOW_LOADED:
            messagebox.showerror("Error",
                                 "PO Flow core not loaded.",
                                 parent=self.root)
            return
        ManageSuppliersDialog(self.root, self)

    def _show_recent(self):
        if not PO_FLOW_LOADED:
            messagebox.showerror("Error",
                                 "PO Flow core not loaded.",
                                 parent=self.root)
            return
        ShowRecentDialog(self.root, self)

    def _settings(self):
        SettingsDialog(self.root, self)

    def _coming_soon(self):
        messagebox.showinfo(
            "Coming in next version",
            "This feature is available in the CLI version (PO_FLOW_v3_1_1.py).\n\n"
            "It will be added to the GUI in the next release.",
            parent=self.root
        )


# ──────────────────────────────────────────────
# CREATE PO DIALOG
# ──────────────────────────────────────────────
class CreatePODialog:
    def __init__(self, parent, app: POFlowApp):
        self.app = app
        self.win = tk.Toplevel(parent)
        self.win.title("Create New Purchase Order")
        self.win.geometry("700x680")
        self.win.resizable(True, True)
        self.win.configure(bg=C["bg_light"])
        self.win.grab_set()
        self.win.focus_set()
        self._build()

    def _build(self):
        # ─ Header ─────────────────────────────
        hdr = tk.Frame(self.win, bg=C["bg_dark"], height=52)
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="📋  Create New Purchase Order",
                 font=FONT["heading"],
                 bg=C["bg_dark"], fg=C["txt_light"]
                 ).place(x=16, rely=0.5, anchor="w")

        # ─ Footer (MUST be packed BEFORE canvas so it reserves space) ──
        footer = tk.Frame(self.win, bg=C["bg_dark"], height=54)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)

        self.submit_btn = tk.Button(
            footer, text="  ✓  Create PO  ",
            font=FONT["btn"],
            bg=C["primary"], fg="white",
            activebackground=C["primary_h"],
            relief="flat", cursor="hand2",
            command=self._submit, padx=20, pady=8)
        self.submit_btn.place(relx=1.0, rely=0.5, anchor="e", x=-16)

        tk.Button(
            footer, text="Cancel",
            font=FONT["body"],
            bg=C["bg_dark"], fg="#7F8C8D",
            activebackground=C["bg_mid"],
            relief="flat", cursor="hand2",
            command=self.win.destroy, padx=12, pady=8
        ).place(relx=1.0, rely=0.5, anchor="e", x=-148)

        self._status_lbl = tk.Label(footer, text="",
                                    font=FONT["small"],
                                    bg=C["bg_dark"], fg="#7F8C8D")
        self._status_lbl.place(x=16, rely=0.5, anchor="w")

        # ─ Scrollable form (fills remaining space between header+footer) ─
        scroll_wrap = tk.Frame(self.win, bg=C["bg_light"])
        scroll_wrap.pack(fill="both", expand=True, side="top")

        canvas = tk.Canvas(scroll_wrap, bg=C["bg_light"],
                           highlightthickness=0)
        scrollbar = ttk.Scrollbar(scroll_wrap, orient="vertical",
                                  command=canvas.yview)
        self.form_frame = tk.Frame(canvas, bg=C["bg_light"])

        self.form_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")))

        canvas.create_window((0, 0), window=self.form_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True, padx=16, pady=8)

        # Mousewheel scrolling — only when mouse is over the form area.
        # bind_all would catch wheel events even when over dropdowns, which
        # we don't want. Enter/Leave activate/deactivate scrolling instead.
        def _on_wheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")

        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>",
                                                          _on_wheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        self._build_form()

    # ── AUTOCOMPLETE HANDLERS ──────────────────
    def _on_machine_change(self, event=None):
        """When machine changes, update area dropdown with relevant areas."""
        machine = self._get_field(self.machine)
        areas = get_areas_for_machine(machine)
        self.area["values"] = areas

    def _on_supplier_type(self, event=None):
        """Filter supplier dropdown to suppliers matching what user typed."""
        # Don't filter on arrow keys / control keys
        if event and event.keysym in ("Up", "Down", "Left", "Right",
                                      "Return", "Tab", "Escape",
                                      "Shift_L", "Shift_R",
                                      "Control_L", "Control_R"):
            return
        typed = self._get_field(self.supplier).lower()
        if not typed:
            filtered = self._all_suppliers
        else:
            # Show suppliers that start with typed text OR contain it
            starts_with = [s for s in self._all_suppliers
                           if s.lower().startswith(typed)]
            contains = [s for s in self._all_suppliers
                        if typed in s.lower() and s not in starts_with]
            filtered = starts_with + contains
        self.supplier["values"] = filtered
        # Note: user clicks dropdown arrow or presses Alt+Down to see list

    # ── DYNAMIC ITEM ROWS ─────────────────────
    def _add_item_row(self):
        """Add a new Qty | Part# | Description | Price | Mode row."""
        row_frame = tk.Frame(self.items_container, bg=C["bg_card"])
        row_frame.pack(fill="x", pady=2)

        qty_var = tk.StringVar(value="1")
        qty_e = ttk.Entry(row_frame, textvariable=qty_var,
                          width=5, font=FONT["body"])
        qty_e.pack(side="left")

        pn_var = tk.StringVar()
        ttk.Entry(row_frame, textvariable=pn_var,
                  width=14, font=FONT["body"]
                  ).pack(side="left", padx=4)

        desc_var = tk.StringVar()
        desc_e = ttk.Entry(row_frame, textvariable=desc_var,
                           font=FONT["body"])
        desc_e.pack(side="left", fill="x", expand=True, padx=4)

        price_var = tk.StringVar()
        ttk.Entry(row_frame, textvariable=price_var,
                  width=8, font=FONT["body"]
                  ).pack(side="left", padx=4)

        # Mode dropdown: "each" = unit price (× qty), "total" = line total
        mode_var = tk.StringVar(value="each")
        mode_combo = ttk.Combobox(row_frame, textvariable=mode_var,
                                   values=["each", "total"],
                                   width=6, font=FONT["body"],
                                   state="readonly")
        mode_combo.pack(side="left", padx=4)

        remove_btn = tk.Button(row_frame, text="✕",
                               font=("Segoe UI", 9),
                               bg=C["bg_card"], fg=C["border"],
                               activebackground=C["bg_light"],
                               relief="flat", cursor="hand2",
                               width=2)
        remove_btn.pack(side="left", padx=(4, 0))

        row_data = (row_frame, qty_var, pn_var, desc_var,
                    price_var, mode_var, remove_btn)
        self.item_rows.append(row_data)
        remove_btn.config(
            command=lambda r=row_data: self._remove_item_row(r))

        # Live total recalculation when these fields change
        for var in (qty_var, price_var, mode_var):
            var.trace_add("write", lambda *a: self._update_total_label())

        self._update_remove_buttons()
        self._update_total_label()
        desc_e.focus_set()

    def _remove_item_row(self, row_data):
        if len(self.item_rows) <= 1:
            return
        row_data[0].destroy()
        self.item_rows.remove(row_data)
        self._update_remove_buttons()
        self._update_total_label()

    def _update_remove_buttons(self):
        active = len(self.item_rows) > 1
        for r in self.item_rows:
            btn = r[6]  # remove button is last
            btn.config(
                state="normal" if active else "disabled",
                fg=C["error"] if active else C["border"])

    def _calculate_net_total(self) -> float:
        """Sum line totals across all item rows.
        Per row: if mode='each' → qty × price.  If mode='total' → price as-is.
        Empty/invalid rows contribute 0.
        """
        total = 0.0
        for r in self.item_rows:
            _, qty_var, _, desc_var, price_var, mode_var, _ = r
            desc = desc_var.get().strip()
            if not desc:
                continue
            try:
                qty = int(qty_var.get().strip() or "1")
            except ValueError:
                qty = 1
            try:
                price = float(
                    price_var.get().strip().replace("£", "").replace(",", "")
                    or "0")
            except ValueError:
                price = 0.0
            mode = (mode_var.get() or "each").lower()
            line = price * qty if mode == "each" else price
            total += line
        return round(total, 2)

    def _update_total_label(self):
        """Refresh the live 'Net Total' label in the Cost & Dates section."""
        if not hasattr(self, "total_label"):
            return  # form not built yet
        total = self._calculate_net_total()
        n_items = sum(1 for r in self.item_rows
                      if r[3].get().strip())  # desc not empty
        self.total_label.config(
            text=f"£{total:,.2f}",
            fg=C["success"] if total > 0 else C["txt_mid"])
        self.total_subtitle.config(
            text=f"auto-calculated from {n_items} item(s)"
                 if n_items else "add at least one item with a price")

    def _build_items_summary(self) -> str:
        parts = []
        for r in self.item_rows:
            _, qty_var, pn_var, desc_var, _, _, _ = r
            qty = qty_var.get().strip() or "1"
            pn = pn_var.get().strip()
            desc = desc_var.get().strip()
            if not desc:
                continue
            parts.append(f"x{qty} [{pn}] {desc}" if pn
                         else f"x{qty} {desc}")
        return "; ".join(parts)

    def _section(self, title: str) -> tk.Frame:
        outer = tk.Frame(self.form_frame, bg=C["bg_light"])
        outer.pack(fill="x", pady=(8, 0))

        tk.Label(outer, text=title.upper(),
                 font=("Segoe UI", 8, "bold"),
                 bg=C["bg_light"], fg=C["txt_mid"]
                 ).pack(anchor="w", padx=2, pady=(4, 4))

        card = tk.Frame(outer, bg=C["bg_card"],
                        relief="flat", bd=0,
                        highlightthickness=1,
                        highlightbackground=C["border"])
        card.pack(fill="x")
        return card

    def _row(self, parent, label: str, required=False) -> tk.Frame:
        """Create a form row with label on left."""
        row = tk.Frame(parent, bg=C["bg_card"])
        row.pack(fill="x", padx=14, pady=5)

        lbl_text = f"{label} {'*' if required else ''}"
        tk.Label(row, text=lbl_text,
                 font=FONT["body"],
                 bg=C["bg_card"], fg=C["txt_dark"],
                 width=22, anchor="w"
                 ).pack(side="left")
        return row

    def _entry(self, parent, **kw) -> ttk.Entry:
        e = ttk.Entry(parent, font=FONT["body"], **kw)
        e.pack(side="left", fill="x", expand=True)
        return e

    def _combo(self, parent, values, **kw) -> ttk.Combobox:
        c = ttk.Combobox(parent, values=values,
                         font=FONT["body"], state="normal", **kw)
        c.pack(side="left", fill="x", expand=True)
        return c

    def _build_form(self):
        suppliers = get_suppliers_list()
        machines = get_machines_list()
        self._all_suppliers = suppliers  # for autocomplete filtering

        # ── Section: Machine & Area ────────────
        sec1 = self._section("Machine & Location")
        r1 = self._row(sec1, "Machine", required=True)
        self.machine = self._combo(r1, machines, width=30)
        self.machine.bind("<<ComboboxSelected>>", self._on_machine_change)
        self.machine.bind("<KeyRelease>", self._on_machine_change)

        r2 = self._row(sec1, "Area")
        self.area = self._combo(r2, get_areas_for_machine(""))

        r3 = self._row(sec1, "Reason", required=True)
        self.reason = self._combo(r3, REASONS)

        r4 = self._row(sec1, "Mech / Elec", required=True)
        self.mech_elec = self._combo(r4, MECH_ELEC)
        self.mech_elec.set("Mech")

        # ── Section: Supplier ─────────────────
        sec2 = self._section("Supplier")
        r5 = self._row(sec2, "Supplier name", required=True)
        self.supplier = self._combo(r5, suppliers, width=35)
        # Bind autocomplete: filter list as user types
        self.supplier.bind("<KeyRelease>", self._on_supplier_type)

        r5b = self._row(sec2, "Job No")
        self.job_no = self._entry(r5b)
        tk.Label(r5b, text="(optional)",
                 font=FONT["small"], bg=C["bg_card"],
                 fg=C["txt_mid"]).pack(side="left", padx=6)

        # ── Section: Parts ────────────────────
        sec3 = self._section("Parts Ordered")

        # Column headers
        col_hdr = tk.Frame(sec3, bg=C["bg_card"])
        col_hdr.pack(fill="x", padx=14, pady=(10, 2))
        tk.Label(col_hdr, text="Qty",
                 font=("Segoe UI", 8, "bold"),
                 bg=C["bg_card"], fg=C["txt_mid"],
                 width=5, anchor="w").pack(side="left")
        tk.Label(col_hdr, text="Part Number",
                 font=("Segoe UI", 8, "bold"),
                 bg=C["bg_card"], fg=C["txt_mid"],
                 width=14, anchor="w").pack(side="left", padx=(4, 0))
        tk.Label(col_hdr, text="Description  *",
                 font=("Segoe UI", 8, "bold"),
                 bg=C["bg_card"], fg=C["txt_mid"],
                 anchor="w").pack(side="left", padx=(8, 0),
                                  fill="x", expand=True)
        tk.Label(col_hdr, text="Price £  *",
                 font=("Segoe UI", 8, "bold"),
                 bg=C["bg_card"], fg=C["txt_mid"],
                 width=8, anchor="w").pack(side="left", padx=(0, 0))
        tk.Label(col_hdr, text="Mode",
                 font=("Segoe UI", 8, "bold"),
                 bg=C["bg_card"], fg=C["txt_mid"],
                 width=6, anchor="w").pack(side="left", padx=(8, 0))
        tk.Label(col_hdr, text="",
                 bg=C["bg_card"],
                 width=3).pack(side="left", padx=(0, 0))

        # Dynamic items container
        self.item_rows = []
        self.items_container = tk.Frame(sec3, bg=C["bg_card"])
        self.items_container.pack(fill="x", padx=14, pady=(0, 4))

        # Start with one empty row
        self._add_item_row()

        # "Add item" button
        tk.Button(sec3, text="＋  Add another item",
                  font=FONT["small"],
                  bg=C["bg_card"], fg=C["primary"],
                  activebackground=C["bg_light"],
                  relief="flat", cursor="hand2",
                  command=self._add_item_row
                  ).pack(anchor="w", padx=14, pady=(2, 10))

        # ── Section: Cost & Dates ─────────────
        sec4 = self._section("Cost & Dates")

        # Net Total — auto-calculated, read-only display
        total_row = tk.Frame(sec4, bg=C["bg_card"])
        total_row.pack(fill="x", padx=14, pady=8)
        tk.Label(total_row, text="Net Total:",
                 font=FONT["body"],
                 bg=C["bg_card"], fg=C["txt_dark"],
                 width=22, anchor="w").pack(side="left")
        total_inner = tk.Frame(total_row, bg=C["bg_card"])
        total_inner.pack(side="left", fill="x", expand=True)
        self.total_label = tk.Label(total_inner, text="£0.00",
                                    font=("Segoe UI", 14, "bold"),
                                    bg=C["bg_card"], fg=C["txt_mid"],
                                    anchor="w")
        self.total_label.pack(anchor="w")
        self.total_subtitle = tk.Label(
            total_inner, text="add at least one item with a price",
            font=FONT["small"],
            bg=C["bg_card"], fg=C["txt_mid"],
            anchor="w")
        self.total_subtitle.pack(anchor="w")

        # Hidden compatibility: keep self.cost attribute so other code that
        # reads it still works. We expose it as a fake widget with .get()
        # returning the calculated total as string.
        class _CostShim:
            def __init__(self, dialog):
                self.dialog = dialog
            def get(self):
                return f"{self.dialog._calculate_net_total():.2f}"
        self.cost = _CostShim(self)

        r7 = self._row(sec4, "Due date")
        self.due_date = self._entry(r7, width=16)
        self.due_date.insert(0, "dd/mm/yyyy")
        self.due_date.bind("<FocusIn>", lambda e: (
            self.due_date.delete(0, "end")
            if self.due_date.get() == "dd/mm/yyyy" else None))
        tk.Label(r7, text="optional — for service contracts / future months",
                 font=FONT["small"], bg=C["bg_card"],
                 fg=C["txt_mid"]).pack(side="left", padx=8)

        # ── Section: Optional extras ──────────
        sec5 = self._section("Optional")
        r8 = self._row(sec5, "Savings (£)")
        self.savings = self._entry(r8, width=16)

        r9 = self._row(sec5, "Original supplier")
        self.orig_supplier = self._entry(r9)

    # ── SUBMIT ────────────────────────────────
    def _get_field(self, widget) -> str:
        if isinstance(widget, scrolledtext.ScrolledText):
            return widget.get("1.0", "end-1c").strip()
        return widget.get().strip()

    def _validate(self) -> str | None:
        """Return error message or None if valid."""
        if not self._get_field(self.machine):
            return "Machine is required."
        if not self._get_field(self.reason):
            return "Reason is required."
        if not self._get_field(self.mech_elec):
            return "Mech/Elec is required."
        if not self._get_field(self.supplier):
            return "Supplier name is required."
        # Check items
        items_summary = self._build_items_summary()
        if not items_summary:
            return "At least one item with a description is required."
        # Check that items have prices
        for i, r in enumerate(self.item_rows, 1):
            _, qty_var, _, desc_var, price_var, _, _ = r
            if not desc_var.get().strip():
                continue
            price_raw = price_var.get().strip().replace("£", "").replace(",", "")
            if not price_raw:
                return f"Item {i}: price is required."
            try:
                p = float(price_raw)
                if p < 0:
                    return f"Item {i}: price cannot be negative."
            except ValueError:
                return f"Item {i}: invalid price '{price_var.get()}'."
        # Net total must be > 0
        if self._calculate_net_total() <= 0:
            return "Net total must be greater than zero."
        due = self._get_field(self.due_date)
        if due and due != "dd/mm/yyyy":
            try:
                datetime.strptime(due, "%d/%m/%Y")
            except ValueError:
                return f"Invalid due date '{due}'. Use dd/mm/yyyy format."
        return None

    def _submit(self):
        err = self._validate()
        if err:
            messagebox.showerror("Validation error", err,
                                 parent=self.win)
            return

        self.submit_btn.config(state="disabled", text="  Creating PO...  ")
        self._status_lbl.config(text="⏳  Please wait...", fg=C["warning"])

        threading.Thread(target=self._do_create, daemon=True).start()

    def _do_create(self):
        # CRITICAL: Outlook is a COM application. When we run in a background
        # thread (not the main thread), we MUST call pythoncom.CoInitialize()
        # before any COM calls. Without it, win32com.client.Dispatch silently
        # fails to connect to Outlook. The main thread handles COM init
        # automatically, but threading.Thread does not.
        com_initialized = False
        try:
            try:
                import pythoncom
                pythoncom.CoInitialize()
                com_initialized = True
            except ImportError:
                pass  # pywin32 not installed — Outlook will be skipped anyway

            result = self._run_create_po()
            self.win.after(0, lambda: self._on_success(result))
        except Exception as e:
            err_msg = str(e)
            self.win.after(0, lambda m=err_msg: self._on_error(m))
        finally:
            if com_initialized:
                try:
                    import pythoncom
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

    def _run_create_po(self) -> dict:
        """Build order dict and call PO Flow core functions."""
        pf = _pf

        # Net total is calculated from items (no longer a separate field)
        net_total = self._calculate_net_total()

        due_raw = self._get_field(self.due_date)
        if due_raw == "dd/mm/yyyy":
            due_raw = ""

        initials = self.app.user_cfg.get("initials", "GF")

        # Build items list from dynamic rows with per-item pricing.
        # Each item knows its own unit_price and line_total based on mode:
        #   mode="each"  → unit_price = entered, line_total = unit × qty
        #   mode="total" → line_total = entered, unit_price = line / qty
        # Keys must match generate_po_pdf expectations:
        #   qty, part_no, desc, unit_price, line_total
        items = []
        for r in self.item_rows:
            _, qty_var, pn_var, desc_var, price_var, mode_var, _ = r
            desc = desc_var.get().strip()
            if not desc:
                continue
            try:
                qty = int(qty_var.get().strip() or "1")
            except ValueError:
                qty = 1
            try:
                price = float(
                    price_var.get().strip().replace("£", "").replace(",", "")
                    or "0")
            except ValueError:
                price = 0.0
            mode = (mode_var.get() or "each").lower()
            if mode == "each":
                unit_price = price
                line_total = round(price * qty, 2)
            else:
                line_total = price
                unit_price = round(price / qty, 2) if qty else 0.0
            items.append({
                "qty":        qty,
                "part_no":    pn_var.get().strip(),
                "desc":       desc,
                "unit_price": unit_price,
                "line_total": line_total,
            })

        if not items:
            items = [{"qty": 1, "part_no": "", "desc": "See attached PO",
                      "unit_price": net_total, "line_total": net_total}]

        # Build items_summary for Excel column F
        items_summary = self._build_items_summary()

        # Build order dict
        # If user selected "(machine itself)" → convert to empty string
        area_val = self._get_field(self.area)
        if area_val == GENERAL_AREA_LABEL:
            area_val = ""

        order = {
            "machine":       self._get_field(self.machine),
            "area":          area_val,
            "reason":        self._get_field(self.reason),
            "mech_elec":     self._get_field(self.mech_elec),
            "supplier":      self._get_field(self.supplier),
            "job_no":        self._get_field(self.job_no),
            "due_raw":       due_raw,
            "savings_raw":   self._get_field(self.savings),
            "orig_supplier": self._get_field(self.orig_supplier),
            "items_summary": items_summary,
            "part_number":   "",
        }

        # Normalise machine / area / cost centre
        order["machine"] = pf.normalize_machine_name(order["machine"])
        order["area"] = pf.normalize_area_name(
            order["area"], order["machine"])

        wb = pf.load_or_create_workbook(pf.WORKBOOK_PATH)
        cost_centre = pf.resolve_cost_centre(
            wb, order["machine"], order["area"], order["mech_elec"])
        order["cost_centre"] = cost_centre

        # Look up full supplier object from JSON (address, phone, email, etc.)
        try:
            db = pf.load_suppliers_db(pf.SUPPLIERS_JSON_PATH)
            suppliers = db.get("suppliers", [])
            supplier_obj = next(
                (s for s in suppliers
                 if (s.get("name", "") or "").strip().lower() ==
                    order["supplier"].strip().lower()),
                {"name": order["supplier"]}  # fallback: just the name
            )
        except Exception:
            supplier_obj = {"name": order["supplier"]}

        # Determine target monthly sheet
        target_first = pf.get_target_month_for_order(order)
        pf.ensure_month_sheets_through(wb, target_first)
        ws_parts = pf.get_or_create_sheet_for_date(wb, target_first)

        header_row, colmap = pf.ensure_parts_column_map(ws_parts)
        pf.validate_required_parts_columns(colmap)

        ws_recv = pf.ensure_to_receive_sheet(wb)

        # Generate PO number with the GUI user's initials
        po_number = pf.generate_po_number(cost_centre, initials)

        # Build totals dict with proper VAT calculation
        vat_rate = getattr(pf, "VAT_RATE_DEFAULT", 0.20)
        vat_amount = round(net_total * vat_rate, 2)
        gross_total = round(net_total + vat_amount, 2)
        totals = {
            "net_total":   net_total,
            "subtotal":    net_total,
            "vat_rate":    vat_rate,
            "vat_amount":  vat_amount,
            "vat":         vat_amount,
            "gross_total": gross_total,
            "grand_total": gross_total,
        }

        # Write to local workbook
        parts_row = pf.write_parts_order_row(
            ws_parts, header_row, colmap, order,
            po_number, cost_centre, net_total)
        recv_row = pf.add_to_receive_row(
            ws_recv, po_number, order,
            cost_centre, net_total, totals)

        pf.update_month_totals(ws_parts, header_row, colmap)
        pf.auto_fit_columns(ws_parts)
        pf.auto_fit_columns(ws_recv)

        # Generate PDF
        pf.ensure_dir(pf.PO_PDF_OUTPUT_DIR)
        pdf_path = os.path.join(pf.PO_PDF_OUTPUT_DIR,
                                f"{po_number}.pdf")
        pdf_ok = pf.generate_po_pdf(
            pdf_path, po_number, supplier_obj, order, items, totals)

        # Save local workbook
        pf.safe_save_workbook(wb, pf.WORKBOOK_PATH)

        # Open Outlook with PDF attached
        # CRITICAL: The CLI version's validate_supplier_email() calls input()
        # to ask the user. With pythonw.exe there's no terminal — input()
        # blocks/fails silently. We monkey-patch it with a silent version:
        #   - If supplier has email → returns it
        #   - If no email → returns "" (Outlook opens with no To: field,
        #                   user types it manually)
        email_ok = False
        _orig_validate = getattr(pf, "validate_supplier_email", None)
        if _orig_validate is not None:
            def _gui_validate(supplier_obj):
                email = (supplier_obj.get("email") or "").strip()
                # Skip CLI-style validation prompts entirely
                if email and ("@" not in email or
                              "." not in email.split("@")[-1]):
                    return ""  # invalid format → blank To:
                return email   # empty or valid both pass through

            pf.validate_supplier_email = _gui_validate
        try:
            email_ok = bool(pf.open_outlook_with_po(
                pdf_path, po_number, supplier_obj, order, totals))
        except Exception as e:
            print(f"[gui] Outlook error: {e}")
        finally:
            if _orig_validate is not None:
                pf.validate_supplier_email = _orig_validate

        # Push to main SharePoint
        push_ok = pf.push_po_to_main(
            order, po_number, net_total, ws_parts.title)

        return {
            "po_number": po_number,
            "sheet":     ws_parts.title,
            "row":       parts_row,
            "pdf_ok":    pdf_ok,
            "push_ok":   push_ok,
            "email_ok":  email_ok,
            "pdf_path":  pdf_path,
        }

    def _on_success(self, result: dict):
        po = result["po_number"]
        sheet = result["sheet"]
        push_status = "✓ Pushed to main SharePoint" if result["push_ok"] \
            else "⚠ Push failed — manual copy required"
        pdf_status = "✓ PDF generated" if result["pdf_ok"] \
            else "⚠ PDF not generated"
        email_status = "✓ Outlook email prepared" if result.get("email_ok") \
            else "⚠ Outlook not opened (check manually)"

        self._status_lbl.config(
            text=f"✓  PO created: {po}", fg=C["success"])

        messagebox.showinfo(
            "Purchase Order Created",
            f"PO Number:   {po}\n"
            f"Sheet:       {sheet}\n"
            f"Row:         {result['row']}\n\n"
            f"{pdf_status}\n"
            f"{email_status}\n"
            f"{push_status}",
            parent=self.win
        )
        self.app.set_status(f"✓  Last PO: {po}  ({sheet})")
        # Refresh dashboard to reflect the new PO
        self.app.refresh_metrics()
        self.win.destroy()

    def _on_error(self, msg: str):
        self.submit_btn.config(state="normal",
                               text="  Create PO  ")
        self._status_lbl.config(text="❌  Error", fg=C["error"])
        messagebox.showerror("Error creating PO",
                             f"Something went wrong:\n\n{msg}",
                             parent=self.win)


# ──────────────────────────────────────────────
# SHOW RECENT POs DIALOG
# ──────────────────────────────────────────────
# ──────────────────────────────────────────────
# MARK AS RECEIVED DIALOG
# ──────────────────────────────────────────────
class MarkAsReceivedDialog:
    def __init__(self, parent, app):
        self.app = app
        self.win = tk.Toplevel(parent)
        self.win.title("Mark as Received")
        self.win.geometry("980x600")
        self.win.configure(bg=C["bg_light"])
        self.win.grab_set()
        self._all_pending = []
        self._build()
        self._load_pending()

    def _build(self):
        # Header
        hdr = tk.Frame(self.win, bg=C["bg_dark"], height=52)
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="✅  Mark Purchase Order as Received",
                 font=FONT["heading"],
                 bg=C["bg_dark"], fg=C["txt_light"]
                 ).place(x=16, rely=0.5, anchor="w")

        # Footer
        footer = tk.Frame(self.win, bg=C["bg_dark"], height=64)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)

        # Invoice number on left of footer
        inv_frame = tk.Frame(footer, bg=C["bg_dark"])
        inv_frame.place(x=16, rely=0.5, anchor="w")
        tk.Label(inv_frame, text="Invoice Number (optional):",
                 font=FONT["small"], bg=C["bg_dark"],
                 fg="#BDC3C7").pack(side="left", padx=(0, 8))
        self.invoice_var = tk.StringVar()
        ttk.Entry(inv_frame, textvariable=self.invoice_var,
                  width=20, font=FONT["body"]).pack(side="left")

        # Buttons on right
        self.mark_btn = tk.Button(
            footer, text="  ✓  Mark as Received  ",
            font=FONT["btn"],
            bg=C["success"], fg="white",
            activebackground="#196F3D",
            relief="flat", cursor="hand2",
            command=self._on_mark, padx=20, pady=8,
            state="disabled")
        self.mark_btn.place(relx=1.0, rely=0.5, anchor="e", x=-16)

        tk.Button(
            footer, text="Cancel",
            font=FONT["body"],
            bg=C["bg_dark"], fg="#7F8C8D",
            activebackground=C["bg_mid"],
            relief="flat", cursor="hand2",
            command=self.win.destroy, padx=12, pady=8
        ).place(relx=1.0, rely=0.5, anchor="e", x=-200)

        # Main content area
        main = tk.Frame(self.win, bg=C["bg_light"])
        main.pack(fill="both", expand=True, padx=16, pady=10)

        # Filter / search bar
        ctrl = tk.Frame(main, bg=C["bg_light"])
        ctrl.pack(fill="x", pady=(0, 8))
        tk.Label(ctrl, text="🔍 Filter:",
                 font=FONT["body"], bg=C["bg_light"],
                 fg=C["txt_dark"]).pack(side="left")
        self.filter_var = tk.StringVar()
        self.filter_var.trace_add("write", self._on_filter)
        ttk.Entry(ctrl, textvariable=self.filter_var,
                  font=FONT["body"], width=30
                  ).pack(side="left", padx=8)

        ttk.Button(ctrl, text="Refresh",
                   command=self._load_pending).pack(side="left", padx=8)

        self._summary_lbl = tk.Label(ctrl, text="Loading...",
                                     font=FONT["body"],
                                     bg=C["bg_light"], fg=C["txt_mid"])
        self._summary_lbl.pack(side="left", padx=12)

        # Treeview of pending POs
        tree_frame = tk.Frame(main, bg=C["bg_light"])
        tree_frame.pack(fill="both", expand=True)

        cols = ("po", "date", "supplier", "machine", "area", "cost")
        self.tree = ttk.Treeview(tree_frame, columns=cols,
                                 show="headings", height=14,
                                 selectmode="browse")
        col_cfg = [
            ("po",       "PO Number",     140, "w"),
            ("date",     "Date Ordered",   95, "center"),
            ("supplier", "Supplier",      180, "w"),
            ("machine",  "Machine",       130, "w"),
            ("area",     "Area",          120, "w"),
            ("cost",     "Cost",           90, "e"),
        ]
        for cid, h, w, anchor in col_cfg:
            self.tree.heading(cid, text=h)
            self.tree.column(cid, width=w, anchor=anchor)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Selection enables Mark button
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        # Double-click also marks
        self.tree.bind("<Double-Button-1>",
                       lambda e: self._on_mark() if self.tree.selection() else None)

    def _load_pending(self):
        for r in self.tree.get_children():
            self.tree.delete(r)
        self._summary_lbl.config(text="Loading...")
        threading.Thread(target=self._do_load, daemon=True).start()

    def _do_load(self):
        pending = get_pending_pos()
        self.win.after(0, lambda: self._populate(pending))

    def _populate(self, pending):
        self._all_pending = pending
        self._show_filtered(pending)

    def _show_filtered(self, items):
        for r in self.tree.get_children():
            self.tree.delete(r)
        for p in items:
            self.tree.insert("", "end", values=(
                p["po_number"],
                p["date_ordered"],
                p["supplier"][:30],
                p["machine"][:20],
                p["area"][:18],
                f"£{p['cost']:,.2f}",
            ))
        if items:
            self._summary_lbl.config(
                text=f"{len(items)} pending PO(s)", fg=C["txt_mid"])
        else:
            self._summary_lbl.config(
                text="No pending POs found", fg=C["txt_mid"])

    def _on_filter(self, *args):
        q = self.filter_var.get().strip().lower()
        if not q:
            self._show_filtered(self._all_pending)
            return
        filtered = [p for p in self._all_pending
                    if q in p["po_number"].lower()
                    or q in p["supplier"].lower()
                    or q in p["machine"].lower()]
        self._show_filtered(filtered)

    def _on_select(self, event=None):
        if self.tree.selection():
            self.mark_btn.config(state="normal")
        else:
            self.mark_btn.config(state="disabled")

    def _on_mark(self):
        sel = self.tree.selection()
        if not sel:
            return
        item = self.tree.item(sel[0])
        po = item["values"][0]
        invoice = self.invoice_var.get().strip()

        self.mark_btn.config(state="disabled", text="  Working...  ")
        self._summary_lbl.config(text="⏳ Updating workbooks...",
                                 fg=C["warning"])

        threading.Thread(target=self._do_mark,
                         args=(po, invoice), daemon=True).start()

    def _do_mark(self, po, invoice):
        com_initialized = False
        try:
            try:
                import pythoncom
                pythoncom.CoInitialize()
                com_initialized = True
            except ImportError:
                pass
            result = mark_received_workflow(po, invoice)
            self.win.after(0, lambda: self._on_done(result))
        except Exception as e:
            err = str(e)
            self.win.after(0, lambda m=err: self._on_error(m))
        finally:
            if com_initialized:
                try:
                    import pythoncom
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

    def _on_done(self, result):
        if not result["success"]:
            self._on_error(result.get("error", "Unknown error"))
            return

        push_status = "✓ Pushed to main SharePoint" if result["push_ok"] \
            else "⚠ Push to main failed — check manually"
        recv_status = "✓ To Receive sheet updated" if result["to_recv_ok"] \
            else "⚠ To Receive entry not found"

        messagebox.showinfo(
            "Marked as Received",
            f"PO {result['po_number']} marked as received.\n\n"
            f"Sheet:  {result['sheet']}\n"
            f"Row:    {result['row']}\n\n"
            f"✓ Delivered date set + green fill applied\n"
            f"{recv_status}\n"
            f"{push_status}",
            parent=self.win
        )
        self.app.set_status(
            f"✓ Received: {result['po_number']} ({result['sheet']})")
        # Reload the list (the marked PO should disappear)
        self._load_pending()
        # Refresh dashboard — Mark as Received doesn't change spend but
        # it's good UX to show "Updated: now" so user knows app is live.
        self.app.refresh_metrics()
        self.mark_btn.config(state="disabled", text="  ✓  Mark as Received  ")
        self.invoice_var.set("")

    def _on_error(self, msg):
        self.mark_btn.config(state="normal",
                             text="  ✓  Mark as Received  ")
        self._summary_lbl.config(text="❌ Error", fg=C["error"])
        messagebox.showerror("Error",
                             f"Could not mark as received:\n\n{msg}",
                             parent=self.win)


# ──────────────────────────────────────────────
# SHOW RECENT POs DIALOG
# ──────────────────────────────────────────────
# ──────────────────────────────────────────────
# MANAGE SUPPLIERS DIALOG
# ──────────────────────────────────────────────
class ManageSuppliersDialog:
    """List, add, edit and delete suppliers stored in suppliers.json.
    Changes propagate automatically to outgoing Outlook emails
    (the email sender reads the JSON via load_suppliers_db, so updating
    the contact_name or email here changes the next email's greeting
    and recipient)."""

    def __init__(self, parent, app):
        self.app = app
        self.win = tk.Toplevel(parent)
        self.win.title("Manage Suppliers")
        self.win.geometry("980x640")
        self.win.configure(bg=C["bg_light"])
        self.win.grab_set()
        self._all = []          # full list of supplier dicts
        self._build()
        self._reload()

    def _build(self):
        # Header
        hdr = tk.Frame(self.win, bg=C["bg_dark"], height=52)
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="🏢  Manage Suppliers",
                 font=FONT["heading"],
                 bg=C["bg_dark"], fg=C["txt_light"]
                 ).place(x=16, rely=0.5, anchor="w")

        # Footer
        footer = tk.Frame(self.win, bg=C["bg_dark"], height=64)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)
        tk.Button(footer, text="Close",
                  font=FONT["body"],
                  bg=C["bg_dark"], fg="#BDC3C7",
                  activebackground=C["bg_mid"],
                  relief="flat", cursor="hand2",
                  command=self.win.destroy, padx=20, pady=8
                  ).place(relx=1.0, rely=0.5, anchor="e", x=-16)

        # Main content area
        main = tk.Frame(self.win, bg=C["bg_light"])
        main.pack(fill="both", expand=True, padx=16, pady=10)

        # Control bar: search + action buttons
        ctrl = tk.Frame(main, bg=C["bg_light"])
        ctrl.pack(fill="x", pady=(0, 8))
        tk.Label(ctrl, text="🔍 Search:",
                 font=FONT["body"], bg=C["bg_light"],
                 fg=C["txt_dark"]).pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self._on_filter)
        ttk.Entry(ctrl, textvariable=self.search_var,
                  font=FONT["body"], width=30
                  ).pack(side="left", padx=8)

        # Action buttons on right
        tk.Button(ctrl, text="🗑  Delete",
                  font=FONT["body"], bg=C["bg_light"],
                  fg=C["error"], activebackground="#FADBD8",
                  relief="flat", cursor="hand2",
                  command=self._on_delete, padx=12, pady=4
                  ).pack(side="right", padx=2)
        tk.Button(ctrl, text="✎  Edit",
                  font=FONT["body"], bg=C["bg_light"],
                  fg=C["primary"], activebackground="#D6EAF8",
                  relief="flat", cursor="hand2",
                  command=self._on_edit, padx=12, pady=4
                  ).pack(side="right", padx=2)
        tk.Button(ctrl, text="+  Add New",
                  font=FONT["btn"], bg=C["success"], fg="white",
                  activebackground="#196F3D",
                  relief="flat", cursor="hand2",
                  command=self._on_add, padx=14, pady=4
                  ).pack(side="right", padx=2)

        # Treeview
        tree_frame = tk.Frame(main, bg=C["bg_light"])
        tree_frame.pack(fill="both", expand=True)

        cols = ("name", "contact", "email", "phone")
        self.tree = ttk.Treeview(tree_frame, columns=cols,
                                 show="headings", height=18,
                                 selectmode="browse")
        col_cfg = [
            ("name",    "Name",         260, "w"),
            ("contact", "Contact",      180, "w"),
            ("email",   "Email",        240, "w"),
            ("phone",   "Phone",        150, "w"),
        ]
        for cid, h, w, anchor in col_cfg:
            self.tree.heading(cid, text=h)
            self.tree.column(cid, width=w, anchor=anchor)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.tree.bind("<Double-Button-1>", lambda e: self._on_edit())

        # Status label
        self._summary = tk.Label(main, text="",
                                 font=FONT["small"],
                                 bg=C["bg_light"], fg=C["txt_mid"])
        self._summary.pack(anchor="w", pady=(6, 0))

    def _reload(self):
        self._all = load_full_suppliers_db()
        # Sort alphabetically by name
        self._all.sort(key=lambda s: (s.get("name") or "").lower())
        self._show_filtered(self._all)

    def _show_filtered(self, items):
        def safe(v):
            if v is None:
                return ""
            if isinstance(v, (list, tuple)):
                return ", ".join(str(x) for x in v if x is not None)
            return str(v)
        for r in self.tree.get_children():
            self.tree.delete(r)
        for s in items:
            try:
                self.tree.insert("", "end", values=(
                    safe(s.get("name")),
                    safe(s.get("contact_name")),
                    safe(s.get("email")),
                    safe(s.get("phone")),
                ))
            except Exception as e:
                print(f"[gui] tree.insert failed for "
                      f"{s.get('name','?')}: {e}")
        self._summary.config(
            text=f"{len(items)} of {len(self._all)} supplier(s)")

    def _on_filter(self, *args):
        q = self.search_var.get().strip().lower()
        if not q:
            self._show_filtered(self._all)
            return
        filtered = []
        for s in self._all:
            name = (s.get("name") or "").lower()
            email = (s.get("email") or "").lower()
            contact = (s.get("contact_name") or "").lower()
            aliases = " ".join(s.get("aliases") or []).lower()
            sid = (s.get("supplier_id") or "").lower()
            if (q in name or q in email or q in contact
                    or q in aliases or q in sid):
                filtered.append(s)
        self._show_filtered(filtered)

    def _selected_supplier(self):
        sel = self.tree.selection()
        if not sel:
            return None
        item = self.tree.item(sel[0])
        name = item["values"][0]
        for s in self._all:
            if s.get("name") == name:
                return s
        return None

    def _on_add(self):
        SupplierEditDialog(self.win, self, supplier=None)

    def _on_edit(self):
        s = self._selected_supplier()
        if not s:
            messagebox.showinfo(
                "No selection",
                "Please select a supplier first.",
                parent=self.win)
            return
        SupplierEditDialog(self.win, self, supplier=s)

    def _on_delete(self):
        s = self._selected_supplier()
        if not s:
            messagebox.showinfo(
                "No selection",
                "Please select a supplier first.",
                parent=self.win)
            return
        name = s.get("name", "(unnamed)")
        if not messagebox.askyesno(
                "Confirm delete",
                f"Delete supplier '{name}'?\n\n"
                f"This cannot be undone (but a .bak file is kept).",
                parent=self.win):
            return
        # Remove from list and save
        new_list = [x for x in self._all if x is not s]
        ok, err = save_suppliers_db(new_list)
        if not ok:
            messagebox.showerror(
                "Save failed", f"Could not save:\n\n{err}",
                parent=self.win)
            return
        self._reload()
        self.app.set_status(f"✓ Deleted supplier '{name}'")

    def save_supplier(self, original, updated) -> tuple:
        """
        Called by SupplierEditDialog when user clicks Save.
        original=None means new addition. Returns (success, error).
        """
        # Build new list
        new_list = list(self._all)
        if original is None:
            # Add new — check for duplicate name
            new_name = (updated.get("name") or "").strip().lower()
            for x in new_list:
                if (x.get("name") or "").strip().lower() == new_name:
                    return False, f"Supplier '{updated['name']}' already exists."
            new_list.append(updated)
        else:
            # Replace existing by identity
            replaced = False
            for i, x in enumerate(new_list):
                if x is original:
                    new_list[i] = updated
                    replaced = True
                    break
            if not replaced:
                # Fallback: replace by name match
                orig_name = (original.get("name") or "").lower()
                for i, x in enumerate(new_list):
                    if (x.get("name") or "").lower() == orig_name:
                        new_list[i] = updated
                        replaced = True
                        break
            if not replaced:
                return False, "Could not locate original supplier in list."

        ok, err = save_suppliers_db(new_list)
        if not ok:
            return False, err
        self._reload()
        action = "Added" if original is None else "Updated"
        self.app.set_status(
            f"✓ {action} supplier '{updated.get('name', '')}'")
        return True, ""


# ──────────────────────────────────────────────
# SUPPLIER EDIT (Add / Edit) DIALOG
# ──────────────────────────────────────────────
class SupplierEditDialog:
    """Form for adding a new supplier or editing an existing one.
    The contact_name field controls the email greeting ('Hi {first_name},').
    The email field controls the recipient. Saving immediately writes the
    full suppliers.json so the next PO email picks up the changes."""

    def __init__(self, parent, manage_dialog, supplier=None):
        self.manage = manage_dialog
        self.original = supplier  # None for add, dict for edit
        is_edit = supplier is not None

        self.win = tk.Toplevel(parent)
        self.win.title("Edit Supplier" if is_edit else "Add New Supplier")
        self.win.geometry("640x720")
        self.win.configure(bg=C["bg_light"])
        self.win.grab_set()
        self._build(is_edit)
        if is_edit:
            self._populate(supplier)

    def _build(self, is_edit):
        # Header
        hdr = tk.Frame(self.win, bg=C["bg_dark"], height=52)
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)
        icon, title = ("✎", "Edit Supplier") if is_edit \
            else ("+", "Add New Supplier")
        tk.Label(hdr, text=f"{icon}  {title}",
                 font=FONT["heading"],
                 bg=C["bg_dark"], fg=C["txt_light"]
                 ).place(x=16, rely=0.5, anchor="w")

        # Footer
        footer = tk.Frame(self.win, bg=C["bg_dark"], height=64)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)
        tk.Button(footer, text="  ✓  Save  ",
                  font=FONT["btn"], bg=C["success"], fg="white",
                  activebackground="#196F3D",
                  relief="flat", cursor="hand2",
                  command=self._on_save, padx=20, pady=8
                  ).place(relx=1.0, rely=0.5, anchor="e", x=-16)
        tk.Button(footer, text="Cancel",
                  font=FONT["body"],
                  bg=C["bg_dark"], fg="#7F8C8D",
                  activebackground=C["bg_mid"],
                  relief="flat", cursor="hand2",
                  command=self.win.destroy, padx=12, pady=8
                  ).place(relx=1.0, rely=0.5, anchor="e", x=-130)

        # Scrollable form area
        main = tk.Frame(self.win, bg=C["bg_light"])
        main.pack(fill="both", expand=True, padx=20, pady=15)

        card = tk.Frame(main, bg=C["bg_card"],
                        highlightthickness=1,
                        highlightbackground=C["border"])
        card.pack(fill="both", expand=True)

        tk.Label(card, text="SUPPLIER DETAILS",
                 font=("Segoe UI", 9, "bold"),
                 bg=C["bg_card"], fg=C["txt_mid"]
                 ).pack(anchor="w", padx=14, pady=(12, 4))

        # Edit-mode safety hint
        if is_edit:
            hint = tk.Frame(card, bg="#FFF8E1",
                            highlightthickness=1,
                            highlightbackground="#F39C12")
            hint.pack(fill="x", padx=14, pady=(2, 8))
            tk.Label(hint,
                     text="💡  Leave a field blank to keep its current "
                          "value. Only fields you actually change will "
                          "be updated.",
                     font=FONT["small"],
                     bg="#FFF8E1", fg="#7E5109",
                     anchor="w", justify="left",
                     wraplength=560
                     ).pack(anchor="w", padx=10, pady=6)

        # Field builder
        def field(label, width=None, helper=""):
            r = tk.Frame(card, bg=C["bg_card"])
            r.pack(fill="x", padx=14, pady=4)
            tk.Label(r, text=label,
                     font=FONT["body"],
                     bg=C["bg_card"], fg=C["txt_dark"],
                     width=18, anchor="w").pack(side="left")
            e = ttk.Entry(r, font=FONT["body"],
                          **({"width": width} if width else {}))
            e.pack(side="left", fill="x", expand=(width is None), padx=(0, 8))
            if helper:
                tk.Label(r, text=helper,
                         font=FONT["small"],
                         bg=C["bg_card"], fg=C["txt_mid"]
                         ).pack(side="left")
            return e

        self.name_e = field("Name *", helper="company / supplier name")
        self.contact_e = field("Contact name",
                               helper="for email greeting 'Hi {first name},'")
        self.email_e = field("Email", helper="goes to the To: field")
        self.phone_e = field("Phone")
        self.sid_e = field("Supplier ID", width=20,
                           helper="optional internal reference")
        self.web_e = field("Website", helper="e.g. supplier.co.uk")
        self.aliases_e = field("Aliases",
                               helper="comma-separated, for search")

        # Address — multi-line
        addr_lbl = tk.Frame(card, bg=C["bg_card"])
        addr_lbl.pack(fill="x", padx=14, pady=(10, 4))
        tk.Label(addr_lbl, text="Address (one line each)",
                 font=FONT["body"],
                 bg=C["bg_card"], fg=C["txt_dark"]
                 ).pack(side="left")
        tk.Label(addr_lbl, text="appears on PO PDFs",
                 font=FONT["small"],
                 bg=C["bg_card"], fg=C["txt_mid"]
                 ).pack(side="left", padx=8)

        addr_frame = tk.Frame(card, bg=C["bg_card"])
        addr_frame.pack(fill="x", padx=14, pady=(0, 12))
        self.addr_text = tk.Text(addr_frame, height=6,
                                 font=FONT["body"],
                                 bg="white", fg=C["txt_dark"],
                                 relief="solid", bd=1,
                                 highlightthickness=0,
                                 wrap="word")
        self.addr_text.pack(fill="x")

        self.name_e.focus_set()

    def _populate(self, s):
        """Pre-fill form fields from existing supplier dict.

        Defensive: any value in suppliers.json could be None, a list,
        or another non-string type — convert everything to string and
        guard against Tk's 'wrong # args' error when special chars or
        non-strings sneak in.
        """
        def safe_str(value, default=""):
            if value is None:
                return default
            if isinstance(value, (list, tuple)):
                return ", ".join(str(x) for x in value if x is not None)
            return str(value)

        try:
            self.name_e.insert(0,    safe_str(s.get("name")))
            self.contact_e.insert(0, safe_str(s.get("contact_name")))
            self.email_e.insert(0,   safe_str(s.get("email")))
            self.phone_e.insert(0,   safe_str(s.get("phone")))
            self.sid_e.insert(0,     safe_str(s.get("supplier_id")))
            self.web_e.insert(0,     safe_str(s.get("web")))

            # Aliases: list of strings → comma-separated string
            aliases = s.get("aliases") or []
            if isinstance(aliases, list):
                aliases_str = ", ".join(str(a) for a in aliases if a)
            else:
                aliases_str = safe_str(aliases)
            self.aliases_e.insert(0, aliases_str)

            # Address: list of strings → newline-separated
            addr_lines = s.get("address_lines") or []
            if isinstance(addr_lines, list):
                addr_str = "\n".join(str(ln) for ln in addr_lines if ln)
            else:
                addr_str = safe_str(addr_lines)
            if addr_str:
                self.addr_text.insert("1.0", addr_str)
        except Exception as e:
            print(f"[gui] _populate error: {e}")
            messagebox.showwarning(
                "Display warning",
                f"Some fields couldn't be displayed cleanly:\n\n{e}\n\n"
                f"You can still edit the supplier.",
                parent=self.win)

    def _collect(self) -> dict:
        """Build updated supplier dict from form fields.

        Behaviour:
          - Add mode (self.original is None): take whatever was entered.
          - Edit mode: if a field is left BLANK, keep the original value.
            This protects against accidental data loss — if you only
            want to change name + email, you don't need to retype
            phone, address, etc.

        Any extra keys present in the original supplier dict that aren't
        editable in the form (e.g. legacy fields) are preserved untouched.
        """
        out = dict(self.original) if self.original else {}
        is_edit = self.original is not None

        # Simple text fields: form_field_name → entry widget
        text_fields = [
            ("name",         self.name_e),
            ("contact_name", self.contact_e),
            ("email",        self.email_e),
            ("phone",        self.phone_e),
            ("supplier_id",  self.sid_e),
            ("web",          self.web_e),
        ]
        for key, widget in text_fields:
            new_val = widget.get().strip()
            if new_val:
                out[key] = new_val
            elif is_edit:
                # Empty + edit mode → keep whatever was there before
                # (don't overwrite original with blank).
                # If original had no value, leave key as-is (likely empty
                # string from the dict() copy above, or absent).
                pass
            else:
                # Add mode → store empty string explicitly
                out[key] = ""

        # Aliases — comma-separated list
        aliases_raw = self.aliases_e.get().strip()
        if aliases_raw:
            out["aliases"] = [a.strip() for a in aliases_raw.split(",")
                              if a.strip()]
        elif not is_edit:
            out["aliases"] = []
        # else: keep original aliases (already in `out` from dict copy)

        # Address — newline-separated list
        addr_raw = self.addr_text.get("1.0", "end").strip()
        if addr_raw:
            out["address_lines"] = [ln.strip()
                                     for ln in addr_raw.split("\n")
                                     if ln.strip()]
        elif not is_edit:
            out["address_lines"] = []
        # else: keep original address_lines

        return out

    def _validate(self, data) -> str:
        # In edit mode `data` already contains preserved values from the
        # original where the form was blank, so we just check the final
        # result is sane.
        if not (data.get("name") or "").strip():
            return "Name is required."
        email = (data.get("email") or "").strip()
        if email and ("@" not in email
                      or "." not in email.split("@")[-1]):
            return f"Email '{email}' looks invalid."
        return ""

    def _on_save(self):
        data = self._collect()
        err = self._validate(data)
        if err:
            messagebox.showerror("Invalid input", err, parent=self.win)
            return

        ok, save_err = self.manage.save_supplier(self.original, data)
        if not ok:
            messagebox.showerror(
                "Save failed", save_err, parent=self.win)
            return

        self.win.destroy()


# ──────────────────────────────────────────────
# SHOW RECENT POs DIALOG
# ──────────────────────────────────────────────
class ShowRecentDialog:
    def __init__(self, parent, app: POFlowApp):
        self.app = app
        self.win = tk.Toplevel(parent)
        self.win.title("Recent Purchase Orders")
        self.win.geometry("1000x520")
        self.win.configure(bg=C["bg_light"])
        self.win.grab_set()
        self._build()
        self._load()

    def _build(self):
        hdr = tk.Frame(self.win, bg=C["bg_dark"], height=52)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="📊  Recent Purchase Orders",
                 font=FONT["heading"],
                 bg=C["bg_dark"], fg=C["txt_light"]
                 ).place(x=16, rely=0.5, anchor="w")

        # Controls row
        ctrl = tk.Frame(self.win, bg=C["bg_light"])
        ctrl.pack(fill="x", padx=16, pady=10)
        tk.Label(ctrl, text="Days back:",
                 font=FONT["body"], bg=C["bg_light"],
                 fg=C["txt_dark"]).pack(side="left")

        self.days_var = tk.StringVar(value="7")
        days_spin = ttk.Spinbox(ctrl, from_=1, to=90,
                                textvariable=self.days_var,
                                width=5, font=FONT["body"])
        days_spin.pack(side="left", padx=6)

        ttk.Button(ctrl, text="Refresh",
                   command=self._load).pack(side="left", padx=8)

        self._summary_lbl = tk.Label(ctrl, text="",
                                     font=FONT["body"],
                                     bg=C["bg_light"],
                                     fg=C["txt_mid"])
        self._summary_lbl.pack(side="left", padx=16)

        # Treeview
        tree_frame = tk.Frame(self.win, bg=C["bg_light"])
        tree_frame.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        cols = ("sheet", "po", "date", "due", "supplier",
                "machine", "cost", "age")
        self.tree = ttk.Treeview(tree_frame, columns=cols,
                                 show="headings", height=16)

        col_cfg = [
            ("sheet",    "Sheet",     110, "w"),
            ("po",       "PO Number",  130, "w"),
            ("date",     "Ordered",     90, "center"),
            ("due",      "Due",         90, "center"),
            ("supplier", "Supplier",   160, "w"),
            ("machine",  "Machine",    130, "w"),
            ("cost",     "Cost",        80, "e"),
            ("age",      "Age",         60, "center"),
        ]
        for col_id, heading, width, anchor in col_cfg:
            self.tree.heading(col_id, text=heading)
            self.tree.column(col_id, width=width, anchor=anchor)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

    def _load(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._summary_lbl.config(text="Loading...")

        try:
            days = int(self.days_var.get())
        except ValueError:
            days = 7

        threading.Thread(target=self._do_load,
                         args=(days,), daemon=True).start()

    def _do_load(self, days: int):
        pf = _pf
        today = datetime.today().date()
        cutoff = today - __import__("datetime").timedelta(days=days - 1)
        results = []
        total_cost = 0.0

        try:
            from openpyxl import load_workbook as lw
            wb = lw(pf.WORKBOOK_PATH, data_only=True)
            for sheet_name in wb.sheetnames:
                if not pf._is_valid_month_sheet(sheet_name):
                    continue
                ws = wb[sheet_name]
                header_row, colmap = pf.detect_header_row_and_columns(ws)
                if not colmap:
                    continue
                po_col = colmap.get("PONumber")
                date_col = colmap.get("DateOrdered")
                if not po_col or not date_col:
                    continue
                supp_col = colmap.get("Supplier")
                mach_col = colmap.get("Machine")
                cost_col = colmap.get("Cost")
                due_col = colmap.get("DueDate")
                data_start = header_row + 1
                for r in range(data_start, (ws.max_row or data_start) + 1):
                    po_val = ws[f"{po_col}{r}"].value
                    if not po_val:
                        continue
                    date_val = ws[f"{date_col}{r}"].value
                    ordered_date = None
                    if isinstance(date_val, datetime):
                        ordered_date = date_val.date()
                    elif isinstance(date_val, date):
                        ordered_date = date_val
                    elif isinstance(date_val, str):
                        ordered_date = pf.parse_ddmmyyyy_to_date(date_val)
                    if not ordered_date:
                        continue
                    if cutoff <= ordered_date <= today:
                        cost = ws[f"{cost_col}{r}"].value if cost_col else 0
                        try:
                            cost_f = float(cost) if cost else 0.0
                        except (TypeError, ValueError):
                            cost_f = 0.0
                        due_val = ws[f"{due_col}{r}"].value if due_col else None
                        due_str = ""
                        if isinstance(due_val, (datetime, date)):
                            due_str = (due_val.date() if isinstance(
                                due_val, datetime) else due_val
                                       ).strftime("%d/%m/%Y")
                        elif isinstance(due_val, str):
                            due_str = due_val[:10]
                        age = (today - ordered_date).days
                        age_str = "TODAY" if age == 0 else \
                            "yest." if age == 1 else f"{age}d"
                        results.append({
                            "sheet": sheet_name,
                            "po": str(po_val),
                            "date": ordered_date.strftime("%d/%m/%Y"),
                            "due": due_str,
                            "supplier": str(ws[f"{supp_col}{r}"].value or "")[:30],
                            "machine": str(ws[f"{mach_col}{r}"].value or "")[:22],
                            "cost": cost_f,
                            "age": age_str,
                            "ordered_date": ordered_date,
                        })
                        total_cost += cost_f
        except Exception as e:
            results = [{"error": str(e)}]

        self.win.after(0, lambda: self._populate(
            results, days, total_cost))

    def _populate(self, results, days, total_cost):
        for row in self.tree.get_children():
            self.tree.delete(row)

        if results and "error" in results[0]:
            self._summary_lbl.config(
                text=f"Error: {results[0]['error']}", fg=C["error"])
            return

        results.sort(key=lambda r: r["ordered_date"])
        for r in results:
            tag = "today" if r["age"] == "TODAY" else ""
            self.tree.insert("", "end", values=(
                r["sheet"], r["po"], r["date"], r["due"],
                r["supplier"], r["machine"],
                f"£{r['cost']:,.2f}", r["age"]
            ), tags=(tag,))

        self.tree.tag_configure("today", background="#EBF5FB")
        summary = (f"{len(results)} PO(s) in last {days} day(s)  •  "
                   f"Total: £{total_cost:,.2f}")
        self._summary_lbl.config(text=summary, fg=C["txt_mid"])


# ──────────────────────────────────────────────
# SETTINGS DIALOG
# ──────────────────────────────────────────────
class SettingsDialog:
    def __init__(self, parent, app: POFlowApp):
        self.app = app
        self.win = tk.Toplevel(parent)
        self.win.title("Settings")
        self.win.geometry("420x280")
        self.win.resizable(False, False)
        self.win.configure(bg=C["bg_light"])
        self.win.grab_set()
        self._build()

    def _build(self):
        hdr = tk.Frame(self.win, bg=C["bg_dark"], height=52)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="⚙️  Settings",
                 font=FONT["heading"],
                 bg=C["bg_dark"], fg=C["txt_light"]
                 ).place(x=16, rely=0.5, anchor="w")

        card = tk.Frame(self.win, bg=C["bg_card"],
                        highlightthickness=1,
                        highlightbackground=C["border"])
        card.pack(fill="both", expand=True, padx=20, pady=16)

        tk.Label(card, text="Your display name",
                 font=FONT["body"], bg=C["bg_card"],
                 fg=C["txt_dark"]).grid(row=0, column=0,
                                        sticky="w", padx=14, pady=(16, 4))
        self.name_var = tk.StringVar(
            value=self.app.user_cfg.get("name", ""))
        ttk.Entry(card, textvariable=self.name_var,
                  font=FONT["body"], width=28
                  ).grid(row=1, column=0, sticky="ew",
                         padx=14, pady=(0, 12))

        tk.Label(card, text="Your initials (used in PO numbers)",
                 font=FONT["body"], bg=C["bg_card"],
                 fg=C["txt_dark"]).grid(row=2, column=0,
                                        sticky="w", padx=14)
        self.init_var = tk.StringVar(
            value=self.app.user_cfg.get("initials", ""))
        ttk.Entry(card, textvariable=self.init_var,
                  font=FONT["body"], width=8
                  ).grid(row=3, column=0, sticky="w",
                         padx=14, pady=(4, 12))
        tk.Label(card,
                 text="Example: George Fiala → GF",
                 font=FONT["small"], bg=C["bg_card"],
                 fg=C["txt_mid"]).grid(row=4, column=0,
                                       sticky="w", padx=14)
        card.columnconfigure(0, weight=1)

        # Footer
        footer = tk.Frame(self.win, bg=C["bg_light"])
        footer.pack(fill="x", padx=20, pady=(0, 16))

        tk.Button(footer, text="Save",
                  font=FONT["btn"],
                  bg=C["success"], fg="white",
                  relief="flat", cursor="hand2",
                  command=self._save, padx=16, pady=6
                  ).pack(side="right")
        tk.Button(footer, text="Cancel",
                  font=FONT["body"],
                  bg=C["bg_light"], fg=C["txt_mid"],
                  relief="flat", cursor="hand2",
                  command=self.win.destroy, padx=12, pady=6
                  ).pack(side="right", padx=8)

    def _save(self):
        cfg = {
            "name":     self.name_var.get().strip(),
            "initials": self.init_var.get().strip().upper(),
        }
        if not cfg["name"] or not cfg["initials"]:
            messagebox.showerror("Error",
                                 "Name and initials are required.",
                                 parent=self.win)
            return
        save_user_config(cfg)
        self.app.user_cfg = cfg
        messagebox.showinfo(
            "Saved",
            f"Settings saved.\n\n"
            f"Name: {cfg['name']}\n"
            f"Initials: {cfg['initials']}\n\n"
            f"Note: PO numbers will use initials '{cfg['initials']}' "
            f"from your next PO.",
            parent=self.win)
        self.win.destroy()


# ──────────────────────────────────────────────
# ENTRY POINT
# ──────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()

    # Hide terminal flash on Windows when launched with pythonw
    try:
        root.wm_attributes("-alpha", 0)
        root.update()
        root.wm_attributes("-alpha", 1)
    except Exception:
        pass

    app = POFlowApp(root)
    root.mainloop()
