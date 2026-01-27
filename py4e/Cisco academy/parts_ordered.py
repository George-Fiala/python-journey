from openpyxl import load_workbook, Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Alignment, Font
from datetime import datetime
from difflib import get_close_matches
import re
import time
import json

# ====== SETTINGS ======
FILE_PATH = r"C:\Users\george\OneDrive - S Sheard\Desktop\Parts ordered Home.xlsx"
SUPPLIERS_PATH = r"C:\Users\george\OneDrive - S Sheard\Desktop\Python project\suppliers.json"

SHEET_COSTS = "Cost centre"   # sheet with Machine / Area / CostCentre table
INITIALS = "GF"               # your initials

# ====== COLUMNS CONFIGURATION (UPDATED) ======
COL_MACHINE      = "A"
COL_AREA         = "B"
COL_REASON       = "C"
COL_MECH_ELEC    = "D"
COL_SUPPLIER     = "E"
COL_PARTS        = "F"
COL_JOB          = "G"
COL_PO           = "H"
COL_COST         = "I"   # Cost £
COL_ORDERED      = "J"   # Date ordered
COL_DUE          = "K"   # Due date
COL_DELIVERED    = "L"   # Delivered date

# --- NEW COLUMNS ADDED ---
COL_SAVINGS       = "M"   # Cost Initiatives (Savings)
COL_ORIG_SUPPLIER = "N"   # Original Supplier

# --- MOVED COLUMN ---
COL_COST_CENTER   = "O"   # Cost Centre (Moved from N to O)

# ====== ALIASES ======
AREA_ALIAS = {
    "cons": "consumables",
    "consum": "consumables",
    "consumable": "consumables",
    "consumables": "consumables",
    "conumables": "consumables",
    "sonsumbales": "consumables",
}

MACHINE_ALIAS = {
    "e1": "Emba 1", "e 1": "Emba 1", "emba1": "Emba 1", "emba 1": "Emba 1",
    "e2": "Emba 2", "e 2": "Emba 2", "emba2": "Emba 2", "emba 2": "Emba 2",
    "e3": "Emba 3", "e 3": "Emba 3", "emba3": "Emba 3", "emba 3": "Emba 3",
    "e4": "Emba 4", "e 4": "Emba 4", "emba4": "Emba 4", "emba 4": "Emba 4",
    "e5": "Emba 5", "e 5": "Emba 5", "emba5": "Emba 5", "emba 5": "Emba 5",
    "e6": "Emba 6", "e 6": "Emba 6", "emba6": "Emba 6", "emba 6": "Emba 6",
    "g1": "Gopfert 1", "g 1": "Gopfert 1", "gopfert1": "Gopfert 1", "gopfert 1": "Gopfert 1",
    "g2": "Gopfert 2", "g 2": "Gopfert 2", "gopfert2": "Gopfert 2", "gopfert 2": "Gopfert 2",
    "b1": "Bobst 1 Visionfold", "b 1": "Bobst 1 Visionfold", "bobst1": "Bobst 1 Visionfold", "bobst 1": "Bobst 1 Visionfold",
    "b2": "Bobst 2 ExpertFold", "b 2": "Bobst 2 ExpertFold", "bobst2": "Bobst 2 ExpertFold", "bobst 2": "Bobst 2 ExpertFold",
    "a1": "Asahi 1", "a 1": "Asahi 1", "asahi1": "Asahi 1", "asahi 1": "Asahi 1",
    "a2": "Asahi 2", "a 2": "Asahi 2", "asahi2": "Asahi 2", "asahi 2": "Asahi 2",
    "vega": "Vega", "avanti": "Avanti conveyor lines", "production": "Workshop/Plant", "prod": "Workshop/Plant",
    "plant": "Workshop/Plant", "workshop": "Workshop/Plant"
}

# ====== NORMALIZATION & HELPERS ======

def normalize_machine_name(machine_raw):
    m = (machine_raw or "").strip().lower()
    return MACHINE_ALIAS.get(m, machine_raw)

def normalize_text(name):
    if name is None: return ""
    return re.sub(r"[^a-z0-9]+", "", str(name).strip().lower())

def is_workshop_plant_electrical(machine_any, mech_elec_raw):
    m = normalize_text(machine_any)
    me = normalize_text(mech_elec_raw)
    workshop_like = m in {"plant", "workshop", "workshopplant", "production", "prod"}
    electrical_like = (me in {"ele", "elec", "electrical"} or "elect" in me or me in {"elerepairs", "elecrepairs"})
    return workshop_like and electrical_like

AREA_ALIAS_NORM = {normalize_text(k): v for k, v in AREA_ALIAS.items()}

def normalize_area_name(area_raw, machine_canon=""):
    raw = (area_raw or "").strip()
    norm_area = normalize_text(raw)
    norm_machine = normalize_text(machine_canon)
    if norm_machine == normalize_text("Bobst 2 ExpertFold") and norm_area in {"ctech", "ctechrobopacker"}:
        return "C-Tech Robopacker"
    return AREA_ALIAS_NORM.get(norm_area, raw)

# ====== EXCEL OPS ======

def load_or_create_workbook(file_path):
    while True:
        try:
            return load_workbook(file_path)
        except FileNotFoundError:
            return Workbook()
        except PermissionError:
            print(f"\n⚠️ Excel file is open: {file_path}")
            input("➡️ Close it and press Enter...")
            time.sleep(0.5)

def get_sheet_for_current_month(wb):
    now = datetime.today()
    full_name = now.strftime("%B %Y")
    if full_name in wb.sheetnames: return wb[full_name]
    ws = wb.create_sheet(full_name) if len(wb.sheetnames) > 1 or wb.active.max_row > 1 else wb.active
    ws.title = full_name
    return ws

def init_parts_sheet(ws):
    if ws[f"{COL_MACHINE}2"].value is None:
        # Row 1: Summary Formulas
        ws["H1"], ws["L1"] = "Total spend £", "Total Savings £"
        ws["I1"], ws["M1"] = f"=SUM({COL_COST}3:{COL_COST}1000)", f"=SUM({COL_SAVINGS}3:{COL_SAVINGS}1000)"
        ws["I1"].font = ws["M1"].font = Font(bold=True)
        ws["M1"].font.color = "006100"

        # Row 2: Headers
        headers = {
            COL_MACHINE: "Machine", COL_AREA: "Area", COL_REASON: "Reason",
            COL_MECH_ELEC: "Mech/Elec/Other", COL_SUPPLIER: "Supplier",
            COL_PARTS: "Parts ordered", COL_JOB: "Job No", COL_PO: "PO number",
            COL_COST: "Cost £", COL_ORDERED: "Date ordered", COL_DUE: "Due date",
            COL_DELIVERED: "Delivered date", COL_SAVINGS: "Cost Initiatives",
            COL_ORIG_SUPPLIER: "Original supplier", COL_COST_CENTER: "Cost Centre"
        }
        for col, title in headers.items():
            cell = ws[f"{col}2"]
            cell.value, cell.font = title, Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            if col == COL_SAVINGS: cell.fill = PatternFill("solid", start_color="92D050")
            if col == COL_ORIG_SUPPLIER: 
                cell.fill = PatternFill("solid", start_color="0070C0")
                cell.font = Font(bold=True, color="FFFFFF")

        widths = {"A": 16, "B": 18, "C": 18, "D": 18, "E": 24, "F": 40, "G": 14, "H": 20, "I": 12, "J": 14, "K": 14, "L": 14, "M": 16, "N": 20, "O": 16}
        for c, w in widths.items(): ws.column_dimensions[c].width = w

def find_cost_centre(ws_costs, machine_input, area_input):
    m_canon = normalize_machine_name(machine_input)
    a_canon = normalize_area_name(area_input, m_canon)
    norm_m, norm_a = normalize_text(m_canon), normalize_text(a_canon)
    
    # Priority matching logic (Simplified for space but keeps your logic)
    candidates = {}
    for row in ws_costs.iter_rows(min_row=2, values_only=True):
        mach, area, cc = row[:3]
        if not mach or not cc: continue
        nm, na = normalize_text(mach), normalize_text(area)
        if nm == norm_m:
            if na == norm_a: return cc # Exact match
            if na in ("", "machine"): candidates['mach'] = cc
    return candidates.get('mach', "UNKNOWN")

def count_orders_today(ws):
    today = datetime.today().date()
    count = 0
    for r in range(3, ws.max_row + 1):
        val = ws[f"{COL_ORDERED}{r}"].value
        if isinstance(val, datetime) and val.date() == today: count += 1
    return count

def generate_po(cost_center, initials, count_today):
    now = datetime.now()
    return f"{cost_center}-{initials}{now.day:02d}{now.month:02d}{count_today + 1}{now.year % 100:02d}"

# ====== SUPPLIER DATABASE LOGIC ======

def supplier_norm(text):
    t = (text or "").strip().lower().replace("&", " and ").replace("-", " ")
    return " ".join(re.sub(r"[^a-z0-9\s]", " ", t).split())

def load_suppliers_db(path):
    try:
        with open(path, "r", encoding="utf-8") as f: return json.load(f)
    except: return {"version": 1, "suppliers": []}

def save_suppliers_db(path, db):
    with open(path, "w", encoding="utf-8") as f: json.dump(db, f, indent=2)

def suggest_suppliers(query, suppliers, limit=8):
    q = supplier_norm(query)
    if not q: return []
    res = [s for s in suppliers if any(q in supplier_norm(str(v)) for v in [s.get("name"), s.get("supplier_id")])]
    return res[:limit]

def choose_supplier_interactive(path, allow_back=True):
    db = load_suppliers_db(path)
    suppliers = db.get("suppliers", [])
    last_suggestions = []
    while True:
        q = input(f"Supplier search (letters, 'add', 'manual', 'back'): ").strip().lower()
        if q == "back" and allow_back: return "BACK"
        if q == "manual": return {"name": input("Manual Name: ")}
        if q == "add":
            name = input("New Supplier Name: ")
            s = {"name": name, "supplier_id": name[:5].upper()}
            suppliers.append(s)
            save_suppliers_db(path, db)
            return s
        if q.isdigit() and last_suggestions and 1 <= int(q) <= len(last_suggestions):
            return last_suggestions[int(q)-1]
        
        last_suggestions = suggest_suppliers(q, suppliers)
        for i, s in enumerate(last_suggestions, 1):
            print(f"  {i}) {s.get('name')} [{s.get('supplier_id')}]")

# ====== INPUT & MAIN ======

def collect_order_data():
    steps = [("machine", "Machine"), ("area", "Area"), ("reason", "Reason"), ("mech_elec", "Mech/Elec"), 
             ("supplier", "Search Supplier"), ("parts", "Parts"), ("job_no", "Job No"), 
             ("cost", "Cost £"), ("savings", "Savings £"), ("orig_sup", "Original Supplier"), ("due", "Due (dd/mm/yyyy)")]
    data, i = {}, 0
    while i < len(steps):
        key, quest = steps[i]
        if key == "supplier":
            res = choose_supplier_interactive(SUPPLIERS_PATH)
            if res == "BACK": i -= 1; continue
            data["supplier"] = res.get("name")
            i += 1; continue
        ans = input(f"{quest}: ").strip()
        if ans.lower() == "back": i = max(0, i - 1); continue
        data[key] = ans
        i += 1
    return data

def main():
    order = collect_order_data()
    wb = load_or_create_workbook(FILE_PATH)
    ws = get_sheet_for_current_month(wb)
    init_parts_sheet(ws)
    
    machine = normalize_machine_name(order["machine"])
    area = normalize_area_name(order["area"], machine)
    
    # Logic for 321B
    if is_workshop_plant_electrical(machine, order["mech_elec"]):
        cost_center = "321B"
    else:
        cost_center = find_cost_centre(wb[SHEET_COSTS], machine, area)

    po = generate_po(cost_center, INITIALS, count_orders_today(ws))
    row = 3
    while ws[f"A{row}"].value: row += 1

    # Writing
    ws[f"{COL_ORDERED}{row}"] = datetime.today().date()
    ws[f"{COL_MACHINE}{row}"], ws[f"{COL_AREA}{row}"] = machine, area
    ws[f"{COL_REASON}{row}"], ws[f"{COL_MECH_ELEC}{row}"] = order["reason"], order["mech_elec"]
    ws[f"{COL_SUPPLIER}{row}"], ws[f"{COL_PARTS}{row}"] = order["supplier"], order["parts"]
    ws[f"{COL_JOB}{row}"], ws[f"{COL_PO}{row}"] = order["job_no"], po
    ws[f"{COL_COST_CENTER}{row}"] = cost_center

    if order["cost"]:
        ws[f"{COL_COST}{row}"] = float(order["cost"].replace("£","").replace(",",""))
        ws[f"{COL_COST}{row}"].number_format = '£#,##0.00'
    if order["savings"]:
        ws[f"{COL_SAVINGS}{row}"] = float(order["savings"].replace("£","").replace(",",""))
        ws[f"{COL_SAVINGS}{row}"].number_format = '£#,##0.00'
    if order["orig_sup"]:
        ws[f"{COL_ORIG_SUPPLIER}{row}"] = order["orig_sup"]
    if order["due"]:
        try: ws[f"{COL_DUE}{row}"] = datetime.strptime(order["due"], "%d/%m/%Y").date()
        except: pass

    wb.save(FILE_PATH)
    print(f"\n✅ Added to row {row}. PO: {po}")

if __name__ == "__main__":
    main()