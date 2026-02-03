# =========================
# PO + GR One Flow (v1.3.1)
# - Create Purchase Order (PO) PDF
# - Write to "Parts Ordered" monthly sheet
# - Add to "To Receive" (Goods Receipt (GR) backlog)
# - Mark as Received later (updates both sheets)
#
# Fix in v1.3.1:
# - Month totals (I1, M1) sum ONLY real rows (must have PO number; fallback Date ordered)
#   This prevents "ghost" costs from old/cleared rows from being included.
# =========================

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import json
import re
import os
from difflib import get_close_matches

# Optional PDF generation (ReportLab)
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False


# ========= CONFIG =========
WORKBOOK_PATH = r"C:\Users\george\OneDrive - S Sheard\Desktop\Parts ordered Home.xlsx"
SUPPLIERS_JSON_PATH = r"C:\Users\george\OneDrive - S Sheard\Desktop\Python project\suppliers.json"

INITIALS = "GF"
SHEET_COST_CENTRE = "Cost centre"
TO_RECEIVE_SHEET = "To Receive"
PO_PDF_OUTPUT_DIR = r"C:\Users\george\OneDrive - S Sheard\Desktop\PO PDFs"

DEBUG = True  # Set False when stable


# ========= VAT SETTINGS =========
VAT_RATE_DEFAULT = 0.20          # 20%
PROMPT_VAT_RATE_EACH_PO = True   # Ask VAT% each PO (Enter keeps default)


# ========= COMPANY DETAILS =========
SHEARD_COMPANY_NAME = "Sheard Packaging"

SHEARD_BILL_TO = [
    "Sheard Packaging",
    "Solar Works, Church Street",
    "Greetland, Halifax",
    "HX4 8EG",
    "United Kingdom",
]

SHEARD_DELIVER_TO = [
    "Sheard Packaging",
    "Solar Works, Church Street",
    "Greetland, Halifax",
    "HX4 8EG",
    "United Kingdom",
]

CONTACT_NAME = "George Fiala"
CONTACT_EMAIL = "fialajir84@gmail.com"
CONTACT_PHONE = "07842 104144"


# ========= INPUT HELPERS =========
BACK = "__BACK__"

def is_back(s: str) -> bool:
    return (s or "").strip().lower() == "back"

def input_text(prompt: str, allow_blank: bool = False, allow_back: bool = True) -> str:
    while True:
        raw = input(prompt).strip()
        if allow_back and is_back(raw):
            return BACK
        if raw == "" and allow_blank:
            return ""
        if raw == "" and not allow_blank:
            print("This field cannot be blank. Type a value or 'back'.")
            continue
        return raw

def input_positive_int(prompt: str, allow_blank: bool = False, allow_back: bool = True) -> str:
    while True:
        raw = input(prompt).strip()
        if allow_back and is_back(raw):
            return BACK
        if raw == "" and allow_blank:
            return ""
        if not raw.isdigit() or int(raw) <= 0:
            print("Enter a whole number (e.g., 1). Or press Enter to finish. Type 'back' to go back.")
            continue
        return raw

def input_money(prompt: str, allow_blank: bool = False, allow_back: bool = True) -> str:
    while True:
        raw = input(prompt).strip()
        if allow_back and is_back(raw):
            return BACK
        if raw == "" and allow_blank:
            return ""
        val = parse_currency_to_float(raw)
        if val is None:
            print("Enter a valid amount (e.g., 12.50). Type 'back' to go back.")
            continue
        return raw


# ========= HELPERS =========
def normalize_text(x) -> str:
    if x is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(x).strip().lower())

def parse_currency_to_float(text: str):
    t = (text or "").strip()
    if not t:
        return None
    t = t.replace("£", "").replace(",", "").strip()
    try:
        return float(t)
    except ValueError:
        return None

def parse_ddmmyyyy_to_date(text: str):
    t = (text or "").strip()
    if not t:
        return None
    try:
        return datetime.strptime(t, "%d/%m/%Y").date()
    except ValueError:
        return None

def ensure_dir(path: str):
    if path and not os.path.isdir(path):
        os.makedirs(path, exist_ok=True)

def load_or_create_workbook(path: str):
    while True:
        try:
            return load_workbook(path)
        except FileNotFoundError:
            return Workbook()
        except PermissionError:
            print(f"\nFile is open and cannot be read: {path}")
            input("Close it and press Enter to retry...")
        except Exception as e:
            print(f"\nError while loading workbook: {e}")
            input("Press Enter to retry...")

def safe_save_workbook(wb, path: str):
    while True:
        try:
            wb.save(path)
            return
        except PermissionError:
            print(f"\nFile is open and cannot be saved: {path}")
            input("Close it and press Enter to retry...")
        except Exception as e:
            print(f"\nError while saving workbook: {e}")
            input("Press Enter to retry...")

def get_month_sheet_name() -> str:
    return datetime.today().strftime("%B %Y")

def get_or_create_month_sheet(wb):
    name = get_month_sheet_name()

    if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet":
        ws = wb["Sheet"]
        ws.title = name
        return ws

    if name not in wb.sheetnames:
        return wb.create_sheet(name)
    return wb[name]


# ========= MONTH SUMMARY + AUTOFIT =========
def to_float_cell(value):
    """Convert Excel cell value to float where possible (handles £ strings)."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip()
        if not s or s.startswith("="):  # openpyxl does not evaluate formulas
            return 0.0
        v = parse_currency_to_float(s)
        return float(v) if v is not None else 0.0
    return 0.0

def is_blank_cell(value) -> bool:
    """True if cell is empty or contains only whitespace (including NBSP)."""
    if value is None:
        return True
    if isinstance(value, str):
        s = value.replace("\u00A0", " ").strip()
        return s == ""
    return False

def ensure_month_summary_cells(ws):
    """
    Matches your layout:
      H1: Total spend £  | I1: value
      L1: Total Savings £ | M1: value
    """
    if ws["H1"].value is None:
        ws["H1"].value = "Total spend £"
        ws["H1"].font = Font(bold=True)
        ws["H1"].alignment = Alignment(horizontal="right")

    if ws["I1"].value is None:
        ws["I1"].value = 0.0
    ws["I1"].number_format = "£#,##0.00"
    ws["I1"].font = Font(bold=True)

    if ws["L1"].value is None:
        ws["L1"].value = "Total Savings £"
        ws["L1"].font = Font(bold=True)
        ws["L1"].alignment = Alignment(horizontal="right")

    if ws["M1"].value is None:
        ws["M1"].value = 0.0
    ws["M1"].number_format = "£#,##0.00"
    ws["M1"].font = Font(bold=True)

def update_month_totals(ws_parts, header_row, colmap):
    """
    Sum Cost and Savings columns and write to I1 and M1.
    FIX: Only sum rows that are real entries (must have PO number; fallback Date ordered).
    """
    ensure_month_summary_cells(ws_parts)

    data_start = header_row + 1
    total_cost = 0.0
    total_savings = 0.0

    cost_col = colmap.get("Cost")
    savings_col = colmap.get("Savings")

    po_col = colmap.get("PONumber")
    date_col = colmap.get("DateOrdered")

    for r in range(data_start, ws_parts.max_row + 1):
        # Determine if this is a "real" row.
        real_row = False

        if po_col:
            po_val = ws_parts[f"{po_col}{r}"].value
            if not is_blank_cell(po_val):
                real_row = True
        elif date_col:
            dt_val = ws_parts[f"{date_col}{r}"].value
            if not is_blank_cell(dt_val):
                real_row = True
        else:
            # No reliable keys detected; fallback to old behavior (but this is risky).
            real_row = True

        if not real_row:
            continue

        if cost_col:
            total_cost += to_float_cell(ws_parts[f"{cost_col}{r}"].value)
        if savings_col:
            total_savings += to_float_cell(ws_parts[f"{savings_col}{r}"].value)

    ws_parts["I1"].value = float(total_cost)
    ws_parts["M1"].value = float(total_savings)

def auto_fit_columns(ws, min_width=8, max_width=70, scan_rows_limit=300):
    """
    Approx Excel AutoFit: sets column width based on longest cell text in each column.
    scan_rows_limit prevents slowdowns on huge sheets.
    """
    max_row = min(ws.max_row, scan_rows_limit)
    max_col = ws.max_column

    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        best = 0

        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            if len(s) > best:
                best = len(s)

        width = best + 2
        width = max(min_width, min(width, max_width))
        ws.column_dimensions[letter].width = width

def set_sheet_view(ws, zoom=90):
    """Sets zoom so the sheet looks more 'fit' when opened."""
    try:
        ws.sheet_view.zoomScale = int(zoom)
    except Exception:
        pass


# ========= HEADER DETECTION (WRONG-COLUMN FIX) =========
HEADER_SYNONYMS = {
    "Machine": {"machine"},
    "Area": {"area", "areaslot", "slot", "areaslotlocation"},
    "Reason": {"reason"},
    "MechElecOther": {"mechelecother", "mechelec", "mech", "elec", "other"},
    "Supplier": {"supplier"},
    "PartsOrdered": {"partsordered", "partsorder", "parts"},
    "JobNo": {"jobno", "job", "jobnumber"},
    "PONumber": {"ponumber", "purchaseorder", "po"},
    "Cost": {"cost", "costgbp", "costpound"},
    "DateOrdered": {"dateordered", "ordereddate", "dateorder"},
    "DueDate": {"duedate", "due"},
    "DeliveredDate": {"delivereddate", "delivered"},
    "Savings": {"costinitiatives", "initiatives", "savings"},
    "OriginalSupplier": {"originalsupplier", "origsupplier"},
    "CostCentre": {"costcentre", "costcenter", "cc"},
}

REQUIRED_PARTS_FIELDS = [
    "Machine", "Area", "Reason", "MechElecOther", "Supplier",
    "PartsOrdered", "JobNo", "PONumber", "Cost", "DateOrdered",
    "DueDate", "Savings", "OriginalSupplier", "CostCentre"
]

def detect_header_row_and_columns(ws, max_scan_rows: int = 12):
    best_row = None
    best_score = -1
    best_map = {}

    max_cols = max(ws.max_column, 30)

    for r in range(1, max_scan_rows + 1):
        row_values = []
        for c in range(1, max_cols + 1):
            v = ws.cell(row=r, column=c).value
            row_values.append((c, normalize_text(v)))

        field_to_col = {}
        score = 0

        for field, synonyms in HEADER_SYNONYMS.items():
            for c, norm in row_values:
                if norm and norm in synonyms:
                    field_to_col[field] = get_column_letter(c)
                    score += 1
                    break

        if score > best_score:
            best_score = score
            best_row = r
            best_map = field_to_col

    if best_score >= 6:
        return best_row, best_map
    return None, {}

def init_parts_sheet_default(ws):
    headers = [
        ("A", "Machine"),
        ("B", "Area"),
        ("C", "Reason"),
        ("D", "Mech/Elec/Other"),
        ("E", "Supplier"),
        ("F", "Parts ordered"),
        ("G", "Job No"),
        ("H", "PO number"),
        ("I", "Cost £"),
        ("J", "Date ordered"),
        ("K", "Due date"),
        ("L", "Delivered date"),
        ("M", "Cost Initiatives"),
        ("N", "Original supplier"),
        ("O", "Cost Centre"),
    ]
    header_row = 2
    for col, label in headers:
        cell = ws[f"{col}{header_row}"]
        cell.value = label
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A3"

def ensure_parts_column_map(ws):
    header_row, colmap = detect_header_row_and_columns(ws)
    if not colmap:
        init_parts_sheet_default(ws)
        header_row, colmap = detect_header_row_and_columns(ws)

    missing = [f for f in REQUIRED_PARTS_FIELDS if f not in colmap]
    if missing:
        print("\nWARNING: Some required headers were not detected:")
        print("  " + ", ".join(missing))
        print("Please ensure your header names match your sheet (or adjust HEADER_SYNONYMS).")

    return header_row or 2, colmap

def find_first_empty_row(ws, key_col_letter: str, start_row: int) -> int:
    r = start_row
    while ws[f"{key_col_letter}{r}"].value:
        r += 1
    return r


# ========= COST CENTRE (CC) LOGIC =========
MACHINE_ALIAS_RAW = {
    "plant": "Workshop/Plant",
    "workshop": "Workshop/Plant",
    "prod": "Workshop/Plant",
    "production": "Workshop/Plant",
    "site": "Workshop/Plant",

    "bobst 1": "Bobst 1 Visionfold",
    "bobst1": "Bobst 1 Visionfold",
    "b1": "Bobst 1 Visionfold",

    "bobst 2": "Bobst 2 ExpertFold",
    "bobst2": "Bobst 2 ExpertFold",
    "b2": "Bobst 2 ExpertFold",
}
MACHINE_ALIAS = {normalize_text(k): v for k, v in MACHINE_ALIAS_RAW.items()}

AREA_ALIAS_RAW = {
    "cons": "Consumables",
    "consum": "Consumables",
    "consumable": "Consumables",
    "consumables": "Consumables",
    "strapper": "Strapper",

    "conveyor": "Conveyors",
    "conveyors": "Conveyors",

    "avanti": "Avanti Conveyors",
    "avanti conveyor": "Avanti Conveyors",
    "avanti conveyors": "Avanti Conveyors",

    "multiple misclics": "Multiple Misc Lics",
    "multiple misc lics": "Multiple Misc Lics",
    "misc lics": "Multiple Misc Lics",
}
AREA_ALIAS = {normalize_text(k): v for k, v in AREA_ALIAS_RAW.items()}

def normalize_machine_name(machine_raw: str) -> str:
    raw = (machine_raw or "").strip()
    return MACHINE_ALIAS.get(normalize_text(raw), raw)

def normalize_area_name(area_raw: str, machine_canon: str = "") -> str:
    raw = (area_raw or "").strip()
    return AREA_ALIAS.get(normalize_text(raw), raw)

def is_workshop_plant_electrical(machine_any: str, mech_elec_raw: str) -> bool:
    m = normalize_text(machine_any)
    me = normalize_text(mech_elec_raw)
    workshop_like = m in {"plant", "workshop", "workshopplant", "production", "prod", "site"}
    electrical_like = any(x in me for x in ["ele", "elec", "electrical"])
    return workshop_like and electrical_like

def is_plant_site_combination(machine_any: str, area_any: str, mech_elec_raw: str) -> bool:
    m = normalize_text(machine_any)
    a = normalize_text(area_any)
    me = normalize_text(mech_elec_raw)

    machine_keywords = {"plant", "plan", "site", "workshop", "workshopplant", "production", "prod"}
    area_keywords = {"plant", "plan", "site", "consumables", "cons", "consum", "consumable"}
    me_keywords = ["mech", "ele", "other", "pneu", "pneumatics"]

    return (m in machine_keywords) and (a in area_keywords) and any(w in me for w in me_keywords)

def _tolerant_match(a: str, b: str) -> bool:
    if not a or not b:
        return False
    return (a == b) or a.startswith(b) or b.startswith(a) or (a in b) or (b in a)

def find_cost_centre(ws_costs, machine_input: str, area_input: str):
    m = normalize_text(machine_input)
    a = normalize_text(area_input)

    for row in ws_costs.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        mach, area, cc = row[0], row[1], row[2]
        if not mach or not cc:
            continue

        mach_norm = normalize_text(mach)
        area_norm = normalize_text(area or "")

        if _tolerant_match(mach_norm, m) and _tolerant_match(area_norm, a):
            return str(cc)
    return None

CC_OVERRIDES = {
    (normalize_text("Bobst 1 Visionfold"), normalize_text("Strapper")): "327A",

    (normalize_text("Workshop/Plant"), normalize_text("Conveyors")): "355",
    (normalize_text("Workshop/Plant"), normalize_text("Avanti Conveyors")): "355",
    (normalize_text("Workshop/Plant"), normalize_text("Multiple Misc Lics")): "355",
}

def resolve_cost_centre(wb, machine: str, area: str, mech_elec: str) -> str:
    m_norm = normalize_text(machine)
    a_norm = normalize_text(area)

    if (m_norm, a_norm) in CC_OVERRIDES:
        return CC_OVERRIDES[(m_norm, a_norm)]

    if is_workshop_plant_electrical(machine, mech_elec):
        return "321B"
    if is_plant_site_combination(machine, area, mech_elec):
        return "321C"

    if SHEET_COST_CENTRE in wb.sheetnames:
        cc = find_cost_centre(wb[SHEET_COST_CENTRE], machine, area)
        return cc or "UNKNOWN"

    return "UNKNOWN"

def count_orders_today(ws, ordered_col: str, start_row: int) -> int:
    today = datetime.today().date()
    count = 0
    for r in range(start_row, ws.max_row + 1):
        val = ws[f"{ordered_col}{r}"].value
        if isinstance(val, datetime) and val.date() == today:
            count += 1
        elif isinstance(val, date) and val == today:
            count += 1
        elif isinstance(val, str):
            try:
                if datetime.strptime(val.strip(), "%d/%m/%Y").date() == today:
                    count += 1
            except ValueError:
                pass
    return count

def generate_po_number(cost_centre: str, initials: str, seq_base: int) -> str:
    now = datetime.now()
    seq = seq_base + 1
    return f"{cost_centre}-{initials}{now.day:02d}{now.month:02d}{seq}{now.year % 100:02d}"


# ========= SUPPLIERS =========
def load_suppliers_db(path: str):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"suppliers": []}

def pick_supplier(path: str, query: str):
    db = load_suppliers_db(path)
    suppliers = db.get("suppliers", [])

    q_raw = (query or "").strip()
    q = q_raw.lower()

    if q == "manual":
        name = input_text("Manual supplier name (or 'back'): ", allow_blank=False, allow_back=True)
        if name == BACK:
            return BACK
        return {"name": name}

    if q == "add":
        name = input_text("New supplier name (or 'back'): ", allow_blank=False, allow_back=True)
        if name == BACK:
            return BACK
        return {"name": name}

    if not q:
        return {"name": ""}

    def matches(s):
        name = (s.get("name") or "").lower()
        aliases = [str(a).lower() for a in (s.get("aliases") or [])]
        return (q in name) or any(q in a for a in aliases)

    found = [s for s in suppliers if matches(s)]
    if not found:
        names = [s.get("name", "") for s in suppliers if s.get("name")]
        suggestion = get_close_matches(q_raw, names, n=3, cutoff=0.65)
        if suggestion:
            print("No direct match. Did you mean:")
            for i, s in enumerate(suggestion, 1):
                print(f"  {i}) {s}")
            pick = input_text("Pick 1-3, Enter for manual, or type 'back': ", allow_blank=True, allow_back=True)
            if pick == BACK:
                return BACK
            if pick.isdigit() and 1 <= int(pick) <= len(suggestion):
                chosen = suggestion[int(pick) - 1]
                for sup in suppliers:
                    if sup.get("name") == chosen:
                        return sup

        manual = input_text("No matches. Manual supplier name (or 'back'): ", allow_blank=False, allow_back=True)
        if manual == BACK:
            return BACK
        return {"name": manual}

    for i, s in enumerate(found[:8], 1):
        print(f"  {i}) {s.get('name')} ({s.get('supplier_id','N/A')})")

    pick = input_text("Pick 1-8, Enter for first, or type 'back': ", allow_blank=True, allow_back=True)
    if pick == BACK:
        return BACK
    if pick.isdigit() and 1 <= int(pick) <= min(8, len(found)):
        return found[int(pick) - 1]
    return found[0]


# ========= TO RECEIVE (GR BACKLOG) =========
TO_RECEIVE_HEADERS = [
    "PO Number",
    "Supplier",
    "Machine",
    "Area",
    "Reason",
    "Mech/Elec/Other",
    "Job No",
    "Items",
    "Total Cost",
    "Date Ordered",
    "Due Date",
    "Cost Centre",
    "Status",
    "Received Date",
    "Invoice Number",
    "Notes",
]

def ensure_to_receive_sheet(wb):
    if TO_RECEIVE_SHEET in wb.sheetnames:
        ws = wb[TO_RECEIVE_SHEET]
    else:
        ws = wb.create_sheet(TO_RECEIVE_SHEET)

    if ws["A1"].value is None:
        for i, h in enumerate(TO_RECEIVE_HEADERS, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    return ws

def to_receive_col_index(header_name: str) -> int:
    return TO_RECEIVE_HEADERS.index(header_name) + 1

def find_to_receive_row_by_po(ws, po_number: str):
    target = (po_number or "").strip()
    if not target:
        return None
    po_col = to_receive_col_index("PO Number")
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=po_col).value or "").strip() == target:
            return r
    return None


# ========= PDF GENERATION =========
def try_register_dejavusans():
    font_paths = [
        r"C:\Windows\Fonts\DejaVuSans.ttf",
        r"C:\Windows\Fonts\dejavusans.ttf",
        r"C:\Windows\Fonts\Arial.ttf",
    ]
    for p in font_paths:
        if os.path.isfile(p):
            try:
                pdfmetrics.registerFont(TTFont("BodyFont", p))
                return "BodyFont"
            except Exception:
                pass
    return None

def generate_po_pdf(pdf_path: str, po_number: str, supplier_obj: dict, order: dict, items: list, totals: dict) -> bool:
    if not REPORTLAB_AVAILABLE:
        print("PDF generation skipped (ReportLab not available).")
        return False

    try:
        ensure_dir(os.path.dirname(pdf_path))

        font_name = try_register_dejavusans()
        c = canvas.Canvas(pdf_path, pagesize=A4)
        width, height = A4

        if font_name:
            c.setFont(font_name, 11)
        else:
            c.setFont("Helvetica", 11)

        def draw_text(x, y, text):
            c.drawString(x, y, str(text))

        y = height - 50
        draw_text(40, y, f"PURCHASE ORDER (PO): {po_number}")
        y -= 20
        draw_text(40, y, f"Date: {datetime.today().strftime('%d/%m/%Y')}")
        y -= 30

        supplier_name = supplier_obj.get("name", order.get("supplier", ""))
        supplier_addr = supplier_obj.get("address_lines") or []
        supplier_email = supplier_obj.get("email", "")
        supplier_phone = supplier_obj.get("phone", "")

        draw_text(40, y, "Supplier:")
        y -= 16
        draw_text(60, y, supplier_name)
        y -= 16
        for line in supplier_addr[:5]:
            draw_text(60, y, line)
            y -= 16
        if supplier_email:
            draw_text(60, y, f"Email: {supplier_email}")
            y -= 16
        if supplier_phone:
            draw_text(60, y, f"Phone: {supplier_phone}")
            y -= 16

        draw_text(300, height - 130, "Bill To:")
        yy = height - 146
        for line in SHEARD_BILL_TO:
            draw_text(320, yy, line)
            yy -= 16

        draw_text(300, height - 230, "Deliver To:")
        yy = height - 246
        for line in SHEARD_DELIVER_TO:
            draw_text(320, yy, line)
            yy -= 16

        draw_text(300, height - 330, "Contact:")
        draw_text(320, height - 346, f"{CONTACT_NAME} | {CONTACT_PHONE} | {CONTACT_EMAIL}")

        y = height - 380
        draw_text(40, y, f"Machine: {order.get('machine','')}")
        y -= 16
        draw_text(40, y, f"Area: {order.get('area','')}")
        y -= 16
        draw_text(40, y, f"Reason: {order.get('reason','')}")
        y -= 16
        draw_text(40, y, f"Mech/Elec/Other: {order.get('mech_elec','')}")
        y -= 16
        draw_text(40, y, f"Job No: {order.get('job_no','')}")
        y -= 16
        draw_text(40, y, f"Cost Centre: {order.get('cost_centre','')}")
        y -= 16
        draw_text(40, y, f"Due Date: {order.get('due_raw','')}")
        y -= 24

        draw_text(40, y, "Items:")
        y -= 16
        draw_text(40, y, "Qty")
        draw_text(80, y, "Description")
        draw_text(420, y, "Unit £")
        draw_text(500, y, "Line £")
        y -= 12
        c.line(40, y, 560, y)
        y -= 16

        for it in items:
            if y < 120:
                c.showPage()
                y = height - 60
                if font_name:
                    c.setFont(font_name, 11)
                else:
                    c.setFont("Helvetica", 11)

            draw_text(40, y, it["qty"])
            draw_text(80, y, (it["desc"] or "")[:55])
            draw_text(420, y, f"{it['unit_price']:.2f}")
            draw_text(500, y, f"{it['line_total']:.2f}")
            y -= 16

        y -= 10
        c.line(40, y, 560, y)
        y -= 18

        net_total = totals.get("net_total", 0.0)
        vat_rate = totals.get("vat_rate", VAT_RATE_DEFAULT)
        vat_amount = totals.get("vat_amount", 0.0)
        gross_total = totals.get("gross_total", net_total + vat_amount)

        draw_text(380, y, "Net Total £:")
        draw_text(520, y, f"{net_total:.2f}")
        y -= 16
        draw_text(380, y, f"VAT ({vat_rate*100:.0f}%) £:")
        draw_text(520, y, f"{vat_amount:.2f}")
        y -= 16
        draw_text(380, y, "Gross Total £:")
        draw_text(520, y, f"{gross_total:.2f}")

        c.save()
        return os.path.isfile(pdf_path) and os.path.getsize(pdf_path) > 0

    except Exception as e:
        print(f"PDF generation failed: {e}")
        return False


# ========= INPUT FLOW =========
def prompt_items():
    print("\nEnter line items. Leave Quantity blank to finish. Type 'back' to go back.")
    items = []

    while True:
        qty_raw = input_positive_int("Quantity (blank to finish): ", allow_blank=True, allow_back=True)

        if qty_raw == BACK:
            if items:
                removed = items.pop()
                print(f"Removed last item: x{removed['qty']} {removed['desc']}")
                continue
            return BACK

        if qty_raw == "":
            break

        qty = int(qty_raw)

        desc = input_text("Item description: ", allow_blank=False, allow_back=True)
        if desc == BACK:
            continue

        unit_raw = input_money("Unit price (£) (blank if you only know line total): ", allow_blank=True, allow_back=True)
        if unit_raw == BACK:
            continue

        if unit_raw == "":
            line_raw = input_money("Line total (£): ", allow_blank=False, allow_back=True)
            if line_raw == BACK:
                continue
            line_total = float(parse_currency_to_float(line_raw) or 0.0)
            unit = (line_total / qty) if qty else 0.0
        else:
            unit = float(parse_currency_to_float(unit_raw) or 0.0)
            line_total = qty * unit

        items.append({"desc": desc, "qty": qty, "unit_price": float(unit), "line_total": float(line_total)})

    if not items:
        items.append({"desc": "N/A", "qty": 1, "unit_price": 0.0, "line_total": 0.0})

    net_total = sum(i["line_total"] for i in items)

    while True:
        if PROMPT_VAT_RATE_EACH_PO:
            raw = input_text(f"VAT rate % (Enter for {int(VAT_RATE_DEFAULT*100)}): ", allow_blank=True, allow_back=True)
            if raw == BACK:
                print("Type VAT rate or press Enter for default.")
                continue
            if raw == "":
                vat_rate = VAT_RATE_DEFAULT
                break
            raw = raw.replace("%", "").strip()
            try:
                vat_rate = float(raw) / 100.0
                if vat_rate < 0:
                    print("VAT cannot be negative.")
                    continue
                break
            except ValueError:
                print("Invalid VAT. Try again.")
        else:
            vat_rate = VAT_RATE_DEFAULT
            break

    vat_amount = net_total * vat_rate
    gross_total = net_total + vat_amount

    totals = {
        "net_total": float(net_total),
        "vat_rate": float(vat_rate),
        "vat_amount": float(vat_amount),
        "gross_total": float(gross_total),
    }
    return items, totals

def collect_po_data():
    print("\n=== CREATE NEW PURCHASE ORDER (PO) ===")
    order = {}
    supplier_obj = {"name": ""}

    steps = [
        ("machine", lambda: input_text("Machine: ", allow_blank=False, allow_back=True)),
        ("area", lambda: input_text("Area: ", allow_blank=False, allow_back=True)),
        ("reason", lambda: input_text("Reason: ", allow_blank=False, allow_back=True)),
        ("mech_elec", lambda: input_text("Mech/Elec/Other: ", allow_blank=False, allow_back=True)),
        ("supplier", None),
        ("items", None),
        ("job_no", lambda: input_text("Job No (optional): ", allow_blank=True, allow_back=True)),
        ("due_raw", lambda: input_text("Due date (dd/mm/yyyy) or blank: ", allow_blank=True, allow_back=True)),
        ("savings_raw", lambda: input_text("Cost initiatives / Savings £ (optional): ", allow_blank=True, allow_back=True)),
        ("orig_supplier", lambda: input_text("Original supplier (optional): ", allow_blank=True, allow_back=True)),
    ]

    i = 0
    items = None
    totals = None

    while i < len(steps):
        key, fn = steps[i]

        if key == "supplier":
            q = input_text("Supplier search (type 'manual' or 'add'): ", allow_blank=False, allow_back=True)
            if q == BACK:
                i = max(0, i - 1)
                continue

            sup = pick_supplier(SUPPLIERS_JSON_PATH, q)
            if sup == BACK:
                continue

            supplier_obj = sup
            order["supplier"] = (supplier_obj.get("name", "") or "").strip()
            i += 1
            continue

        if key == "items":
            res = prompt_items()
            if res == BACK:
                i = max(0, i - 1)
                continue

            items, totals = res
            items_summary = "; ".join([f"x{it['qty']} {it['desc']}" for it in items])[:250]
            order["items_summary"] = items_summary
            i += 1
            continue

        val = fn()
        if val == BACK:
            i = max(0, i - 1)
            continue

        if key == "due_raw" and val:
            if parse_ddmmyyyy_to_date(val) is None:
                print("Invalid date. Use dd/mm/yyyy or leave blank. Type 'back' to return.")
                continue

        order[key] = val
        i += 1

    return order, supplier_obj, items, totals


# ========= WRITE FUNCTIONS =========
def write_parts_order_row(ws_parts, header_row, colmap, order: dict, po_number: str, cost_centre: str, net_total: float):
    data_start = header_row + 1
    row = find_first_empty_row(ws_parts, colmap["Machine"], data_start)

    ws_parts[f"{colmap['DateOrdered']}{row}"] = datetime.today().date()
    ws_parts[f"{colmap['DateOrdered']}{row}"].number_format = "DD/MM/YYYY"

    ws_parts[f"{colmap['Machine']}{row}"] = order["machine"]
    ws_parts[f"{colmap['Area']}{row}"] = order["area"]
    ws_parts[f"{colmap['Reason']}{row}"] = order["reason"]
    ws_parts[f"{colmap['MechElecOther']}{row}"] = order["mech_elec"]
    ws_parts[f"{colmap['Supplier']}{row}"] = order["supplier"]
    ws_parts[f"{colmap['PartsOrdered']}{row}"] = order["items_summary"]

    parts_cell = ws_parts[f"{colmap['PartsOrdered']}{row}"]
    parts_cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_parts[f"{colmap['JobNo']}{row}"] = order.get("job_no", "")
    ws_parts[f"{colmap['PONumber']}{row}"] = po_number
    ws_parts[f"{colmap['CostCentre']}{row}"] = cost_centre

    ws_parts[f"{colmap['Cost']}{row}"] = float(net_total)
    ws_parts[f"{colmap['Cost']}{row}"].number_format = "£#,##0.00"

    savings_val = parse_currency_to_float(order.get("savings_raw", ""))
    if savings_val is not None and "Savings" in colmap:
        ws_parts[f"{colmap['Savings']}{row}"] = float(savings_val)
        ws_parts[f"{colmap['Savings']}{row}"].number_format = "£#,##0.00"

    if order.get("orig_supplier") and "OriginalSupplier" in colmap:
        ws_parts[f"{colmap['OriginalSupplier']}{row}"] = order["orig_supplier"]

    due_dt = parse_ddmmyyyy_to_date(order.get("due_raw", ""))
    if due_dt and "DueDate" in colmap:
        ws_parts[f"{colmap['DueDate']}{row}"] = due_dt
        ws_parts[f"{colmap['DueDate']}{row}"].number_format = "DD/MM/YYYY"
    elif order.get("due_raw") and "DueDate" in colmap:
        ws_parts[f"{colmap['DueDate']}{row}"] = order["due_raw"]

    return row

def add_to_receive_row(ws_recv, po_number: str, order: dict, cost_centre: str, net_total: float, totals: dict):
    row = ws_recv.max_row + 1

    def set_cell(header, value):
        ws_recv.cell(row=row, column=to_receive_col_index(header), value=value)

    set_cell("PO Number", po_number)
    set_cell("Supplier", order["supplier"])
    set_cell("Machine", order["machine"])
    set_cell("Area", order["area"])
    set_cell("Reason", order["reason"])
    set_cell("Mech/Elec/Other", order["mech_elec"])
    set_cell("Job No", order.get("job_no", ""))
    set_cell("Items", order["items_summary"])
    set_cell("Total Cost", float(net_total))

    set_cell("Date Ordered", datetime.today().date())
    set_cell("Due Date", parse_ddmmyyyy_to_date(order.get("due_raw", "")) or order.get("due_raw", ""))

    set_cell("Cost Centre", cost_centre)
    set_cell("Status", "Ordered")
    set_cell("Received Date", "")
    set_cell("Invoice Number", "")
    set_cell("Notes", f"VAT {totals.get('vat_rate', VAT_RATE_DEFAULT)*100:.0f}% | Gross £{totals.get('gross_total', 0.0):.2f}")

    ws_recv.cell(row=row, column=to_receive_col_index("Total Cost")).number_format = "£#,##0.00"
    ws_recv.cell(row=row, column=to_receive_col_index("Date Ordered")).number_format = "DD/MM/YYYY"

    return row

def mark_received(wb, po_number: str, invoice_number: str = ""):
    if TO_RECEIVE_SHEET not in wb.sheetnames:
        print("To Receive sheet not found.")
        return

    ws_recv = wb[TO_RECEIVE_SHEET]
    recv_row = find_to_receive_row_by_po(ws_recv, po_number)
    if not recv_row:
        print("PO number not found in To Receive.")
        return

    ws_recv.cell(row=recv_row, column=to_receive_col_index("Status"), value="Received")
    ws_recv.cell(row=recv_row, column=to_receive_col_index("Received Date"), value=datetime.today().date())
    ws_recv.cell(row=recv_row, column=to_receive_col_index("Received Date")).number_format = "DD/MM/YYYY"
    if invoice_number:
        ws_recv.cell(row=recv_row, column=to_receive_col_index("Invoice Number"), value=invoice_number)

    for sheet_name in wb.sheetnames:
        if sheet_name in {TO_RECEIVE_SHEET, SHEET_COST_CENTRE}:
            continue
        ws = wb[sheet_name]
        header_row, colmap = detect_header_row_and_columns(ws)
        if not colmap or "PONumber" not in colmap or "DeliveredDate" not in colmap:
            continue

        start_row = (header_row or 2) + 1
        for r in range(start_row, ws.max_row + 1):
            val = str(ws[f"{colmap['PONumber']}{r}"].value or "").strip()
            if val == po_number:
                ws[f"{colmap['DeliveredDate']}{r}"] = datetime.today().date()
                ws[f"{colmap['DeliveredDate']}{r}"].number_format = "DD/MM/YYYY"
                return

    print("Received status updated. Delivered date not updated (PO not found in monthly sheets).")


# ========= MAIN MENU =========
def create_new_po():
    wb = load_or_create_workbook(WORKBOOK_PATH)
    ws_parts = get_or_create_month_sheet(wb)
    ws_recv = ensure_to_receive_sheet(wb)

    header_row, colmap = ensure_parts_column_map(ws_parts)
    data_start = header_row + 1

    if DEBUG:
        print("\nDetected Parts Ordered columns:")
        for k in sorted(colmap.keys()):
            print(f"  {k:14} -> {colmap[k]}")
        print(f"Header row: {header_row}, data starts: {data_start}")

    order, supplier_obj, items, totals = collect_po_data()

    order["machine"] = normalize_machine_name(order["machine"])
    order["area"] = normalize_area_name(order["area"], order["machine"])

    cost_centre = resolve_cost_centre(wb, order["machine"], order["area"], order["mech_elec"])

    if DEBUG:
        print(f"\nMachine (canonical): {order['machine']}")
        print(f"Area (canonical): {order['area']}")
        print(f"Cost Centre result: {cost_centre}")

    seq_base = count_orders_today(ws_parts, colmap["DateOrdered"], data_start)
    po_number = generate_po_number(cost_centre, INITIALS, seq_base)

    order["cost_centre"] = cost_centre

    net_total = totals["net_total"]
    parts_row = write_parts_order_row(ws_parts, header_row, colmap, order, po_number, cost_centre, net_total)
    recv_row = add_to_receive_row(ws_recv, po_number, order, cost_centre, net_total, totals)

    # Update month totals + make it "fit" nicely
    update_month_totals(ws_parts, header_row, colmap)
    auto_fit_columns(ws_parts)
    auto_fit_columns(ws_recv)
    set_sheet_view(ws_parts, zoom=90)
    set_sheet_view(ws_recv, zoom=90)

    ensure_dir(PO_PDF_OUTPUT_DIR)
    pdf_path = os.path.join(PO_PDF_OUTPUT_DIR, f"{po_number}.pdf")
    pdf_ok = generate_po_pdf(pdf_path, po_number, supplier_obj, order, items, totals)

    safe_save_workbook(wb, WORKBOOK_PATH)

    print("\nSUCCESS")
    print(f"Purchase Order (PO): {po_number}")
    print(f"Written to sheet '{ws_parts.title}' row {parts_row}")
    print(f"Added to '{TO_RECEIVE_SHEET}' row {recv_row}")

    if REPORTLAB_AVAILABLE:
        if pdf_ok:
            print(f"PDF saved: {pdf_path}")
        else:
            print("PDF was attempted but not created. Check folder permissions/path and ReportLab install.")
    else:
        print("PDF not created (ReportLab not available).")

def receive_po():
    wb = load_or_create_workbook(WORKBOOK_PATH)
    po_number = input_text("Enter PO number to mark as Received (or 'back'): ", allow_blank=False, allow_back=True)
    if po_number == BACK:
        return
    invoice_number = input_text("Invoice number (optional, or 'back'): ", allow_blank=True, allow_back=True)
    if invoice_number == BACK:
        invoice_number = ""
    mark_received(wb, po_number, invoice_number)
    safe_save_workbook(wb, WORKBOOK_PATH)
    print("Saved.")

def main():
    ensure_dir(PO_PDF_OUTPUT_DIR)

    while True:
        print("\n=== MENU ===")
        print("1) Create new Purchase Order (PO)")
        print("2) Mark Purchase Order (PO) as Received (Goods Receipt (GR))")
        print("3) Exit")
        choice = input("Choose 1-3: ").strip()

        if choice == "1":
            create_new_po()
        elif choice == "2":
            receive_po()
        elif choice == "3":
            break
        else:
            print("Invalid choice.")

if __name__ == "__main__":
    main()
